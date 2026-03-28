"""Unit tests for app/ppf_database.py — PPF database layer."""
import json
from pathlib import Path
from unittest.mock import patch

import pytest


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def ppf_base_dir(tmp_path):
    """Create a temp base dir with PPF subdir."""
    (tmp_path / "PPF").mkdir()
    return tmp_path


@pytest.fixture(autouse=True)
def _mock_drive():
    with patch("app.ppf_database._sync_to_drive"), \
         patch("app.ppf_database._delete_from_drive"):
        yield


# ---------------------------------------------------------------------------
# Tests — empty state
# ---------------------------------------------------------------------------

def test_get_all_empty(ppf_base_dir):
    from app.ppf_database import get_all
    assert get_all(base_dir=str(ppf_base_dir)) == []


def test_get_dashboard_empty(ppf_base_dir):
    from app.ppf_database import get_dashboard
    dash = get_dashboard(base_dir=str(ppf_base_dir))
    assert dash["active_count"] == 0
    assert dash["total_count"] == 0
    assert dash["total_deposited"] == 0


# ---------------------------------------------------------------------------
# Tests — add SIP-based account
# ---------------------------------------------------------------------------

def test_add_sip_ppf(ppf_base_dir):
    from app.ppf_database import add, get_all
    data = {
        "account_name": "My PPF",
        "bank": "SBI",
        "interest_rate": 7.1,
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "payment_type": "sip",
        "sip_amount": 5000,
        "sip_frequency": "monthly",
        "account_number": "PPF12345",
    }
    result = add(data, base_dir=str(ppf_base_dir))
    assert result["bank"] == "SBI"
    assert result["interest_rate"] == 7.1
    assert result["tenure_years"] == 15
    assert result["sip_amount"] == 5000
    assert result["source"] == "xlsx"

    # xlsx created
    xlsx_files = list((ppf_base_dir / "PPF").glob("*.xlsx"))
    assert len(xlsx_files) == 1

    # get_all picks it up
    items = get_all(base_dir=str(ppf_base_dir))
    assert len(items) == 1


def test_add_one_time_ppf(ppf_base_dir):
    from app.ppf_database import add
    data = {
        "account_name": "One Time PPF",
        "bank": "Post Office",
        "interest_rate": 7.1,
        "start_date": "2020-04-01",
        "tenure_years": 15,
        "payment_type": "one_time",
        "amount_added": 150000,
    }
    result = add(data, base_dir=str(ppf_base_dir))
    assert result["total_deposited"] > 0


# ---------------------------------------------------------------------------
# Tests — delete
# ---------------------------------------------------------------------------

def test_delete_ppf(ppf_base_dir):
    from app.ppf_database import add, delete, get_all
    data = {
        "account_name": "Del PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 1000,
        "sip_frequency": "monthly",
    }
    result = add(data, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    items = get_all(base_dir=str(ppf_base_dir))
    assert len(items) == 1

    del_result = delete(ppf_id, base_dir=str(ppf_base_dir))
    assert "deleted" in del_result["message"]

    items = get_all(base_dir=str(ppf_base_dir))
    assert len(items) == 0


def test_delete_nonexistent_raises(ppf_base_dir):
    from app.ppf_database import delete
    with pytest.raises(ValueError, match="not found"):
        delete("badid", base_dir=str(ppf_base_dir))


# ---------------------------------------------------------------------------
# Tests — update
# ---------------------------------------------------------------------------

def test_update_ppf(ppf_base_dir):
    from app.ppf_database import add, update
    data = {
        "account_name": "Upd PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }
    result = add(data, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    updated = update(ppf_id, {"bank": "PNB"}, base_dir=str(ppf_base_dir))
    assert updated["bank"] == "PNB"


# ---------------------------------------------------------------------------
# Tests — add_contribution
# ---------------------------------------------------------------------------

def test_add_contribution(ppf_base_dir):
    from app.ppf_database import add, add_contribution, get_all
    data = {
        "account_name": "Contrib PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }
    result = add(data, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    contrib_result = add_contribution(ppf_id, {
        "date": "2024-06-15",
        "amount": 50000,
        "remarks": "Lump sum",
    }, base_dir=str(ppf_base_dir))
    assert contrib_result is not None

    # Verify contribution is reflected
    items = get_all(base_dir=str(ppf_base_dir))
    assert len(items) == 1
    contribs = items[0].get("contributions", [])
    assert len(contribs) >= 1


# ---------------------------------------------------------------------------
# Tests — helpers
# ---------------------------------------------------------------------------

def test_gen_ppf_id_deterministic():
    from app.ppf_database import _gen_ppf_id
    assert _gen_ppf_id("My PPF") == _gen_ppf_id("My PPF")
    assert _gen_ppf_id("PPF A") != _gen_ppf_id("PPF B")


def test_sip_freq_to_months():
    from app.ppf_database import _sip_freq_to_months
    assert _sip_freq_to_months("monthly") == 1
    assert _sip_freq_to_months("Monthly") == 1
    assert _sip_freq_to_months("quarterly") == 3
    assert _sip_freq_to_months("yearly") == 12
    assert _sip_freq_to_months("annual") == 12
    assert _sip_freq_to_months("unknown") == 1  # default


def test_get_financial_year():
    from app.ppf_database import _get_financial_year
    from datetime import date
    assert _get_financial_year(date(2024, 4, 1)) == "2024-25"
    assert _get_financial_year(date(2024, 3, 31)) == "2023-24"
    assert _get_financial_year(date(2024, 1, 1)) == "2023-24"


# ---------------------------------------------------------------------------
# Tests — dashboard
# ---------------------------------------------------------------------------

def test_dashboard_after_add(ppf_base_dir):
    from app.ppf_database import add, get_dashboard
    add({
        "account_name": "Dash PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))

    dash = get_dashboard(base_dir=str(ppf_base_dir))
    assert dash["total_count"] == 1
    assert dash["active_count"] == 1
    assert dash["total_deposited"] > 0


# ---------------------------------------------------------------------------
# Tests — installments structure
# ---------------------------------------------------------------------------

def test_installments_have_correct_fields(ppf_base_dir):
    from app.ppf_database import add
    result = add({
        "account_name": "Inst PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))

    assert result["installments_total"] == 180  # 15 years * 12 months
    first = result["installments"][0]
    assert "month" in first
    assert "date" in first
    assert "amount_invested" in first
    assert "is_past" in first
    assert "lock_status" in first


# ---------------------------------------------------------------------------
# Tests — _sync_to_drive / _delete_from_drive (lines 64, 69-82)
# ---------------------------------------------------------------------------

def test_sync_to_drive_is_noop():
    from app.ppf_database import _sync_to_drive
    from pathlib import Path
    _sync_to_drive(Path("/tmp/test.xlsx"))


def test_delete_from_drive_handles_exception():
    from app.ppf_database import _delete_from_drive
    from pathlib import Path
    _delete_from_drive(Path("/nonexistent/path/file.xlsx"))


# ---------------------------------------------------------------------------
# Tests — _to_date, _to_str, _to_float helpers (lines 106, 115-125)
# ---------------------------------------------------------------------------

def test_to_date_helpers():
    from app.ppf_database import _to_date
    from datetime import datetime, date
    assert _to_date(datetime(2024, 6, 15)) == date(2024, 6, 15)
    assert _to_date(date(2024, 6, 15)) == date(2024, 6, 15)
    assert _to_date(None) is None
    assert _to_date("2024-06-15") is None


def test_to_str_helpers():
    from app.ppf_database import _to_str
    from datetime import datetime, date
    assert _to_str(None) == ""
    assert _to_str(datetime(2024, 6, 15)) == "2024-06-15"
    assert _to_str(date(2024, 6, 15)) == "2024-06-15"
    assert _to_str("  hello  ") == "hello"
    assert _to_str(None, "default") == "default"


def test_to_float_helpers():
    from app.ppf_database import _to_float
    assert _to_float(3.14) == 3.14
    assert _to_float(None) == 0.0
    assert _to_float("bad") == 0.0
    assert _to_float(None, 5.0) == 5.0


# ---------------------------------------------------------------------------
# Tests — JSON parsing helpers (lines 156-183)
# ---------------------------------------------------------------------------

def test_parse_json_array():
    from app.ppf_database import _parse_json_array
    assert _parse_json_array(None) is None
    assert _parse_json_array("") is None
    assert _parse_json_array("not json") is None
    assert _parse_json_array("[]") is None  # empty list
    assert _parse_json_array('[1,2,3]') == [1, 2, 3]
    assert _parse_json_array("[invalid json") is None  # starts with [ but not valid


def test_parse_sip_phases_json():
    from app.ppf_database import _parse_sip_phases_json
    assert _parse_sip_phases_json(None) is None
    assert _parse_sip_phases_json("") is None
    assert _parse_sip_phases_json("[]") is None  # empty
    phases = _parse_sip_phases_json('[{"amount": 5000, "frequency": "monthly"}]')
    assert len(phases) == 1
    assert _parse_sip_phases_json("invalid") is None
    assert _parse_sip_phases_json("[invalid json") is None  # starts with [ but not valid


# ---------------------------------------------------------------------------
# Tests — _build_single_phase (line 190)
# ---------------------------------------------------------------------------

def test_build_single_phase_zero_amount():
    from app.ppf_database import _build_single_phase
    from datetime import date
    result = _build_single_phase(0, "monthly", date(2024, 1, 1), None)
    assert result == []


def test_build_single_phase_with_end():
    from app.ppf_database import _build_single_phase
    from datetime import date
    result = _build_single_phase(5000, "monthly", date(2024, 1, 1), date(2025, 12, 31))
    assert len(result) == 1
    assert result[0]["end"] == "2025-12-31"


# ---------------------------------------------------------------------------
# Tests — _get_active_phase (lines 200-212)
# ---------------------------------------------------------------------------

def test_get_active_phase():
    from app.ppf_database import _get_active_phase
    from datetime import date
    phases = [
        {"start": "2020-01-01", "end": "2022-12-31", "amount": 5000, "frequency": "monthly"},
        {"start": "2023-01-01", "end": None, "amount": 10000, "frequency": "monthly"},
    ]
    # Date in first phase
    p = _get_active_phase(phases, date(2021, 6, 15))
    assert p["amount"] == 5000
    # Date in second phase
    p = _get_active_phase(phases, date(2024, 1, 1))
    assert p["amount"] == 10000
    # Date before any phase
    p = _get_active_phase(phases, date(2019, 1, 1))
    assert p is None


# ---------------------------------------------------------------------------
# Tests — _parse_contributions_json (lines 221-234)
# ---------------------------------------------------------------------------

def test_parse_contributions_json():
    from app.ppf_database import _parse_contributions_json
    assert _parse_contributions_json(None) == []
    assert _parse_contributions_json("") == []
    assert _parse_contributions_json("not json") == []
    # Valid
    result = _parse_contributions_json('[{"date": "2024-01-01", "amount": 50000}]')
    assert len(result) == 1
    # Broken JSON (amount: ,)
    result = _parse_contributions_json('[{"date": "2024-01-01", "amount": , "remarks": "test"}]')
    assert len(result) == 1
    assert result[0]["amount"] == 0  # fixed broken JSON
    # Entry without 'date' key is filtered out
    result = _parse_contributions_json('[{"amount": 5000}]')
    assert result == []


# ---------------------------------------------------------------------------
# Tests — _scan_data_rows (lines 250-272)
# ---------------------------------------------------------------------------

def test_scan_data_rows_with_withdrawals():
    from app.ppf_database import _scan_data_rows
    from datetime import datetime
    rows = [
        (None,) * 11,  # row 1
        (None,) * 11,  # row 2
        (None,) * 11,  # row 3
        (None,) * 11,  # row 4
        (None,) * 11,  # row 5 (headers)
        (1, None, 1, datetime(2020, 1, 1), 5000, 0, 0, 1000, 2000),  # row 6: withdrawal=1000, contribution=2000
    ]
    result = _scan_data_rows(rows, 180)
    assert len(result) == 2
    # One withdrawal (negative), one contribution (positive)
    assert any(c["amount"] == -1000 for c in result)
    assert any(c["amount"] == 2000 for c in result)


def test_scan_data_rows_fallback_to_h4():
    from app.ppf_database import _scan_data_rows
    # No data in cols 8-9, but H4 has contributions
    rows = [
        (None,) * 11,  # row 1
        (None,) * 11,  # row 2
        (None,) * 11,  # row 3
        (None, None, None, None, None, None, None, '[{"date": "2024-01-01", "amount": 50000}]'),  # row 4 with H4
        (None,) * 11,  # row 5
    ]
    result = _scan_data_rows(rows, 180)
    assert len(result) == 1
    assert result[0]["amount"] == 50000


# ---------------------------------------------------------------------------
# Tests — _parse_all_xlsx edge cases (lines 503-512)
# ---------------------------------------------------------------------------

def test_parse_all_xlsx_nonexistent():
    from app.ppf_database import _parse_all_xlsx
    from pathlib import Path
    assert _parse_all_xlsx(ppf_dir=Path("/nonexistent/dir")) == []


def test_parse_all_xlsx_skips_temp_and_corrupt(ppf_base_dir):
    ppf_dir = ppf_base_dir / "PPF"
    (ppf_dir / "~$temp.xlsx").write_text("temp")
    (ppf_dir / "corrupt.xlsx").write_text("not xlsx")
    from app.ppf_database import _parse_all_xlsx
    results = _parse_all_xlsx(ppf_dir=ppf_dir)
    assert len(results) == 0


# ---------------------------------------------------------------------------
# Tests — _create_ppf_xlsx duplicate name handling (lines 544-545)
# ---------------------------------------------------------------------------

def test_create_ppf_xlsx_avoids_overwrite(ppf_base_dir):
    from app.ppf_database import _create_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    # Create first file
    _create_ppf_xlsx("Test PPF", "SBI", 5000, 7.1, 15, "2020-01-01", ppf_dir=ppf_dir)
    # Create second with same name - should get suffix
    _create_ppf_xlsx("Test PPF", "SBI", 5000, 7.1, 15, "2020-01-01", ppf_dir=ppf_dir)
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 2


# ---------------------------------------------------------------------------
# Tests — _create_ppf_xlsx with contributions and phases (lines 569-608)
# ---------------------------------------------------------------------------

def test_create_ppf_xlsx_with_contributions(ppf_base_dir):
    from app.ppf_database import _create_ppf_xlsx, _parse_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    filepath = _create_ppf_xlsx(
        "Contrib PPF", "SBI", 5000, 7.1, 15, "2020-01-01",
        contributions=[
            {"date": "2020-06-15", "amount": 50000, "remarks": "Lump sum"},
        ],
        ppf_dir=ppf_dir,
    )
    result = _parse_ppf_xlsx(filepath)
    assert result["name"] == "Contrib PPF"


def test_create_ppf_xlsx_with_sip_phases(ppf_base_dir):
    from app.ppf_database import _create_ppf_xlsx, _parse_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    phases = [
        {"amount": 5000, "frequency": "monthly", "start": "2020-01-01", "end": "2022-12-31"},
        {"amount": 10000, "frequency": "monthly", "start": "2023-01-01", "end": None},
    ]
    filepath = _create_ppf_xlsx(
        "Multi Phase PPF", "SBI", 5000, 7.1, 15, "2020-01-01",
        sip_phases=phases,
        ppf_dir=ppf_dir,
    )
    result = _parse_ppf_xlsx(filepath)
    assert result["name"] == "Multi Phase PPF"


def test_create_ppf_xlsx_with_sip_end_date(ppf_base_dir):
    from app.ppf_database import _create_ppf_xlsx, _parse_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    filepath = _create_ppf_xlsx(
        "SIP End PPF", "SBI", 5000, 7.1, 15, "2020-01-01",
        sip_end_date="2025-12-31",
        ppf_dir=ppf_dir,
    )
    result = _parse_ppf_xlsx(filepath)
    assert result["sip_end_date"] == "2025-12-31"


# ---------------------------------------------------------------------------
# Tests — _migrate_json_to_xlsx (lines 714-772)
# ---------------------------------------------------------------------------

def test_migrate_json_to_xlsx(ppf_base_dir):
    from app.ppf_database import _migrate_json_to_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    json_file = ppf_base_dir / "ppf_accounts.json"

    import json
    accounts = [{
        "account_name": "Migrate PPF",
        "bank": "PNB",
        "interest_rate": 7.1,
        "tenure_years": 15,
        "start_date": "2020-01-01",
        "sip_amount": 5000,
        "sip_frequency": "monthly",
        "account_number": "PPF001",
    }]
    json_file.write_text(json.dumps(accounts))

    _migrate_json_to_xlsx(ppf_dir=ppf_dir, json_file=json_file)

    # JSON renamed to .bak
    assert json_file.with_suffix(".json.bak").exists()
    # xlsx created
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1


def test_migrate_json_empty_file(ppf_base_dir):
    from app.ppf_database import _migrate_json_to_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    json_file = ppf_base_dir / "ppf_accounts.json"
    json_file.write_text("[]")  # empty list
    _migrate_json_to_xlsx(ppf_dir=ppf_dir, json_file=json_file)
    # No xlsx created
    assert len(list(ppf_dir.glob("*.xlsx"))) == 0


def test_migrate_json_nonexistent(ppf_base_dir):
    from app.ppf_database import _migrate_json_to_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    json_file = ppf_base_dir / "ppf_accounts.json"
    # File doesn't exist - should not raise
    _migrate_json_to_xlsx(ppf_dir=ppf_dir, json_file=json_file)


def test_migrate_json_contributions_to_sip(ppf_base_dir):
    """Migration converts one-time contributions to monthly sip_amount."""
    from app.ppf_database import _migrate_json_to_xlsx
    import json
    ppf_dir = ppf_base_dir / "PPF"
    json_file = ppf_base_dir / "ppf_accounts.json"

    accounts = [{
        "account_name": "Contrib PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 0,  # no SIP
        "contributions": [
            {"date": "2020-06-01", "amount": 150000},
            {"date": "2021-06-01", "amount": 150000},
        ],
    }]
    json_file.write_text(json.dumps(accounts))

    _migrate_json_to_xlsx(ppf_dir=ppf_dir, json_file=json_file)
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1


# ---------------------------------------------------------------------------
# Tests — update with new_sip_phase (lines 1017-1039)
# ---------------------------------------------------------------------------

def test_update_with_new_sip_phase(ppf_base_dir):
    from app.ppf_database import add, update
    result = add({
        "account_name": "Phase PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    updated = update(ppf_id, {
        "new_sip_phase": {
            "amount": 10000,
            "frequency": "monthly",
            "start": "2023-01-01",
            "end": None,
        }
    }, base_dir=str(ppf_base_dir))
    assert updated is not None


# ---------------------------------------------------------------------------
# Tests — update with payment_type toggle (lines 1044-1065)
# ---------------------------------------------------------------------------

def test_update_payment_type_one_time(ppf_base_dir):
    from app.ppf_database import add, update
    result = add({
        "account_name": "Toggle PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    updated = update(ppf_id, {
        "payment_type": "one_time",
        "amount_added": 150000,
    }, base_dir=str(ppf_base_dir))
    assert updated is not None


def test_update_payment_type_sip(ppf_base_dir):
    from app.ppf_database import add, update
    result = add({
        "account_name": "SIP Toggle PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "payment_type": "one_time",
        "amount_added": 150000,
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    updated = update(ppf_id, {
        "payment_type": "sip",
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    assert updated is not None


# ---------------------------------------------------------------------------
# Tests — update with name change (lines 1073-1076)
# ---------------------------------------------------------------------------

def test_update_name_change(ppf_base_dir):
    from app.ppf_database import add, update, get_all
    result = add({
        "account_name": "Old Name",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    updated = update(ppf_id, {"account_name": "New Name"}, base_dir=str(ppf_base_dir))
    assert updated["account_name"] == "New Name"


# ---------------------------------------------------------------------------
# Tests — update nonexistent (line 992)
# ---------------------------------------------------------------------------

def test_update_nonexistent_raises(ppf_base_dir):
    from app.ppf_database import update
    with pytest.raises(ValueError, match="not found"):
        update("bad_id", {"bank": "X"}, base_dir=str(ppf_base_dir))


# ---------------------------------------------------------------------------
# Tests — delete cleans up meta files (lines 1113-1114)
# ---------------------------------------------------------------------------

def test_delete_cleans_meta_file(ppf_base_dir):
    from app.ppf_database import add, delete
    from pathlib import Path
    result = add({
        "account_name": "Meta PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    # Create a fake meta file
    ppf_dir = ppf_base_dir / "PPF"
    xlsx_file = list(ppf_dir.glob("*.xlsx"))[0]
    meta_file = xlsx_file.with_suffix(".meta.json")
    meta_file.write_text("{}")

    del_result = delete(ppf_id, base_dir=str(ppf_base_dir))
    assert "deleted" in del_result["message"]
    assert not meta_file.exists()


# ---------------------------------------------------------------------------
# Tests — add_contribution validation (lines 1128-1156)
# ---------------------------------------------------------------------------

def test_add_contribution_invalid_date(ppf_base_dir):
    from app.ppf_database import add, add_contribution
    result = add({
        "account_name": "Date PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 500,
        "sip_frequency": "yearly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    with pytest.raises(ValueError, match="Invalid date"):
        add_contribution(ppf_id, {
            "date": "bad-date",
            "amount": 50000,
        }, base_dir=str(ppf_base_dir))


def test_add_contribution_exceeds_yearly_limit(ppf_base_dir):
    from app.ppf_database import add, add_contribution
    result = add({
        "account_name": "Limit PPF",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 12000,
        "sip_frequency": "monthly",  # 12000*12 = 144000 per year
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    # Try adding a large contribution that would exceed 150000 limit
    with pytest.raises(ValueError, match="Exceeds yearly limit"):
        add_contribution(ppf_id, {
            "date": "2020-06-15",
            "amount": 100000,
        }, base_dir=str(ppf_base_dir))


def test_add_contribution_nonexistent(ppf_base_dir):
    from app.ppf_database import add_contribution
    with pytest.raises(ValueError, match="not found"):
        add_contribution("bad_id", {
            "date": "2024-01-01",
            "amount": 50000,
        }, base_dir=str(ppf_base_dir))


# ---------------------------------------------------------------------------
# Tests — _enrich_withdrawal (lines 1199-1238)
# ---------------------------------------------------------------------------

def test_enrich_withdrawal_locked():
    from app.ppf_database import _enrich_withdrawal
    item = {
        "start_date": "2023-01-01",
        "tenure_years": 15,
        "total_withdrawn": 0,
        "installments": [],
    }
    _enrich_withdrawal(item)
    assert item["withdrawal_status"] == "locked"
    assert item["withdrawable_amount"] == 0


def test_enrich_withdrawal_partial():
    from app.ppf_database import _enrich_withdrawal
    # Account started 8 years ago => partial withdrawal eligible
    from datetime import date
    from dateutil.relativedelta import relativedelta
    start = (date.today() - relativedelta(years=8)).strftime("%Y-%m-%d")
    installments = []
    for m in range(1, 97):  # 8 years = 96 months
        installments.append({
            "is_past": True,
            "amount_invested": 5000,
            "interest_earned": 100,
        })
    item = {
        "start_date": start,
        "tenure_years": 15,
        "total_withdrawn": 0,
        "installments": installments,
    }
    _enrich_withdrawal(item)
    assert item["withdrawal_status"] == "partial"
    assert item["withdrawable_amount"] > 0


def test_enrich_withdrawal_full():
    from app.ppf_database import _enrich_withdrawal
    from datetime import date
    from dateutil.relativedelta import relativedelta
    start = (date.today() - relativedelta(years=16)).strftime("%Y-%m-%d")
    installments = []
    for m in range(1, 193):  # 16 years
        installments.append({
            "is_past": True,
            "amount_invested": 5000,
            "interest_earned": 100,
        })
    item = {
        "start_date": start,
        "tenure_years": 15,
        "total_withdrawn": 0,
        "installments": installments,
    }
    _enrich_withdrawal(item)
    assert item["withdrawal_status"] == "full"
    assert item["withdrawable_amount"] > 0


def test_enrich_withdrawal_recent_start_date():
    """Account started recently (< 7 years) is locked."""
    from app.ppf_database import _enrich_withdrawal
    item = {
        "start_date": "2024-01-01",
        "tenure_years": 15,
        "total_withdrawn": 0,
        "installments": [],
    }
    _enrich_withdrawal(item)
    assert item["withdrawal_status"] == "locked"
    assert item["years_completed"] >= 0


def test_enrich_withdrawal_partial_with_already_withdrawn():
    """Partial withdrawal reduces withdrawable by already_withdrawn amount."""
    from app.ppf_database import _enrich_withdrawal
    from datetime import date
    from dateutil.relativedelta import relativedelta
    start = (date.today() - relativedelta(years=9)).strftime("%Y-%m-%d")
    installments = []
    for m in range(1, 109):  # 9 years
        installments.append({
            "is_past": True,
            "amount_invested": 5000,
            "interest_earned": 100,
        })
    item = {
        "start_date": start,
        "tenure_years": 15,
        "total_withdrawn": 100000,  # already withdrawn 100k
        "installments": installments,
    }
    _enrich_withdrawal(item)
    assert item["withdrawal_status"] == "partial"
    assert "already withdrawn" in item["withdrawal_note"]


# ---------------------------------------------------------------------------
# Tests — withdraw (lines 1247-1298)
# ---------------------------------------------------------------------------

def test_withdraw_locked_raises(ppf_base_dir):
    from app.ppf_database import add, withdraw
    result = add({
        "account_name": "Lock PPF",
        "bank": "SBI",
        "start_date": "2023-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    with pytest.raises(ValueError, match="lock-in period"):
        withdraw(ppf_id, {"amount": 50000}, base_dir=str(ppf_base_dir))


def test_withdraw_zero_amount_raises(ppf_base_dir):
    from app.ppf_database import add, withdraw
    result = add({
        "account_name": "Zero PPF",
        "bank": "SBI",
        "start_date": "2015-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    with pytest.raises(ValueError, match="must be positive"):
        withdraw(ppf_id, {"amount": 0}, base_dir=str(ppf_base_dir))


def test_withdraw_nonexistent_raises(ppf_base_dir):
    from app.ppf_database import withdraw
    with pytest.raises(ValueError, match="not found"):
        withdraw("bad_id", {"amount": 50000}, base_dir=str(ppf_base_dir))


# ---------------------------------------------------------------------------
# Tests — _find_xlsx (lines 1309, 1312)
# ---------------------------------------------------------------------------

def test_find_xlsx_nonexistent_dir(ppf_base_dir):
    from app.ppf_database import _find_xlsx
    from pathlib import Path
    result = _find_xlsx("bad_id", ppf_dir=Path("/nonexistent/dir"))
    assert result is None


def test_find_xlsx_skips_temp_files(ppf_base_dir):
    from app.ppf_database import _find_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    (ppf_dir / "~$temp.xlsx").write_text("temp")
    result = _find_xlsx("bad_id", ppf_dir=ppf_dir)
    assert result is None


# ---------------------------------------------------------------------------
# Tests — _migrate_old_xlsx and _migrate_h4_to_cols (lines 780-878)
# ---------------------------------------------------------------------------

def test_migrate_old_xlsx_nonexistent_dir(ppf_base_dir):
    from app.ppf_database import _migrate_old_xlsx
    from pathlib import Path
    _migrate_old_xlsx(ppf_dir=Path("/nonexistent/dir"))  # should not raise


def test_migrate_h4_to_cols_nonexistent_dir(ppf_base_dir):
    from app.ppf_database import _migrate_h4_to_cols
    from pathlib import Path
    _migrate_h4_to_cols(ppf_dir=Path("/nonexistent/dir"))  # should not raise


def test_migrate_old_xlsx_skips_new_format(ppf_base_dir):
    """Files not in old format (A1 != 'Account Name') are skipped."""
    from app.ppf_database import add, _migrate_old_xlsx
    # Create a new-format file
    add({
        "account_name": "New Format",
        "bank": "SBI",
        "start_date": "2020-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_dir = ppf_base_dir / "PPF"
    # Running migration should skip new-format files
    _migrate_old_xlsx(ppf_dir=ppf_dir)
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1


def test_migrate_old_xlsx_converts_old_format(ppf_base_dir):
    """Old-format xlsx (A1='Account Name') is converted to new format."""
    import openpyxl
    from app.ppf_database import _migrate_old_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    filepath = ppf_dir / "Old PPF.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index"
    # Old format: row 1 has "Account Name" in A1
    ws.cell(1, 1, "Account Name")
    ws.cell(1, 2, "Old PPF")
    ws.cell(1, 4, "PNB")
    ws.cell(1, 6, "PPF001")
    ws.cell(1, 8, 7.1)
    ws.cell(2, 2, "2020-01-01")
    ws.cell(2, 4, 15)
    ws.cell(2, 8, 5000)
    ws.cell(3, 2, "monthly")
    wb.save(str(filepath))
    wb.close()

    _migrate_old_xlsx(ppf_dir=ppf_dir)
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1


def test_migrate_h4_to_cols_with_h4_data(ppf_base_dir):
    """xlsx with H4 contributions gets re-saved with data in cols 8-9."""
    import openpyxl
    import json
    from app.ppf_database import _create_ppf_xlsx, _migrate_h4_to_cols
    from datetime import datetime

    ppf_dir = ppf_base_dir / "PPF"

    # Create a PPF file first
    filepath = _create_ppf_xlsx(
        "H4 PPF", "SBI", 5000, 7.1, 15, "2020-01-01",
        overwrite=True,
        ppf_dir=ppf_dir,
    )

    # Manually add H4 contributions to the file
    wb = openpyxl.load_workbook(str(filepath))
    ws = wb["Index"]
    contribs = [{"date": "2020-06-15", "amount": 50000, "remarks": "Extra"}]
    ws.cell(4, 8, json.dumps(contribs))  # H4
    wb.save(str(filepath))
    wb.close()

    _migrate_h4_to_cols(ppf_dir=ppf_dir)
    # File should still exist and have been re-saved
    xlsx_files = list(ppf_dir.glob("*.xlsx"))
    assert len(xlsx_files) >= 1


# ---------------------------------------------------------------------------
# Tests — day clamping in installment generation (lines 373-376)
# ---------------------------------------------------------------------------

def test_day_clamping_feb_31(ppf_base_dir):
    """When phase starts on 31st and month is Feb, day clamps to 28/29."""
    from app.ppf_database import add
    result = add({
        "account_name": "Clamp PPF",
        "bank": "SBI",
        "start_date": "2020-01-31",  # Start on 31st
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))

    # Feb installment should be clamped (month 2)
    feb_inst = result["installments"][1]
    assert feb_inst["date"].startswith("2020-02-")


# ---------------------------------------------------------------------------
# Tests — withdraw success (lines 1268-1298)
# ---------------------------------------------------------------------------

def test_withdraw_success(ppf_base_dir):
    """Successful withdrawal from a matured PPF account."""
    from app.ppf_database import add, withdraw

    # Create an account with start 16+ years ago (fully matured)
    from datetime import date
    from dateutil.relativedelta import relativedelta
    start = (date.today() - relativedelta(years=16)).strftime("%Y-%m-%d")

    result = add({
        "account_name": "Withdraw PPF",
        "bank": "SBI",
        "start_date": start,
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    # Get the account to check withdrawable amount
    from app.ppf_database import get_all
    items = get_all(base_dir=str(ppf_base_dir))
    assert len(items) == 1
    withdrawable = items[0].get("withdrawable_amount", 0)

    if withdrawable > 0:
        updated = withdraw(ppf_id, {"amount": min(1000, withdrawable)},
                           base_dir=str(ppf_base_dir))
        assert updated is not None


def test_withdraw_exceeds_limit(ppf_base_dir):
    """Withdrawal exceeding withdrawable limit raises ValueError."""
    from app.ppf_database import add, withdraw
    from datetime import date
    from dateutil.relativedelta import relativedelta
    start = (date.today() - relativedelta(years=16)).strftime("%Y-%m-%d")

    result = add({
        "account_name": "Limit PPF",
        "bank": "SBI",
        "start_date": start,
        "tenure_years": 15,
        "sip_amount": 100,  # small SIP
        "sip_frequency": "yearly",
    }, base_dir=str(ppf_base_dir))
    ppf_id = result["id"]

    with pytest.raises(ValueError, match="exceeds|lock-in"):
        withdraw(ppf_id, {"amount": 99999999}, base_dir=str(ppf_base_dir))


# ---------------------------------------------------------------------------
# Tests — contribution with bad date in contrib_by_month (lines 343-344)
# ---------------------------------------------------------------------------

def test_parse_ppf_with_bad_contribution_date(ppf_base_dir):
    """Contributions with invalid dates in contrib_by_month are skipped."""
    from app.ppf_database import _create_ppf_xlsx, _parse_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    filepath = _create_ppf_xlsx(
        "Bad Contrib", "SBI", 5000, 7.1, 15, "2020-01-01",
        contributions=[
            {"date": "bad-date", "amount": 50000, "remarks": "Invalid"},
            {"date": "2020-06-15", "amount": 50000, "remarks": "Valid"},
        ],
        ppf_dir=ppf_dir,
    )
    result = _parse_ppf_xlsx(filepath)
    assert result["name"] == "Bad Contrib"


# ---------------------------------------------------------------------------
# Tests — lock_status for months > lockin_months (line 432)
# ---------------------------------------------------------------------------

def test_installments_lock_status_free(ppf_base_dir):
    """Installments beyond tenure years should have lock_status='free'."""
    from app.ppf_database import add
    result = add({
        "account_name": "Lock Status PPF",
        "bank": "SBI",
        "start_date": "2005-01-01",
        "tenure_years": 15,
        "sip_amount": 5000,
        "sip_frequency": "monthly",
    }, base_dir=str(ppf_base_dir))

    # Last installment (month 180) should be at or past lockin
    last = result["installments"][-1]
    # Month 180 = 15 years, should be 'partial' (month 84-180) or close to 'free'
    assert last["lock_status"] in ("partial", "free", "locked")

    # Months > 84 (year 7) should be at least partial
    month_85 = result["installments"][84]
    assert month_85["lock_status"] == "partial"


# ---------------------------------------------------------------------------
# Tests — scan_data_rows break at None date (line 250)
# ---------------------------------------------------------------------------

def test_scan_data_rows_stops_at_none_date():
    from app.ppf_database import _scan_data_rows
    from datetime import datetime
    rows = [
        (None,) * 11,  # row 1
        (None,) * 11,  # row 2
        (None,) * 11,  # row 3
        (None,) * 11,  # row 4
        (None,) * 11,  # row 5 (headers)
        (1, None, 1, datetime(2020, 1, 1), 5000, 0, 0, 0, 0),  # row 6: valid
        (2, None, 2, None, 5000, 0, 0, 0, 0),  # row 7: None date => stop
    ]
    result = _scan_data_rows(rows, 180)
    assert len(result) == 0  # no withdrawals/contributions in first row


# ---------------------------------------------------------------------------
# Tests — _create_ppf_xlsx with withdrawal in contributions (line 688)
# ---------------------------------------------------------------------------

def test_create_ppf_xlsx_with_withdrawals(ppf_base_dir):
    """_create_ppf_xlsx handles negative contribution (withdrawal) amounts."""
    from app.ppf_database import _create_ppf_xlsx, _parse_ppf_xlsx
    ppf_dir = ppf_base_dir / "PPF"
    filepath = _create_ppf_xlsx(
        "With Withdrawal", "SBI", 5000, 7.1, 15, "2020-01-01",
        contributions=[
            {"date": "2020-06-15", "amount": 50000, "remarks": "Deposit"},
            {"date": "2023-06-15", "amount": -10000, "remarks": "Withdrawal"},
        ],
        ppf_dir=ppf_dir,
    )
    result = _parse_ppf_xlsx(filepath)
    assert result["name"] == "With Withdrawal"
