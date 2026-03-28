"""Unit tests for app/si_database.py — Standing Instructions database layer."""
import json
from pathlib import Path
from unittest.mock import patch

import pytest


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def si_base_dir(tmp_path):
    """Create a temp base dir with Standing Instructions subdir."""
    (tmp_path / "Standing Instructions").mkdir()
    return tmp_path


@pytest.fixture(autouse=True)
def _mock_drive():
    with patch("app.si_database._sync_to_drive"), \
         patch("app.si_database._delete_from_drive"):
        yield


# ---------------------------------------------------------------------------
# Tests — empty state
# ---------------------------------------------------------------------------

def test_get_all_empty(si_base_dir):
    from app.si_database import get_all
    assert get_all(base_dir=str(si_base_dir)) == []


def test_get_dashboard_empty(si_base_dir):
    from app.si_database import get_dashboard
    dash = get_dashboard(base_dir=str(si_base_dir))
    assert dash["active_count"] == 0
    assert dash["total_count"] == 0
    assert dash["total_monthly_outflow"] == 0


# ---------------------------------------------------------------------------
# Tests — add
# ---------------------------------------------------------------------------

def test_add_si(si_base_dir):
    from app.si_database import add, get_all
    data = {
        "bank": "HDFC",
        "beneficiary": "Mutual Fund X",
        "amount": 5000,
        "frequency": "Monthly",
        "purpose": "SIP",
        "mandate_type": "NACH",
        "account_number": "1234567890",
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
        "alert_days": 30,
        "status": "Active",
    }
    result = add(data, base_dir=str(si_base_dir))
    assert result["bank"] == "HDFC"
    assert result["beneficiary"] == "Mutual Fund X"
    assert result["amount"] == 5000
    assert result["id"]

    # Persisted
    items = get_all(base_dir=str(si_base_dir))
    assert len(items) == 1
    assert items[0]["bank"] == "HDFC"


def test_add_multiple_sis(si_base_dir):
    from app.si_database import add, get_all
    for i in range(3):
        add({
            "bank": f"Bank{i}",
            "beneficiary": f"Ben{i}",
            "amount": 1000 * (i + 1),
            "start_date": "2024-01-01",
            "expiry_date": "2027-01-01",
        }, base_dir=str(si_base_dir))

    items = get_all(base_dir=str(si_base_dir))
    assert len(items) == 3


# ---------------------------------------------------------------------------
# Tests — update
# ---------------------------------------------------------------------------

def test_update_si(si_base_dir):
    from app.si_database import add, update
    result = add({
        "bank": "SBI",
        "beneficiary": "Insurance Co",
        "amount": 10000,
        "start_date": "2024-01-01",
        "expiry_date": "2025-01-01",
    }, base_dir=str(si_base_dir))
    si_id = result["id"]

    updated = update(si_id, {"amount": 15000, "frequency": "Quarterly"}, base_dir=str(si_base_dir))
    assert updated["amount"] == 15000
    assert updated["frequency"] == "Quarterly"


def test_update_nonexistent_raises(si_base_dir):
    from app.si_database import update
    with pytest.raises(ValueError, match="not found"):
        update("bad", {"amount": 0}, base_dir=str(si_base_dir))


# ---------------------------------------------------------------------------
# Tests — delete
# ---------------------------------------------------------------------------

def test_delete_si(si_base_dir):
    from app.si_database import add, delete, get_all
    result = add({
        "bank": "ICICI",
        "beneficiary": "EMI",
        "amount": 25000,
        "start_date": "2024-01-01",
        "expiry_date": "2026-01-01",
    }, base_dir=str(si_base_dir))
    si_id = result["id"]

    items = get_all(base_dir=str(si_base_dir))
    assert len(items) == 1

    del_result = delete(si_id, base_dir=str(si_base_dir))
    assert "deleted" in del_result["message"]

    items = get_all(base_dir=str(si_base_dir))
    assert len(items) == 0


def test_delete_nonexistent_raises(si_base_dir):
    from app.si_database import delete
    with pytest.raises(ValueError, match="not found"):
        delete("bad", base_dir=str(si_base_dir))


# ---------------------------------------------------------------------------
# Tests — dashboard
# ---------------------------------------------------------------------------

def test_dashboard_monthly_outflow(si_base_dir):
    from app.si_database import add, get_dashboard
    # Add a monthly SI
    add({
        "bank": "HDFC",
        "beneficiary": "MF",
        "amount": 5000,
        "frequency": "Monthly",
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
    }, base_dir=str(si_base_dir))
    # Add a quarterly SI
    add({
        "bank": "SBI",
        "beneficiary": "Insurance",
        "amount": 12000,
        "frequency": "Quarterly",
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
    }, base_dir=str(si_base_dir))

    dash = get_dashboard(base_dir=str(si_base_dir))
    assert dash["active_count"] == 2
    assert dash["total_count"] == 2
    # Monthly: 5000; Quarterly 12000/3 = 4000; total = 9000
    assert dash["total_monthly_outflow"] == 9000.0


# ---------------------------------------------------------------------------
# Tests — days_to_expiry computed field
# ---------------------------------------------------------------------------

def test_days_to_expiry_computed(si_base_dir):
    from app.si_database import add, get_all
    result = add({
        "bank": "BOB",
        "beneficiary": "Utility",
        "amount": 500,
        "start_date": "2020-01-01",
        "expiry_date": "2030-12-31",
    }, base_dir=str(si_base_dir))

    items = get_all(base_dir=str(si_base_dir))
    assert len(items) == 1
    # Expiry is far in the future, so days_to_expiry should be positive
    assert items[0]["days_to_expiry"] > 0


# ---------------------------------------------------------------------------
# Tests — _to_date_str helper
# ---------------------------------------------------------------------------

def test_to_date_str():
    from app.si_database import _to_date_str
    from datetime import datetime, date
    assert _to_date_str(datetime(2024, 6, 15)) == "2024-06-15"
    assert _to_date_str(date(2024, 6, 15)) == "2024-06-15"
    assert _to_date_str("2024-06-15") == "2024-06-15"
    assert _to_date_str(None) == ""
    assert _to_date_str("") == ""


# ---------------------------------------------------------------------------
# Tests — _sync_to_drive / _delete_from_drive (lines 40, 45-57)
# ---------------------------------------------------------------------------

def test_sync_to_drive_is_noop():
    """_sync_to_drive is a no-op that should not raise."""
    from app.si_database import _sync_to_drive
    from pathlib import Path
    _sync_to_drive(Path("/tmp/test.xlsx"))


def test_delete_from_drive_handles_exception():
    """_delete_from_drive swallows exceptions."""
    from app.si_database import _delete_from_drive
    from pathlib import Path
    _delete_from_drive(Path("/nonexistent/path/file.xlsx"))


# ---------------------------------------------------------------------------
# Tests — _load with corrupt file (lines 125-127)
# ---------------------------------------------------------------------------

def test_load_corrupt_file(si_base_dir):
    """_load returns [] when xlsx file is corrupt."""
    from app.si_database import _load
    si_dir = si_base_dir / "Standing Instructions"
    si_file = si_dir / "Standing Instructions.xlsx"
    si_file.write_text("NOT AN XLSX FILE")

    items = _load(si_file)
    assert items == []


# ---------------------------------------------------------------------------
# Tests — _load skips rows without id (line 138)
# ---------------------------------------------------------------------------

def test_load_skips_rows_without_id(si_base_dir):
    """Rows in xlsx without an id in column A are skipped."""
    import openpyxl
    from app.si_database import _load, _HEADERS
    si_dir = si_base_dir / "Standing Instructions"
    si_file = si_dir / "Standing Instructions.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standing Instructions"
    for i, h in enumerate(_HEADERS, 1):
        ws.cell(1, i, h)
    # Row 2: valid row
    ws.cell(2, 1, "si001")
    ws.cell(2, 2, "HDFC")
    ws.cell(2, 3, "MF X")
    ws.cell(2, 4, 5000)
    # Row 3: missing id (None)
    ws.cell(3, 1, None)
    ws.cell(3, 2, "SBI")
    wb.save(str(si_file))
    wb.close()

    items = _load(si_file)
    assert len(items) == 1
    assert items[0]["id"] == "si001"


# ---------------------------------------------------------------------------
# Tests — _save with invalid dates (lines 183-184, 187-188)
# ---------------------------------------------------------------------------

def test_save_with_invalid_dates(si_base_dir):
    """_save handles invalid date strings gracefully."""
    from app.si_database import _save, _load
    si_dir = si_base_dir / "Standing Instructions"
    si_file = si_dir / "Standing Instructions.xlsx"

    items = [{
        "id": "bad_dates",
        "bank": "Test",
        "beneficiary": "X",
        "amount": 100,
        "frequency": "Monthly",
        "purpose": "SIP",
        "mandate_type": "NACH",
        "account_number": "",
        "start_date": "not-a-date",
        "expiry_date": "also-not-a-date",
        "alert_days": 30,
        "status": "Active",
        "remarks": "",
    }]
    _save(items, si_dir, si_file)

    loaded = _load(si_file)
    assert len(loaded) == 1
    assert loaded[0]["id"] == "bad_dates"


# ---------------------------------------------------------------------------
# Tests — get_all with invalid expiry_date (lines 213-214)
# ---------------------------------------------------------------------------

def test_get_all_invalid_expiry_date(si_base_dir):
    """get_all handles invalid expiry_date by setting days_to_expiry=0."""
    from app.si_database import _save, get_all
    si_dir = si_base_dir / "Standing Instructions"
    si_file = si_dir / "Standing Instructions.xlsx"
    from datetime import datetime

    items = [{
        "id": "inv_exp",
        "bank": "Test",
        "beneficiary": "X",
        "amount": 100,
        "frequency": "Monthly",
        "purpose": "SIP",
        "mandate_type": "NACH",
        "account_number": "",
        "start_date": "2024-01-01",
        "expiry_date": "bad-date",
        "alert_days": 30,
        "status": "Active",
        "remarks": "",
    }]
    _save(items, si_dir, si_file)

    result = get_all(base_dir=str(si_base_dir))
    assert result[0]["days_to_expiry"] == 0


# ---------------------------------------------------------------------------
# Tests — dashboard frequency variants (lines 232-237)
# ---------------------------------------------------------------------------

def test_dashboard_half_yearly_and_annual_outflow(si_base_dir):
    """Dashboard normalizes Half-Yearly and Annual SIs to monthly equivalent."""
    from app.si_database import add, get_dashboard
    add({
        "bank": "A",
        "beneficiary": "X",
        "amount": 6000,
        "frequency": "Half-Yearly",
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
    }, base_dir=str(si_base_dir))
    add({
        "bank": "B",
        "beneficiary": "Y",
        "amount": 12000,
        "frequency": "Annually",
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
    }, base_dir=str(si_base_dir))
    add({
        "bank": "C",
        "beneficiary": "Z",
        "amount": 500,
        "frequency": "Weekly",  # unknown freq defaults to monthly
        "start_date": "2024-01-01",
        "expiry_date": "2027-01-01",
    }, base_dir=str(si_base_dir))

    dash = get_dashboard(base_dir=str(si_base_dir))
    # Half-Yearly: 6000/6 = 1000; Annually: 12000/12 = 1000; Weekly(default): 500
    assert dash["total_monthly_outflow"] == 2500.0


# ---------------------------------------------------------------------------
# Tests — dashboard expiring_soon (line 239-242)
# ---------------------------------------------------------------------------

def test_dashboard_expiring_soon(si_base_dir):
    """Dashboard counts SIs expiring within alert_days."""
    from app.si_database import _save, get_dashboard
    from datetime import datetime, timedelta
    si_dir = si_base_dir / "Standing Instructions"
    si_file = si_dir / "Standing Instructions.xlsx"
    soon = (datetime.now().date() + timedelta(days=15)).strftime("%Y-%m-%d")
    items = [{
        "id": "soon001",
        "bank": "Test",
        "beneficiary": "X",
        "amount": 1000,
        "frequency": "Monthly",
        "purpose": "SIP",
        "mandate_type": "NACH",
        "account_number": "",
        "start_date": "2020-01-01",
        "expiry_date": soon,
        "alert_days": 30,
        "status": "Active",
        "remarks": "",
    }]
    _save(items, si_dir, si_file)

    dash = get_dashboard(base_dir=str(si_base_dir))
    assert dash["expiring_soon"] == 1
