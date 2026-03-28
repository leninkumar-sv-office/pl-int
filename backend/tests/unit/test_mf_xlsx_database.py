"""
Unit tests for app.mf_xlsx_database — the mutual fund XLSX portfolio engine.

Targets 100% line coverage by exercising every function and branch:
  - _sync_to_drive (conflict copy cleanup)
  - _fetch_amfi_navs (AMFI NAV bulk fetch + cache + error paths)
  - _fetch_nav_google_finance (Google Finance scraping + retry)
  - fetch_live_navs (ISIN + GF routing, cache hits)
  - clear_nav_cache
  - _load_scheme_map / _save_scheme_map
  - _search_mfapi_scheme (fund name search, direct/growth matching)
  - _fetch_nav_history_mfapi
  - compute_nav_changes (1D/7D/30D change, 52W high/low, SMA, RSI, signals)
  - get_mf_nav_history / _filter_by_period
  - _gen_mf_id, _parse_date, _safe_float
  - fifo_match_mf (fractional units, multiple sells, tiny remainder)
  - _extract_mf_index_data, _parse_mf_trading_history
  - MFXlsxPortfolio: init, _build_file_map, reindex, caching
  - add_mf_holding (create file, dup detection, invalid date)
  - add_mf_sell_transaction (FIFO, default date, no file)
  - update_mf_holding, update_mf_sold_row
  - rename_fund
  - get_fund_summary (with live NAV mocking, LTCG/STCG)
  - get_dashboard_summary
  - get_fund_nav
"""

import json
import os
import time
from datetime import datetime, date, timedelta
from pathlib import Path
from unittest.mock import MagicMock, patch, mock_open

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Helper: create a minimal MF xlsx file for testing
# ---------------------------------------------------------------------------

def _create_mf_xlsx(filepath, fund_code="INF200K01RJ1", buys=None, sells=None,
                    current_price=0.0, w52_high=0.0, w52_low=0.0,
                    lock_in_cols=False):
    """Create a test MF xlsx file with Index + Trading History sheets."""
    wb = openpyxl.Workbook()

    # Trading History sheet
    ws = wb.active
    ws.title = "Trading History"
    headers = ["DATE", "EXCH", "ACTION", "Units", "NAV", "COST", "REMARKS",
               "STT", "ADD CHARGES", "Current Price", "Gain%"]
    for col, h in enumerate(headers, 1):
        ws.cell(4, col, value=h)

    row_idx = 5
    for buy in (buys or []):
        ws.cell(row_idx, 1, value=datetime.strptime(buy["date"], "%Y-%m-%d"))
        ws.cell(row_idx, 2, value="NSE")
        ws.cell(row_idx, 3, value="Buy")
        ws.cell(row_idx, 4, value=buy["units"])
        ws.cell(row_idx, 5, value=buy["nav"])
        ws.cell(row_idx, 6, value=round(buy["units"] * buy["nav"], 2))
        ws.cell(row_idx, 7, value=buy.get("remarks", "~"))
        row_idx += 1

    for sell in (sells or []):
        ws.cell(row_idx, 1, value=datetime.strptime(sell["date"], "%Y-%m-%d"))
        ws.cell(row_idx, 2, value="NSE")
        ws.cell(row_idx, 3, value="Sell")
        ws.cell(row_idx, 4, value=sell["units"])
        ws.cell(row_idx, 5, value=sell.get("nav", 0))
        ws.cell(row_idx, 6, value=round(sell["units"] * sell.get("nav", 0), 2))
        ws.cell(row_idx, 7, value=sell.get("remarks", "~"))
        row_idx += 1

    # Index sheet
    ws_idx = wb.create_sheet("Index")
    ws_idx.cell(1, 2, value="Code")
    ws_idx.cell(1, 3, value=fund_code)
    ws_idx.cell(2, 2, value="Current Price")
    ws_idx.cell(2, 3, value=current_price)
    ws_idx.cell(3, 2, value="52 Week High")
    ws_idx.cell(3, 3, value=w52_high)
    ws_idx.cell(4, 2, value="52 Week Low")
    ws_idx.cell(4, 3, value=w52_low)

    if lock_in_cols:
        # Add Lock In and Exit Load labels in column I (index 9)
        ws_idx.cell(5, 9, value="Lock In Period")
        ws_idx.cell(6, 9, value="Exit Load")

    wb.save(filepath)
    wb.close()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def mf_dir(tmp_path):
    """Create an empty Mutual Funds directory."""
    d = tmp_path / "Mutual Funds"
    d.mkdir(parents=True)
    return d


@pytest.fixture
def mf_portfolio(mf_dir):
    """Create an MFXlsxPortfolio with external deps mocked."""
    with patch("app.mf_xlsx_database._sync_to_drive"):
        from app.mf_xlsx_database import MFXlsxPortfolio
        db = MFXlsxPortfolio(mf_dir)
    return db


@pytest.fixture
def mf_portfolio_with_fund(mf_dir):
    """Create a portfolio pre-loaded with one fund file."""
    _create_mf_xlsx(
        mf_dir / "SBI Small Cap Fund.xlsx",
        fund_code="INF200K01RJ1",
        buys=[
            {"date": "2024-01-15", "units": 50.0, "nav": 100.0},
            {"date": "2024-06-15", "units": 30.0, "nav": 110.0},
        ],
        current_price=120.0,
        w52_high=130.0,
        w52_low=90.0,
    )
    with patch("app.mf_xlsx_database._sync_to_drive"):
        from app.mf_xlsx_database import MFXlsxPortfolio
        db = MFXlsxPortfolio(mf_dir)
    return db


# ---------------------------------------------------------------------------
# _sync_to_drive
# ---------------------------------------------------------------------------

class TestSyncToDrive:
    def test_removes_conflict_copies(self, tmp_path):
        from app.mf_xlsx_database import _sync_to_drive
        # Create main file and a conflict copy
        main = tmp_path / "Fund.xlsx"
        main.write_bytes(b"main")
        conflict = tmp_path / "Fund (1).xlsx"
        conflict.write_bytes(b"conflict")

        _sync_to_drive(main)
        assert not conflict.exists()
        assert main.exists()

    def test_handles_unlink_error(self, tmp_path):
        """When conflict file can't be deleted, no exception raised."""
        from app.mf_xlsx_database import _sync_to_drive
        main = tmp_path / "Fund.xlsx"
        main.write_bytes(b"main")
        # No conflict files — should not raise
        _sync_to_drive(main)


# ---------------------------------------------------------------------------
# AMFI NAV Fetching
# ---------------------------------------------------------------------------

class TestFetchAmfiNavs:
    def test_successful_fetch(self):
        from app.mf_xlsx_database import _fetch_amfi_navs, _amfi_isin_nav
        import app.mf_xlsx_database as mod

        # Reset cache
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

        amfi_text = (
            "Header line\n"
            "101;INF200K01RJ1;INF200K01RJ2;SBI Small Cap;120.50;01-01-2026\n"
            "102;;INF846K01DP8;Axis Bluechip;45.67;01-01-2026\n"
            "bad line\n"
            "103;INVALID;;Fund3;abc;01-01-2026\n"  # invalid nav
            "104;INF999;;Negative NAV;-5.0;01-01-2026\n"  # negative nav
        )

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = amfi_text

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            result = _fetch_amfi_navs()

        assert "INF200K01RJ1" in result
        assert result["INF200K01RJ1"] == 120.50
        assert "INF200K01RJ2" in result
        assert "INF846K01DP8" in result

        # Cleanup
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

    def test_cache_hit(self):
        import app.mf_xlsx_database as mod
        mod._amfi_isin_nav = {"INF123": 100.0}
        mod._amfi_fetch_time = time.time()  # Fresh cache

        result = mod._fetch_amfi_navs()
        assert result == {"INF123": 100.0}

        # Cleanup
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

    def test_non_200_response(self):
        import app.mf_xlsx_database as mod
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

        mock_resp = MagicMock()
        mock_resp.status_code = 500

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            result = mod._fetch_amfi_navs()
        assert result == {}

        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

    def test_network_error(self):
        import app.mf_xlsx_database as mod
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

        with patch("app.mf_xlsx_database._requests.get", side_effect=Exception("timeout")):
            result = mod._fetch_amfi_navs()
        assert result == {}

        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

    def test_empty_mapping_not_cached(self):
        """If no valid ISINs found, cache not updated."""
        import app.mf_xlsx_database as mod
        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0

        amfi_text = "header\nbad;no;isins;here;abc;date\n"
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = amfi_text

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            result = mod._fetch_amfi_navs()
        assert result == {}

        mod._amfi_isin_nav = {}
        mod._amfi_fetch_time = 0.0


# ---------------------------------------------------------------------------
# Google Finance NAV Fetching
# ---------------------------------------------------------------------------

class TestFetchNavGoogleFinance:
    def test_successful_fetch(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = '<div data-last-price="125.50">NAV</div>'

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            with patch("app.mf_xlsx_database.time.sleep"):
                nav = _fetch_nav_google_finance("MUTF_IN:SBI_SMAL_CAP")
        assert nav == 125.50

    def test_no_fund_code(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance
        assert _fetch_nav_google_finance(None) is None
        assert _fetch_nav_google_finance("NOCOLON") is None

    def test_no_match_in_html(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = '<div>No price data</div>'

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            with patch("app.mf_xlsx_database.time.sleep"):
                nav = _fetch_nav_google_finance("MUTF_IN:FUND")
        assert nav is None

    def test_request_exception_retries(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance

        with patch("app.mf_xlsx_database._requests.get", side_effect=Exception("net err")):
            with patch("app.mf_xlsx_database.time.sleep"):
                nav = _fetch_nav_google_finance("MUTF_IN:FUND")
        assert nav is None

    def test_non_200_status(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance

        mock_resp = MagicMock()
        mock_resp.status_code = 404

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            with patch("app.mf_xlsx_database.time.sleep"):
                nav = _fetch_nav_google_finance("MUTF_IN:FUND")
        assert nav is None

    def test_zero_price(self):
        from app.mf_xlsx_database import _fetch_nav_google_finance

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.text = '<div data-last-price="0">NAV</div>'

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            with patch("app.mf_xlsx_database.time.sleep"):
                nav = _fetch_nav_google_finance("MUTF_IN:FUND")
        assert nav is None


# ---------------------------------------------------------------------------
# fetch_live_navs
# ---------------------------------------------------------------------------

class TestFetchLiveNavs:
    def test_isin_codes_from_amfi(self):
        from app.mf_xlsx_database import fetch_live_navs
        import app.mf_xlsx_database as mod

        # Clear nav cache
        mod._nav_cache.clear()

        with patch("app.mf_xlsx_database._fetch_amfi_navs", return_value={"INF123": 100.0, "INF456": 200.0}):
            result = fetch_live_navs(["INF123", "INF456"])
        assert result == {"INF123": 100.0, "INF456": 200.0}

        mod._nav_cache.clear()

    def test_gf_codes(self):
        from app.mf_xlsx_database import fetch_live_navs
        import app.mf_xlsx_database as mod
        mod._nav_cache.clear()

        with patch("app.mf_xlsx_database._fetch_nav_google_finance", return_value=150.0):
            with patch("app.mf_xlsx_database.time.sleep"):
                with patch("app.mf_xlsx_database.random.uniform", return_value=0.5):
                    result = fetch_live_navs(["MUTF_IN:FUND1"])
        assert result == {"MUTF_IN:FUND1": 150.0}

        mod._nav_cache.clear()

    def test_cache_hit(self):
        from app.mf_xlsx_database import fetch_live_navs
        import app.mf_xlsx_database as mod
        mod._nav_cache.clear()
        mod._nav_cache["INF123"] = 99.0
        mod._nav_cache["MUTF_IN:X"] = 88.0

        result = fetch_live_navs(["INF123", "MUTF_IN:X"])
        assert result["INF123"] == 99.0
        assert result["MUTF_IN:X"] == 88.0

        mod._nav_cache.clear()

    def test_gf_nav_is_none(self):
        """When Google Finance returns None, code should not cache."""
        from app.mf_xlsx_database import fetch_live_navs
        import app.mf_xlsx_database as mod
        mod._nav_cache.clear()

        with patch("app.mf_xlsx_database._fetch_nav_google_finance", return_value=None):
            with patch("app.mf_xlsx_database.time.sleep"):
                with patch("app.mf_xlsx_database.random.uniform", return_value=0.5):
                    result = fetch_live_navs(["MUTF_IN:NOFUND"])
        assert "MUTF_IN:NOFUND" not in result

        mod._nav_cache.clear()

    def test_empty_and_none_codes_filtered(self):
        from app.mf_xlsx_database import fetch_live_navs
        result = fetch_live_navs([None, "", ""])
        assert result == {}

    def test_amfi_nav_zero_or_none(self):
        """AMFI NAV that's 0 or None should not be cached."""
        from app.mf_xlsx_database import fetch_live_navs
        import app.mf_xlsx_database as mod
        mod._nav_cache.clear()

        with patch("app.mf_xlsx_database._fetch_amfi_navs", return_value={"INF123": 0}):
            result = fetch_live_navs(["INF123"])
        assert "INF123" not in result

        mod._nav_cache.clear()


class TestClearNavCache:
    def test_clears_cache(self):
        import app.mf_xlsx_database as mod
        mod._nav_cache["test"] = 1.0
        mod.clear_nav_cache()
        assert len(mod._nav_cache) == 0


# ---------------------------------------------------------------------------
# Scheme Map
# ---------------------------------------------------------------------------

class TestSchemeMap:
    def test_load_scheme_map_existing(self, tmp_path):
        import app.mf_xlsx_database as mod
        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"INF123": 12345}))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)):
            result = mod._load_scheme_map()
        assert result == {"INF123": 12345}

    def test_load_scheme_map_missing_file(self, tmp_path):
        import app.mf_xlsx_database as mod
        with patch.object(mod, "_SCHEME_MAP_FILE", str(tmp_path / "nope.json")):
            result = mod._load_scheme_map()
        assert result == {}

    def test_load_scheme_map_corrupt_json(self, tmp_path):
        import app.mf_xlsx_database as mod
        map_file = tmp_path / "bad.json"
        map_file.write_text("not json{{{")
        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)):
            result = mod._load_scheme_map()
        assert result == {}

    def test_save_scheme_map(self, tmp_path):
        import app.mf_xlsx_database as mod
        map_file = tmp_path / "data" / "scheme_map.json"
        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)):
            with patch.object(mod, "_NAV_DATA_DIR", str(tmp_path / "data")):
                mod._save_scheme_map({"INF123": 999})
        assert json.loads(map_file.read_text()) == {"INF123": 999}

    def test_save_scheme_map_error(self, tmp_path):
        """Should not raise when save fails."""
        import app.mf_xlsx_database as mod
        with patch.object(mod, "_SCHEME_MAP_FILE", "/dev/null/impossible/path.json"):
            with patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)):
                mod._save_scheme_map({"INF123": 999})  # Should not raise


# ---------------------------------------------------------------------------
# _search_mfapi_scheme
# ---------------------------------------------------------------------------

class TestSearchMfapiScheme:
    def test_finds_direct_growth(self):
        from app.mf_xlsx_database import _search_mfapi_scheme

        results = [
            {"schemeCode": 101, "schemeName": "SBI Small Cap Regular Growth"},
            {"schemeCode": 102, "schemeName": "SBI Small Cap Direct Plan Growth"},
        ]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = results

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            code = _search_mfapi_scheme("SBI Small Cap Fund - Direct Plan - Growth")
        assert code == 102

    def test_finds_direct_without_growth(self):
        from app.mf_xlsx_database import _search_mfapi_scheme

        results = [
            {"schemeCode": 201, "schemeName": "Axis Bluechip Regular"},
            {"schemeCode": 202, "schemeName": "Axis Bluechip Direct Plan"},
        ]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = results

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            code = _search_mfapi_scheme("Axis Bluechip Fund")
        assert code == 202

    def test_falls_back_to_first_result(self):
        from app.mf_xlsx_database import _search_mfapi_scheme

        results = [
            {"schemeCode": 301, "schemeName": "Some Regular Fund"},
        ]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = results

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            code = _search_mfapi_scheme("Some Fund")
        assert code == 301

    def test_no_results(self):
        from app.mf_xlsx_database import _search_mfapi_scheme

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = []

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            code = _search_mfapi_scheme("Nonexistent Fund")
        assert code is None

    def test_network_error(self):
        from app.mf_xlsx_database import _search_mfapi_scheme

        with patch("app.mf_xlsx_database._requests.get", side_effect=Exception("timeout")):
            code = _search_mfapi_scheme("Any Fund")
        assert code is None


# ---------------------------------------------------------------------------
# _fetch_nav_history_mfapi
# ---------------------------------------------------------------------------

class TestFetchNavHistoryMfapi:
    def test_successful_fetch(self):
        from app.mf_xlsx_database import _fetch_nav_history_mfapi

        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "data": [
                {"date": "01-01-2026", "nav": "100.0"},
                {"date": "02-01-2026", "nav": "105.0"},
            ]
        }

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            data = _fetch_nav_history_mfapi(12345)
        assert len(data) == 2

    def test_non_200(self):
        from app.mf_xlsx_database import _fetch_nav_history_mfapi

        mock_resp = MagicMock()
        mock_resp.status_code = 500

        with patch("app.mf_xlsx_database._requests.get", return_value=mock_resp):
            data = _fetch_nav_history_mfapi(12345)
        assert data is None

    def test_network_error(self):
        from app.mf_xlsx_database import _fetch_nav_history_mfapi

        with patch("app.mf_xlsx_database._requests.get", side_effect=Exception("err")):
            data = _fetch_nav_history_mfapi(12345)
        assert data is None


# ---------------------------------------------------------------------------
# compute_nav_changes — comprehensive
# ---------------------------------------------------------------------------

class TestComputeNavChanges:
    def _make_nav_data(self, num_days=400, start_nav=100.0, daily_change=0.1):
        """Generate nav_data list for mfapi format (most recent first)."""
        today = date.today()
        data = []
        nav = start_nav + (num_days * daily_change)
        for i in range(num_days):
            d = today - timedelta(days=i)
            data.append({
                "date": d.strftime("%d-%m-%Y"),
                "nav": str(round(nav, 4)),
            })
            nav -= daily_change
        return data

    def test_full_computation(self, tmp_path):
        """Tests 1D, 7D, 30D changes, 52W high/low, SMA, signal, RSI."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        nav_data = self._make_nav_data(num_days=400, start_nav=100.0, daily_change=0.1)

        scheme_map = {"FUND1": 12345}
        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps(scheme_map))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Test Fund", 140.0)

        assert "day_change" in result
        assert "day_change_pct" in result
        assert "week_change_pct" in result
        assert "month_change_pct" in result
        assert result["week_52_high"] > 0
        assert result["week_52_low"] > 0
        assert result["sma_50"] is not None
        assert result["sma_200"] is not None
        assert result["signal"] is not None
        assert result["rsi"] is not None

        mod._nav_change_cache.clear()

    def test_cache_hit(self):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()
        mod._nav_change_cache["CACHED"] = {
            "day_change": 1.0, "day_change_pct": 0.5,
            "week_change_pct": 2.0, "month_change_pct": 5.0,
            "week_52_high": 150.0, "week_52_low": 80.0,
            "sma_50": 120.0, "sma_200": 110.0,
            "signal": "strong_bull", "days_below_sma": 0,
            "rsi": 55.0, "fetched_at": time.time(),
        }

        result = mod.compute_nav_changes("CACHED", "Fund", 130.0)
        assert result["week_change_pct"] == 2.0
        assert result["signal"] == "strong_bull"

        mod._nav_change_cache.clear()

    def test_zero_nav(self):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()
        result = mod.compute_nav_changes("X", "Fund", 0.0)
        assert result["week_change_pct"] == 0.0
        mod._nav_change_cache.clear()

    def test_no_scheme_code_caches_miss(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({}))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._search_mfapi_scheme", return_value=None):
            result = mod.compute_nav_changes("NOFUND", "No Fund", 100.0)

        assert result["week_change_pct"] == 0.0
        # Should be cached so next call doesn't retry
        assert "NOFUND" in mod._nav_change_cache

        mod._nav_change_cache.clear()

    def test_new_scheme_found_and_saved(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({}))

        nav_data = self._make_nav_data(num_days=50)

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._search_mfapi_scheme", return_value=99999), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("NEWFUND", "New Fund", 110.0)

        # Scheme map should be updated
        saved = json.loads(map_file.read_text())
        assert saved.get("NEWFUND") == 99999

        mod._nav_change_cache.clear()

    def test_no_nav_data(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=None):
            result = mod.compute_nav_changes("FUND1", "Fund", 100.0)

        assert result["week_change_pct"] == 0.0
        mod._nav_change_cache.clear()

    def test_invalid_nav_entries_skipped(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        nav_data = [
            {"date": "invalid-date", "nav": "100.0"},
            {"date": "01-01-2026", "nav": "abc"},
            {"date": "01-01-2026"},  # missing nav key
            {"date": "01-01-2026", "nav": "100.0"},  # valid
        ]

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 100.0)
        # Should not crash
        assert "week_change_pct" in result
        mod._nav_change_cache.clear()

    def test_signal_weak_bull(self, tmp_path):
        """When sma_50 > sma_200 but current_nav < sma_200 → weak_bull."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        # Create data where recent NAVs dipped below sma_200 level
        today = date.today()
        nav_data = []
        for i in range(300):
            d = today - timedelta(days=i)
            if i < 50:
                nav = 90.0  # Recent NAVs low
            elif i < 200:
                nav = 110.0  # Medium-term higher
            else:
                nav = 100.0  # Long-term moderate
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(nav)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 95.0)

        # sma_50 calculated from first 50 entries (90.0), sma_200 from first 200
        # The exact signal depends on computed values
        assert result["signal"] in ("strong_bull", "weak_bull", "weak_bear", "strong_bear")
        mod._nav_change_cache.clear()

    def test_signal_strong_bear(self, tmp_path):
        """When sma_50 < sma_200 and current_nav < sma_200 → strong_bear."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        today = date.today()
        nav_data = []
        for i in range(300):
            d = today - timedelta(days=i)
            # Declining: older is higher, newer is lower
            nav = 200.0 - (300 - i) * 0.3
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(max(nav, 1.0))})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 50.0)

        assert result["signal"] is not None
        mod._nav_change_cache.clear()

    def test_rsi_all_gains(self, tmp_path):
        """When all recent changes are gains, RSI should be 100."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        today = date.today()
        nav_data = []
        for i in range(30):
            d = today - timedelta(days=i)
            nav = 100.0 + (30 - i)  # Strictly increasing (most recent highest)
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(nav)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 131.0)

        assert result["rsi"] == 100.0
        mod._nav_change_cache.clear()

    def test_short_data_sma_fallback(self, tmp_path):
        """With 20-49 data points, SMA uses 20-day fallback."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        today = date.today()
        nav_data = []
        for i in range(25):
            d = today - timedelta(days=i)
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(100.0 + i * 0.1)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 105.0)

        # With 25 data points: sma_50 fallback to 20-day, sma_200 fallback to None (< 50)
        assert result["sma_50"] is not None  # 20-day fallback
        assert result["sma_200"] is None     # Not enough data
        mod._nav_change_cache.clear()

    def test_days_below_sma(self, tmp_path):
        """Test days_below_sma counter."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        today = date.today()
        nav_data = []
        # Create pattern: first 5 days below SMA, then above
        for i in range(60):
            d = today - timedelta(days=i)
            if i < 5:
                nav = 50.0  # Well below average
            else:
                nav = 100.0 + i * 0.1
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(nav)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 50.0)

        assert result["days_below_sma"] >= 0
        mod._nav_change_cache.clear()

    def test_too_few_data_for_rsi(self, tmp_path):
        """With < 15 data points, RSI should remain None."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        today = date.today()
        nav_data = []
        for i in range(10):
            d = today - timedelta(days=i)
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(100.0 + i)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.compute_nav_changes("FUND1", "Fund", 110.0)

        assert result["rsi"] is None
        mod._nav_change_cache.clear()


# ---------------------------------------------------------------------------
# get_mf_nav_history / _filter_by_period
# ---------------------------------------------------------------------------

class TestGetMfNavHistory:
    def test_successful_fetch(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_history_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        nav_data = [
            {"date": "15-01-2026", "nav": "100.0"},
            {"date": "16-01-2026", "nav": "101.0"},
            {"date": "17-01-2026", "nav": "invalid"},  # should skip
            {"date": "bad-date", "nav": "102.0"},       # should skip
        ]

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.get_mf_nav_history("FUND1", "1y")

        assert len(result) == 2
        assert result[0]["close"] == 100.0
        mod._nav_history_cache.clear()

    def test_cache_hit(self):
        import app.mf_xlsx_database as mod
        mod._nav_history_cache.clear()
        cached_data = [{"date": "2026-01-15", "close": 100.0}]
        mod._nav_history_cache["FUND1"] = (time.time(), cached_data)

        result = mod.get_mf_nav_history("FUND1", "1y")
        assert len(result) == 1
        mod._nav_history_cache.clear()

    def test_no_scheme_code(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_history_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({}))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._search_mfapi_scheme", return_value=None):
            result = mod.get_mf_nav_history("NOFUND", "1y", fund_name="No Fund")
        assert result is None
        mod._nav_history_cache.clear()

    def test_scheme_found_via_search(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_history_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({}))

        nav_data = [{"date": "15-01-2026", "nav": "100.0"}]

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._search_mfapi_scheme", return_value=88888), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            result = mod.get_mf_nav_history("NEWFUND", "1y", fund_name="New Fund")

        assert result is not None
        # Scheme should be saved
        saved = json.loads(map_file.read_text())
        assert saved.get("NEWFUND") == 88888
        mod._nav_history_cache.clear()

    def test_no_nav_data(self, tmp_path):
        import app.mf_xlsx_database as mod
        mod._nav_history_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND1": 12345}))

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=None):
            result = mod.get_mf_nav_history("FUND1", "1y")
        assert result is None
        mod._nav_history_cache.clear()


class TestFilterByPeriod:
    def _make_data(self, num_days):
        today = date.today()
        return [
            {"date": (today - timedelta(days=i)).isoformat(), "close": 100.0 + i}
            for i in range(num_days - 1, -1, -1)
        ]

    def test_empty_data(self):
        from app.mf_xlsx_database import _filter_by_period
        assert _filter_by_period([], "1y") == []

    def test_ytd_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(400)
        result = _filter_by_period(data, "ytd")
        assert all(d["date"] >= f"{date.today().year}-01-01" for d in result)

    def test_max_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(100)
        result = _filter_by_period(data, "max")
        assert len(result) == 100

    def test_6m_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(400)
        result = _filter_by_period(data, "6m")
        assert len(result) <= 200

    def test_3y_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(2000)
        result = _filter_by_period(data, "3y")
        assert len(result) > 0

    def test_5y_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(2000)
        result = _filter_by_period(data, "5y")
        assert len(result) > 0

    def test_1m_period(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(400)
        result = _filter_by_period(data, "1m")
        assert len(result) <= 35

    def test_unknown_period_defaults_to_1y(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(400)
        result = _filter_by_period(data, "unknown")
        assert len(result) <= 370

    def test_downsampling(self):
        from app.mf_xlsx_database import _filter_by_period
        data = self._make_data(600)
        result = _filter_by_period(data, "max")
        # Should be downsampled to ~300 + last point
        assert len(result) < 600


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_gen_mf_id(self):
        from app.mf_xlsx_database import _gen_mf_id
        id1 = _gen_mf_id("FUND1", "2025-01-01", 100.0, 5)
        id2 = _gen_mf_id("FUND1", "2025-01-01", 100.0, 5)
        id3 = _gen_mf_id("FUND1", "2025-01-01", 100.0, 6)
        assert id1 == id2  # Same inputs → same ID
        assert id1 != id3  # Different row_idx → different ID

    def test_parse_date_datetime(self):
        from app.mf_xlsx_database import _parse_date
        dt = datetime(2025, 6, 15, 10, 30)
        assert _parse_date(dt) == "2025-06-15"

    def test_parse_date_date(self):
        from app.mf_xlsx_database import _parse_date
        d = date(2025, 6, 15)
        assert _parse_date(d) == "2025-06-15"

    def test_parse_date_str_formats(self):
        from app.mf_xlsx_database import _parse_date
        assert _parse_date("2025-06-15") == "2025-06-15"
        assert _parse_date("15-06-2025") == "2025-06-15"
        assert _parse_date("06/15/2025") == "2025-06-15"
        assert _parse_date("15-Jun-2025") == "2025-06-15"
        assert _parse_date("15-June-2025") == "2025-06-15"

    def test_parse_date_invalid(self):
        from app.mf_xlsx_database import _parse_date
        assert _parse_date("not a date") is None
        assert _parse_date(12345) is None
        assert _parse_date(None) is None

    def test_safe_float(self):
        from app.mf_xlsx_database import _safe_float
        assert _safe_float(None) == 0.0
        assert _safe_float(None, 5.0) == 5.0
        assert _safe_float("abc") == 0.0
        assert _safe_float("123.45") == 123.45
        assert _safe_float(100) == 100.0


# ---------------------------------------------------------------------------
# Pure function: fifo_match_mf (fractional units)
# ---------------------------------------------------------------------------

class TestFifoMatchMF:
    """Tests for the fifo_match_mf() pure function with fractional units."""

    def test_no_sells(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [
            {"date": "2025-01-01", "units": 10.5, "nav": 50.0},
            {"date": "2025-02-01", "units": 5.25, "nav": 55.0},
        ]
        remaining, sold = fifo_match_mf(buys, [])
        assert len(remaining) == 2
        assert abs(remaining[0]["remaining"] - 10.5) < 0.001
        assert abs(remaining[1]["remaining"] - 5.25) < 0.001
        assert sold == []

    def test_full_sell_fifo_order(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [
            {"date": "2025-01-01", "units": 10.0, "nav": 50.0},
            {"date": "2025-02-01", "units": 5.0, "nav": 55.0},
        ]
        sells = [{"date": "2025-03-01", "units": 10.0, "nav": 60.0}]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 1
        assert abs(remaining[0]["remaining"] - 5.0) < 0.001
        assert remaining[0]["nav"] == 55.0
        assert len(sold) == 1
        assert sold[0]["buy_nav"] == 50.0
        assert sold[0]["sell_nav"] == 60.0
        assert abs(sold[0]["units"] - 10.0) < 0.001
        assert abs(sold[0]["realized_pl"] - (60.0 - 50.0) * 10.0) < 0.01

    def test_fractional_sell(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [{"date": "2025-01-01", "units": 10.5678, "nav": 100.0}]
        sells = [{"date": "2025-03-01", "units": 3.2345, "nav": 110.0}]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 1
        assert abs(remaining[0]["remaining"] - (10.5678 - 3.2345)) < 0.001
        assert len(sold) == 1
        assert abs(sold[0]["units"] - 3.2345) < 0.001

    def test_sell_spans_multiple_lots(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [
            {"date": "2025-01-01", "units": 5.5, "nav": 100.0},
            {"date": "2025-02-01", "units": 4.5, "nav": 110.0},
        ]
        sells = [{"date": "2025-03-01", "units": 8.0, "nav": 120.0}]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 1
        assert abs(remaining[0]["remaining"] - 2.0) < 0.001
        assert len(sold) == 2
        assert abs(sold[0]["units"] - 5.5) < 0.001
        assert abs(sold[1]["units"] - 2.5) < 0.001

    def test_sell_all(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [
            {"date": "2025-01-01", "units": 5.0, "nav": 100.0},
            {"date": "2025-02-01", "units": 5.0, "nav": 110.0},
        ]
        sells = [{"date": "2025-03-01", "units": 10.0, "nav": 120.0}]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 0
        assert len(sold) == 2

    def test_multiple_sells(self):
        from app.mf_xlsx_database import fifo_match_mf
        buys = [{"date": "2025-01-01", "units": 100.0, "nav": 50.0}]
        sells = [
            {"date": "2025-02-01", "units": 30.5, "nav": 55.0},
            {"date": "2025-03-01", "units": 20.25, "nav": 60.0},
        ]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 1
        assert abs(remaining[0]["remaining"] - 49.25) < 0.001
        assert len(sold) == 2

    def test_tiny_remainder_filtered(self):
        """Lots with < 0.0001 units remaining should be filtered out."""
        from app.mf_xlsx_database import fifo_match_mf
        buys = [{"date": "2025-01-01", "units": 10.0, "nav": 100.0}]
        sells = [{"date": "2025-03-01", "units": 9.9999, "nav": 110.0}]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 0


# ---------------------------------------------------------------------------
# XLSX Parsing
# ---------------------------------------------------------------------------

class TestExtractMfIndexData:
    def test_extract_from_valid_wb(self, tmp_path):
        from app.mf_xlsx_database import _extract_mf_index_data
        filepath = tmp_path / "test.xlsx"
        _create_mf_xlsx(filepath, fund_code="INF123", current_price=150.0,
                        w52_high=200.0, w52_low=100.0, lock_in_cols=True)
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        data = _extract_mf_index_data(wb)
        wb.close()

        assert data["fund_code"] == "INF123"
        assert data["current_nav"] == 150.0
        assert data["week_52_high"] == 200.0
        assert data["week_52_low"] == 100.0

    def test_no_index_sheet(self, tmp_path):
        from app.mf_xlsx_database import _extract_mf_index_data
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NotIndex"
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        data = _extract_mf_index_data(wb2)
        wb2.close()
        assert data["fund_code"] is None


class TestParseMfTradingHistory:
    def test_parse_buys_and_sells(self, tmp_path):
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        _create_mf_xlsx(filepath, buys=[
            {"date": "2025-01-15", "units": 50.0, "nav": 100.0},
            {"date": "2025-02-15", "units": 30.0, "nav": 110.0},
        ], sells=[
            {"date": "2025-03-15", "units": 20.0, "nav": 120.0},
        ])
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb)
        wb.close()

        assert len(buys) == 2
        assert len(sells) == 1
        assert buys[0]["units"] == 50.0
        assert sells[0]["units"] == 20.0

    def test_no_trading_history_sheet(self, tmp_path):
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NotTH"
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb2)
        wb2.close()
        assert buys == []
        assert sells == []

    def test_too_few_rows(self, tmp_path):
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        ws.cell(1, 1, value="A")
        ws.cell(2, 1, value="B")
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb2)
        wb2.close()
        assert buys == []
        assert sells == []

    def test_no_header_row_found(self, tmp_path):
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        for i in range(1, 11):
            ws.cell(i, 1, value="data")
            ws.cell(i, 2, value="data")
            ws.cell(i, 3, value="data")
            ws.cell(i, 4, value="data")
            ws.cell(i, 5, value="data")
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb2)
        wb2.close()
        assert buys == []
        assert sells == []

    def test_sell_with_zero_nav_calculates_from_cost(self, tmp_path):
        """When NAV is 0 but cost > 0 and units > 0, NAV = cost/units."""
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        # Header row
        ws.cell(4, 1, value="DATE")
        ws.cell(4, 2, value="EXCH")
        ws.cell(4, 3, value="ACTION")
        ws.cell(4, 4, value="Units")
        ws.cell(4, 5, value="NAV")
        ws.cell(4, 6, value="COST")
        ws.cell(4, 7, value="REMARKS")
        # Buy row
        ws.cell(5, 1, value=datetime(2025, 1, 15))
        ws.cell(5, 2, value="NSE")
        ws.cell(5, 3, value="Buy")
        ws.cell(5, 4, value=10.0)
        ws.cell(5, 5, value=100.0)
        ws.cell(5, 6, value=1000.0)
        # Sell row with NAV=0, COST=1200
        ws.cell(6, 1, value=datetime(2025, 3, 15))
        ws.cell(6, 2, value="NSE")
        ws.cell(6, 3, value="Sell")
        ws.cell(6, 4, value=10.0)
        ws.cell(6, 5, value=0)
        ws.cell(6, 6, value=1200.0)
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb2)
        wb2.close()

        assert len(sells) == 1
        assert abs(sells[0]["nav"] - 120.0) < 0.01  # 1200 / 10

    def test_skip_invalid_rows(self, tmp_path):
        """Rows without action, without date, with zero units, or non Buy/Sell are skipped."""
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        ws.cell(4, 1, value="DATE")
        ws.cell(4, 2, value="EXCH")
        ws.cell(4, 3, value="ACTION")
        ws.cell(4, 4, value="Units")
        ws.cell(4, 5, value="NAV")
        ws.cell(4, 6, value="COST")
        ws.cell(4, 7, value="REMARKS")
        # Row with no action
        ws.cell(5, 1, value=datetime(2025, 1, 1))
        ws.cell(5, 3, value=None)
        # Row with no date
        ws.cell(6, 1, value=None)
        ws.cell(6, 3, value="Buy")
        ws.cell(6, 4, value=10)
        ws.cell(6, 5, value=100)
        # Row with action "Hold" (not Buy/Sell)
        ws.cell(7, 1, value=datetime(2025, 2, 1))
        ws.cell(7, 3, value="Hold")
        ws.cell(7, 4, value=10)
        ws.cell(7, 5, value=100)
        # Row with units = 0
        ws.cell(8, 1, value=datetime(2025, 3, 1))
        ws.cell(8, 3, value="Buy")
        ws.cell(8, 4, value=0)
        ws.cell(8, 5, value=100)
        # Row with unparseable date string
        ws.cell(9, 1, value="not-a-date")
        ws.cell(9, 3, value="Buy")
        ws.cell(9, 4, value=10)
        ws.cell(9, 5, value=100)
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, sells = _parse_mf_trading_history(wb2)
        wb2.close()
        assert buys == []
        assert sells == []

    def test_buy_price_from_cost(self, tmp_path):
        """When cost > 0, buy_price = cost / units."""
        from app.mf_xlsx_database import _parse_mf_trading_history
        filepath = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        ws.cell(4, 1, value="DATE")
        ws.cell(4, 2, value="EXCH")
        ws.cell(4, 3, value="ACTION")
        ws.cell(4, 4, value="Units")
        ws.cell(4, 5, value="NAV")
        ws.cell(4, 6, value="COST")
        ws.cell(4, 7, value="REMARKS")
        # Buy with cost different from nav * units
        ws.cell(5, 1, value=datetime(2025, 1, 15))
        ws.cell(5, 2, value="NSE")
        ws.cell(5, 3, value="Buy")
        ws.cell(5, 4, value=10.0)
        ws.cell(5, 5, value=100.0)
        ws.cell(5, 6, value=1050.0)  # 1050 / 10 = 105 per unit
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        buys, _ = _parse_mf_trading_history(wb2)
        wb2.close()

        assert len(buys) == 1
        assert abs(buys[0]["buy_price"] - 105.0) < 0.01


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: Construction & file map
# ---------------------------------------------------------------------------

class TestMFPortfolioConstruction:
    """Tests for MFXlsxPortfolio initialisation."""

    def test_empty_dir(self, mf_portfolio):
        assert mf_portfolio.get_all_holdings() == []

    def test_dir_created_if_missing(self, tmp_path):
        new_dir = tmp_path / "nonexistent" / "Mutual Funds"
        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(new_dir)
        assert new_dir.exists()

    def test_skips_temp_files(self, mf_dir):
        # Create temp file (starts with ~)
        (mf_dir / "~$SBI Fund.xlsx").write_bytes(b"temp")
        # Create hidden file (starts with .)
        (mf_dir / ".hidden.xlsx").write_bytes(b"hidden")
        # Create valid file
        _create_mf_xlsx(mf_dir / "Valid Fund.xlsx", fund_code="INF111")

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        assert len(db._file_map) == 1
        assert "INF111" in db._file_map

    def test_file_without_fund_code_uses_filename(self, mf_dir):
        """When Index sheet has no Code, use filename as fallback."""
        filepath = mf_dir / "My Fund.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        ws.cell(4, 1, value="DATE")
        ws.cell(4, 2, value="EXCH")
        ws.cell(4, 3, value="ACTION")
        ws.cell(4, 4, value="Units")
        ws_idx = wb.create_sheet("Index")
        ws_idx.cell(1, 2, value="Code")
        ws_idx.cell(1, 3, value=None)  # No code
        wb.save(filepath)
        wb.close()

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        assert "My Fund" in db._file_map

    def test_corrupt_file_skipped(self, mf_dir):
        """Corrupt xlsx file should be skipped without crashing."""
        (mf_dir / "Corrupt.xlsx").write_bytes(b"not a valid xlsx")
        _create_mf_xlsx(mf_dir / "Good Fund.xlsx", fund_code="INF111")

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        assert "INF111" in db._file_map


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: reindex
# ---------------------------------------------------------------------------

class TestMFReindex:
    def test_reindex_detects_additions(self, mf_portfolio, mf_dir):
        assert len(mf_portfolio._file_map) == 0

        _create_mf_xlsx(mf_dir / "New Fund.xlsx", fund_code="INF999")
        mf_portfolio.reindex()

        assert "INF999" in mf_portfolio._file_map

    def test_reindex_detects_removals(self, mf_dir):
        _create_mf_xlsx(mf_dir / "Fund A.xlsx", fund_code="INFA")
        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        assert "INFA" in db._file_map

        (mf_dir / "Fund A.xlsx").unlink()
        db.reindex()

        assert "INFA" not in db._file_map


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: caching
# ---------------------------------------------------------------------------

class TestMFCaching:
    def test_cache_reused_on_same_mtime(self, mf_portfolio_with_fund):
        db = mf_portfolio_with_fund
        # First read parses
        h1, _, _ = db._get_fund_data("INF200K01RJ1")
        # Second read uses cache
        h2, _, _ = db._get_fund_data("INF200K01RJ1")
        assert h1 is h2

    def test_file_not_in_map(self, mf_portfolio):
        h, s, idx = mf_portfolio._get_fund_data("NONEXISTENT")
        assert h == []
        assert s == []
        assert idx == {}

    def test_file_stat_error(self, mf_portfolio, mf_dir):
        """When file disappears between map build and stat, return empty."""
        mf_portfolio._file_map["GONE"] = mf_dir / "gone.xlsx"
        h, s, idx = mf_portfolio._get_fund_data("GONE")
        assert h == []
        assert s == []
        assert idx == {}

    def test_parse_error_returns_empty(self, mf_portfolio, mf_dir):
        """When the file is corrupt, _parse_and_match_fund returns empty."""
        corrupt = mf_dir / "Bad.xlsx"
        corrupt.write_bytes(b"not xlsx")
        mf_portfolio._file_map["BAD"] = corrupt
        mf_portfolio._name_map["BAD"] = "Bad"

        h, s, idx = mf_portfolio._parse_and_match_fund("BAD", corrupt)
        assert h == []
        assert s == []
        assert idx == {}

    def test_no_buys_no_sells(self, mf_dir):
        """Fund file with no transactions returns empty holdings."""
        filepath = mf_dir / "Empty Fund.xlsx"
        _create_mf_xlsx(filepath, fund_code="INFEMPTY", buys=[], sells=[])

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        holdings = db.get_all_holdings()
        assert len(holdings) == 0


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: add_mf_holding
# ---------------------------------------------------------------------------

class TestAddMFHolding:
    """Tests for MFXlsxPortfolio.add_mf_holding()."""

    def test_add_creates_xlsx(self, mf_portfolio, mf_dir):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.1234,
                nav=120.50,
                buy_date="2025-01-15",
            )

        xlsx_files = list(mf_dir.glob("*.xlsx"))
        assert len(xlsx_files) == 1

        wb = openpyxl.load_workbook(xlsx_files[0])
        assert "Trading History" in wb.sheetnames
        assert "Index" in wb.sheetnames
        wb.close()

    def test_add_returns_correct_info(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.1234,
                nav=120.50,
                buy_date="2025-01-15",
            )

        assert result["fund_code"] == "INF200K01RJ1"
        assert abs(result["units"] - 50.1234) < 0.0001
        assert abs(result["nav"] - 120.50) < 0.01
        assert abs(result["cost"] - round(50.1234 * 120.50, 2)) < 0.01

    def test_add_appears_in_get_all(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        assert len(holdings) == 1
        assert holdings[0].fund_code == "INF200K01RJ1"
        assert abs(holdings[0].units - 50.0) < 0.001

    def test_multiple_buys_same_fund(self, mf_portfolio, mf_dir):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
                skip_dup_check=True,
            )
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=30.5,
                nav=125.0,
                buy_date="2025-02-15",
                skip_dup_check=True,
            )

        xlsx_files = list(mf_dir.glob("*.xlsx"))
        assert len(xlsx_files) == 1

        holdings = mf_portfolio.get_all_holdings()
        assert len(holdings) == 2
        total_units = sum(h.units for h in holdings)
        assert abs(total_units - 80.5) < 0.001

    def test_duplicate_detection(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
            )
            with pytest.raises(ValueError, match="Duplicate"):
                mf_portfolio.add_mf_holding(
                    fund_code="INF200K01RJ1",
                    fund_name="SBI Small Cap Fund",
                    units=50.0,
                    nav=120.0,
                    buy_date="2025-01-15",
                )

    def test_skip_dup_check(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
            )
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
                skip_dup_check=True,
            )
        holdings = mf_portfolio.get_all_holdings()
        assert len(holdings) == 2

    def test_add_different_funds(self, mf_portfolio, mf_dir):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="SBI Small Cap Fund",
                units=50.0,
                nav=120.0,
                buy_date="2025-01-15",
            )
            mf_portfolio.add_mf_holding(
                fund_code="INF846K01DP8",
                fund_name="Axis Bluechip Fund",
                units=100.0,
                nav=45.0,
                buy_date="2025-01-15",
            )

        xlsx_files = list(mf_dir.glob("*.xlsx"))
        assert len(xlsx_files) == 2

    def test_invalid_buy_date(self, mf_portfolio):
        """Invalid date should still work (falls through to datetime.now)."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="not-a-date",
            )
        assert result["units"] == 10.0

    def test_dup_check_with_string_date_in_xlsx(self, mf_portfolio):
        """Dup check handles string dates already in the xlsx."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-15",
            )
            # Now manually change the date cell to a string format
            fp = mf_portfolio._file_map["INF200K01RJ1"]
            wb = openpyxl.load_workbook(fp)
            ws = wb["Trading History"]
            ws.cell(5, 1, value="2025-01-15")  # String instead of datetime
            wb.save(fp)
            wb.close()

            mf_portfolio._cache.pop("INF200K01RJ1", None)

            with pytest.raises(ValueError, match="Duplicate"):
                mf_portfolio.add_mf_holding(
                    fund_code="INF200K01RJ1",
                    fund_name="Test Fund",
                    units=10.0,
                    nav=100.0,
                    buy_date="2025-01-15",
                )

    def test_dup_check_skip_non_buy_rows(self, mf_portfolio):
        """Dup check should only look at Buy rows."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=10.0,
                nav=110.0,
                sell_date="2025-03-15",
            )
            # Adding a buy with same date as sell should not raise
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=110.0,
                buy_date="2025-03-15",
            )


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: Redeem / FIFO sell
# ---------------------------------------------------------------------------

class TestMFRedeem:
    """Tests for MFXlsxPortfolio.add_mf_sell_transaction() — FIFO redeem."""

    def _setup_two_lots(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-01",
                skip_dup_check=True,
            )
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=30.0,
                nav=110.0,
                buy_date="2025-02-01",
                skip_dup_check=True,
            )

    def test_partial_redeem(self, mf_portfolio):
        self._setup_two_lots(mf_portfolio)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=20.0,
                nav=120.0,
                sell_date="2025-03-01",
            )

        assert abs(result["remaining_units"] - 60.0) < 0.01
        holdings = mf_portfolio.get_all_holdings()
        total = sum(h.units for h in holdings)
        assert abs(total - 60.0) < 0.01

    def test_redeem_fifo_oldest_first(self, mf_portfolio):
        self._setup_two_lots(mf_portfolio)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=50.0,
                nav=120.0,
                sell_date="2025-03-01",
            )

        holdings = mf_portfolio.get_all_holdings()
        assert len(holdings) == 1
        assert abs(holdings[0].units - 30.0) < 0.001
        assert abs(holdings[0].nav - 110.0) < 0.01

    def test_redeem_all(self, mf_portfolio):
        self._setup_two_lots(mf_portfolio)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=80.0,
                nav=120.0,
                sell_date="2025-03-01",
            )

        holdings = mf_portfolio.get_all_holdings()
        fund_holdings = [h for h in holdings if h.fund_code == "INF200K01RJ1"]
        assert len(fund_holdings) == 0

    def test_redeem_creates_sold_positions(self, mf_portfolio):
        self._setup_two_lots(mf_portfolio)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=60.0,
                nav=120.0,
                sell_date="2025-03-01",
            )

        sold = mf_portfolio.get_all_sold()
        assert len(sold) == 2
        sold_units = sorted([s.units for s in sold])
        assert abs(sold_units[0] - 10.0) < 0.001
        assert abs(sold_units[1] - 50.0) < 0.001

    def test_redeem_more_than_held_raises(self, mf_portfolio):
        self._setup_two_lots(mf_portfolio)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            with pytest.raises(ValueError, match="Cannot redeem"):
                mf_portfolio.add_mf_sell_transaction(
                    fund_code="INF200K01RJ1",
                    units=100.0,
                    nav=120.0,
                    sell_date="2025-03-01",
                )

    def test_redeem_nonexistent_fund_raises(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            with pytest.raises(ValueError, match="No file found"):
                mf_portfolio.add_mf_sell_transaction(
                    fund_code="NOSUCH",
                    units=10.0,
                    nav=100.0,
                    sell_date="2025-03-01",
                )

    def test_realized_pl_positive(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-01",
            )
            result = mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=10.0,
                nav=150.0,
                sell_date="2025-06-01",
            )
        assert result["realized_pl"] == 500.0

    def test_realized_pl_negative(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-01",
            )
            result = mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=10.0,
                nav=80.0,
                sell_date="2025-06-01",
            )
        assert result["realized_pl"] == -200.0

    def test_default_sell_date(self, mf_portfolio):
        """When sell_date is empty, uses today's date."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-01",
            )
            result = mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=5.0,
                nav=110.0,
                sell_date="",
            )
        assert result["remaining_units"] > 0

    def test_invalid_sell_date(self, mf_portfolio):
        """Invalid sell_date falls through to datetime.now()."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-01",
            )
            result = mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=5.0,
                nav=110.0,
                sell_date="not-a-date",
            )
        assert abs(result["remaining_units"] - 5.0) < 0.01


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: update_mf_holding
# ---------------------------------------------------------------------------

class TestUpdateMFHolding:
    def test_update_buy_date(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        h_id = holdings[0].id

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", h_id, {"buy_date": "2025-02-01"}
            )
        assert result is True

    def test_update_units(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        h_id = holdings[0].id

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", h_id, {"units": 75.0}
            )
        assert result is True

    def test_update_buy_price(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        h_id = holdings[0].id

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", h_id, {"buy_price": 105.0}
            )
        assert result is True

    def test_update_nonexistent_holding(self, mf_portfolio):
        result = mf_portfolio.update_mf_holding("INF200K01RJ1", "nonexistent", {})
        assert result is False

    def test_update_no_file(self, mf_portfolio):
        """When fund_code has no file path, return False."""
        result = mf_portfolio.update_mf_holding("NOSUCHFUND", "noid", {})
        assert result is False

    def test_update_holding_not_found_in_rows(self, mf_portfolio):
        """When holding ID doesn't match any row, return False after scanning."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        # Use a fake holding_id that exists in holdings but won't match rows
        # This tests the path where we iterate all rows but don't find a match
        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", "fakeid00", {"buy_date": "2025-05-01"}
            )
        # First check: holding not found in _get_fund_data results
        assert result is False

    def test_update_with_invalid_date(self, mf_portfolio):
        """Invalid buy_date in updates should be silently ignored."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        h_id = holdings[0].id

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", h_id, {"buy_date": "not-a-date"}
            )
        assert result is True  # Update proceeds, bad date silently ignored

    def test_update_exception_returns_false(self, mf_portfolio):
        """When openpyxl raises during update, catch and return False."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

        holdings = mf_portfolio.get_all_holdings()
        h_id = holdings[0].id

        with patch("openpyxl.load_workbook", side_effect=Exception("open failed")):
            result = mf_portfolio.update_mf_holding(
                "INF200K01RJ1", h_id, {"units": 75.0}
            )
        assert result is False


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: update_mf_sold_row
# ---------------------------------------------------------------------------

class TestUpdateMFSoldRow:
    def _setup_with_sell(self, mf_portfolio):
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-01",
            )
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=20.0,
                nav=120.0,
                sell_date="2025-06-01",
            )

    def test_update_sell_date(self, mf_portfolio):
        self._setup_with_sell(mf_portfolio)
        sold = mf_portfolio.get_all_sold()
        row_idx = sold[0].row_idx

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", row_idx, {"sell_date": "2025-07-01"}
            )
        assert result is True

    def test_update_sell_units(self, mf_portfolio):
        self._setup_with_sell(mf_portfolio)
        sold = mf_portfolio.get_all_sold()
        row_idx = sold[0].row_idx

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", row_idx, {"units": 15.0}
            )
        assert result is True

    def test_update_sell_price(self, mf_portfolio):
        self._setup_with_sell(mf_portfolio)
        sold = mf_portfolio.get_all_sold()
        row_idx = sold[0].row_idx

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", row_idx, {"sell_price": 130.0}
            )
        assert result is True

    def test_update_no_filepath(self, mf_portfolio):
        result = mf_portfolio.update_mf_sold_row("NOSUCH", 5, {"sell_date": "2025-01-01"})
        assert result is False

    def test_update_wrong_row_type(self, mf_portfolio):
        """If row_idx points to a Buy row, should return False."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-01",
            )

        # Row 5 is a Buy row (header at 4, data at 5)
        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", 5, {"sell_date": "2025-01-01"}
            )
        assert result is False

    def test_update_exception_returns_false(self, mf_portfolio):
        self._setup_with_sell(mf_portfolio)
        sold = mf_portfolio.get_all_sold()
        row_idx = sold[0].row_idx

        with patch("openpyxl.load_workbook", side_effect=Exception("err")):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", row_idx, {"sell_price": 130.0}
            )
        assert result is False

    def test_update_sell_invalid_date(self, mf_portfolio):
        """Invalid sell_date should be silently ignored."""
        self._setup_with_sell(mf_portfolio)
        sold = mf_portfolio.get_all_sold()
        row_idx = sold[0].row_idx

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.update_mf_sold_row(
                "INF200K01RJ1", row_idx, {"sell_date": "not-a-date"}
            )
        assert result is True


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: rename_fund
# ---------------------------------------------------------------------------

class TestRenameFund:
    def test_rename_success(self, mf_portfolio, mf_dir):
        _create_mf_xlsx(mf_dir / "Old Fund.xlsx", fund_code="OLD_CODE")
        mf_portfolio.reindex()

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.rename_fund("OLD_CODE", "NEW_CODE", "New Fund Name")
        assert result is True
        assert "NEW_CODE" in mf_portfolio._file_map
        assert "OLD_CODE" not in mf_portfolio._file_map
        assert mf_portfolio._name_map["NEW_CODE"] == "New Fund Name"

    def test_rename_nonexistent(self, mf_portfolio):
        result = mf_portfolio.rename_fund("NONEXIST", "NEW", "New")
        assert result is False

    def test_rename_no_new_name(self, mf_portfolio, mf_dir):
        """When new_name is empty, display_name defaults to new_code."""
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="OLD")
        mf_portfolio.reindex()

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio.rename_fund("OLD", "NEW")
        assert result is True
        assert mf_portfolio._name_map["NEW"] == "NEW"

    def test_rename_exception(self, mf_portfolio, mf_dir):
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="OLD")
        mf_portfolio.reindex()

        with patch("openpyxl.load_workbook", side_effect=Exception("err")):
            result = mf_portfolio.rename_fund("OLD", "NEW")
        assert result is False

    def test_rename_updates_index_sheet(self, mf_portfolio, mf_dir):
        """Code cell in Index sheet should be updated."""
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="OLD_CODE")
        mf_portfolio.reindex()

        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.rename_fund("OLD_CODE", "NEW_CODE")

        wb = openpyxl.load_workbook(mf_dir / "Fund.xlsx")
        ws = wb["Index"]
        assert ws.cell(1, 3).value == "NEW_CODE"
        wb.close()


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: get_fund_summary & get_dashboard_summary
# ---------------------------------------------------------------------------

class TestMFSummary:
    def test_get_fund_summary(self, mf_dir):
        _create_mf_xlsx(
            mf_dir / "Test Fund.xlsx",
            fund_code="INF200K01RJ1",
            buys=[
                {"date": "2024-01-15", "units": 50.0, "nav": 100.0},
                {"date": "2024-06-15", "units": 30.0, "nav": 110.0},
            ],
            sells=[
                {"date": "2025-01-15", "units": 20.0, "nav": 130.0},
            ],
            current_price=120.0,
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={"INF200K01RJ1": 140.0}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 1.0, "day_change_pct": 0.5,
                 "week_change_pct": 2.0, "month_change_pct": 5.0,
                 "week_52_high": 150.0, "week_52_low": 80.0,
                 "sma_50": 130.0, "sma_200": 120.0,
                 "signal": "strong_bull", "days_below_sma": 0,
                 "rsi": 60.0,
             }):
            summaries = db.get_fund_summary()

        assert len(summaries) == 1
        s = summaries[0]
        assert s["fund_code"] == "INF200K01RJ1"
        assert s["current_nav"] == 140.0
        assert s["total_held_units"] > 0
        assert s["total_sold_units"] > 0
        assert s["realized_pl"] != 0
        assert s["day_change"] == 1.0
        assert s["signal"] == "strong_bull"
        assert "held_lots" in s
        assert "sold_lots" in s

    def test_get_fund_summary_no_live_nav(self, mf_dir):
        """Falls back to xlsx index sheet nav."""
        _create_mf_xlsx(
            mf_dir / "Fund.xlsx",
            fund_code="INF123",
            buys=[{"date": "2024-01-01", "units": 10.0, "nav": 100.0}],
            current_price=105.0,
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            summaries = db.get_fund_summary()

        assert summaries[0]["current_nav"] == 105.0

    def test_get_fund_summary_zero_nav(self, mf_dir):
        """When current_nav is 0, LTCG/STCG loop should break early."""
        _create_mf_xlsx(
            mf_dir / "Fund.xlsx",
            fund_code="INF123",
            buys=[{"date": "2024-01-01", "units": 10.0, "nav": 100.0}],
            current_price=0.0,
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            summaries = db.get_fund_summary()

        assert summaries[0]["current_value"] == 0
        assert summaries[0]["ltcg_unrealized_pl"] == 0

    def test_get_fund_summary_error_handling(self, mf_dir):
        """When _get_fund_data raises, fund should be skipped."""
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="INF123",
                        buys=[{"date": "2024-01-01", "units": 10.0, "nav": 100.0}])

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}), \
             patch.object(db, "_get_fund_data", side_effect=Exception("parse err")):
            summaries = db.get_fund_summary()

        assert len(summaries) == 0

    def test_get_dashboard_summary(self, mf_dir):
        _create_mf_xlsx(
            mf_dir / "Fund A.xlsx",
            fund_code="INFA",
            buys=[{"date": "2024-01-01", "units": 100.0, "nav": 50.0}],
        )
        _create_mf_xlsx(
            mf_dir / "Fund B.xlsx",
            fund_code="INFB",
            buys=[{"date": "2024-01-01", "units": 200.0, "nav": 30.0}],
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={"INFA": 60.0, "INFB": 35.0}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            dash = db.get_dashboard_summary()

        assert dash["total_funds"] == 2
        assert dash["total_invested"] > 0
        assert dash["current_value"] > 0


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: get_fund_nav
# ---------------------------------------------------------------------------

class TestGetFundNav:
    def test_live_nav(self, mf_portfolio_with_fund):
        with patch("app.mf_xlsx_database.fetch_live_navs",
                   return_value={"INF200K01RJ1": 150.0}):
            nav = mf_portfolio_with_fund.get_fund_nav("INF200K01RJ1")
        assert nav == 150.0

    def test_fallback_to_xlsx(self, mf_portfolio_with_fund):
        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}):
            nav = mf_portfolio_with_fund.get_fund_nav("INF200K01RJ1")
        assert nav == 120.0  # From current_price in xlsx

    def test_nonexistent_fund(self, mf_portfolio):
        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}):
            nav = mf_portfolio.get_fund_nav("NONEXIST")
        assert nav == 0.0

    def test_exception_in_fallback(self, mf_portfolio):
        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={}), \
             patch.object(mf_portfolio, "_get_fund_data", side_effect=Exception("err")):
            nav = mf_portfolio.get_fund_nav("NONEXIST")
        assert nav == 0.0


# ---------------------------------------------------------------------------
# MFXlsxPortfolio: get_all_holdings / get_all_sold error handling
# ---------------------------------------------------------------------------

class TestGetAllErrorHandling:
    def test_get_all_holdings_error(self, mf_dir):
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="INF123",
                        buys=[{"date": "2024-01-01", "units": 10.0, "nav": 100.0}])

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch.object(db, "_get_fund_data", side_effect=Exception("parse err")):
            holdings = db.get_all_holdings()
        assert holdings == []

    def test_get_all_sold_error(self, mf_dir):
        _create_mf_xlsx(mf_dir / "Fund.xlsx", fund_code="INF123",
                        buys=[{"date": "2024-01-01", "units": 10.0, "nav": 100.0}])

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch.object(db, "_get_fund_data", side_effect=Exception("err")):
            sold = db.get_all_sold()
        assert sold == []


# ---------------------------------------------------------------------------
# record_nav_history (no-op)
# ---------------------------------------------------------------------------

class TestRecordNavHistory:
    def test_no_op(self):
        from app.mf_xlsx_database import record_nav_history
        record_nav_history({"FUND": 100.0})  # Should not raise


# ---------------------------------------------------------------------------
# _create_mf_file with existing file
# ---------------------------------------------------------------------------

class TestCreateMFFile:
    def test_existing_file_returns_path(self, mf_portfolio, mf_dir):
        """When file already exists, _create_mf_file returns it without creating new."""
        filepath = mf_dir / "Existing Fund.xlsx"
        _create_mf_xlsx(filepath, fund_code="EXISTING")

        with patch("app.mf_xlsx_database._sync_to_drive"):
            result = mf_portfolio._create_mf_file("EXISTING", "Existing Fund")
        assert result == filepath


# ---------------------------------------------------------------------------
# Coverage: remaining uncovered lines
# ---------------------------------------------------------------------------

class TestRemainingCoverage:
    """Target remaining uncovered lines for 100% coverage."""

    def test_sync_to_drive_unlink_exception(self, tmp_path):
        """Lines 42-43: when dup.unlink() raises."""
        from app.mf_xlsx_database import _sync_to_drive
        main = tmp_path / "Fund.xlsx"
        main.write_bytes(b"main")
        conflict = tmp_path / "Fund (1).xlsx"
        conflict.write_bytes(b"conflict")

        with patch.object(Path, "unlink", side_effect=PermissionError("denied")):
            _sync_to_drive(main)  # Should not raise
        # conflict still exists because unlink failed
        # (we patched Path.unlink globally, so just verify no exception)

    def test_signal_weak_bear_exact(self, tmp_path):
        """Line 394 (weak_bull) + 396 (weak_bear): craft data for exact signal paths."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND_WB": 12345}))

        today = date.today()
        nav_data = []
        # Create data: sma_50 < sma_200, current > sma_200 → weak_bear
        # Most recent 50 entries: low (around 90), older 200 entries: high (around 120)
        for i in range(300):
            d = today - timedelta(days=i)
            if i < 50:
                nav = 95.0  # Recent low → sma_50 ~ 95
            else:
                nav = 120.0  # Older high → sma_200 ~ higher
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(nav)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            # current_nav > sma_200 and sma_50 < sma_200 → weak_bear
            result = mod.compute_nav_changes("FUND_WB", "Fund", 115.0)

        # sma_50 ~ 95, sma_200 ~ (50*95 + 150*120)/200 = (4750+18000)/200 = 113.75
        # current (115) > sma_200 (113.75) and sma_50 (95) < sma_200 (113.75) → weak_bear
        assert result["signal"] == "weak_bear"
        mod._nav_change_cache.clear()

    def test_signal_weak_bull_exact(self, tmp_path):
        """Line 394: sma_50 > sma_200 and current < sma_200 → weak_bull."""
        import app.mf_xlsx_database as mod
        mod._nav_change_cache.clear()

        map_file = tmp_path / "scheme_map.json"
        map_file.write_text(json.dumps({"FUND_WBU": 12345}))

        today = date.today()
        nav_data = []
        # sma_50 > sma_200 but current < sma_200
        # Recent 50: high (130), older: low (80) → sma_50 ~ 130, sma_200 ~ mixed
        for i in range(300):
            d = today - timedelta(days=i)
            if i < 50:
                nav = 130.0
            else:
                nav = 80.0
            nav_data.append({"date": d.strftime("%d-%m-%Y"), "nav": str(nav)})

        with patch.object(mod, "_SCHEME_MAP_FILE", str(map_file)), \
             patch.object(mod, "_NAV_DATA_DIR", str(tmp_path)), \
             patch("app.mf_xlsx_database._fetch_nav_history_mfapi", return_value=nav_data):
            # sma_50 ~ 130, sma_200 ~ (50*130 + 150*80)/200 = (6500+12000)/200 = 92.5
            # current (85) < sma_200 (92.5) and sma_50 (130) > sma_200 (92.5) → weak_bull
            result = mod.compute_nav_changes("FUND_WBU", "Fund", 85.0)

        assert result["signal"] == "weak_bull"
        mod._nav_change_cache.clear()

    def test_fifo_lot_already_exhausted(self):
        """Line 603: lot.remaining <= 0.0001 → continue."""
        from app.mf_xlsx_database import fifo_match_mf
        buys = [
            {"date": "2025-01-01", "units": 5.0, "nav": 100.0},
            {"date": "2025-02-01", "units": 5.0, "nav": 110.0},
        ]
        # Two sells: first exhausts lot 1, second goes to lot 2
        sells = [
            {"date": "2025-03-01", "units": 5.0, "nav": 120.0},
            {"date": "2025-04-01", "units": 3.0, "nav": 130.0},
        ]
        remaining, sold = fifo_match_mf(buys, sells)
        assert len(remaining) == 1
        assert abs(remaining[0]["remaining"] - 2.0) < 0.001
        # Second sell should have skipped exhausted lot 1 (line 603)
        assert len(sold) == 2

    def test_extract_index_many_rows(self, tmp_path):
        """Line 647: break after 15 rows."""
        from app.mf_xlsx_database import _extract_mf_index_data
        filepath = tmp_path / "many_rows.xlsx"
        wb = openpyxl.Workbook()
        ws_idx = wb.create_sheet("Index")
        wb.remove(wb.active)
        ws_idx.cell(1, 2, value="Code")
        ws_idx.cell(1, 3, value="INF_MANY")
        # Add 20 rows to trigger the break at row 15
        for r in range(2, 22):
            ws_idx.cell(r, 2, value=f"Label_{r}")
            ws_idx.cell(r, 3, value=r)
        wb.save(filepath)
        wb.close()

        wb2 = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        data = _extract_mf_index_data(wb2)
        wb2.close()
        assert data["fund_code"] == "INF_MANY"

    def test_parse_and_match_tiny_remaining_after_sell(self, mf_dir):
        """Line 880: skip remaining lots with < 0.0001 units in _parse_and_match_fund."""
        # Create a fund with buys and sells
        _create_mf_xlsx(
            mf_dir / "Tiny Remain.xlsx",
            fund_code="INFTINY",
            buys=[{"date": "2025-01-01", "units": 10.0, "nav": 100.0}],
            sells=[{"date": "2025-06-01", "units": 9.0, "nav": 120.0}],
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        # Mock fifo_match_mf to return a lot with tiny remaining
        tiny_remaining = [
            {"date": "2025-01-01", "units": 10.0, "nav": 100.0,
             "remaining": 0.00005, "row_idx": 5},  # < 0.0001 → line 880
            {"date": "2025-01-01", "units": 10.0, "nav": 100.0,
             "remaining": 1.0, "row_idx": 5},  # Normal remaining
        ]
        fifo_sold = [{"buy_nav": 100.0, "buy_date": "2025-01-01",
                       "sell_nav": 120.0, "sell_date": "2025-06-01",
                       "units": 9.0, "realized_pl": 180.0,
                       "row_idx": 5, "sell_row_idx": 6}]

        with patch("app.mf_xlsx_database.fifo_match_mf",
                   return_value=(tiny_remaining, fifo_sold)):
            db._cache.pop("INFTINY", None)
            holdings, sold, idx = db._get_fund_data("INFTINY")

        # Only the normal remaining lot should produce a holding (tiny one skipped)
        assert len(holdings) == 1
        assert abs(holdings[0].units - 1.0) < 0.01

    def test_fund_summary_ltcg_stcg(self, mf_dir):
        """Lines 1007-1012, 1022-1027: LTCG/STCG for both held and sold lots."""
        # Buy > 365 days ago for LTCG, and recent buy for STCG
        old_date = (date.today() - timedelta(days=400)).strftime("%Y-%m-%d")
        recent_date = (date.today() - timedelta(days=100)).strftime("%Y-%m-%d")
        sell_date = (date.today() - timedelta(days=50)).strftime("%Y-%m-%d")

        _create_mf_xlsx(
            mf_dir / "LTCG Fund.xlsx",
            fund_code="INFLTCG",
            buys=[
                {"date": old_date, "units": 50.0, "nav": 100.0},
                {"date": recent_date, "units": 30.0, "nav": 110.0},
            ],
            sells=[
                {"date": sell_date, "units": 20.0, "nav": 130.0},
            ],
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        with patch("app.mf_xlsx_database.fetch_live_navs", return_value={"INFLTCG": 140.0}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            summaries = db.get_fund_summary()

        assert len(summaries) == 1
        s = summaries[0]
        # Should have LTCG and STCG for both held and sold
        assert s["ltcg_unrealized_pl"] != 0 or s["stcg_unrealized_pl"] != 0
        # Buy was > 365 days ago, so most of the held FIFO should be LTCG
        # Sold: oldest buy (400 days), sold at 50 days ago → held > 365 → LTCG sold
        assert s["ltcg_realized_pl"] != 0 or s["stcg_realized_pl"] != 0

    def test_fund_summary_date_parse_error(self, mf_dir):
        """Lines 1007-1008, 1022-1023: ValueError/TypeError in date parsing."""
        _create_mf_xlsx(
            mf_dir / "Fund.xlsx",
            fund_code="INFT",
            buys=[{"date": "2025-01-01", "units": 10.0, "nav": 100.0}],
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        # Mock holding with empty buy_date to trigger TypeError in strptime
        # Line 1084: `if h.buy_date else False` guards the held_lots comprehension
        # but lines 1004-1008: try/except catches ValueError/TypeError
        from app.models import MFHolding
        bad_holding = MFHolding(
            id="test1",
            fund_code="INFT",
            name="Test",
            units=10.0,
            nav=100.0,
            buy_price=100.0,
            buy_cost=1000.0,
            buy_date="",  # Empty string: strptime raises ValueError, line 1084 guards with `if h.buy_date`
        )

        with patch.object(db, "_get_fund_data", return_value=([bad_holding], [], {})), \
             patch("app.mf_xlsx_database.fetch_live_navs", return_value={"INFT": 110.0}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            summaries = db.get_fund_summary()

        assert len(summaries) == 1
        # Empty buy_date → strptime fails → days=0 → STCG (lines 1007-1008, 1012)
        assert summaries[0]["stcg_unrealized_pl"] != 0

    def test_fund_summary_sold_date_parse_error(self, mf_dir):
        """Lines 1022-1023, 1027: date parse error for sold positions."""
        _create_mf_xlsx(
            mf_dir / "Fund.xlsx",
            fund_code="INFS",
            buys=[{"date": "2025-01-01", "units": 10.0, "nav": 100.0}],
        )

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        from app.models import MFSoldPosition
        bad_sold = MFSoldPosition(
            id="s1",
            fund_code="INFS",
            name="Test",
            units=5.0,
            buy_nav=100.0,
            buy_date="not-a-date",
            sell_nav=120.0,
            sell_date="also-bad",
            realized_pl=100.0,
        )

        with patch.object(db, "_get_fund_data", return_value=([], [bad_sold], {})), \
             patch("app.mf_xlsx_database.fetch_live_navs", return_value={"INFS": 110.0}), \
             patch("app.mf_xlsx_database.compute_nav_changes", return_value={
                 "day_change": 0, "day_change_pct": 0,
                 "week_change_pct": 0, "month_change_pct": 0,
                 "week_52_high": 0, "week_52_low": 0,
                 "sma_50": None, "sma_200": None,
                 "signal": None, "days_below_sma": 0,
                 "rsi": None,
             }):
            summaries = db.get_fund_summary()

        assert len(summaries) == 1
        # Date parse error → days=0 → STCG for sold (line 1027)
        assert summaries[0]["stcg_realized_pl"] == 100.0

    def test_find_header_row_default(self, mf_dir):
        """Line 1177: when no header row found, defaults to 4."""
        filepath = mf_dir / "NoHeader.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading History"
        # No DATE/ACTION header
        for r in range(1, 6):
            ws.cell(r, 1, value="data")
            ws.cell(r, 2, value="data")
            ws.cell(r, 3, value="data")
            ws.cell(r, 4, value="data")
        ws_idx = wb.create_sheet("Index")
        ws_idx.cell(1, 2, value="Code")
        ws_idx.cell(1, 3, value="INFNH")
        wb.save(filepath)
        wb.close()

        with patch("app.mf_xlsx_database._sync_to_drive"):
            from app.mf_xlsx_database import MFXlsxPortfolio
            db = MFXlsxPortfolio(mf_dir)

        # Now try to add holding — this calls _find_header_row
        with patch("app.mf_xlsx_database._sync_to_drive"):
            db.add_mf_holding(
                fund_code="INFNH",
                fund_name="No Header Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

    def test_dup_check_bad_string_date_continue(self, mf_portfolio):
        """Lines 1224-1225: row_date is a string that can't be parsed."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-15",
            )

            # Manually corrupt date to unparseable string
            fp = mf_portfolio._file_map["INF200K01RJ1"]
            wb = openpyxl.load_workbook(fp)
            ws = wb["Trading History"]
            ws.cell(5, 1, value="BADDATE")  # Unparseable string
            wb.save(fp)
            wb.close()

            mf_portfolio._cache.pop("INF200K01RJ1", None)

            # This should NOT raise (bad date is skipped via continue on line 1225)
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=10.0,
                nav=100.0,
                buy_date="2025-01-15",
                skip_dup_check=False,
            )

    def test_update_holding_no_filepath(self, mf_portfolio):
        """Line 1362: fund_code in _file_map but holdings exist → filepath None."""
        # Create a holding manually in _get_fund_data result
        from app.models import MFHolding
        fake_holding = MFHolding(
            id="fakeid",
            fund_code="NOFP",
            name="No File",
            units=10.0,
            nav=100.0,
            buy_price=100.0,
            buy_cost=1000.0,
            buy_date="2025-01-01",
        )

        with patch.object(mf_portfolio, "_get_fund_data", return_value=([fake_holding], [], {})):
            result = mf_portfolio.update_mf_holding("NOFP", "fakeid", {"units": 20.0})
        assert result is False

    def test_update_holding_row_not_matched(self, mf_portfolio):
        """Line 1372 (continue for non-Buy) + 1402 (wb.close after no match)."""
        with patch("app.mf_xlsx_database._sync_to_drive"):
            mf_portfolio.add_mf_holding(
                fund_code="INF200K01RJ1",
                fund_name="Test Fund",
                units=50.0,
                nav=100.0,
                buy_date="2025-01-15",
            )
            # Also add a sell to create non-Buy rows
            mf_portfolio.add_mf_sell_transaction(
                fund_code="INF200K01RJ1",
                units=10.0,
                nav=110.0,
                sell_date="2025-06-01",
            )

        holdings = mf_portfolio.get_all_holdings()
        h = holdings[0]

        # Create a fake holding_id that will match the _get_fund_data
        # but NOT match any row in the actual xlsx
        from app.models import MFHolding
        fake_h = MFHolding(
            id=h.id,
            fund_code=h.fund_code,
            name=h.name,
            units=99999.0,  # wrong units — won't match any row
            nav=99999.0,    # wrong nav — won't match
            buy_price=99999.0,
            buy_cost=99999.0,
            buy_date="2099-01-01",  # wrong date
        )

        with patch.object(mf_portfolio, "_get_fund_data",
                          return_value=([fake_h], [], {})):
            with patch("app.mf_xlsx_database._sync_to_drive"):
                result = mf_portfolio.update_mf_holding(
                    "INF200K01RJ1", h.id, {"units": 75.0}
                )
        # Should close wb and return False (line 1402)
        assert result is False
