"""Unit tests for app/nps_database.py — NPS database layer."""
import json
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def nps_base_dir(tmp_path):
    """Create a temp base dir with NPS subdir."""
    (tmp_path / "NPS").mkdir()
    return tmp_path


@pytest.fixture(autouse=True)
def _mock_drive_and_import():
    """Mock drive + skip PDF import (no PDF_IMPORT_DIR in tests)."""
    with patch("app.nps_database._sync_to_drive"), \
         patch("app.nps_database._delete_from_drive"), \
         patch("app.nps_database._imported", True):
        yield


def _get_nps_id(nps_base_dir):
    """Helper: get the actual persisted ID from get_all (may differ from add's return)."""
    from app.nps_database import get_all
    items = get_all(base_dir=str(nps_base_dir))
    assert len(items) >= 1
    return items[-1]["id"]


# ---------------------------------------------------------------------------
# Tests — empty state
# ---------------------------------------------------------------------------

def test_get_all_empty(nps_base_dir):
    from app.nps_database import get_all
    assert get_all(base_dir=str(nps_base_dir)) == []


def test_get_dashboard_empty(nps_base_dir):
    from app.nps_database import get_dashboard
    dash = get_dashboard(base_dir=str(nps_base_dir))
    assert dash["active_count"] == 0
    assert dash["total_count"] == 0
    assert dash["total_contributed"] == 0
    assert dash["current_value"] == 0


# ---------------------------------------------------------------------------
# Tests — add
# ---------------------------------------------------------------------------

def test_add_nps_account(nps_base_dir):
    from app.nps_database import add, get_all
    data = {
        "account_name": "My NPS",
        "pran": "110012345678",
        "tier": "Tier I",
        "fund_manager": "SBI Pension Fund",
        "scheme_preference": "Auto Choice",
        "start_date": "2019-01-01",
        "current_value": 500000,
        "status": "Active",
    }
    result = add(data, base_dir=str(nps_base_dir))
    assert result["pran"] == "110012345678"
    assert result["account_name"] == "My NPS"
    assert result["current_value"] == 500000

    # xlsx created
    xlsx_files = list((nps_base_dir / "NPS").glob("*.xlsx"))
    assert len(xlsx_files) == 1

    items = get_all(base_dir=str(nps_base_dir))
    assert len(items) == 1
    assert items[0]["pran"] == "110012345678"


# ---------------------------------------------------------------------------
# Tests — update
# ---------------------------------------------------------------------------

def test_update_nps(nps_base_dir):
    from app.nps_database import add, update, get_all
    add({
        "account_name": "Upd NPS",
        "pran": "110099887766",
        "start_date": "2020-01-01",
        "current_value": 100000,
    }, base_dir=str(nps_base_dir))
    # Re-read to get the actual persisted ID
    nps_id = _get_nps_id(nps_base_dir)

    updated = update(nps_id, {"current_value": 200000}, base_dir=str(nps_base_dir))
    assert updated["current_value"] == 200000


def test_update_nonexistent_raises(nps_base_dir):
    from app.nps_database import update
    with pytest.raises(ValueError, match="not found"):
        update("bad_id", {"current_value": 0}, base_dir=str(nps_base_dir))


# ---------------------------------------------------------------------------
# Tests — delete
# ---------------------------------------------------------------------------

def test_delete_nps(nps_base_dir):
    from app.nps_database import add, delete, get_all
    add({
        "account_name": "Del NPS",
        "pran": "110000000001",
        "start_date": "2020-01-01",
    }, base_dir=str(nps_base_dir))
    nps_id = _get_nps_id(nps_base_dir)

    items = get_all(base_dir=str(nps_base_dir))
    assert len(items) == 1

    del_result = delete(nps_id, base_dir=str(nps_base_dir))
    assert "deleted" in del_result["message"]

    items = get_all(base_dir=str(nps_base_dir))
    assert len(items) == 0


def test_delete_nonexistent_raises(nps_base_dir):
    from app.nps_database import delete
    with pytest.raises(ValueError, match="not found"):
        delete("bad_id", base_dir=str(nps_base_dir))


# ---------------------------------------------------------------------------
# Tests — add_contribution
# ---------------------------------------------------------------------------

def test_add_contribution(nps_base_dir):
    from app.nps_database import add, add_contribution, get_all
    add({
        "account_name": "Contrib NPS",
        "pran": "110000000002",
        "start_date": "2020-01-01",
        "current_value": 100000,
    }, base_dir=str(nps_base_dir))
    nps_id = _get_nps_id(nps_base_dir)

    updated = add_contribution(nps_id, {
        "date": "2024-06-15",
        "amount": 50000,
        "remarks": "FY 2024-25",
    }, base_dir=str(nps_base_dir))
    assert len(updated["contributions"]) == 1
    assert updated["contributions"][0]["amount"] == 50000

    # Verify on reload
    items = get_all(base_dir=str(nps_base_dir))
    assert len(items[0]["contributions"]) == 1


def test_add_contribution_invalid_date(nps_base_dir):
    from app.nps_database import add, add_contribution
    add({
        "account_name": "Date NPS",
        "pran": "110000000003",
        "start_date": "2020-01-01",
    }, base_dir=str(nps_base_dir))
    nps_id = _get_nps_id(nps_base_dir)

    with pytest.raises(ValueError, match="Invalid date format"):
        add_contribution(nps_id, {
            "date": "not-a-date",
            "amount": 1000,
        }, base_dir=str(nps_base_dir))


# ---------------------------------------------------------------------------
# Tests — enrich computes gains
# ---------------------------------------------------------------------------

def test_enrich_calculates_gains(nps_base_dir):
    from app.nps_database import add, add_contribution, get_all
    add({
        "account_name": "Gain NPS",
        "pran": "110000000004",
        "start_date": "2020-01-01",
        "current_value": 200000,
    }, base_dir=str(nps_base_dir))
    nps_id = _get_nps_id(nps_base_dir)

    add_contribution(nps_id, {
        "date": "2020-06-01",
        "amount": 100000,
    }, base_dir=str(nps_base_dir))

    items = get_all(base_dir=str(nps_base_dir))
    item = items[0]
    assert item["total_contributed"] == 100000
    assert item["gain"] == 100000  # 200000 - 100000
    assert item["gain_pct"] == 100.0


# ---------------------------------------------------------------------------
# Tests — dashboard
# ---------------------------------------------------------------------------

def test_dashboard_after_add(nps_base_dir):
    from app.nps_database import add, get_dashboard
    add({
        "account_name": "Dash NPS",
        "pran": "110000000005",
        "start_date": "2020-01-01",
        "current_value": 300000,
    }, base_dir=str(nps_base_dir))

    dash = get_dashboard(base_dir=str(nps_base_dir))
    assert dash["total_count"] == 1
    assert dash["active_count"] == 1
    assert dash["current_value"] == 300000


# ---------------------------------------------------------------------------
# Tests — helpers
# ---------------------------------------------------------------------------

def test_gen_id_deterministic():
    from app.nps_database import _gen_id
    assert _gen_id("pran123") == _gen_id("pran123")
    assert _gen_id("a") != _gen_id("b")


def test_parse_num():
    from app.nps_database import _parse_num
    assert _parse_num("1,234.56") == 1234.56
    assert _parse_num("(500.00)") == -500.0
    assert _parse_num("abc") == 0.0


def test_to_float():
    from app.nps_database import _to_float
    assert _to_float(3.14) == 3.14
    assert _to_float(None) == 0.0
    assert _to_float("bad") == 0.0


# ---------------------------------------------------------------------------
# Tests — _sync_to_drive / _delete_from_drive (lines 45, 50-62)
# ---------------------------------------------------------------------------

def test_sync_to_drive_is_noop():
    from app.nps_database import _sync_to_drive
    from pathlib import Path
    _sync_to_drive(Path("/tmp/test.xlsx"))


def test_delete_from_drive_handles_exception():
    from app.nps_database import _delete_from_drive
    from pathlib import Path
    _delete_from_drive(Path("/nonexistent/path/file.xlsx"))


# ---------------------------------------------------------------------------
# Tests — _parse_date (lines 92-97)
# ---------------------------------------------------------------------------

def test_parse_date_formats():
    from app.nps_database import _parse_date
    from datetime import date
    # DD-Mon-YYYY format
    assert _parse_date("15-Jan-2024") == date(2024, 1, 15)
    # DD-Mon-YY format
    assert _parse_date("15-Jan-24") == date(2024, 1, 15)
    # Invalid
    assert _parse_date("invalid") is None
    assert _parse_date("2024-01-15") is None


# ---------------------------------------------------------------------------
# Tests — _to_str (lines 111)
# ---------------------------------------------------------------------------

def test_to_str():
    from app.nps_database import _to_str
    from datetime import datetime, date
    assert _to_str(None) == ""
    assert _to_str(datetime(2024, 1, 15)) == "2024-01-15"
    assert _to_str(date(2024, 1, 15)) == "2024-01-15"
    assert _to_str("  hello  ") == "hello"
    assert _to_str(None, "default") == "default"


# ---------------------------------------------------------------------------
# Tests — _extract_pdf_text (lines 121-128)
# ---------------------------------------------------------------------------

def test_extract_pdf_text(tmp_path):
    """_extract_pdf_text uses pdfplumber to read PDF."""
    from app.nps_database import _extract_pdf_text
    from unittest.mock import MagicMock

    mock_page = MagicMock()
    mock_page.extract_text.return_value = "Page 1 text"
    mock_pdf = MagicMock()
    mock_pdf.pages = [mock_page]
    mock_pdf.__enter__ = lambda s: s
    mock_pdf.__exit__ = MagicMock(return_value=False)

    with patch("pdfplumber.open", return_value=mock_pdf):
        result = _extract_pdf_text("/fake/path.pdf")
    assert result == "Page 1 text"


def test_extract_pdf_text_empty_page(tmp_path):
    """Pages returning None text are skipped."""
    from app.nps_database import _extract_pdf_text

    mock_page = MagicMock()
    mock_page.extract_text.return_value = None
    mock_pdf = MagicMock()
    mock_pdf.pages = [mock_page]
    mock_pdf.__enter__ = lambda s: s
    mock_pdf.__exit__ = MagicMock(return_value=False)

    with patch("pdfplumber.open", return_value=mock_pdf):
        result = _extract_pdf_text("/fake/path.pdf")
    assert result == ""


# ---------------------------------------------------------------------------
# Tests — _parse_subscriber_info (lines 133-193)
# ---------------------------------------------------------------------------

def test_parse_subscriber_info():
    from app.nps_database import _parse_subscriber_info
    text = """
PRAN 110012345678
Registration Date 01-Jan-2019
Subscriber Name John Doe Tier I
Tier I Status Active
Scheme Choice - Auto Choice (LC50)
SBI PENSION FUND SCHEME E - TIER I
SCHEME E - TIER I 50%
SCHEME C - TIER I 30%
SCHEME G - TIER I 20%
Nominee Name/s Percentage
Jane Doe 100%
₹ 5,00,000.50 100 ₹ 3,00,000.25
    """
    info = _parse_subscriber_info(text)
    assert info["pran"] == "110012345678"
    assert info["registration_date"] == "2019-01-01"
    assert info["subscriber_name"] == "John Doe"
    assert info["status"] == "Active"
    assert info["scheme_preference"] == "Auto Choice"
    assert info["fund_manager"] == "SBI Pension Fund"
    assert info.get("nominee") == "Jane Doe (100%)"
    assert info["holdings_value"] == 500000.50
    assert info["total_contribution"] == 300000.25


def test_parse_subscriber_info_scheme_choice_multiline():
    from app.nps_database import _parse_subscriber_info
    text = "Scheme Choice - Active Choice\nSome other text"
    info = _parse_subscriber_info(text)
    assert info["scheme_preference"] == "Active Choice"


def test_parse_subscriber_info_minimal():
    from app.nps_database import _parse_subscriber_info
    info = _parse_subscriber_info("No useful info here")
    assert info == {}


# ---------------------------------------------------------------------------
# Tests — _parse_scheme_transactions (lines 201-311)
# ---------------------------------------------------------------------------

def test_parse_scheme_transactions():
    from app.nps_database import _parse_scheme_transactions
    text = """01-Jan-2024 Opening balance 100.0000
01-Feb-2024 By Contribution 5,000.00 50.0000 100.0000
01-Mar-2024 Closing Balance 200.0000"""
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 3
    # Opening balance
    assert txns[0]["type"] == "opening_balance"
    # Contribution
    assert txns[1]["type"] == "contribution"
    assert txns[1]["amount"] == 5000.0
    # Closing balance
    assert txns[2]["type"] == "closing_balance"


def test_parse_scheme_transactions_with_billing():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 Billing Charges 100.00 50.0000 2.0000"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1
    assert txns[0]["type"] == "billing"


def test_parse_scheme_transactions_switch():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 Switch out from E (1,000.00) 50.0000 (20.0000)\n01-Feb-2024 Switch In to C 1,000.00 25.0000 40.0000"
    txns = _parse_scheme_transactions(text)
    assert any(t["type"] == "switch_out" for t in txns)
    assert any(t["type"] == "switch_in" for t in txns)


def test_parse_scheme_transactions_persistency():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 Persistency charge 50.00 100.0000 0.5000"
    txns = _parse_scheme_transactions(text)
    assert txns[0]["type"] == "persistency_charge"


def test_parse_scheme_transactions_trail_commission():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 Trail Commission (100.00) 50.0000 (2.0000)"
    txns = _parse_scheme_transactions(text)
    assert txns[0]["type"] == "trail_commission"


def test_parse_scheme_transactions_other_type():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 Unknown transaction 100.00 50.0000 2.0000"
    txns = _parse_scheme_transactions(text)
    assert txns[0]["type"] == "other"


def test_parse_scheme_transactions_multiline_desc():
    """Suffix lines are merged with the date line."""
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution\nthrough online portal 5,000.00 50.0000 100.0000"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1
    assert "portal" in txns[0]["description"]


def test_parse_scheme_transactions_prefix_lines():
    """Non-date lines before a date line are treated as prefix."""
    from app.nps_database import _parse_scheme_transactions
    text = "By Some prefix text\n01-Jan-2024 Contribution 5,000.00 50.0000 100.0000"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1


def test_parse_scheme_transactions_invalid_date():
    """Lines with invalid dates are skipped."""
    from app.nps_database import _parse_scheme_transactions
    text = "99-Xyz-2024 Bad date 100.00 50.0000 2.0000"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 0


# ---------------------------------------------------------------------------
# Tests — _parse_pdf (lines 323-366)
# ---------------------------------------------------------------------------

def test_parse_pdf():
    from app.nps_database import _parse_pdf
    text = """
PRAN 110012345678
Registration Date 01-Jan-2019
Subscriber Name Test User Tier I

SBI PENSION FUND SCHEME E - TIER I
Date Description Amount NAV Units
01-Jan-2024 Opening balance 100.0000
01-Feb-2024 By Contribution 5,000.00 50.0000 100.0000
01-Mar-2024 Closing Balance 200.0000

Contribution/Redemption Details
01-Feb-2024 By Contribution some text 5,000.00 50.00 5,000.00
"""
    mock_page = MagicMock()
    mock_page.extract_text.return_value = text
    mock_pdf = MagicMock()
    mock_pdf.pages = [mock_page]
    mock_pdf.__enter__ = lambda s: s
    mock_pdf.__exit__ = MagicMock(return_value=False)

    with patch("pdfplumber.open", return_value=mock_pdf):
        result = _parse_pdf("/fake/path.pdf")

    assert "subscriber_info" in result
    assert result["subscriber_info"]["pran"] == "110012345678"


# ---------------------------------------------------------------------------
# Tests — _merge_pdf_data (lines 380-443)
# ---------------------------------------------------------------------------

def test_merge_pdf_data():
    from app.nps_database import _merge_pdf_data
    pdf1 = {
        "subscriber_info": {"pran": "110012345678", "holdings_value": 100000, "status": "Active"},
        "scheme_transactions": {
            "E": [
                {"date": "2023-01-01", "description": "Contribution", "amount": 5000, "nav": 50, "units": 100, "type": "contribution"},
                {"date": "2023-06-01", "description": "Closing Balance", "amount": 0, "nav": 0, "units": 200, "type": "closing_balance"},
            ]
        },
        "contributions": [{"date": "2023-01-01", "amount": 5000, "remarks": "By Contribution"}],
    }
    pdf2 = {
        "subscriber_info": {"pran": "110012345678", "holdings_value": 200000, "total_contribution": 10000, "status": "Active"},
        "scheme_transactions": {
            "E": [
                {"date": "2023-01-01", "description": "Contribution", "amount": 5000, "nav": 50, "units": 100, "type": "contribution"},  # duplicate
                {"date": "2024-01-01", "description": "Contribution", "amount": 5000, "nav": 55, "units": 90.91, "type": "contribution"},
                {"date": "2024-06-01", "description": "Closing Balance", "amount": 0, "nav": 0, "units": 300, "type": "closing_balance"},
            ]
        },
        "contributions": [
            {"date": "2023-01-01", "amount": 5000, "remarks": "By Contribution"},  # dup
            {"date": "2024-01-01", "amount": 5000, "remarks": "By Contribution"},
        ],
    }

    merged = _merge_pdf_data([pdf1, pdf2])
    assert merged["current_value"] == 200000
    # Duplicate contributions should be deduped
    assert len(merged["contributions"]) == 2
    # Duplicate transactions should be deduped
    txns = merged["transactions"]
    assert len(txns) == 2  # 2 unique contributions (closing/opening skipped)


# ---------------------------------------------------------------------------
# Tests — _write_xlsx / _read_xlsx (lines 491-595)
# ---------------------------------------------------------------------------

def test_write_and_read_xlsx(nps_base_dir):
    from app.nps_database import _write_xlsx, _read_xlsx
    from pathlib import Path
    nps_dir = nps_base_dir / "NPS"
    filepath = nps_dir / "test.xlsx"

    account = {
        "pran": "110012345678",
        "current_value": 500000,
        "start_date": "2019-01-01",
        "account_name": "Test NPS",
        "tier": "Tier I",
        "xirr": "12.5",
        "scheme_preference": "Auto Choice",
        "fund_manager": "SBI Pension Fund",
        "status": "Active",
        "nominee": "Jane Doe (100%)",
        "remarks": "Test account",
        "scheme_splits": [{"scheme": "E", "pct": 50}],
        "contributions": [{"date": "2020-01-01", "amount": 50000, "remarks": "FY20"}],
        "schemes_summary": [{"scheme": "E", "units": 100, "nav": 50, "value": 5000}],
    }
    transactions = [
        {"date": "2020-01-01", "scheme": "E", "description": "Contribution", "amount": 50000, "nav": 50, "units": 1000},
    ]

    _write_xlsx(filepath, account, transactions)
    assert filepath.exists()

    read_back = _read_xlsx(filepath)
    assert read_back["pran"] == "110012345678"
    assert read_back["current_value"] == 500000
    assert read_back["account_name"] == "Test NPS"
    assert len(read_back["contributions"]) == 1
    assert len(read_back["_transactions"]) == 1


def test_read_xlsx_with_string_date(nps_base_dir):
    """_read_xlsx handles string dates in transaction rows."""
    from app.nps_database import _write_xlsx, _read_xlsx
    import openpyxl

    nps_dir = nps_base_dir / "NPS"
    filepath = nps_dir / "strdate.xlsx"

    # Create xlsx with string date
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPS"
    ws.cell(1, 2, "110099887766")
    ws.cell(1, 5, 100000)
    ws.cell(1, 8, "2020-01-01")
    ws.cell(1, 11, "Str Date NPS")
    ws.cell(2, 2, "Tier I")
    ws.cell(3, 2, "Active")
    ws.cell(4, 5, "[]")
    headers = ["S.No", "Date", "Scheme", "Description", "Amount", "NAV", "Units"]
    for ci, h in enumerate(headers, 1):
        ws.cell(5, ci, h)
    ws.cell(6, 1, 1)
    ws.cell(6, 2, "2020-01-15")  # string date
    ws.cell(6, 3, "E")
    ws.cell(6, 4, "Contribution")
    ws.cell(6, 5, 10000)
    ws.cell(6, 6, 50.0)
    ws.cell(6, 7, 200.0)
    wb.save(str(filepath))
    wb.close()

    read_back = _read_xlsx(filepath)
    assert len(read_back["_transactions"]) == 1
    assert read_back["_transactions"][0]["date"] == "2020-01-15"


# ---------------------------------------------------------------------------
# Tests — _enrich (lines 692-693)
# ---------------------------------------------------------------------------

def test_enrich_with_invalid_start_date():
    from app.nps_database import _enrich
    account = {
        "contributions": [{"amount": 10000}],
        "current_value": 15000,
        "start_date": "invalid-date",
        "status": "Active",
    }
    result = _enrich(account)
    assert result["years_active"] == 0
    assert result["gain"] == 5000


def test_enrich_zero_contributions():
    from app.nps_database import _enrich
    account = {
        "contributions": [],
        "current_value": 0,
        "start_date": "2020-01-01",
        "status": "Active",
    }
    result = _enrich(account)
    assert result["gain_pct"] == 0


# ---------------------------------------------------------------------------
# Tests — _load_all_xlsx error handling (lines 721-722)
# ---------------------------------------------------------------------------

def test_load_all_xlsx_skips_corrupt(nps_base_dir):
    from app.nps_database import _load_all_xlsx
    nps_dir = nps_base_dir / "NPS"
    (nps_dir / "corrupt.xlsx").write_text("not xlsx")
    items = _load_all_xlsx(nps_dir=nps_dir)
    assert len(items) == 0


# ---------------------------------------------------------------------------
# Tests — _save_account (lines 735-739)
# ---------------------------------------------------------------------------

def test_save_account_reads_existing_transactions(nps_base_dir):
    """_save_account reads existing transactions when not provided."""
    from app.nps_database import add, _save_account, _load_all_xlsx
    add({
        "account_name": "Save Test",
        "pran": "110077777777",
        "start_date": "2020-01-01",
        "current_value": 100000,
    }, base_dir=str(nps_base_dir))

    items = _load_all_xlsx(nps_dir=nps_base_dir / "NPS")
    assert len(items) == 1

    # Save without providing transactions (should read from existing xlsx)
    account = items[0]
    account["current_value"] = 200000
    _save_account(account, nps_dir=nps_base_dir / "NPS")

    items2 = _load_all_xlsx(nps_dir=nps_base_dir / "NPS")
    assert items2[0]["current_value"] == 200000


# ---------------------------------------------------------------------------
# Tests — add_contribution non-existent (lines 844, 849)
# ---------------------------------------------------------------------------

def test_add_contribution_nonexistent_raises(nps_base_dir):
    from app.nps_database import add_contribution
    with pytest.raises(ValueError, match="not found"):
        add_contribution("bad_id", {
            "date": "2024-01-01",
            "amount": 1000,
        }, base_dir=str(nps_base_dir))


# ---------------------------------------------------------------------------
# Tests — _parse_scheme_transactions suffix handling (lines 214-241)
# ---------------------------------------------------------------------------

def test_parse_scheme_transactions_empty_lines_skipped():
    """Empty lines between transactions are skipped."""
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution 5,000.00 50.0000 100.0000\n\n\n01-Feb-2024 By Contribution 5,000.00 55.0000 90.9091"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 2


def test_parse_scheme_transactions_stops_at_to_prefix():
    """Suffix collection stops at 'To ' prefix lines."""
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution 5,000.00 50.0000 100.0000\nTo Withdrawal details"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1


def test_parse_scheme_transactions_stops_at_scheme_header():
    """Suffix collection stops at PENSION FUND SCHEME header."""
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution 5,000.00 50.0000 100.0000\nSBI PENSION FUND SCHEME C - TIER I"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1


def test_parse_scheme_transactions_stops_at_notes():
    """Suffix collection stops at Notes or View More markers."""
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution 5,000.00 50.0000 100.0000\nNotes: some text here"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1


def test_parse_scheme_transactions_stops_at_view_more():
    from app.nps_database import _parse_scheme_transactions
    text = "01-Jan-2024 By Contribution 5,000.00 50.0000 100.0000\nView More details"
    txns = _parse_scheme_transactions(text)
    assert len(txns) == 1


# ---------------------------------------------------------------------------
# Tests — _write_xlsx with invalid/non-string date (lines 496-499)
# ---------------------------------------------------------------------------

def test_write_xlsx_with_invalid_date(nps_base_dir):
    """_write_xlsx handles invalid date strings by writing as-is."""
    from app.nps_database import _write_xlsx, _read_xlsx
    nps_dir = nps_base_dir / "NPS"
    filepath = nps_dir / "invdate.xlsx"

    account = {
        "pran": "110099999999",
        "current_value": 100000,
        "start_date": "2020-01-01",
        "account_name": "Inv Date NPS",
        "tier": "Tier I",
        "status": "Active",
        "scheme_splits": [],
        "contributions": [],
        "schemes_summary": [],
    }
    transactions = [
        {"date": "not-a-date", "scheme": "E", "description": "Test", "amount": 100, "nav": 50, "units": 2},
        {"date": "", "scheme": "E", "description": "Empty date", "amount": 0, "nav": 0, "units": 0},
        {"date": 12345, "scheme": "E", "description": "Int date", "amount": 0, "nav": 0, "units": 0},
    ]

    _write_xlsx(filepath, account, transactions)
    assert filepath.exists()

    read_back = _read_xlsx(filepath)
    assert len(read_back["_transactions"]) >= 1


# ---------------------------------------------------------------------------
# Tests — _read_xlsx JSON parsing edge cases (lines 557-574, 580)
# ---------------------------------------------------------------------------

def test_read_xlsx_invalid_json_in_metadata(nps_base_dir):
    """_read_xlsx handles invalid JSON in scheme splits, contributions, etc."""
    import openpyxl
    nps_dir = nps_base_dir / "NPS"
    filepath = nps_dir / "badjson.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPS"
    ws.cell(1, 2, "110088888888")
    ws.cell(1, 5, 50000)
    ws.cell(1, 8, "2020-01-01")
    ws.cell(1, 11, "Bad JSON NPS")
    ws.cell(2, 2, "Tier I")
    ws.cell(3, 2, "Active")
    # Bad JSON in H4, E4, K4
    ws.cell(4, 8, "NOT JSON")  # scheme splits
    ws.cell(4, 5, "NOT JSON")  # contributions
    ws.cell(4, 11, "NOT JSON")  # schemes summary
    headers = ["S.No", "Date", "Scheme", "Description", "Amount", "NAV", "Units"]
    for ci, h in enumerate(headers, 1):
        ws.cell(5, ci, h)
    # Row without date
    ws.cell(6, 1, 1)
    ws.cell(6, 2, None)  # no date
    wb.save(str(filepath))
    wb.close()

    from app.nps_database import _read_xlsx
    result = _read_xlsx(filepath)
    assert result["pran"] == "110088888888"
    # Bad JSON should be gracefully handled
    assert result.get("contributions", []) == []
    assert len(result["_transactions"]) == 0  # row without date is skipped


def test_read_xlsx_schemes_summary_not_array(nps_base_dir):
    """K4 that doesn't start with [ is skipped."""
    import openpyxl
    nps_dir = nps_base_dir / "NPS"
    filepath = nps_dir / "notarray.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPS"
    ws.cell(1, 2, "110077777777")
    ws.cell(1, 5, 50000)
    ws.cell(1, 8, "2020-01-01")
    ws.cell(1, 11, "No Array NPS")
    ws.cell(2, 2, "Tier I")
    ws.cell(3, 2, "Active")
    ws.cell(4, 8, "[]")
    ws.cell(4, 5, "[]")
    ws.cell(4, 11, '{"key": "not an array"}')  # Not an array
    headers = ["S.No", "Date", "Scheme", "Description", "Amount", "NAV", "Units"]
    for ci, h in enumerate(headers, 1):
        ws.cell(5, ci, h)
    wb.save(str(filepath))
    wb.close()

    from app.nps_database import _read_xlsx
    result = _read_xlsx(filepath)
    assert "schemes_summary" not in result


# ---------------------------------------------------------------------------
# Tests — _import_from_pdfs (lines 613-670)
# ---------------------------------------------------------------------------

def test_import_from_pdfs_already_imported(nps_base_dir):
    """_import_from_pdfs returns early if already imported."""
    from app.nps_database import _import_from_pdfs
    # The autouse mock sets _imported=True, so it should return early
    _import_from_pdfs(nps_dir=nps_base_dir / "NPS")


def test_import_from_pdfs_runs_when_not_imported(nps_base_dir):
    """_import_from_pdfs processes PDFs when _imported=False and PDFs exist."""
    import app.nps_database as mod

    nps_dir = nps_base_dir / "NPS"
    pdf_dir = nps_base_dir / "pdf_import"
    pdf_dir.mkdir()

    # Create a fake PDF
    (pdf_dir / "statement.pdf").write_text("fake pdf")

    mock_parsed = {
        "subscriber_info": {
            "pran": "110012345678",
            "subscriber_name": "Test User",
            "registration_date": "2019-01-01",
            "status": "Active",
            "holdings_value": 500000,
        },
        "scheme_transactions": {},
        "contributions": [{"date": "2020-01-01", "amount": 50000, "remarks": "By Contribution"}],
    }

    with patch.object(mod, "_imported", False), \
         patch.object(mod, "PDF_IMPORT_DIR", pdf_dir), \
         patch.object(mod, "_parse_pdf", return_value=mock_parsed), \
         patch.object(mod, "_sync_to_drive"):
        mod._imported = False
        _import_result = mod._import_from_pdfs(nps_dir=nps_dir)

    # xlsx should be created
    xlsx_files = list(nps_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1


def test_import_from_pdfs_skips_if_xlsx_exists(nps_base_dir):
    """_import_from_pdfs skips if xlsx files already exist."""
    import app.nps_database as mod
    nps_dir = nps_base_dir / "NPS"
    pdf_dir = nps_base_dir / "pdf_import"
    pdf_dir.mkdir()
    (pdf_dir / "statement.pdf").write_text("fake pdf")

    # Create existing xlsx
    (nps_dir / "existing.xlsx").write_text("existing")

    with patch.object(mod, "_imported", False), \
         patch.object(mod, "PDF_IMPORT_DIR", pdf_dir), \
         patch.object(mod, "_parse_pdf") as mock_parse:
        mod._imported = False
        mod._import_from_pdfs(nps_dir=nps_dir)
    # Should not parse any PDFs
    mock_parse.assert_not_called()


def test_import_from_pdfs_handles_parse_error(nps_base_dir):
    """_import_from_pdfs handles PDF parse errors gracefully."""
    import app.nps_database as mod
    nps_dir = nps_base_dir / "NPS"
    pdf_dir = nps_base_dir / "pdf_import"
    pdf_dir.mkdir()
    (pdf_dir / "bad.pdf").write_text("fake pdf")

    with patch.object(mod, "_imported", False), \
         patch.object(mod, "PDF_IMPORT_DIR", pdf_dir), \
         patch.object(mod, "_parse_pdf", side_effect=Exception("parse error")), \
         patch.object(mod, "_sync_to_drive"):
        mod._imported = False
        mod._import_from_pdfs(nps_dir=nps_dir)

    # No xlsx created since parsing failed
    xlsx_files = list(nps_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 0


def test_import_from_pdfs_no_pdfs(nps_base_dir):
    """_import_from_pdfs does nothing if no PDFs found."""
    import app.nps_database as mod
    nps_dir = nps_base_dir / "NPS"
    pdf_dir = nps_base_dir / "pdf_import"
    pdf_dir.mkdir()

    with patch.object(mod, "_imported", False), \
         patch.object(mod, "PDF_IMPORT_DIR", pdf_dir):
        mod._imported = False
        mod._import_from_pdfs(nps_dir=nps_dir)


def test_import_from_pdfs_dir_not_exists(nps_base_dir):
    """_import_from_pdfs returns if PDF_IMPORT_DIR doesn't exist."""
    import app.nps_database as mod
    nps_dir = nps_base_dir / "NPS"

    with patch.object(mod, "_imported", False), \
         patch.object(mod, "PDF_IMPORT_DIR", nps_base_dir / "nonexistent"):
        mod._imported = False
        mod._import_from_pdfs(nps_dir=nps_dir)


# ---------------------------------------------------------------------------
# Tests — _save_account without transactions (line 739)
# ---------------------------------------------------------------------------

def test_save_account_new_without_xlsx(nps_base_dir):
    """_save_account creates new file when xlsx doesn't exist."""
    from app.nps_database import _save_account, _load_all_xlsx
    nps_dir = nps_base_dir / "NPS"

    account = {
        "id": "new001",
        "pran": "110055555555",
        "account_name": "New Account",
        "tier": "Tier I",
        "status": "Active",
        "start_date": "2020-01-01",
        "current_value": 0,
        "contributions": [],
    }
    with patch("app.nps_database._sync_to_drive"):
        _save_account(account, nps_dir=nps_dir)

    items = _load_all_xlsx(nps_dir=nps_dir)
    assert len(items) == 1
