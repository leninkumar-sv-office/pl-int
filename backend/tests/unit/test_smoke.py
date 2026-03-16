"""Smoke test to verify the test infrastructure is working."""


def test_tmp_data_dir_has_required_files(tmp_data_dir):
    """Verify the tmp_data_dir fixture creates all expected JSON files."""
    expected_files = [
        "users.json",
        "stock_prices.json",
        "market_ticker.json",
        "market_ticker_history.json",
        "portfolio.json",
        "mf_scheme_map.json",
        "nav_history.json",
        "manual_prices.json",
        "symbol_cache.json",
        ".jwt_secret",
    ]
    for fname in expected_files:
        assert (tmp_data_dir / fname).exists(), f"Missing: {fname}"


def test_tmp_dumps_dir_has_subdirectories(tmp_dumps_dir):
    """Verify the tmp_dumps_dir fixture creates the user asset directories."""
    user_dir = tmp_dumps_dir / "test@example.com" / "TestUser"
    for sub in ["Stocks", "Mutual Funds", "FD", "RD", "PPF", "NPS", "Standing Instructions"]:
        assert (user_dir / sub).is_dir(), f"Missing subdir: {sub}"


def test_auth_token_is_valid_jwt(auth_token):
    """Verify the auth_token fixture produces a decodable JWT."""
    import jwt as pyjwt

    payload = pyjwt.decode(auth_token, "test-jwt-secret-for-pytest-only-00", algorithms=["HS256"])
    assert payload["email"] == "test@example.com"
    assert payload["name"] == "TestUser"


def test_sample_stock_xlsx_exists(sample_stock_xlsx):
    """Verify the sample_stock_xlsx fixture creates a valid XLSX file."""
    assert sample_stock_xlsx.exists()
    assert sample_stock_xlsx.suffix == ".xlsx"

    import openpyxl

    wb = openpyxl.load_workbook(sample_stock_xlsx)
    assert "Trading History" in wb.sheetnames
    ws = wb["Trading History"]
    assert ws.cell(row=1, column=1).value == "Date"
    assert ws.cell(row=2, column=2).value == "Buy"
    assert ws.cell(row=2, column=3).value == 10
