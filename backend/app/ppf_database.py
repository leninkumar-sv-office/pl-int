"""
PPF (Public Provident Fund) database layer.

Reads xlsx metadata from dumps/PPF/ (rows 1-4 only) and computes ALL
installments in Python -- never trusts cached formula values from xlsx.

Also supports migration from legacy JSON format and creates xlsx files
for new entries.

xlsx layout (matches FD/RD format):
    Row 1: [_, start_date(B), _, _, total_invested(E), _, _, maturity_years(H), _, _, bank(K)]
    Row 2: [_, end_date(B), _, _, interest_earned(E), _, _, "Annually"(H), _, _, sip_frequency(K)]
    Row 3: [_, rate_decimal(B), _, _, maturity_amount(E), _, _, sip_amount(H), _, _, account_number(K)]
    Row 4: [_, sip_end_date(B), _, _, remarks(E), _, _, contributions_json(H), _, _, sip_phases_json(K)]
    Row 5: headers [S.No, _, #, Date, Amount Invested, Interest Earned, Interest to be Earned]
    Row 6+: monthly data rows

SIP phases (K4):
    JSON array storing multiple SIP configurations over time.
    Each phase: {"amount": float, "frequency": str, "start": "YYYY-MM-DD", "end": "YYYY-MM-DD"|null}
    If K4 is empty/missing, a single phase is derived from H3/K2/B4 for backward compatibility.

One-time contributions (H4):
    JSON array storing ad-hoc lump-sum deposits made between SIP installments.
    Each: {"date": "YYYY-MM-DD", "amount": float, "remarks": str}
    These are injected into the installment schedule at the matching month.

Interest logic (PPF annual compounding):
    - PPF rate is annual (e.g. 7.1%)
    - Interest compounds annually: each year, interest = balance * rate
    - SIP deposits occur at sip_frequency intervals during SIP period
    - After SIP ends, no more deposits, but compounding continues until maturity
    - is_compound_month = True every 12th month from start (annual compounding)
"""

import json
import hashlib
import threading
import uuid
import calendar
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path
import openpyxl

from app.config import DUMPS_DIR


# ===================================================================
#  FILE PATHS
# ===================================================================

PPF_DIR = DUMPS_DIR / "PPF"
PPF_JSON_FILE = DUMPS_DIR / "ppf_accounts.json"  # legacy, for migration

_lock = threading.Lock()

# PPF constants
PPF_DEFAULT_RATE = 7.1   # current rate (%)
PPF_TENURE_YEARS = 15
PPF_YEARLY_MAX = 150000
PPF_YEARLY_MIN = 500


# ===================================================================
#  HELPERS
# ===================================================================

def _gen_ppf_id(name: str) -> str:
    """Deterministic ID from PPF account name (filename stem)."""
    return hashlib.md5(name.encode()).hexdigest()[:8]


def _to_date(val) -> date | None:
    """Convert openpyxl cell value to date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_str(val, default=""):
    """Safely convert cell value to string."""
    if val is None:
        return default
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _to_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _sip_freq_to_months(freq: str) -> int:
    """Convert SIP frequency string to interval in months.
    monthly -> 1, quarterly -> 3, yearly -> 12.
    """
    f = (freq or "monthly").strip().lower()
    if "month" in f:
        return 1
    if "quarter" in f:
        return 3
    if "year" in f or "annual" in f:
        return 12
    return 1  # default monthly


def _get_financial_year(dt: date) -> str:
    """Get financial year string like '2023-24' for a given date."""
    if dt.month >= 4:
        return f"{dt.year}-{str(dt.year + 1)[-2:]}"
    else:
        return f"{dt.year - 1}-{str(dt.year)[-2:]}"


# ===================================================================
#  XLSX PARSER -- Python-computed installments
# ===================================================================

def _parse_json_array(raw_val) -> list | None:
    """Try to parse a JSON array from a cell value. Returns list or None."""
    if raw_val is None:
        return None
    s = str(raw_val).strip()
    if not s or not s.startswith("["):
        return None
    try:
        arr = json.loads(s)
        if isinstance(arr, list) and len(arr) > 0:
            return arr
    except (json.JSONDecodeError, TypeError):
        pass
    return None


def _parse_sip_phases_json(raw_val) -> list | None:
    """Try to parse SIP phases JSON from cell K4. Returns list of phase dicts or None."""
    if raw_val is None:
        return None
    s = str(raw_val).strip()
    if not s or not s.startswith("["):
        return None
    try:
        phases = json.loads(s)
        if isinstance(phases, list) and len(phases) > 0:
            return phases
    except (json.JSONDecodeError, TypeError):
        pass
    return None


def _build_single_phase(sip_amount: float, sip_frequency: str,
                        start_dt: date, sip_end_dt: date | None) -> list:
    """Build a single-phase list from legacy single-SIP metadata."""
    if sip_amount <= 0:
        return []
    return [{
        "amount": sip_amount,
        "frequency": sip_frequency,
        "start": start_dt.strftime("%Y-%m-%d"),
        "end": sip_end_dt.strftime("%Y-%m-%d") if sip_end_dt else None,
    }]


def _get_active_phase(phases: list, inst_date: date) -> dict | None:
    """Find the active SIP phase for a given date.
    A phase is active if inst_date >= phase.start and (phase.end is None or inst_date <= phase.end).
    If multiple phases overlap, the last one wins (most recently added).
    """
    active = None
    for phase in phases:
        p_start = datetime.strptime(phase["start"], "%Y-%m-%d").date()
        p_end = None
        if phase.get("end"):
            p_end = datetime.strptime(phase["end"], "%Y-%m-%d").date()
        if inst_date >= p_start and (p_end is None or inst_date <= p_end):
            active = phase
    return active


def _scan_data_rows(ws, tenure_months: int) -> list:
    """Scan xlsx data rows (row 6+) for withdrawals (col 8) and
    one-time contributions (col 9). Returns a contributions list."""
    contributions = []
    for r in range(6, 6 + tenure_months):
        row_date = _to_date(ws.cell(r, 4).value)  # column 4: date
        if row_date is None:
            break
        withdrawn = _to_float(ws.cell(r, 8).value, 0)   # column 8: Withdrawn
        extra = _to_float(ws.cell(r, 9).value, 0)       # column 9: Contribution
        if withdrawn > 0:
            contributions.append({
                "date": row_date.strftime("%Y-%m-%d"),
                "amount": -withdrawn,
                "remarks": "Withdrawal",
            })
        if extra > 0:
            contributions.append({
                "date": row_date.strftime("%Y-%m-%d"),
                "amount": extra,
                "remarks": "Contribution",
            })
    return contributions


def _parse_ppf_xlsx(filepath: Path) -> dict:
    """Parse a single PPF xlsx file.

    Reads metadata from rows 1-4, scans data rows for withdrawals/contributions
    (columns 8-9), and computes ALL monthly installments in Python.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["Index"]
    name = filepath.stem

    # -- Metadata rows 1-4 --
    start_date_raw = ws.cell(1, 2).value                          # B1: start date
    maturity_years = _to_float(ws.cell(1, 8).value, PPF_TENURE_YEARS)  # H1: years
    bank = _to_str(ws.cell(1, 11).value, "Post Office")          # K1: bank

    # H2: interest payout type (always "Annually" for PPF)
    sip_frequency = _to_str(ws.cell(2, 11).value, "monthly")     # K2: sip_frequency

    rate_decimal = _to_float(ws.cell(3, 2).value, 0)             # B3: rate as decimal
    sip_amount = _to_float(ws.cell(3, 8).value, 0)               # H3: SIP amount
    account_number = _to_str(ws.cell(3, 11).value)               # K3: account number

    sip_end_date_raw = ws.cell(4, 2).value                        # B4: sip end date
    remarks = _to_str(ws.cell(4, 5).value)                        # E4: remarks

    # K4: SIP phases JSON (only for multi-phase accounts)
    sip_phases_raw = ws.cell(4, 11).value

    # -- Convert & derive --
    start_dt = _to_date(start_date_raw) or date.today()
    rate_decimal = float(rate_decimal)
    sip_amount = float(sip_amount)
    maturity_years = float(maturity_years)
    tenure_months = int(maturity_years * 12)
    rate_pct = round(rate_decimal * 100, 4) if rate_decimal < 1 else round(rate_decimal, 4)
    rate_for_calc = rate_decimal if rate_decimal < 1 else rate_decimal / 100

    end_dt = start_dt + relativedelta(months=tenure_months)

    sip_end_dt = _to_date(sip_end_date_raw)
    sip_end_str = sip_end_dt.strftime("%Y-%m-%d") if sip_end_dt else None

    # -- SIP phases: from K4 (multi-phase) or derive from metadata (single-phase) --
    sip_phases = _parse_sip_phases_json(sip_phases_raw)
    if sip_phases is None:
        sip_phases = _build_single_phase(sip_amount, sip_frequency, start_dt, sip_end_dt)

    # -- Contributions & withdrawals: scan data rows (cols 8-9) --
    contributions = _scan_data_rows(ws, tenure_months)
    wb.close()

    # Build a lookup: (year, month) -> total contribution amount for that month
    contrib_by_month = {}
    for c in contributions:
        try:
            c_date = datetime.strptime(c["date"], "%Y-%m-%d").date()
            key = (c_date.year, c_date.month)
            contrib_by_month[key] = contrib_by_month.get(key, 0) + float(c.get("amount", 0))
        except (ValueError, KeyError, TypeError):
            pass

    # -- Generate monthly installments using phases + contributions --
    today = date.today()
    installments = []
    running_balance = 0.0
    total_deposited = 0.0
    total_withdrawn = 0.0
    cumulative_deposited = 0.0
    total_interest_earned = 0.0
    total_interest_projected = 0.0
    cumulative_interest = 0.0

    lockin_months = int(maturity_years) * 12   # typically 180 (15 years)
    partial_months = 7 * 12                     # partial withdrawal from year 7 (month 84)

    for m in range(1, tenure_months + 1):
        # Base date from account start (month counter for compounding)
        base_date = start_dt + relativedelta(months=m - 1)

        # Find active phase and adjust day-of-month to match phase start
        active_phase = _get_active_phase(sip_phases, base_date)
        if active_phase:
            phase_start = datetime.strptime(active_phase["start"], "%Y-%m-%d").date()
            try:
                inst_date = base_date.replace(day=phase_start.day)
            except ValueError:
                # Day doesn't exist in this month (e.g. 31st in Feb) -- clamp to last day
                last_day = calendar.monthrange(base_date.year, base_date.month)[1]
                inst_date = base_date.replace(day=last_day)
        else:
            inst_date = base_date

        is_past = inst_date <= today

        # SIP deposit for this month
        invested = 0.0
        if active_phase:
            phase_amount = float(active_phase.get("amount", 0))
            phase_freq = active_phase.get("frequency", "monthly")
            phase_interval = _sip_freq_to_months(phase_freq)

            # Align frequency to phase start date
            months_since_start = (inst_date.year - phase_start.year) * 12 + (inst_date.month - phase_start.month)
            if months_since_start >= 0 and (months_since_start % phase_interval == 0):
                invested = phase_amount

        # Add one-time contributions for this month
        contrib_key = (inst_date.year, inst_date.month)
        if contrib_key in contrib_by_month:
            invested += contrib_by_month[contrib_key]

        running_balance += invested
        cumulative_deposited += invested
        if invested >= 0:
            total_deposited += invested
        else:
            total_withdrawn += abs(invested)

        # Annual compounding: interest credited every 12 months from start
        is_compound_month = (m % 12 == 0)
        interest = 0.0
        if is_compound_month:
            interest = round(running_balance * rate_for_calc, 2)
            running_balance += interest
            cumulative_interest += interest

        if is_past:
            total_interest_earned += interest
        else:
            total_interest_projected += interest

        # Lock-in status for this month
        if m > lockin_months:
            lock_status = "free"
        elif m > partial_months:
            lock_status = "partial"
        else:
            lock_status = "locked"

        deposit = max(invested, 0)
        withdrawn = abs(min(invested, 0))

        installments.append({
            "month": m,
            "date": inst_date.strftime("%Y-%m-%d"),
            "amount_invested": round(deposit, 2),
            "amount_withdrawn": round(withdrawn, 2),
            "cumulative_deposited": round(cumulative_deposited, 2),
            "interest_earned": round(interest, 2) if is_past else 0.0,
            "interest_projected": round(interest, 2) if not is_past else 0.0,
            "cumulative_interest": round(cumulative_interest, 2),
            "cumulative_amount": round(running_balance, 2),
            "is_compound_month": is_compound_month,
            "is_past": is_past,
            "lock_status": lock_status,
        })

    # -- Totals --
    maturity_amount = round(running_balance, 2)
    status = "Matured" if end_dt <= today else "Active"
    days_to_maturity = max(0, (end_dt - today).days)

    installments_paid = sum(1 for i in installments if i["is_past"])

    return {
        "id": _gen_ppf_id(name),
        "name": name,
        "account_name": name,
        "bank": bank,
        "account_number": account_number,
        "interest_rate": rate_pct,
        "interest_payout": "Annually",
        "sip_amount": round(sip_amount, 2),
        "sip_frequency": sip_frequency,
        "sip_end_date": sip_end_str,
        "sip_phases": sip_phases,
        "contributions": contributions,
        "tenure_years": int(maturity_years),
        "tenure_months": tenure_months,
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "maturity_date": end_dt.strftime("%Y-%m-%d"),
        "maturity_amount": maturity_amount,
        "total_deposited": round(total_deposited, 2),
        "total_withdrawn": round(total_withdrawn, 2),
        "total_interest_accrued": round(total_interest_earned, 2),
        "interest_earned": round(total_interest_earned, 2),
        "interest_projected": round(total_interest_projected, 2),
        "status": status,
        "days_to_maturity": days_to_maturity,
        "source": "xlsx",
        "remarks": remarks,
        "installments": installments,
        "installments_paid": installments_paid,
        "installments_total": len(installments),
    }


def _parse_all_xlsx() -> list:
    """Parse all xlsx files from dumps/PPF/ directory."""
    results = []
    if not PPF_DIR.exists():
        return results

    for f in sorted(PPF_DIR.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            parsed = _parse_ppf_xlsx(f)
            results.append(parsed)
        except Exception as e:
            print(f"[PPF] Error parsing {f.name}: {e}")
    return results


# ===================================================================
#  XLSX CREATION (for new entries from UI)
# ===================================================================

def _create_ppf_xlsx(name: str, bank: str, sip_amount: float,
                     rate_pct: float, maturity_years: int,
                     start_date: str, sip_frequency: str = "monthly",
                     sip_end_date: str = None, account_number: str = "",
                     remarks: str = "", overwrite: bool = False,
                     sip_phases: list = None, contributions: list = None):
    """Create an xlsx file following the FD template structure for PPF.

    If overwrite=False (default for new entries) and a file with the same name
    already exists, a numeric suffix is appended to avoid clobbering.
    If overwrite=True (used by update/contribution), the existing file is replaced.

    sip_phases: optional list of phase dicts stored in K4.
    contributions: optional list of one-time contribution dicts stored in H4.
    """
    PPF_DIR.mkdir(parents=True, exist_ok=True)
    filepath = PPF_DIR / f"{name}.xlsx"

    if not overwrite:
        # Avoid overwriting existing files -- append suffix
        counter = 2
        while filepath.exists():
            filepath = PPF_DIR / f"{name} ({counter}).xlsx"
            counter += 1
        name = filepath.stem

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index"

    rate_dec = rate_pct / 100
    tenure_months = maturity_years * 12
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = start_dt + relativedelta(months=tenure_months)

    sip_end_dt = None
    if sip_end_date:
        try:
            sip_end_dt = datetime.strptime(sip_end_date, "%Y-%m-%d")
        except (ValueError, TypeError):
            pass

    # Build phases for computation
    if sip_phases and len(sip_phases) > 0:
        phases = sip_phases
    else:
        # Single phase from parameters
        phases = []
        if sip_amount > 0:
            phases.append({
                "amount": sip_amount,
                "frequency": sip_frequency,
                "start": start_date,
                "end": sip_end_date,
            })

    # Build contribution lookup: (year, month) -> total amount
    contrib_by_month = {}
    if contributions:
        for c in contributions:
            try:
                c_date = datetime.strptime(c["date"], "%Y-%m-%d").date()
                key = (c_date.year, c_date.month)
                contrib_by_month[key] = contrib_by_month.get(key, 0) + float(c.get("amount", 0))
            except (ValueError, KeyError, TypeError):
                pass

    # -- Compute installments for xlsx data rows --
    running_balance = 0.0
    total_deposited = 0.0
    cumulative_interest = 0.0
    row_data = []

    for m in range(1, tenure_months + 1):
        base_date = start_dt + relativedelta(months=m - 1)
        inst_d = base_date.date() if isinstance(base_date, datetime) else base_date

        # Adjust day-of-month to match active phase's start date
        active_phase = _get_active_phase(phases, inst_d)
        if active_phase:
            phase_start = datetime.strptime(active_phase["start"], "%Y-%m-%d").date()
            try:
                inst_d = inst_d.replace(day=phase_start.day)
            except ValueError:
                # Clamp to last day of month (e.g. 31st in Feb -> 28/29)
                last_day = calendar.monthrange(inst_d.year, inst_d.month)[1]
                inst_d = inst_d.replace(day=last_day)
        inst_date = datetime(inst_d.year, inst_d.month, inst_d.day)

        sip_deposit = 0.0
        if active_phase:
            phase_amount = float(active_phase.get("amount", 0))
            phase_freq = active_phase.get("frequency", "monthly")
            phase_interval = _sip_freq_to_months(phase_freq)
            months_since_start = (inst_d.year - phase_start.year) * 12 + (inst_d.month - phase_start.month)
            if months_since_start >= 0 and (months_since_start % phase_interval == 0):
                sip_deposit = phase_amount

        # One-time contributions/withdrawals for this month
        extra = 0.0
        contrib_key = (inst_d.year, inst_d.month)
        if contrib_key in contrib_by_month:
            extra = contrib_by_month[contrib_key]

        invested = sip_deposit + extra
        running_balance += invested
        if invested >= 0:
            total_deposited += invested

        is_compound_month = (m % 12 == 0)
        interest = 0.0
        if is_compound_month:
            interest = round(running_balance * rate_dec, 2)
            running_balance += interest
            cumulative_interest += interest

        # Split into columns: deposit (col 5), withdrawn (col 8), contribution (col 9)
        deposit_col = sip_deposit + max(extra, 0)  # SIP + positive contributions
        withdrawn_col = abs(min(extra, 0))          # negative contributions = withdrawals
        contrib_col = max(extra, 0)                 # positive one-time contributions
        row_data.append((m, inst_date, deposit_col, withdrawn_col, contrib_col, interest))

    maturity_amount = round(running_balance, 2)

    # -- Row 1: start_date, total_invested, maturity_years, bank --
    ws.cell(1, 2, start_dt)                     # B1
    ws.cell(1, 5, total_deposited)              # E1
    ws.cell(1, 8, float(maturity_years))        # H1
    ws.cell(1, 11, bank)                        # K1

    # -- Row 2: end_date, interest_earned, "Annually", sip_frequency --
    ws.cell(2, 2, end_dt)                       # B2
    ws.cell(2, 5, cumulative_interest)          # E2
    ws.cell(2, 8, "Annually")                   # H2
    ws.cell(2, 11, sip_frequency)               # K2

    # -- Row 3: rate_decimal, maturity_amount, sip_amount, account_number --
    ws.cell(3, 2, rate_dec)                     # B3
    ws.cell(3, 5, maturity_amount)              # E3
    ws.cell(3, 8, sip_amount)                   # H3
    ws.cell(3, 11, account_number)              # K3

    # -- Row 4: sip_end_date, remarks, SIP phases (only multi-phase) --
    if sip_end_dt:
        ws.cell(4, 2, sip_end_dt)              # B4
    ws.cell(4, 5, remarks)                      # E4
    if phases and len(phases) > 1:
        ws.cell(4, 11, json.dumps(phases))     # K4: only for multi-phase

    # -- Row 5: Headers --
    headers = ['S.No', '', '#', 'Date', 'Amount Invested',
               'Interest Earned', 'Interest to be Earned',
               'Withdrawn', 'Contribution']
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)

    # -- Row 6+: Monthly data rows --
    for m, inst_date, deposit, withdrawn, contrib, interest in row_data:
        row = m + 5
        ws.cell(row, 1, m)
        ws.cell(row, 3, m)
        ws.cell(row, 4, inst_date)
        ws.cell(row, 5, deposit)
        ws.cell(row, 6, interest)
        ws.cell(row, 7, interest)
        if withdrawn > 0:
            ws.cell(row, 8, withdrawn)
        if contrib > 0:
            ws.cell(row, 9, contrib)

    wb.save(str(filepath))
    wb.close()

    # Clean up any legacy sidecar meta files
    meta_file = filepath.with_suffix(".meta.json")
    if meta_file.exists():
        meta_file.unlink()

    return filepath


# ===================================================================
#  LEGACY JSON MIGRATION
# ===================================================================

def _migrate_json_to_xlsx():
    """One-time migration: convert legacy ppf_accounts.json to xlsx files."""
    if not PPF_JSON_FILE.exists():
        return
    try:
        with open(PPF_JSON_FILE, "r") as f:
            data = json.load(f)
        if not isinstance(data, list) or len(data) == 0:
            return
        PPF_DIR.mkdir(parents=True, exist_ok=True)
        for account in data:
            name = account.get("account_name", "PPF Account")
            bank = account.get("bank", "Post Office")
            rate = account.get("interest_rate", PPF_DEFAULT_RATE)
            tenure = account.get("tenure_years", PPF_TENURE_YEARS)
            start_date = account.get("start_date", date.today().strftime("%Y-%m-%d"))
            sip_amount = account.get("sip_amount", 0) or 0
            sip_frequency = account.get("sip_frequency", "monthly")
            sip_end_date = account.get("sip_end_date")
            account_number = account.get("account_number", "")
            remarks_val = account.get("remarks", "")

            # If there were one-time contributions in old format but no sip,
            # sum them into the sip_amount as a yearly equivalent
            contributions = account.get("contributions", [])
            if sip_amount == 0 and contributions:
                total_contrib = sum(c.get("amount", 0) for c in contributions)
                # Treat total contributions as yearly deposit
                years_elapsed = max(1, (date.today() - datetime.strptime(start_date, "%Y-%m-%d").date()).days // 365)
                sip_amount = round(total_contrib / years_elapsed / 12, 2)  # monthly equivalent
                sip_frequency = "monthly"

            sip_phases_val = None
            if sip_amount > 0:
                sip_phases_val = [{
                    "amount": sip_amount,
                    "frequency": sip_frequency,
                    "start": start_date,
                    "end": sip_end_date,
                }]

            _create_ppf_xlsx(
                name=name,
                bank=bank,
                sip_amount=sip_amount,
                rate_pct=rate,
                maturity_years=tenure,
                start_date=start_date,
                sip_frequency=sip_frequency,
                sip_end_date=sip_end_date,
                account_number=account_number,
                remarks=remarks_val,
                overwrite=True,
                sip_phases=sip_phases_val,
            )
            print(f"[PPF] Migrated '{name}' to xlsx")

        # Rename old file so migration doesn't run again
        PPF_JSON_FILE.rename(PPF_JSON_FILE.with_suffix(".json.bak"))
        print("[PPF] Migration complete, renamed old JSON to .json.bak")
    except Exception as e:
        print(f"[PPF] Migration error: {e}")


def _migrate_old_xlsx():
    """Migrate old-format PPF xlsx files (with separate Index/Contributions sheets)
    to the new FD-style format. Detects old format by checking for 'Account Name' in A1."""
    if not PPF_DIR.exists():
        return
    for f in sorted(PPF_DIR.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(str(f), data_only=True)
            ws = wb["Index"]
            a1_val = _to_str(ws.cell(1, 1).value)
            # Old format has "Account Name" label in A1
            if a1_val != "Account Name":
                wb.close()
                continue

            # Read old-format metadata
            account_name = _to_str(ws.cell(1, 2).value, "PPF Account")
            bank = _to_str(ws.cell(1, 4).value, "Post Office")
            account_number = _to_str(ws.cell(1, 6).value)
            interest_rate = _to_float(ws.cell(1, 8).value, PPF_DEFAULT_RATE)

            start_date = _to_str(ws.cell(2, 2).value)
            tenure_years = int(_to_float(ws.cell(2, 4).value, PPF_TENURE_YEARS))
            sip_amount = _to_float(ws.cell(2, 8).value, 0)

            sip_frequency = _to_str(ws.cell(3, 2).value, "monthly")
            sip_end_date = _to_str(ws.cell(3, 4).value) or None
            remarks_val = _to_str(ws.cell(3, 6).value)

            wb.close()

            print(f"[PPF] Migrating old-format xlsx: {f.name}")

            # Remove old file and create new-format xlsx
            f.unlink()

            _create_ppf_xlsx(
                name=account_name,
                bank=bank,
                sip_amount=sip_amount,
                rate_pct=interest_rate,
                maturity_years=tenure_years,
                start_date=start_date,
                sip_frequency=sip_frequency,
                sip_end_date=sip_end_date if sip_end_date else None,
                account_number=account_number,
                remarks=remarks_val,
                overwrite=True,
            )
            print(f"[PPF] Migrated '{account_name}' to new xlsx format")

        except Exception as e:
            print(f"[PPF] Error migrating {f.name}: {e}")


# ===================================================================
#  PUBLIC API
# ===================================================================

def get_all() -> list:
    """Return all PPF accounts with computed monthly installments."""
    with _lock:
        _migrate_json_to_xlsx()
        _migrate_old_xlsx()
        items = _parse_all_xlsx()

    for item in items:
        _enrich_withdrawal(item)

    return items


def get_dashboard() -> dict:
    """Aggregate PPF summary for dashboard."""
    items = get_all()
    active = [i for i in items if i.get("status") == "Active"]

    return {
        "total_deposited": round(sum(i.get("total_deposited", 0) for i in active), 2),
        "total_withdrawn": round(sum(i.get("total_withdrawn", 0) for i in active), 2),
        "total_interest": round(sum(i.get("total_interest_accrued", 0) for i in active), 2),
        "current_balance": round(sum(i.get("maturity_amount", 0) for i in active), 2),
        "active_count": len(active),
        "total_count": len(items),
    }


def add(data: dict) -> dict:
    """Add a new PPF account -- creates xlsx file."""
    account_name = data.get("account_name", "PPF Account")
    bank = data.get("bank", "Post Office")
    rate = data.get("interest_rate", PPF_DEFAULT_RATE)
    tenure = data.get("tenure_years", PPF_TENURE_YEARS)
    start_date = data["start_date"]
    sip_amount = data.get("sip_amount", 0) or 0
    sip_frequency = data.get("sip_frequency", "monthly")
    sip_end_date = data.get("sip_end_date")
    account_number = data.get("account_number", "")
    remarks_val = data.get("remarks", "")

    # If one-time payment, deposit once on start date then stop
    payment_type = data.get("payment_type", "sip")
    initial_amount = data.get("amount_added", 0) or 0
    if payment_type == "one_time" and initial_amount > 0:
        sip_amount = initial_amount
        sip_frequency = "yearly"
        sip_end_date = start_date  # only deposit on month 1, then stop

    name = account_name

    # Build initial phase
    sip_phases = None
    if sip_amount > 0:
        sip_phases = [{
            "amount": sip_amount,
            "frequency": sip_frequency,
            "start": start_date,
            "end": sip_end_date,
        }]

    filepath = _create_ppf_xlsx(
        name=name,
        bank=bank,
        sip_amount=sip_amount,
        rate_pct=rate,
        maturity_years=tenure,
        start_date=start_date,
        sip_frequency=sip_frequency,
        sip_end_date=sip_end_date,
        account_number=account_number,
        remarks=remarks_val,
        sip_phases=sip_phases,
    )

    # Return the freshly parsed result
    return _parse_ppf_xlsx(filepath)


def update(ppf_id: str, data: dict) -> dict:
    """Update an existing PPF account -- rewrites xlsx file.

    If data contains 'new_sip_phase', the existing phases are preserved
    and a new phase is appended. The last existing phase gets its end date
    set to the new phase's start date.
    """
    with _lock:
        filepath = _find_xlsx(ppf_id)
        if filepath is None:
            raise ValueError(f"PPF account {ppf_id} not found")

        # Parse current values
        current = _parse_ppf_xlsx(filepath)
        old_name = current["name"]

        # Apply updates
        account_name = data.get("account_name", current["account_name"])
        bank = data.get("bank", current["bank"])
        rate = data.get("interest_rate", current["interest_rate"])
        tenure = data.get("tenure_years", current["tenure_years"])
        start_date = data.get("start_date", current["start_date"])
        sip_amount = data.get("sip_amount", current["sip_amount"])
        sip_frequency = data.get("sip_frequency", current["sip_frequency"])
        sip_end_date = data.get("sip_end_date", current.get("sip_end_date"))
        account_number = data.get("account_number", current["account_number"])
        remarks_val = data.get("remarks", current["remarks"])

        # Get existing phases
        existing_phases = current.get("sip_phases", [])

        # Check if we're adding a new SIP phase
        new_phase = data.get("new_sip_phase")
        if new_phase:
            # Append new phase to existing phases
            new_phase_start = new_phase.get("start")
            new_phase_amount = float(new_phase.get("amount", 0))
            new_phase_freq = new_phase.get("frequency", "monthly")
            new_phase_end = new_phase.get("end")

            if new_phase_amount > 0 and new_phase_start:
                # End the last existing phase at the new phase's start
                if existing_phases:
                    existing_phases[-1]["end"] = new_phase_start

                existing_phases.append({
                    "amount": new_phase_amount,
                    "frequency": new_phase_freq,
                    "start": new_phase_start,
                    "end": new_phase_end,
                })

                # Update H3/K2/B4 to reflect latest phase for backward compat
                sip_amount = new_phase_amount
                sip_frequency = new_phase_freq
                sip_end_date = new_phase_end

            sip_phases = existing_phases
        else:
            # Handle payment_type toggle from edit form
            payment_type = data.get("payment_type")
            if payment_type == "one_time":
                initial_amount = data.get("amount_added", 0) or 0
                if initial_amount > 0:
                    sip_amount = initial_amount
                    sip_frequency = "yearly"
                    sip_end_date = start_date  # single deposit on start date
                sip_phases = [{
                    "amount": sip_amount,
                    "frequency": sip_frequency,
                    "start": start_date,
                    "end": sip_end_date,
                }] if sip_amount > 0 else []
            elif payment_type == "sip":
                sip_end_date = data.get("sip_end_date") or None
                sip_phases = [{
                    "amount": sip_amount,
                    "frequency": sip_frequency,
                    "start": start_date,
                    "end": sip_end_date,
                }] if sip_amount > 0 else []
            else:
                # No payment_type change — preserve existing phases
                sip_phases = existing_phases

        # Preserve existing contributions
        existing_contributions = current.get("contributions", []) or []

        # If name changed, delete old file
        new_name = account_name
        if new_name != old_name:
            old_meta = filepath.with_suffix(".meta.json")
            filepath.unlink()
            if old_meta.exists():
                old_meta.unlink()

        _create_ppf_xlsx(
            name=new_name,
            bank=bank,
            sip_amount=sip_amount,
            rate_pct=rate,
            maturity_years=tenure,
            start_date=start_date,
            sip_frequency=sip_frequency,
            sip_end_date=sip_end_date,
            account_number=account_number,
            remarks=remarks_val,
            overwrite=True,
            sip_phases=sip_phases,
            contributions=existing_contributions,
        )

        new_filepath = PPF_DIR / f"{new_name}.xlsx"
        return _parse_ppf_xlsx(new_filepath)


def delete(ppf_id: str) -> dict:
    """Delete a PPF account -- removes xlsx file."""
    with _lock:
        filepath = _find_xlsx(ppf_id)
        if filepath is None:
            raise ValueError(f"PPF account {ppf_id} not found")

        account = _parse_ppf_xlsx(filepath)
        meta_file = filepath.with_suffix(".meta.json")
        filepath.unlink()
        if meta_file.exists():
            meta_file.unlink()
        return {"message": f"PPF {ppf_id} deleted", "item": account}


def add_contribution(ppf_id: str, contribution: dict) -> dict:
    """Add a one-time contribution to a PPF account.

    The contribution is stored in the H4 JSON array and reflected in the
    installment schedule at the matching month. Validated against yearly limits.
    """
    with _lock:
        filepath = _find_xlsx(ppf_id)
        if filepath is None:
            raise ValueError(f"PPF account {ppf_id} not found")

        account = _parse_ppf_xlsx(filepath)

        amount = contribution.get("amount", 0)
        contrib_date = contribution.get("date", datetime.now().strftime("%Y-%m-%d"))
        contrib_remarks = contribution.get("remarks", "")

        try:
            c_date = datetime.strptime(contrib_date, "%Y-%m-%d").date()
        except ValueError:
            raise ValueError("Invalid date format")

        # Validate yearly limit by checking projected deposits for this FY
        fy = _get_financial_year(c_date)

        # Calculate deposits already scheduled for this FY
        fy_start_year = c_date.year if c_date.month >= 4 else c_date.year - 1
        fy_start = date(fy_start_year, 4, 1)
        fy_end = date(fy_start_year + 1, 3, 31)

        # Sum up all deposits (SIP + existing contributions) in this FY
        fy_deposits = sum(
            i["amount_invested"] for i in account["installments"]
            if fy_start.strftime("%Y-%m-%d") <= i["date"] <= fy_end.strftime("%Y-%m-%d")
        )

        if fy_deposits + amount > PPF_YEARLY_MAX:
            raise ValueError(
                f"Exceeds yearly limit of Rs.{PPF_YEARLY_MAX:,.0f}. "
                f"Already deposited/projected Rs.{fy_deposits:,.0f} in FY {fy}"
            )

        # Append to contributions list (stored in H4)
        existing_contributions = account.get("contributions", []) or []
        existing_contributions.append({
            "date": contrib_date,
            "amount": amount,
            "remarks": contrib_remarks,
        })

        _create_ppf_xlsx(
            name=account["name"],
            bank=account["bank"],
            sip_amount=account["sip_amount"],
            rate_pct=account["interest_rate"],
            maturity_years=account["tenure_years"],
            start_date=account["start_date"],
            sip_frequency=account["sip_frequency"],
            sip_end_date=account.get("sip_end_date"),
            account_number=account["account_number"],
            remarks=account.get("remarks", ""),
            overwrite=True,
            sip_phases=account.get("sip_phases"),
            contributions=existing_contributions,
        )

        return _parse_ppf_xlsx(filepath)


def _enrich_withdrawal(item: dict):
    """Compute withdrawal eligibility fields for a parsed PPF account.

    Subtracts previously withdrawn amounts so the same money can't be
    withdrawn twice.
    """
    today = date.today()
    try:
        start = datetime.strptime(item["start_date"], "%Y-%m-%d").date()
        yc = max(0, (today - start).days // 365)
    except (ValueError, KeyError):
        yc = 0
    item["years_completed"] = yc
    installments = item.get("installments", [])

    # Total already withdrawn (stored as negative contributions)
    already_withdrawn = item.get("total_withdrawn", 0)

    if yc >= item.get("tenure_years", 15):
        past_insts = [i for i in installments if i["is_past"]]
        current_balance = (
            sum(i["amount_invested"] for i in past_insts)
            + sum(i["interest_earned"] for i in past_insts)
        )
        item["withdrawable_amount"] = round(max(0, current_balance), 2)
        item["withdrawal_status"] = "full"
        item["withdrawal_note"] = "Fully withdrawable — lock-in complete"
    elif yc >= 7:
        preceding_year = yc - 4
        months_cutoff = preceding_year * 12
        balance_at_cutoff = 0.0
        for inst in installments[:months_cutoff]:
            balance_at_cutoff += inst.get("amount_invested", 0)
            balance_at_cutoff += inst.get("interest_earned", 0)
        withdrawable = round(max(0, balance_at_cutoff * 0.5 - already_withdrawn), 2)
        item["withdrawable_amount"] = withdrawable
        item["withdrawal_status"] = "partial"
        item["withdrawal_note"] = (
            f"Partial withdrawal eligible — up to 50% of balance "
            f"at end of year {preceding_year}"
            + (f" (already withdrawn {already_withdrawn:,.0f})" if already_withdrawn > 0 else "")
        )
    else:
        unlock_date = datetime.strptime(item["start_date"], "%Y-%m-%d").date() + relativedelta(years=7)
        item["withdrawable_amount"] = 0
        item["withdrawal_status"] = "locked"
        item["withdrawal_note"] = (
            f"Locked — partial withdrawal from year 7 "
            f"({unlock_date.strftime('%b %Y')})"
        )


def withdraw(ppf_id: str, data: dict) -> dict:
    """Withdraw from a PPF account.

    Records the withdrawal as a negative contribution in H4 so it is
    reflected in the installment schedule and running balance.
    """
    with _lock:
        filepath = _find_xlsx(ppf_id)
        if filepath is None:
            raise ValueError(f"PPF account {ppf_id} not found")

        account = _parse_ppf_xlsx(filepath)
        _enrich_withdrawal(account)

        amount = float(data.get("amount", 0))
        withdraw_date = data.get("date", datetime.now().strftime("%Y-%m-%d"))

        if amount <= 0:
            raise ValueError("Withdrawal amount must be positive")

        # Validate against withdrawable amount
        withdrawable = account.get("withdrawable_amount", 0)
        status = account.get("withdrawal_status", "locked")

        if status == "locked":
            raise ValueError("Account is still in lock-in period")
        if amount > withdrawable:
            raise ValueError(
                f"Amount exceeds withdrawable limit of Rs.{withdrawable:,.0f}"
            )

        # Store as negative contribution
        existing_contributions = account.get("contributions", []) or []
        existing_contributions.append({
            "date": withdraw_date,
            "amount": -amount,  # negative = withdrawal
            "remarks": data.get("remarks", "Withdrawal"),
        })

        _create_ppf_xlsx(
            name=account["name"],
            bank=account["bank"],
            sip_amount=account["sip_amount"],
            rate_pct=account["interest_rate"],
            maturity_years=account["tenure_years"],
            start_date=account["start_date"],
            sip_frequency=account["sip_frequency"],
            sip_end_date=account.get("sip_end_date"),
            account_number=account["account_number"],
            remarks=account.get("remarks", ""),
            overwrite=True,
            sip_phases=account.get("sip_phases"),
            contributions=existing_contributions,
        )

        return _parse_ppf_xlsx(filepath)


# ===================================================================
#  INTERNAL HELPERS (file lookup)
# ===================================================================

def _find_xlsx(ppf_id: str) -> Path | None:
    """Find the xlsx file for a given PPF ID (based on filename hash)."""
    if not PPF_DIR.exists():
        return None
    for f in PPF_DIR.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        if _gen_ppf_id(f.stem) == ppf_id:
            return f
    return None
