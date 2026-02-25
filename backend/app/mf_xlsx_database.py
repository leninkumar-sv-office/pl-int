"""
XLSX-per-fund database layer for Mutual Funds.

Each mutual fund is stored as an individual .xlsx file under dumps/Mutual Funds/.
Holdings and redeemed positions are derived on-the-fly via FIFO matching
of Buy/Sell rows in each file's "Trading History" sheet.

Key differences from stock xlsx:
  - Column D is "Units" (float) instead of "QTY" (int)
  - Column E is "NAV" instead of "PRICE"
  - No STT or additional charges (columns H-I empty)
  - Fund code format: MUTF_IN:xxx (Google Finance code)
  - Lock In Period and Exit Load fields in Index sheet
"""

import hashlib
import re
import threading
import time
import random
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests as _requests
import openpyxl

from .models import MFHolding, MFSoldPosition


# ═══════════════════════════════════════════════════════════
#  LIVE NAV FETCHING (Google Finance)
# ═══════════════════════════════════════════════════════════

_GOOGLE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
}

_nav_cache: Dict[str, float] = {}
_nav_cache_lock = threading.Lock()


def _fetch_nav_google_finance(fund_code: str) -> Optional[float]:
    """Fetch current NAV from Google Finance for a mutual fund.

    fund_code is like 'MUTF_IN:SBI_SMAL_CAP_HY56CY'.
    Google Finance URL: /finance/quote/SBI_SMAL_CAP_HY56CY:MUTF_IN
    """
    if not fund_code or ":" not in fund_code:
        return None
    parts = fund_code.split(":", 1)
    gf_symbol = f"{parts[1]}:{parts[0]}"
    url = f"https://www.google.com/finance/quote/{gf_symbol}"
    for attempt in range(2):
        try:
            resp = _requests.get(url, headers=_GOOGLE_HEADERS, timeout=10)
            if resp.status_code == 200:
                match = re.search(r'data-last-price="([\d,.]+)"', resp.text)
                if match:
                    price = float(match.group(1).replace(",", ""))
                    if price > 0:
                        return price
        except Exception:
            pass
        if attempt < 1:
            time.sleep(1)
    return None


def fetch_live_navs(fund_codes: List[str]) -> Dict[str, float]:
    """Fetch live NAVs for a list of fund codes. Returns {fund_code: nav}."""
    results: Dict[str, float] = {}
    for code in fund_codes:
        with _nav_cache_lock:
            if code in _nav_cache:
                results[code] = _nav_cache[code]
                continue
        nav = _fetch_nav_google_finance(code)
        if nav and nav > 0:
            with _nav_cache_lock:
                _nav_cache[code] = nav
            results[code] = nav
        time.sleep(random.uniform(0.3, 0.8))
    return results


def clear_nav_cache():
    """Clear the NAV cache to force fresh fetches."""
    with _nav_cache_lock:
        _nav_cache.clear()


# ═══════════════════════════════════════════════════════════
#  NAV HISTORY via mfapi.in  (7D / 1M change tracking)
# ═══════════════════════════════════════════════════════════

import os, json

_NAV_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
_SCHEME_MAP_FILE = os.path.join(_NAV_DATA_DIR, "mf_scheme_map.json")

# In-memory cache: {fund_code: {week_change_pct, month_change_pct, fetched_at}}
_nav_change_cache: Dict[str, dict] = {}
_nav_change_cache_lock = threading.Lock()
_NAV_CHANGE_CACHE_TTL = 6 * 3600  # 6 hours


def _load_scheme_map() -> Dict[str, int]:
    """Load {fund_code: amfi_scheme_code} mapping from disk."""
    try:
        with open(_SCHEME_MAP_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_scheme_map(mapping: Dict[str, int]):
    """Persist scheme mapping to disk."""
    os.makedirs(_NAV_DATA_DIR, exist_ok=True)
    try:
        with open(_SCHEME_MAP_FILE, "w") as f:
            json.dump(mapping, f, indent=2)
    except Exception as e:
        print(f"[MF-MFAPI] Failed to save scheme map: {e}")


def _search_mfapi_scheme(fund_name: str) -> Optional[int]:
    """Search mfapi.in for a scheme code by fund name.
    Returns AMFI scheme code (int) or None."""
    # Clean the fund name for search
    q = fund_name.strip()
    # Remove common suffixes that might differ
    for suffix in [" - Direct Plan", " Direct Plan", " - Direct", " Direct"]:
        q = q.replace(suffix, "")
    # Use first few significant words
    words = q.split()[:5]
    search_q = " ".join(words)
    try:
        resp = _requests.get(
            f"https://api.mfapi.in/mf/search?q={search_q}",
            timeout=10,
        )
        if resp.status_code == 200:
            results = resp.json()
            if results and len(results) > 0:
                # Try to find "Direct" + "Growth" variant first
                name_upper = fund_name.upper()
                for r in results:
                    rn = r.get("schemeName", "").upper()
                    if "DIRECT" in rn and "GROWTH" in rn:
                        return int(r["schemeCode"])
                # Fall back to first match with "Direct"
                for r in results:
                    rn = r.get("schemeName", "").upper()
                    if "DIRECT" in rn:
                        return int(r["schemeCode"])
                # Fall back to first result
                return int(results[0]["schemeCode"])
    except Exception as e:
        print(f"[MF-MFAPI] Search error for '{search_q}': {e}")
    return None


def _fetch_nav_history_mfapi(scheme_code: int) -> Optional[list]:
    """Fetch NAV history from mfapi.in for a scheme.
    Returns list of {date: DD-MM-YYYY, nav: str} or None."""
    try:
        resp = _requests.get(
            f"https://api.mfapi.in/mf/{scheme_code}",
            timeout=15,
        )
        if resp.status_code == 200:
            data = resp.json()
            return data.get("data", [])
    except Exception as e:
        print(f"[MF-MFAPI] Fetch error for scheme {scheme_code}: {e}")
    return None


def compute_nav_changes(fund_code: str, fund_name: str, current_nav: float) -> Dict[str, float]:
    """Compute 7D and 30D NAV change % using mfapi.in historical data.
    Returns {week_change_pct, month_change_pct}."""
    result = {"week_change_pct": 0.0, "month_change_pct": 0.0}
    if current_nav <= 0:
        return result

    # Check in-memory cache
    now = time.time()
    with _nav_change_cache_lock:
        cached = _nav_change_cache.get(fund_code)
        if cached and (now - cached.get("fetched_at", 0)) < _NAV_CHANGE_CACHE_TTL:
            return {
                "week_change_pct": cached["week_change_pct"],
                "month_change_pct": cached["month_change_pct"],
            }

    # Look up AMFI scheme code (cached on disk)
    scheme_map = _load_scheme_map()
    scheme_code = scheme_map.get(fund_code)

    if not scheme_code:
        scheme_code = _search_mfapi_scheme(fund_name)
        if scheme_code:
            scheme_map[fund_code] = scheme_code
            _save_scheme_map(scheme_map)
            print(f"[MF-MFAPI] Mapped {fund_name[:40]} → scheme {scheme_code}")
        else:
            print(f"[MF-MFAPI] No scheme found for {fund_name[:40]}")
            # Cache the miss so we don't retry every call
            with _nav_change_cache_lock:
                _nav_change_cache[fund_code] = {
                    "week_change_pct": 0.0, "month_change_pct": 0.0,
                    "fetched_at": now,
                }
            return result

    # Fetch historical NAV data
    nav_data = _fetch_nav_history_mfapi(scheme_code)
    if not nav_data:
        return result

    # Parse dates and find NAVs ~7d and ~30d ago
    from datetime import timedelta
    today = date.today()
    target_7d = today - timedelta(days=7)
    target_30d = today - timedelta(days=30)

    nav_7d = 0.0
    nav_30d = 0.0
    best_7d_date = None
    best_30d_date = None

    for entry in nav_data:
        try:
            d = datetime.strptime(entry["date"], "%d-%m-%Y").date()
            nav_val = float(entry["nav"])
        except (ValueError, KeyError, TypeError):
            continue

        # Find closest date on or before target_7d
        if d <= target_7d:
            if best_7d_date is None or d > best_7d_date:
                best_7d_date = d
                nav_7d = nav_val
        # Find closest date on or before target_30d
        if d <= target_30d:
            if best_30d_date is None or d > best_30d_date:
                best_30d_date = d
                nav_30d = nav_val

    if nav_7d > 0:
        result["week_change_pct"] = round((current_nav - nav_7d) / nav_7d * 100, 2)
    if nav_30d > 0:
        result["month_change_pct"] = round((current_nav - nav_30d) / nav_30d * 100, 2)

    # Cache result
    with _nav_change_cache_lock:
        _nav_change_cache[fund_code] = {
            "week_change_pct": result["week_change_pct"],
            "month_change_pct": result["month_change_pct"],
            "fetched_at": now,
        }

    return result


def record_nav_history(live_navs: Dict[str, float]):
    """No-op — kept for compatibility. mfapi.in provides history directly."""
    pass


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _gen_mf_id(fund_code: str, buy_date: str, nav: float, row_idx: int) -> str:
    """Deterministic holding ID from fund data + row position."""
    key = f"MF|{fund_code}|{buy_date}|{nav:.4f}|{row_idx}"
    return hashlib.md5(key.encode()).hexdigest()[:8]


def _parse_date(val) -> Optional[str]:
    """Convert xlsx cell value to YYYY-MM-DD string."""
    _FMT = "%Y-%m-%d"
    if isinstance(val, datetime):
        return val.strftime(_FMT)
    if isinstance(val, date):
        return val.strftime(_FMT)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%B-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime(_FMT)
            except ValueError:
                continue
    return None


def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════════════════════
#  FIFO MATCHING (adapted for fractional units)
# ═══════════════════════════════════════════════════════════

def fifo_match_mf(buys: list, sells: list):
    """FIFO-match sells against buys with fractional units.

    Returns (remaining_lots, sold_positions).
    """
    buys_sorted = sorted(buys, key=lambda x: x["date"])
    sells_sorted = sorted(sells, key=lambda x: x["date"])

    buy_lots = [{**b, "remaining": b["units"]} for b in buys_sorted]
    sold_positions = []

    for sell in sells_sorted:
        sell_units = sell["units"]
        for lot in buy_lots:
            if sell_units <= 0.0001:
                break
            if lot["remaining"] <= 0.0001:
                continue
            matched = min(lot["remaining"], sell_units)
            realized_pl = (sell["nav"] - lot["nav"]) * matched
            sold_positions.append({
                "buy_nav": lot["nav"],
                "buy_date": lot["date"],
                "sell_nav": sell["nav"],
                "sell_date": sell["date"],
                "units": round(matched, 6),
                "realized_pl": round(realized_pl, 2),
                "row_idx": lot.get("row_idx", 0),
            })
            lot["remaining"] = round(lot["remaining"] - matched, 6)
            sell_units = round(sell_units - matched, 6)

    remaining = [l for l in buy_lots if l["remaining"] > 0.0001]
    return remaining, sold_positions


# ═══════════════════════════════════════════════════════════
#  XLSX PARSING
# ═══════════════════════════════════════════════════════════

def _extract_mf_index_data(wb) -> dict:
    """Read metadata from the Index sheet of a mutual fund xlsx."""
    data = {
        "fund_code": None,
        "current_nav": 0.0,
        "week_52_high": 0.0,
        "week_52_low": 0.0,
        "lock_in_period": "",
        "exit_load": "",
    }
    if "Index" not in wb.sheetnames:
        return data
    ws = wb["Index"]
    max_row = ws.max_row or 0
    if max_row == 0:
        return data

    for row in ws.iter_rows(min_row=1, max_row=min(15, max_row), values_only=False):
        vals = [c.value for c in row]
        if len(vals) >= 3:
            label = vals[1]
            value = vals[2]
            if label == "Code" and value:
                data["fund_code"] = str(value)
            elif label == "Current Price" and isinstance(value, (int, float)):
                data["current_nav"] = float(value)
            elif label == "52 Week High" and isinstance(value, (int, float)):
                data["week_52_high"] = float(value)
            elif label == "52 Week Low" and isinstance(value, (int, float)):
                data["week_52_low"] = float(value)

        # Lock In Period and Exit Load are in column I (index 8)
        if len(vals) >= 9:
            label_i = vals[8]
            if label_i and "Lock In" in str(label_i):
                # Value is in the next row or same row col J
                pass  # usually just a label
            elif label_i and "Exit Load" in str(label_i):
                pass

    return data


def _parse_mf_trading_history(wb) -> Tuple[list, list]:
    """Parse Buy and Sell rows from MF Trading History sheet.

    Returns (buy_lots, sell_rows).
    """
    buys, sells = [], []
    if "Trading History" not in wb.sheetnames:
        return buys, sells
    ws = wb["Trading History"]
    max_row = ws.max_row or 0
    if max_row < 5:
        return buys, sells

    # Find header row
    header_row = None
    for r in range(1, min(11, max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, 5)]
        if "DATE" in vals and "ACTION" in vals:
            header_row = r
            break
    if header_row is None:
        return buys, sells

    for row_idx in range(header_row + 1, max_row + 1):
        date_val = ws.cell(row_idx, 1).value      # A: DATE
        exch = ws.cell(row_idx, 2).value           # B: EXCH
        action = ws.cell(row_idx, 3).value         # C: ACTION
        units_val = ws.cell(row_idx, 4).value      # D: Units (float)
        nav_val = ws.cell(row_idx, 5).value        # E: NAV
        cost_val = ws.cell(row_idx, 6).value       # F: COST
        remarks_val = ws.cell(row_idx, 7).value    # G: REMARKS

        if not action or not date_val:
            continue

        action = str(action).strip()
        if action not in ("Buy", "Sell"):
            continue

        tx_date = _parse_date(date_val)
        if not tx_date:
            continue

        units = _safe_float(units_val)
        nav = _safe_float(nav_val)
        cost = _safe_float(cost_val)

        # For Sell rows, NAV might be a formula =F/D; use data_only=True handles this
        if nav <= 0 and cost > 0 and units > 0:
            nav = cost / units

        if units <= 0:
            continue

        remarks = str(remarks_val or "").strip()

        if action == "Buy":
            buy_price = cost / units if cost > 0 and units > 0 else nav
            buys.append({
                "date": tx_date,
                "units": round(units, 6),
                "nav": round(nav, 4),
                "buy_price": round(buy_price, 4),
                "cost": round(cost, 2),
                "row_idx": row_idx,
                "remarks": remarks,
            })
        elif action == "Sell":
            sells.append({
                "date": tx_date,
                "units": round(units, 6),
                "nav": round(nav, 4),
                "cost": round(cost, 2),
                "row_idx": row_idx,
                "remarks": remarks,
            })

    return buys, sells


# ═══════════════════════════════════════════════════════════
#  MAIN CLASS
# ═══════════════════════════════════════════════════════════

class MFXlsxPortfolio:
    """File-per-fund xlsx database with FIFO-derived holdings."""

    def __init__(self, mf_dir: str | Path):
        self.mf_dir = Path(mf_dir)
        if not self.mf_dir.exists():
            print(f"[MF-XlsxDB] Mutual Funds directory not found: {self.mf_dir}")
            self.mf_dir.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()

        # fund_code → {holdings, sold, index_data}
        self._cache: Dict[str, Tuple[float, list, list, dict]] = {}
        # fund_code → filepath
        self._file_map: Dict[str, Path] = {}
        # fund_code → fund name (from filename)
        self._name_map: Dict[str, str] = {}

        self._build_file_map()

    def _build_file_map(self):
        """Scan xlsx files and build fund_code ↔ filepath map."""
        count = 0
        for fp in sorted(self.mf_dir.glob("*.xlsx")):
            if fp.name.startswith("~") or fp.name.startswith("."):
                continue

            try:
                wb = openpyxl.load_workbook(fp, data_only=True)
                idx = _extract_mf_index_data(wb)
                wb.close()
            except Exception as e:
                print(f"[MF-XlsxDB] Failed to read {fp.name}: {e}")
                continue

            fund_code = idx.get("fund_code")
            if not fund_code:
                # Use filename as fallback code
                fund_code = fp.stem
                print(f"[MF-XlsxDB] No fund code for {fp.name}, using filename")

            fund_name = fp.stem
            self._file_map[fund_code] = fp
            self._name_map[fund_code] = fund_name
            count += 1

        print(f"[MF-XlsxDB] Indexed {count} mutual fund files")

    def reindex(self):
        """Re-scan the MF directory for new/modified files."""
        with self._lock:
            old_codes = set(self._file_map.keys())
            self._file_map.clear()
            self._name_map.clear()
            self._cache.clear()
            self._build_file_map()
            new_codes = set(self._file_map.keys())
            added = new_codes - old_codes
            removed = old_codes - new_codes
            if added or removed:
                parts = []
                if added:
                    parts.append(f"+{len(added)}")
                if removed:
                    parts.append(f"-{len(removed)}")
                print(f"[MF-XlsxDB] Reindex: {len(new_codes)} funds ({', '.join(parts)} changed)")

    # ── Cache Layer ───────────────────────────────────────

    def _get_fund_data(self, fund_code: str) -> Tuple[List[MFHolding], List[MFSoldPosition], dict]:
        """Get holdings/sold for a fund with mtime caching."""
        fp = self._file_map.get(fund_code)
        if not fp:
            return [], [], {}

        try:
            mtime = fp.stat().st_mtime
        except OSError:
            return [], [], {}

        if fund_code in self._cache:
            cached_mtime, cached_h, cached_s, cached_idx = self._cache[fund_code]
            if cached_mtime == mtime:
                return cached_h, cached_s, cached_idx

        holdings, sold, idx_data = self._parse_and_match_fund(fund_code, fp)
        self._cache[fund_code] = (mtime, holdings, sold, idx_data)
        return holdings, sold, idx_data

    def _parse_and_match_fund(self, fund_code: str, filepath: Path):
        """Parse a fund's xlsx file and FIFO-match sells."""
        name = self._name_map.get(fund_code, fund_code)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"[MF-XlsxDB] Failed to open {filepath.name}: {e}")
            return [], [], {}

        idx_data = _extract_mf_index_data(wb)
        buy_lots, sell_rows = _parse_mf_trading_history(wb)
        wb.close()

        if not buy_lots and not sell_rows:
            return [], [], idx_data

        # FIFO match sells against buys
        if sell_rows and buy_lots:
            remaining, fifo_sold = fifo_match_mf(
                [{"date": b["date"], "units": b["units"], "nav": b["buy_price"],
                  "row_idx": b["row_idx"]} for b in buy_lots],
                [{"date": s["date"], "units": s["units"], "nav": s["nav"],
                  "row_idx": s["row_idx"]} for s in sell_rows],
            )

            # Build held lots from remaining
            holdings = []
            for r in remaining:
                if r["remaining"] <= 0.0001:
                    continue
                # Find original buy lot to get cost
                orig = next((b for b in buy_lots
                             if b["row_idx"] == r["row_idx"]), None)
                cost = round(r["nav"] * r["remaining"], 2) if not orig else round(
                    orig["cost"] * r["remaining"] / orig["units"], 2)
                h_id = _gen_mf_id(fund_code, r["date"], r["nav"], r["row_idx"])
                holdings.append(MFHolding(
                    id=h_id,
                    fund_code=fund_code,
                    name=name,
                    units=round(r["remaining"], 6),
                    nav=round(r["nav"], 4),
                    buy_price=round(r["nav"], 4),
                    buy_cost=round(cost, 2),
                    buy_date=r["date"],
                ))

            # Build sold positions
            sold = []
            for fs in fifo_sold:
                s_id = _gen_mf_id(fund_code + "_S", fs["sell_date"],
                                   fs["sell_nav"], fs["row_idx"])
                sold.append(MFSoldPosition(
                    id=s_id,
                    fund_code=fund_code,
                    name=name,
                    units=fs["units"],
                    buy_nav=fs["buy_nav"],
                    buy_date=fs["buy_date"],
                    sell_nav=fs["sell_nav"],
                    sell_date=fs["sell_date"],
                    realized_pl=fs["realized_pl"],
                ))
        else:
            # No sells — all buys are held
            holdings = []
            for b in buy_lots:
                h_id = _gen_mf_id(fund_code, b["date"], b["buy_price"], b["row_idx"])
                holdings.append(MFHolding(
                    id=h_id,
                    fund_code=fund_code,
                    name=name,
                    units=b["units"],
                    nav=round(b["nav"], 4),
                    buy_price=round(b["buy_price"], 4),
                    buy_cost=round(b["cost"], 2),
                    buy_date=b["date"],
                    remarks=b.get("remarks", ""),
                ))
            sold = []

        return holdings, sold, idx_data

    # ── Public READ API ───────────────────────────────────

    def get_all_holdings(self) -> List[MFHolding]:
        """Get all current MF holdings across every fund file."""
        all_holdings: List[MFHolding] = []
        for fund_code in list(self._file_map.keys()):
            try:
                holdings, _, _ = self._get_fund_data(fund_code)
                all_holdings.extend(holdings)
            except Exception as e:
                print(f"[MF-XlsxDB] Error reading {fund_code}: {e}")
        return all_holdings

    def get_all_sold(self) -> List[MFSoldPosition]:
        """Get all redeemed positions across every fund."""
        all_sold: List[MFSoldPosition] = []
        for fund_code in list(self._file_map.keys()):
            try:
                _, sold, _ = self._get_fund_data(fund_code)
                all_sold.extend(sold)
            except Exception as e:
                print(f"[MF-XlsxDB] Error reading sold for {fund_code}: {e}")
        return all_sold

    def get_fund_summary(self) -> List[dict]:
        """Get per-fund aggregated summary with current NAV."""
        from datetime import datetime as dt

        summaries = []
        today = dt.now()

        # Pre-fetch live NAVs for all funds
        all_codes = list(self._file_map.keys())
        live_navs = fetch_live_navs(all_codes)

        # Record today's NAVs for 7d/30d tracking
        record_nav_history(live_navs)

        for fund_code in all_codes:
            try:
                holdings, sold, idx_data = self._get_fund_data(fund_code)
            except Exception as e:
                print(f"[MF-XlsxDB] Error for {fund_code}: {e}")
                continue

            name = self._name_map.get(fund_code, fund_code)
            # Prefer live NAV, fall back to xlsx Index sheet value
            current_nav = live_navs.get(fund_code, 0.0) or idx_data.get("current_nav", 0.0)
            w52_high = idx_data.get("week_52_high", 0.0)
            w52_low = idx_data.get("week_52_low", 0.0)

            total_held_units = sum(h.units for h in holdings)
            total_invested = sum(h.buy_cost for h in holdings)
            current_value = round(current_nav * total_held_units, 2) if current_nav > 0 else 0
            unrealized_pl = round(current_value - total_invested, 2)
            unrealized_pl_pct = round((unrealized_pl / total_invested) * 100, 2) if total_invested > 0 else 0

            avg_nav = round(total_invested / total_held_units, 4) if total_held_units > 0 else 0

            total_sold_units = sum(s.units for s in sold)
            realized_pl = round(sum(s.realized_pl for s in sold), 2)

            # LTCG/STCG for held lots
            ltcg_upl = 0.0
            stcg_upl = 0.0
            for h in holdings:
                if current_nav <= 0:
                    break
                lot_pl = (current_nav - h.buy_price) * h.units
                try:
                    buy_dt = dt.strptime(h.buy_date, "%Y-%m-%d")
                    days = (today - buy_dt).days
                except (ValueError, TypeError):
                    days = 0
                if days > 365:
                    ltcg_upl += lot_pl
                else:
                    stcg_upl += lot_pl

            # LTCG/STCG for sold lots
            ltcg_rpl = 0.0
            stcg_rpl = 0.0
            for s in sold:
                try:
                    buy_dt = dt.strptime(s.buy_date, "%Y-%m-%d")
                    sell_dt = dt.strptime(s.sell_date, "%Y-%m-%d")
                    days = (sell_dt - buy_dt).days
                except (ValueError, TypeError):
                    days = 0
                if days > 365:
                    ltcg_rpl += s.realized_pl
                else:
                    stcg_rpl += s.realized_pl

            # 7D / 1M NAV changes
            nav_changes = compute_nav_changes(fund_code, name, current_nav)

            summaries.append({
                "fund_code": fund_code,
                "name": name,
                "total_held_units": round(total_held_units, 6),
                "total_sold_units": round(total_sold_units, 6),
                "avg_nav": avg_nav,
                "total_invested": round(total_invested, 2),
                "current_nav": current_nav,
                "current_value": current_value,
                "unrealized_pl": unrealized_pl,
                "unrealized_pl_pct": unrealized_pl_pct,
                "realized_pl": realized_pl,
                "ltcg_unrealized_pl": round(ltcg_upl, 2),
                "stcg_unrealized_pl": round(stcg_upl, 2),
                "ltcg_realized_pl": round(ltcg_rpl, 2),
                "stcg_realized_pl": round(stcg_rpl, 2),
                "num_held_lots": len(holdings),
                "num_sold_lots": len(sold),
                "week_52_high": w52_high,
                "week_52_low": w52_low,
                "week_change_pct": nav_changes["week_change_pct"],
                "month_change_pct": nav_changes["month_change_pct"],
                "is_above_avg_nav": current_nav > avg_nav if current_nav > 0 and avg_nav > 0 else False,
                # Include held lots and sold lots for detail view
                "held_lots": [
                    {
                        "id": h.id,
                        "buy_date": h.buy_date,
                        "units": h.units,
                        "nav": h.nav,
                        "buy_price": h.buy_price,
                        "buy_cost": h.buy_cost,
                        "current_value": round(current_nav * h.units, 2) if current_nav > 0 else 0,
                        "pl": round((current_nav - h.buy_price) * h.units, 2) if current_nav > 0 else 0,
                        "pl_pct": round(((current_nav / h.buy_price) - 1) * 100, 2) if h.buy_price > 0 and current_nav > 0 else 0,
                        "is_ltcg": (today - dt.strptime(h.buy_date, "%Y-%m-%d")).days > 365
                                   if h.buy_date else False,
                    }
                    for h in holdings
                ],
                "sold_lots": [
                    {
                        "id": s.id,
                        "buy_date": s.buy_date,
                        "sell_date": s.sell_date,
                        "units": s.units,
                        "buy_nav": s.buy_nav,
                        "sell_nav": s.sell_nav,
                        "realized_pl": s.realized_pl,
                    }
                    for s in sold
                ],
            })

        return summaries

    def get_dashboard_summary(self) -> dict:
        """Get aggregated MF portfolio summary for the dashboard."""
        fund_summaries = self.get_fund_summary()

        total_invested = sum(f["total_invested"] for f in fund_summaries)
        current_value = sum(f["current_value"] for f in fund_summaries)
        unrealized_pl = round(current_value - total_invested, 2)
        realized_pl = sum(f["realized_pl"] for f in fund_summaries)

        funds_in_profit = sum(1 for f in fund_summaries if f["unrealized_pl"] > 0)
        funds_in_loss = sum(1 for f in fund_summaries if f["unrealized_pl"] < 0)

        return {
            "total_invested": round(total_invested, 2),
            "current_value": round(current_value, 2),
            "unrealized_pl": unrealized_pl,
            "unrealized_pl_pct": round((unrealized_pl / total_invested) * 100, 2) if total_invested > 0 else 0,
            "realized_pl": round(realized_pl, 2),
            "total_funds": len(fund_summaries),
            "funds_in_profit": funds_in_profit,
            "funds_in_loss": funds_in_loss,
        }

    # ── Public WRITE API ─────────────────────────────────

    def _create_mf_file(self, fund_code: str, fund_name: str) -> Path:
        """Create a new xlsx file for a mutual fund."""
        # Sanitise name for filename
        safe_name = fund_name.replace("/", "-").replace("\\", "-").replace(":", "-")
        filepath = self.mf_dir / f"{safe_name}.xlsx"
        if filepath.exists():
            return filepath

        wb = openpyxl.Workbook()

        # Trading History sheet
        ws_th = wb.active
        ws_th.title = "Trading History"
        # Row 4 headers (matching existing format)
        headers = ["DATE", "EXCH", "ACTION", "Units", "NAV", "COST", "REMARKS",
                    "STT", "ADD CHARGES", "Current Price", "Gain%"]
        for col, h in enumerate(headers, 1):
            ws_th.cell(4, col, value=h)

        # Index sheet
        ws_idx = wb.create_sheet("Index")
        ws_idx.cell(1, 2, value="Code")
        ws_idx.cell(1, 3, value=fund_code)
        ws_idx.cell(2, 2, value="Current Price")
        ws_idx.cell(2, 3, value=0.0)
        ws_idx.cell(3, 2, value="52 Week High")
        ws_idx.cell(3, 3, value=0.0)
        ws_idx.cell(4, 2, value="52 Week Low")
        ws_idx.cell(4, 3, value=0.0)

        wb.save(filepath)
        wb.close()

        # Register in maps
        self._file_map[fund_code] = filepath
        self._name_map[fund_code] = fund_name
        print(f"[MF-XlsxDB] Created new MF file: {filepath.name}")
        return filepath

    def _find_header_row(self, ws):
        """Find the header row in a Trading History sheet."""
        max_row = ws.max_row or 0
        for r in range(1, min(11, max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, 5)]
            if "DATE" in vals and "ACTION" in vals:
                return r
        return 4  # default

    def add_mf_holding(
        self,
        fund_code: str,
        fund_name: str,
        units: float,
        nav: float,
        buy_date: str,
        remarks: str = "",
    ) -> dict:
        """Add a Buy transaction for a mutual fund.

        If the fund doesn't exist, creates a new xlsx file.
        Returns a dict with the new holding info.
        """
        with self._lock:
            # Find or create file
            filepath = self._file_map.get(fund_code)
            if not filepath or not filepath.exists():
                filepath = self._create_mf_file(fund_code, fund_name)

            wb = openpyxl.load_workbook(filepath)
            ws = wb["Trading History"]
            header_row = self._find_header_row(ws)

            # ── Duplicate check: same date + units + NAV ──
            try:
                dt_check = datetime.strptime(buy_date, "%Y-%m-%d")
            except ValueError:
                dt_check = None

            if dt_check:
                for row in range(header_row + 1, ws.max_row + 1):
                    action = ws.cell(row, 3).value
                    if action != "Buy":
                        continue
                    row_date = ws.cell(row, 1).value
                    row_units = ws.cell(row, 4).value
                    row_nav = ws.cell(row, 5).value
                    # Normalize date comparison
                    if isinstance(row_date, datetime):
                        row_date = row_date.date()
                    elif isinstance(row_date, str):
                        try:
                            row_date = datetime.strptime(row_date.strip(), "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    if (row_date == dt_check.date()
                            and abs(float(row_units or 0) - units) < 1e-4
                            and abs(float(row_nav or 0) - nav) < 1e-2):
                        wb.close()
                        raise ValueError(
                            f"Duplicate: {units:.4f} units @ NAV {nav:.4f} on {buy_date} already exists for {fund_name}"
                        )

            # Insert Buy row at top (below header)
            insert_at = header_row + 1
            ws.insert_rows(insert_at)

            # Parse date
            try:
                dt_obj = datetime.strptime(buy_date, "%Y-%m-%d")
            except ValueError:
                dt_obj = datetime.now()

            cost = round(units * nav, 2)

            ws.cell(insert_at, 1, value=dt_obj)      # A: DATE
            ws.cell(insert_at, 2, value="NSE")        # B: EXCH
            ws.cell(insert_at, 3, value="Buy")        # C: ACTION
            ws.cell(insert_at, 4, value=round(units, 6))  # D: Units
            ws.cell(insert_at, 5, value=round(nav, 4))    # E: NAV
            ws.cell(insert_at, 6, value=cost)              # F: COST
            ws.cell(insert_at, 7, value=remarks or "~")   # G: REMARKS

            wb.save(filepath)
            wb.close()

            # Invalidate cache to force re-parse
            self._cache.pop(fund_code, None)

            print(f"[MF-XlsxDB] Added Buy: {units:.4f} units of {fund_name} @ NAV {nav:.4f}")

            return {
                "fund_code": fund_code,
                "fund_name": fund_name,
                "units": round(units, 6),
                "nav": round(nav, 4),
                "cost": cost,
                "buy_date": buy_date,
                "message": f"Added {units:.4f} units of {fund_name}",
            }

    def add_mf_sell_transaction(
        self,
        fund_code: str,
        units: float,
        nav: float,
        sell_date: str = "",
        remarks: str = "",
    ) -> dict:
        """Record a Sell (redemption) for a mutual fund.

        Inserts a Sell row, uses FIFO matching to compute realized P&L.
        Returns dict with realized_pl and remaining_units.
        """
        with self._lock:
            filepath = self._file_map.get(fund_code)
            if not filepath or not filepath.exists():
                raise ValueError(f"No file found for fund {fund_code}")

            if not sell_date:
                sell_date = datetime.now().strftime("%Y-%m-%d")

            # First, get current holdings to validate
            holdings, _, _ = self._get_fund_data(fund_code)
            total_held = sum(h.units for h in holdings)
            if units > total_held + 0.0001:
                raise ValueError(
                    f"Cannot redeem {units:.4f} units. Only {total_held:.4f} held."
                )

            wb = openpyxl.load_workbook(filepath)
            ws = wb["Trading History"]
            header_row = self._find_header_row(ws)

            # Insert Sell row at top
            insert_at = header_row + 1
            ws.insert_rows(insert_at)

            try:
                dt_obj = datetime.strptime(sell_date, "%Y-%m-%d")
            except ValueError:
                dt_obj = datetime.now()

            cost = round(units * nav, 2)

            ws.cell(insert_at, 1, value=dt_obj)           # A: DATE
            ws.cell(insert_at, 2, value="NSE")             # B: EXCH
            ws.cell(insert_at, 3, value="Sell")            # C: ACTION
            ws.cell(insert_at, 4, value=round(units, 6))   # D: Units
            ws.cell(insert_at, 5, value=round(nav, 4))     # E: NAV
            ws.cell(insert_at, 6, value=cost)               # F: COST
            ws.cell(insert_at, 7, value=remarks or "~")    # G: REMARKS

            wb.save(filepath)
            wb.close()

            # Invalidate cache and re-parse to get FIFO-matched realized P&L
            self._cache.pop(fund_code, None)

            # Re-read to compute realized P&L via FIFO
            new_holdings, sold_positions, _ = self._get_fund_data(fund_code)
            remaining_units = sum(h.units for h in new_holdings)

            # The sell we just inserted — find matching sold positions
            # that include the sell_date
            realized_pl = 0.0
            for sp in sold_positions:
                if sp.sell_date == sell_date and abs(sp.sell_nav - nav) < 0.01:
                    realized_pl += sp.realized_pl

            fund_name = self._name_map.get(fund_code, fund_code)
            print(f"[MF-XlsxDB] Redeemed {units:.4f} units of {fund_name} @ NAV {nav:.4f}, P&L={realized_pl:.2f}")

            return {
                "message": f"Redeemed {units:.4f} units",
                "realized_pl": round(realized_pl, 2),
                "remaining_units": round(remaining_units, 6),
            }

    def get_fund_nav(self, fund_code: str) -> float:
        """Get the current NAV for a fund — live first, then xlsx fallback."""
        live = fetch_live_navs([fund_code])
        if fund_code in live and live[fund_code] > 0:
            return live[fund_code]
        try:
            _, _, idx_data = self._get_fund_data(fund_code)
            return idx_data.get("current_nav", 0.0)
        except Exception:
            return 0.0


# ═══════════════════════════════════════════════════════════
#  MODULE-LEVEL SINGLETON
# ═══════════════════════════════════════════════════════════

from app.config import DUMPS_DIR as _DUMPS_DIR
_MF_DIR = _DUMPS_DIR / "Mutual Funds"

mf_db = MFXlsxPortfolio(_MF_DIR)
