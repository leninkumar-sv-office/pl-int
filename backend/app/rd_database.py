"""
Recurring Deposit database layer.

Reads xlsx metadata from dumps/RD/ (rows 1-3 only) and computes ALL
installments with quarterly compound interest in Python.

Also supports manual CRUD via JSON file and creates xlsx files for new entries.

xlsx layout:
    Row 1: [_, start_date(B), _, _, =SUM(E6:E66), _, _, tenure_months(H), _, _, bank(K)]
    Row 2: [_, =end_date(B), _, _, =interest(E), _, _, payout(H), _, _, frequency(K)]
    Row 3: [_, rate_decimal(B), _, _, =maturity(E), _, _, sip(H)]
    Row 5: headers
    Row 6+: monthly data rows

Compound interest formula (quarterly, freq=4):
    Every 4th month: interest = (SIP * month_num + cumulative_interest) * rate * freq / 12
    This compounds because cumulative_interest includes all previous interest.

Special cases:
    - File 020123343379: E6=10000 (double first payment, hardcoded)
    - All others: E6 formula = IF(D6-NOW()<=0, $H$3, "")
"""

import json
import hashlib
import threading
import uuid
import re
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from .models import RDItem

import logging
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════
#  FILE PATHS
# ═══════════════════════════════════════════════════════════

from app.config import DUMPS_DIR
RD_XLSX_DIR = DUMPS_DIR / "RD"
RD_JSON_FILE = DUMPS_DIR / "recurring_deposits.json"

_lock = threading.Lock()


def _sync_to_drive(filepath: Path):
    try:
        from .config import DUMPS_BASE
        from . import drive_service
        rel = filepath.resolve().relative_to(DUMPS_BASE.resolve())
        drive_service.sync_dumps_file(str(rel))
    except Exception:
        pass


def _delete_from_drive(filepath: Path):
    """Delete a file from Google Drive by its local path."""
    try:
        from .config import DUMPS_BASE
        from . import drive_service
        rel = filepath.resolve().relative_to(DUMPS_BASE.resolve())
        parts = Path(rel).parts
        email = next((p for p in parts if "@" in p), "")
        if len(parts) > 1:
            subfolder = "dumps/" + "/".join(parts[:-1])
        else:
            subfolder = "dumps"
        drive_service.delete_file(filepath.name, subfolder=subfolder, email=email)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _gen_rd_id(name: str) -> str:
    """Deterministic ID from RD name."""
    return hashlib.md5(name.encode()).hexdigest()[:8]


def _extract_account_number(filename: str) -> str:
    """Extract account number from filename like 'Post Office RD - 020123343379'."""
    match = re.search(r'(\d{10,})', filename)
    return match.group(1) if match else ""


def _to_date(val) -> date | None:
    """Convert openpyxl cell value to date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


# ═══════════════════════════════════════════════════════════
#  XLSX PARSER — Python-computed installments
# ═══════════════════════════════════════════════════════════

def _parse_rd_xlsx(filepath: Path) -> dict:
    """Parse a single RD xlsx file.

    Reads metadata from rows 1-3 only. Checks row 6 for special
    first-payment overrides (hardcoded values). Computes all
    installments and compound interest in Python.
    """
    name = filepath.stem
    account_number = _extract_account_number(name)

    # ── Read metadata (data_only=True for cached non-formula values) ──
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb["Index"]

    # Preload rows for read_only mode
    all_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    wb.close()

    row1 = all_rows[0] if len(all_rows) > 0 else ()
    row2 = all_rows[1] if len(all_rows) > 1 else ()
    row3 = all_rows[2] if len(all_rows) > 2 else ()

    start_date_raw = row1[1] if len(row1) > 1 else None       # B1
    maturity_months_raw = (row1[7] if len(row1) > 7 else None) or 60   # H1
    bank = (row1[10] if len(row1) > 10 else None) or "Post Office"    # K1

    interest_payout = (row2[7] if len(row2) > 7 else None) or "Maturity"  # H2
    frequency = (row2[10] if len(row2) > 10 else None) or 4              # K2

    rate_decimal = (row3[1] if len(row3) > 1 else None) or 0   # B3
    sip = (row3[7] if len(row3) > 7 else None) or 0            # H3

    # ── Check for special first-payment (read WITHOUT data_only) ──
    first_payment = None
    try:
        wb_formula = openpyxl.load_workbook(str(filepath), read_only=True)
        ws_formula = wb_formula["Index"]
        formula_rows = list(ws_formula.iter_rows(min_row=6, max_row=6, values_only=True))
        wb_formula.close()
        if formula_rows:
            e6_val = formula_rows[0][4] if len(formula_rows[0]) > 4 else None  # E6 (col 5, 0-indexed=4)
            # If E6 is a number (not a formula string), it's hardcoded
            if isinstance(e6_val, (int, float)):
                first_payment = float(e6_val)
    except Exception:
        pass

    # ── Convert & derive ─────────────────────────────────
    start_dt = _to_date(start_date_raw) or date.today()
    rate_decimal = float(rate_decimal)
    sip = float(sip)
    freq = int(frequency)
    tenure_months = int(maturity_months_raw)
    rate_pct = round(rate_decimal * 100, 4) if rate_decimal < 1 else round(rate_decimal, 4)
    rate_for_calc = rate_decimal if rate_decimal < 1 else rate_decimal / 100

    end_dt = start_dt + relativedelta(months=tenure_months)

    # If first_payment differs from SIP, use it; otherwise ignore
    if first_payment is not None and abs(first_payment - sip) > 0.01:
        month1_amount = first_payment
    else:
        month1_amount = sip

    # ── Generate installments with compound interest ──────
    today = date.today()
    installments = []
    cumulative_interest = 0.0
    total_deposited_past = 0.0
    total_deposited_all = 0.0
    total_interest_earned = 0.0
    total_interest_projected = 0.0

    for m in range(1, tenure_months + 1):
        inst_date = start_dt + relativedelta(months=m - 1)
        is_past = inst_date <= today

        # Monthly deposit
        invested = month1_amount if m == 1 else sip
        total_deposited_all += invested
        if is_past:
            total_deposited_past += invested

        # Compound interest at every freq-th month
        is_compound_month = (freq > 0 and m % freq == 0)
        compound_interest = 0.0

        if is_compound_month:
            # Formula: (SIP * month + cumulative_interest) * rate * freq / 12
            compound_interest = round(
                (sip * m + cumulative_interest) * rate_for_calc * freq / 12, 2
            )
            cumulative_interest += compound_interest

        if is_past:
            total_interest_earned += compound_interest
        else:
            total_interest_projected += compound_interest

        installments.append({
            "month": m,
            "date": inst_date.strftime("%Y-%m-%d"),
            "amount_invested": round(invested, 2),
            "interest_earned": round(compound_interest, 2) if is_past else 0.0,
            "interest_projected": round(compound_interest, 2) if not is_past else 0.0,
            "is_compound_month": is_compound_month,
            "cumulative_interest": round(cumulative_interest, 2),
            "is_past": is_past,
        })

    # ── Totals ────────────────────────────────────────────
    maturity_amount = round(total_deposited_all + cumulative_interest, 2)
    status = "Matured" if end_dt <= today else "Active"
    days_to_maturity = max(0, (end_dt - today).days)

    installments_paid = sum(1 for i in installments if i["is_past"])

    return {
        "id": _gen_rd_id(name),
        "name": name,
        "account_number": account_number,
        "bank": bank,
        "monthly_amount": round(sip, 2),
        "interest_rate": rate_pct,
        "compounding_frequency": freq,
        "interest_payout": interest_payout,
        "tenure_months": tenure_months,
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "maturity_date": end_dt.strftime("%Y-%m-%d"),
        "maturity_amount": maturity_amount,
        "total_deposited": round(total_deposited_past, 2),
        "total_interest_accrued": round(total_interest_earned, 2),
        "total_interest_projected": round(total_interest_projected, 2),
        "interest_earned": round(total_interest_earned, 2),
        "status": status,
        "days_to_maturity": days_to_maturity,
        "source": "xlsx",
        "remarks": "",
        "installments": installments,
        "installments_paid": installments_paid,
        "installments_total": len(installments),
    }


def _parse_all_xlsx(xlsx_dir: Path = None) -> list:
    """Parse all xlsx files from dumps/RD/ directory."""
    xlsx_dir = xlsx_dir or RD_XLSX_DIR
    results = []
    if not xlsx_dir.exists():
        return results

    for f in sorted(xlsx_dir.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        if "_Archive" in str(f):
            continue
        try:
            parsed = _parse_rd_xlsx(f)
            results.append(parsed)
        except Exception as e:
            logger.error(f"[RD] Error parsing {f.name}: {e}")
    return results


# ═══════════════════════════════════════════════════════════
#  XLSX CREATION (for new entries from UI)
# ═══════════════════════════════════════════════════════════

def _create_rd_xlsx(name: str, bank: str, monthly_amount: float,
                    rate_pct: float, tenure_months: int, start_date: str,
                    frequency: int = 4, xlsx_dir: Path = None):
    """Create an xlsx file following the RD template structure."""
    xlsx_dir = xlsx_dir or RD_XLSX_DIR
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    filepath = xlsx_dir / f"{name}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index"

    rate_dec = rate_pct / 100
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")

    # Row 1: Start Date, Latest Amount, Tenure Months, Bank
    ws.cell(1, 2, start_dt)
    ws.cell(1, 8, tenure_months)       # H1
    ws.cell(1, 11, bank)               # K1

    # Row 2: End Date, Payout, Frequency
    end_dt = start_dt + relativedelta(months=tenure_months)
    ws.cell(2, 2, end_dt)
    ws.cell(2, 8, "Maturity")          # H2
    ws.cell(2, 11, float(frequency))   # K2

    # Row 3: Rate (decimal), SIP
    ws.cell(3, 2, rate_dec)            # B3
    ws.cell(3, 8, monthly_amount)      # H3

    # Row 5: Headers
    headers = ['S.No', '', '#', 'Date', 'Amount Invested', 'Interest Earned', 'Interest to be Earned']
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)

    # Data rows: compute values
    cumulative_interest = 0.0
    total_deposited = 0.0
    sip = monthly_amount

    for m in range(1, tenure_months + 1):
        row = m + 5
        inst_date = start_dt + relativedelta(months=m - 1)
        ws.cell(row, 1, m)
        ws.cell(row, 3, m)
        ws.cell(row, 4, inst_date)
        ws.cell(row, 5, sip)
        total_deposited += sip

        is_compound = (frequency > 0 and m % frequency == 0)
        interest = 0.0
        if is_compound:
            interest = round((sip * m + cumulative_interest) * rate_dec * frequency / 12, 2)
            cumulative_interest += interest

        ws.cell(row, 6, interest)  # Interest earned
        ws.cell(row, 7, interest)  # Interest projected

    # Set E1 (total deposited) and E3 (maturity amount)
    ws.cell(1, 5, total_deposited)
    ws.cell(2, 5, cumulative_interest)
    ws.cell(3, 5, round(total_deposited + cumulative_interest, 2))

    wb.save(str(filepath))
    wb.close()
    _sync_to_drive(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════
#  JSON LOAD / SAVE (manual entries)
# ═══════════════════════════════════════════════════════════

def _load_json(json_file: Path = None) -> list:
    json_file = json_file or RD_JSON_FILE
    if not json_file.exists():
        return []
    try:
        with open(json_file, "r") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except (json.JSONDecodeError, Exception):
        return []


def _save_json(data: list, json_file: Path = None, dumps_dir: Path = None):
    dumps_dir = dumps_dir or DUMPS_DIR
    json_file = json_file or RD_JSON_FILE
    dumps_dir.mkdir(parents=True, exist_ok=True)
    with open(json_file, "w") as f:
        json.dump(data, f, indent=2)
    _sync_to_drive(json_file)


# ═══════════════════════════════════════════════════════════
#  CALCULATIONS (for manual/JSON entries)
# ═══════════════════════════════════════════════════════════

def _compute_rd_installments(monthly_amount: float, rate_pct: float,
                             tenure_months: int, start_date: str,
                             frequency: int = 4) -> list:
    """Generate full installment schedule with compound interest."""
    today = date.today()
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    rate_dec = rate_pct / 100
    cumulative_interest = 0.0
    installments = []

    for m in range(1, tenure_months + 1):
        inst_date = start_dt + relativedelta(months=m - 1)
        is_past = inst_date <= today
        is_compound = (frequency > 0 and m % frequency == 0)

        compound_interest = 0.0
        if is_compound:
            compound_interest = round(
                (monthly_amount * m + cumulative_interest) * rate_dec * frequency / 12, 2
            )
            cumulative_interest += compound_interest

        installments.append({
            "month": m,
            "date": inst_date.strftime("%Y-%m-%d"),
            "amount_invested": round(monthly_amount, 2),
            "interest_earned": round(compound_interest, 2) if is_past else 0.0,
            "interest_projected": round(compound_interest, 2) if not is_past else 0.0,
            "is_compound_month": is_compound,
            "cumulative_interest": round(cumulative_interest, 2),
            "is_past": is_past,
        })

    return installments


def _calc_maturity_date(start_date: str, tenure_months: int) -> str:
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        maturity = start + relativedelta(months=tenure_months)
        return maturity.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return ""


def _enrich_json_item(item: dict) -> dict:
    """Add computed fields to a JSON-based RD item."""
    today = date.today()
    monthly = item.get("monthly_amount", 0)
    rate = item.get("interest_rate", 0)
    tenure = item.get("tenure_months", 60)
    start_date = item.get("start_date", "")
    freq = item.get("compounding_frequency", 4)

    # Generate installments
    if start_date and monthly > 0:
        installments = _compute_rd_installments(monthly, rate, tenure, start_date, freq)
    else:
        installments = item.get("installments", [])

    total_deposited = sum(i.get("amount_invested", 0) for i in installments if i.get("is_past"))
    total_interest_earned = sum(i.get("interest_earned", 0) for i in installments)
    total_interest_projected = sum(i.get("interest_projected", 0) for i in installments)
    cumulative = max((i.get("cumulative_interest", 0) for i in installments), default=0)
    total_all_deposits = sum(i.get("amount_invested", 0) for i in installments)

    try:
        mat = datetime.strptime(item.get("maturity_date", ""), "%Y-%m-%d").date()
        days_to_maturity = max(0, (mat - today).days)
        status = "Matured" if mat <= today else "Active"
    except (ValueError, KeyError):
        days_to_maturity = 0
        status = item.get("status", "Active")

    item["source"] = "manual"
    item["name"] = item.get("name", f"{item.get('bank', 'RD')} RD")
    item["account_number"] = item.get("account_number", "")
    item["compounding_frequency"] = freq
    item["interest_payout"] = item.get("interest_payout", "Maturity")
    item["total_deposited"] = round(total_deposited, 2)
    item["total_interest_accrued"] = round(total_interest_earned, 2)
    item["total_interest_projected"] = round(total_interest_projected, 2)
    item["interest_earned"] = round(total_interest_earned, 2)
    item["maturity_amount"] = round(total_all_deposits + cumulative, 2)
    item["installments"] = installments
    item["installments_paid"] = sum(1 for i in installments if i.get("is_past"))
    item["installments_total"] = len(installments)
    item["days_to_maturity"] = days_to_maturity
    item["status"] = status
    return item


# ═══════════════════════════════════════════════════════════
#  PUBLIC API
# ═══════════════════════════════════════════════════════════

def get_all(base_dir=None) -> list:
    """Return all RDs: xlsx-parsed + JSON manual entries."""
    xlsx_dir = (Path(base_dir) / "RD") if base_dir else RD_XLSX_DIR
    json_file = (Path(base_dir) / "recurring_deposits.json") if base_dir else RD_JSON_FILE
    with _lock:
        xlsx_items = _parse_all_xlsx(xlsx_dir=xlsx_dir)
        json_items = _load_json(json_file=json_file)

    for item in json_items:
        _enrich_json_item(item)

    return xlsx_items + json_items


def get_dashboard(base_dir=None) -> dict:
    """Aggregate RD summary for dashboard."""
    items = get_all(base_dir=base_dir)
    active = [i for i in items if i.get("status") == "Active"]

    monthly_commitment = sum(i.get("monthly_amount", 0) for i in active)
    total_deposited = sum(i.get("total_deposited", 0) for i in active)
    total_maturity = sum(i.get("maturity_amount", 0) for i in active)
    total_interest_accrued = sum(i.get("total_interest_accrued", 0) for i in active)

    return {
        "total_deposited": round(total_deposited, 2),
        "total_maturity_value": round(total_maturity, 2),
        "total_interest_accrued": round(total_interest_accrued, 2),
        "monthly_commitment": round(monthly_commitment, 2),
        "active_count": len(active),
        "total_count": len(items),
    }


def add(data: dict, base_dir=None) -> dict:
    """Add a new RD — creates xlsx file + JSON entry."""
    xlsx_dir = (Path(base_dir) / "RD") if base_dir else RD_XLSX_DIR
    monthly = data["monthly_amount"]
    rate = data["interest_rate"]
    tenure = data["tenure_months"]
    start_date = data["start_date"]
    bank = data["bank"]
    freq = data.get("compounding_frequency", 4)
    name = data.get("name", f"{bank} RD")
    account_number = data.get("account_number", "")

    if account_number:
        name = f"{bank} RD - {account_number}"

    maturity_date = data.get("maturity_date") or _calc_maturity_date(start_date, tenure)

    # Compute maturity
    installments = _compute_rd_installments(monthly, rate, tenure, start_date, freq)
    total_all_deposits = monthly * tenure
    cumulative = max((i.get("cumulative_interest", 0) for i in installments), default=0)
    maturity_amount = round(total_all_deposits + cumulative, 2)

    # Create xlsx file (this is the only storage — parser picks it up on next load)
    _create_rd_xlsx(
        name=name,
        bank=bank,
        monthly_amount=monthly,
        rate_pct=rate,
        tenure_months=tenure,
        start_date=start_date,
        frequency=freq,
        xlsx_dir=xlsx_dir,
    )

    return {
        "id": _gen_rd_id(name),
        "name": name,
        "account_number": account_number,
        "bank": bank,
        "monthly_amount": monthly,
        "interest_rate": rate,
        "compounding_frequency": freq,
        "tenure_months": tenure,
        "start_date": start_date,
        "maturity_date": maturity_date,
        "maturity_amount": maturity_amount,
        "total_deposited": 0,
        "status": data.get("status", "Active"),
        "remarks": data.get("remarks", ""),
    }


def update(rd_id: str, data: dict, base_dir=None) -> dict:
    """Update an existing RD — supports both manual (JSON) and xlsx-imported."""
    xlsx_dir = (Path(base_dir) / "RD") if base_dir else RD_XLSX_DIR
    json_file = (Path(base_dir) / "recurring_deposits.json") if base_dir else RD_JSON_FILE
    dumps_dir = Path(base_dir) if base_dir else DUMPS_DIR

    # Try xlsx first
    if xlsx_dir.exists():
        for f in xlsx_dir.glob("*.xlsx"):
            if f.name.startswith("~$"):
                continue
            if _gen_rd_id(f.stem) == rd_id:
                return _update_xlsx_rd(f, data, dumps_dir)

    # Fall back to JSON (manual)
    with _lock:
        items = _load_json(json_file=json_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == rd_id), None)
        if idx is None:
            raise ValueError(f"RD {rd_id} not found")

        item = items[idx]
        for key, val in data.items():
            if val is not None and key in item and key != "installments":
                item[key] = val

        # Recompute maturity
        freq = item.get("compounding_frequency", 4)
        installments = _compute_rd_installments(
            item["monthly_amount"], item["interest_rate"],
            item["tenure_months"], item["start_date"], freq
        )
        total_all_deposits = item["monthly_amount"] * item["tenure_months"]
        cumulative = max((i.get("cumulative_interest", 0) for i in installments), default=0)
        item["maturity_amount"] = round(total_all_deposits + cumulative, 2)

        if "start_date" in data or "tenure_months" in data:
            if not data.get("maturity_date"):
                item["maturity_date"] = _calc_maturity_date(item["start_date"], item["tenure_months"])

        items[idx] = item
        _save_json(items, json_file=json_file, dumps_dir=dumps_dir)
        return item


def _update_xlsx_rd(filepath: Path, data: dict, dumps_dir: Path) -> dict:
    """Update an xlsx-imported RD by editing cells in the Index sheet.
    Supports: bank, monthly_amount, interest_rate, start_date, tenure_months, compounding_frequency."""
    import openpyxl as xl

    wb = xl.load_workbook(str(filepath))
    ws = wb["Index"]

    if "start_date" in data:
        try:
            dt = datetime.strptime(data["start_date"], "%Y-%m-%d")
            ws.cell(1, 2, value=dt)  # B1
        except ValueError:
            pass
    if "tenure_months" in data:
        ws.cell(1, 8, value=int(data["tenure_months"]))  # H1 (months for RD)
    if "bank" in data:
        ws.cell(1, 11, value=data["bank"])  # K1
    if "interest_payout" in data:
        ws.cell(2, 8, value=data["interest_payout"])  # H2
    if "compounding_frequency" in data:
        ws.cell(2, 11, value=int(data["compounding_frequency"]))  # K2
    if "interest_rate" in data:
        rate = float(data["interest_rate"])
        ws.cell(3, 2, value=rate / 100 if rate > 1 else rate)  # B3
    if "monthly_amount" in data:
        ws.cell(3, 8, value=float(data["monthly_amount"]))  # H3

    wb.save(str(filepath))
    wb.close()

    # Sync to drive
    try:
        from .config import DUMPS_BASE
        from . import drive_service
        rel = filepath.resolve().relative_to(DUMPS_BASE.resolve())
        drive_service.sync_dumps_file(str(rel))
    except Exception:
        pass

    return _parse_rd_xlsx(filepath)


def delete(rd_id: str, base_dir=None) -> dict:
    """Delete an RD — removes xlsx file and/or JSON entry."""
    xlsx_dir = (Path(base_dir) / "RD") if base_dir else RD_XLSX_DIR
    json_file = (Path(base_dir) / "recurring_deposits.json") if base_dir else RD_JSON_FILE
    dumps_dir = Path(base_dir) if base_dir else DUMPS_DIR
    # Try xlsx first
    if xlsx_dir.exists():
        for f in xlsx_dir.glob("*.xlsx"):
            if f.name.startswith("~$"):
                continue
            if _gen_rd_id(f.stem) == rd_id:
                _delete_from_drive(f)
                f.unlink()
                return {"message": f"RD {rd_id} deleted (xlsx)", "item": {"id": rd_id, "name": f.stem}}

    # Try JSON
    with _lock:
        items = _load_json(json_file=json_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == rd_id), None)
        if idx is None:
            raise ValueError(f"RD {rd_id} not found")
        removed = items.pop(idx)
        _save_json(items, json_file=json_file, dumps_dir=dumps_dir)
        return {"message": f"RD {rd_id} deleted", "item": removed}


def add_installment(rd_id: str, installment: dict, base_dir=None) -> dict:
    """Add an installment to a manual RD (for tracking extra payments)."""
    json_file = (Path(base_dir) / "recurring_deposits.json") if base_dir else RD_JSON_FILE
    dumps_dir = Path(base_dir) if base_dir else DUMPS_DIR
    with _lock:
        items = _load_json(json_file=json_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == rd_id), None)
        if idx is None:
            raise ValueError(f"RD {rd_id} not found")

        item = items[idx]
        if "extra_installments" not in item:
            item["extra_installments"] = []

        item["extra_installments"].append({
            "date": installment["date"],
            "amount": installment["amount"],
            "remarks": installment.get("remarks", ""),
        })

        items[idx] = item
        _save_json(items, json_file=json_file, dumps_dir=dumps_dir)
        return item
