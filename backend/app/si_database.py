"""
XLSX-based database layer for Standing Instructions (NACH/ECS mandates).

Stores all SIs in a single xlsx file:
    dumps/Standing Instructions/Standing Instructions.xlsx

xlsx layout:
    Row 1: Headers
        A=ID, B=Bank, C=Beneficiary, D=Amount, E=Frequency, F=Purpose,
        G=Mandate Type, H=Account Number, I=Start Date, J=Expiry Date,
        K=Alert Days, L=Status, M=Remarks
    Row 2+: Data rows (one per SI)
"""

import hashlib
import threading
import uuid
from datetime import datetime, date
from pathlib import Path

import openpyxl

from .models import SIItem

import logging
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════
#  FILE PATH
# ═══════════════════════════════════════════════════════════

from app.config import DUMPS_DIR
SI_DIR = DUMPS_DIR / "Standing Instructions"
SI_FILE = SI_DIR / "Standing Instructions.xlsx"


def _sync_to_drive(filepath: Path):
    """No-op — Google Drive desktop sync handles file upload automatically."""
    pass


def _delete_from_drive(filepath: Path):
    """Delete a file from Google Drive by its local path."""
    try:
        from .config import DUMPS_BASE
        from . import drive_service
        rel = filepath.resolve().relative_to(DUMPS_BASE.resolve())
        parts = Path(rel).parts
        email = next((p for p in parts if "@" in p), "")
        if len(parts) > 1:
            subfolder = "dumps/" + "/".join(parts[:-1])
        else:
            subfolder = "dumps"
        drive_service.delete_file(filepath.name, subfolder=subfolder, email=email)
    except Exception:
        pass

_lock = threading.Lock()

# Column mapping (1-indexed for openpyxl)
_COLS = {
    "id": 1,
    "bank": 2,
    "beneficiary": 3,
    "amount": 4,
    "frequency": 5,
    "purpose": 6,
    "mandate_type": 7,
    "account_number": 8,
    "start_date": 9,
    "expiry_date": 10,
    "alert_days": 11,
    "status": 12,
    "remarks": 13,
}

_HEADERS = ["ID", "Bank", "Beneficiary", "Amount", "Frequency", "Purpose",
            "Mandate Type", "Account Number", "Start Date", "Expiry Date",
            "Alert Days", "Status", "Remarks"]


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _to_date_str(val) -> str:
    """Convert openpyxl cell value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str) and val.strip():
        return val.strip()
    return ""


def _ensure_file(si_dir: Path = None, si_file: Path = None):
    """Create the xlsx file with headers if it doesn't exist."""
    si_dir = si_dir or SI_DIR
    si_file = si_file or SI_FILE
    si_dir.mkdir(parents=True, exist_ok=True)
    if si_file.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standing Instructions"
    for i, h in enumerate(_HEADERS, 1):
        ws.cell(1, i, h)
    wb.save(str(si_file))
    wb.close()


# ═══════════════════════════════════════════════════════════
#  LOAD / SAVE
# ═══════════════════════════════════════════════════════════

def _load(si_file: Path = None) -> list:
    """Read all SI rows from the xlsx file."""
    si_file = si_file or SI_FILE
    if not si_file.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(si_file), data_only=True, read_only=True)
    except Exception as e:
        logger.error(f"[SI] Error loading {si_file}: {e}")
        return []

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    items = []
    for row in all_rows[1:]:  # skip header row (index 0)
        # _COLS values are 1-indexed; convert to 0-indexed
        row_id = row[_COLS["id"] - 1] if len(row) >= _COLS["id"] else None
        if not row_id:
            continue
        items.append({
            "id": str(row_id),
            "bank": str((row[_COLS["bank"] - 1] if len(row) >= _COLS["bank"] else None) or ""),
            "beneficiary": str((row[_COLS["beneficiary"] - 1] if len(row) >= _COLS["beneficiary"] else None) or ""),
            "amount": float((row[_COLS["amount"] - 1] if len(row) >= _COLS["amount"] else None) or 0),
            "frequency": str((row[_COLS["frequency"] - 1] if len(row) >= _COLS["frequency"] else None) or "Monthly"),
            "purpose": str((row[_COLS["purpose"] - 1] if len(row) >= _COLS["purpose"] else None) or "SIP"),
            "mandate_type": str((row[_COLS["mandate_type"] - 1] if len(row) >= _COLS["mandate_type"] else None) or "NACH"),
            "account_number": str((row[_COLS["account_number"] - 1] if len(row) >= _COLS["account_number"] else None) or ""),
            "start_date": _to_date_str(row[_COLS["start_date"] - 1] if len(row) >= _COLS["start_date"] else None),
            "expiry_date": _to_date_str(row[_COLS["expiry_date"] - 1] if len(row) >= _COLS["expiry_date"] else None),
            "alert_days": int((row[_COLS["alert_days"] - 1] if len(row) >= _COLS["alert_days"] else None) or 30),
            "status": str((row[_COLS["status"] - 1] if len(row) >= _COLS["status"] else None) or "Active"),
            "remarks": str((row[_COLS["remarks"] - 1] if len(row) >= _COLS["remarks"] else None) or ""),
        })
    return items


def _save(items: list, si_dir: Path = None, si_file: Path = None):
    """Write all SI rows to the xlsx file (full rewrite)."""
    si_dir = si_dir or SI_DIR
    si_file = si_file or SI_FILE
    _ensure_file(si_dir, si_file)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standing Instructions"

    # Header row
    for i, h in enumerate(_HEADERS, 1):
        ws.cell(1, i, h)

    # Data rows
    for r, item in enumerate(items, 2):
        ws.cell(r, _COLS["id"], item["id"])
        ws.cell(r, _COLS["bank"], item["bank"])
        ws.cell(r, _COLS["beneficiary"], item["beneficiary"])
        ws.cell(r, _COLS["amount"], item["amount"])
        ws.cell(r, _COLS["frequency"], item["frequency"])
        ws.cell(r, _COLS["purpose"], item["purpose"])
        ws.cell(r, _COLS["mandate_type"], item["mandate_type"])
        ws.cell(r, _COLS["account_number"], item["account_number"])
        # Write dates as date objects for proper Excel formatting
        try:
            ws.cell(r, _COLS["start_date"], datetime.strptime(item["start_date"], "%Y-%m-%d"))
        except (ValueError, TypeError):
            ws.cell(r, _COLS["start_date"], item["start_date"])
        try:
            ws.cell(r, _COLS["expiry_date"], datetime.strptime(item["expiry_date"], "%Y-%m-%d"))
        except (ValueError, TypeError):
            ws.cell(r, _COLS["expiry_date"], item["expiry_date"])
        ws.cell(r, _COLS["alert_days"], item["alert_days"])
        ws.cell(r, _COLS["status"], item["status"])
        ws.cell(r, _COLS["remarks"], item.get("remarks", ""))

    wb.save(str(si_file))
    wb.close()
    _sync_to_drive(Path(si_file))


# ═══════════════════════════════════════════════════════════
#  CRUD
# ═══════════════════════════════════════════════════════════

def get_all(base_dir=None) -> list:
    """Return all SIs with computed fields."""
    si_dir = (Path(base_dir) / "Standing Instructions") if base_dir else SI_DIR
    si_file = si_dir / "Standing Instructions.xlsx"
    with _lock:
        items = _load(si_file)
    today = datetime.now().date()
    for item in items:
        try:
            expiry = datetime.strptime(item["expiry_date"], "%Y-%m-%d").date()
            item["days_to_expiry"] = (expiry - today).days
        except (ValueError, KeyError):
            item["days_to_expiry"] = 0
    return items


def get_dashboard(base_dir=None) -> dict:
    """Aggregate SI summary for dashboard."""
    items = get_all(base_dir=base_dir)
    active = [i for i in items if i.get("status") == "Active"]

    # Normalize all frequencies to monthly equivalent
    total_monthly = 0.0
    for i in active:
        amt = i.get("amount", 0)
        freq = i.get("frequency", "Monthly")
        if freq == "Monthly":
            total_monthly += amt
        elif freq == "Quarterly":
            total_monthly += amt / 3
        elif freq == "Half-Yearly":
            total_monthly += amt / 6
        elif freq == "Annually":
            total_monthly += amt / 12
        else:
            total_monthly += amt  # default to monthly

    expiring_soon = sum(
        1 for i in active
        if 0 < i.get("days_to_expiry", 0) <= i.get("alert_days", 30)
    )

    return {
        "active_count": len(active),
        "total_count": len(items),
        "total_monthly_outflow": round(total_monthly, 2),
        "expiring_soon": expiring_soon,
    }


def add(data: dict, base_dir=None) -> dict:
    """Add a new standing instruction."""
    si_dir = (Path(base_dir) / "Standing Instructions") if base_dir else SI_DIR
    si_file = si_dir / "Standing Instructions.xlsx"
    with _lock:
        items = _load(si_file)

        si = {
            "id": str(uuid.uuid4())[:8],
            "bank": data["bank"],
            "beneficiary": data["beneficiary"],
            "amount": data["amount"],
            "frequency": data.get("frequency", "Monthly"),
            "purpose": data.get("purpose", "SIP"),
            "mandate_type": data.get("mandate_type", "NACH"),
            "account_number": data.get("account_number", ""),
            "start_date": data["start_date"],
            "expiry_date": data["expiry_date"],
            "alert_days": data.get("alert_days", 30),
            "status": data.get("status", "Active"),
            "remarks": data.get("remarks", ""),
        }

        items.append(si)
        _save(items, si_dir, si_file)
        return si


def update(si_id: str, data: dict, base_dir=None) -> dict:
    """Update an existing standing instruction."""
    si_dir = (Path(base_dir) / "Standing Instructions") if base_dir else SI_DIR
    si_file = si_dir / "Standing Instructions.xlsx"
    with _lock:
        items = _load(si_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == si_id), None)
        if idx is None:
            raise ValueError(f"Standing instruction {si_id} not found")

        item = items[idx]
        for key, val in data.items():
            if val is not None and key in item:
                item[key] = val

        items[idx] = item
        _save(items, si_dir, si_file)
        return item


def delete(si_id: str, base_dir=None) -> dict:
    """Delete a standing instruction."""
    si_dir = (Path(base_dir) / "Standing Instructions") if base_dir else SI_DIR
    si_file = si_dir / "Standing Instructions.xlsx"
    with _lock:
        items = _load(si_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == si_id), None)
        if idx is None:
            raise ValueError(f"Standing instruction {si_id} not found")
        removed = items.pop(idx)
        _save(items, si_dir, si_file)
        return {"message": f"SI {si_id} deleted", "item": removed}
