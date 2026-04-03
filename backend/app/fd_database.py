"""
Fixed Deposit / MIS database layer.

Reads xlsx metadata from dumps/FD/ (rows 1-3 only) and computes ALL
installments in Python — never trusts cached formula values from xlsx.

Also supports manual CRUD via JSON file and creates xlsx files for new entries.

xlsx layout (common to FD & MIS):
    Row 1: [_, start_date(B), _, _, =SUM(E6:E66), _, _, maturity_years(H), _, _, bank(K)]
    Row 2: [_, =end_date(B), _, _, =interest_earned(E), _, _, payout_type(H), _, _, frequency(K)]
    Row 3: [_, rate_decimal(B), _, _, =maturity_amt(E), _, _, sip(H)]
    Row 5: headers
    Row 6+: monthly data rows (formulas that reference NOW() — unreliable when read)

Interest logic (payout frequency driven):
    Monthly   → SIP * rate / 12 every month from month 2
    Quarterly → SIP * rate / 4  every 3 months from month 3
    Half-Yearly → SIP * rate / 2 every 6 months from month 6
    Annually  → SIP * rate every 12 months from month 12
    Month 1 = deposit month, no interest.

    xlsx FD files: payout read from H2 cell.
    MIS files always default to Monthly.
"""

import json
import hashlib
import threading
import uuid
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from .models import FDItem

import logging
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════
#  FILE PATHS
# ═══════════════════════════════════════════════════════════

from app.config import DUMPS_DIR
FD_XLSX_DIR = DUMPS_DIR / "FD"
FD_JSON_FILE = DUMPS_DIR / "fixed_deposits.json"

_lock = threading.Lock()


def _sync_to_drive(filepath: Path):
    """No-op — Google Drive desktop sync handles file upload automatically."""
    pass


def _delete_from_drive(filepath: Path):
    """Delete a file from Google Drive by its local path."""
    try:
        from .config import DUMPS_BASE
        from . import drive_service
        rel = filepath.resolve().relative_to(DUMPS_BASE.resolve())
        parts = Path(rel).parts
        email = next((p for p in parts if "@" in p), "")
        if len(parts) > 1:
            subfolder = "dumps/" + "/".join(parts[:-1])
        else:
            subfolder = "dumps"
        drive_service.delete_file(filepath.name, subfolder=subfolder, email=email)
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _gen_fd_id(name: str) -> str:
    """Deterministic ID from FD name."""
    return hashlib.md5(name.encode()).hexdigest()[:8]


def _payout_to_period(payout: str) -> int:
    """Convert interest payout string to period in months.
    Monthly → 1, Quarterly → 3, Half-Yearly → 6, Annually → 12.
    """
    p = (payout or "").strip().lower()
    if "month" in p:
        return 1
    if "quarter" in p or "quartely" in p:  # handle typo in xlsx
        return 3
    if "half" in p or "semi" in p:
        return 6
    if "annual" in p or "year" in p:
        return 12
    return 3  # default quarterly


def _payout_periods_per_year(period_months: int) -> int:
    """Number of interest periods per year."""
    return 12 // period_months if period_months > 0 else 4


def _to_date(val) -> date | None:
    """Convert openpyxl cell value to date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


# ═══════════════════════════════════════════════════════════
#  XLSX PARSER — Python-computed installments
# ═══════════════════════════════════════════════════════════

def _parse_fd_xlsx(filepath: Path) -> dict:
    """Parse a single FD/MIS xlsx file.

    Reads ONLY metadata from rows 1-3.
    Computes ALL installments and totals in Python.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    ws = wb["Index"]
    name = filepath.stem

    # ── Preload rows for read_only mode ───────────────────
    all_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    wb.close()

    # ── Metadata rows 1-3 ────────────────────────────────
    row1 = all_rows[0] if len(all_rows) > 0 else ()
    row2 = all_rows[1] if len(all_rows) > 1 else ()
    row3 = all_rows[2] if len(all_rows) > 2 else ()

    start_date_raw = row1[1] if len(row1) > 1 else None      # B1: start date
    maturity_years = (row1[7] if len(row1) > 7 else None) or 5   # H1: years
    bank = (row1[10] if len(row1) > 10 else None) or "Unknown"  # K1: bank

    interest_payout = (row2[7] if len(row2) > 7 else None) or "Quarterly"  # H2: payout type

    rate_decimal = (row3[1] if len(row3) > 1 else None) or 0   # B3: rate as decimal (0.0625)
    sip = (row3[7] if len(row3) > 7 else None) or 0            # H3: SIP / principal

    # ── Convert & derive ─────────────────────────────────
    start_dt = _to_date(start_date_raw) or date.today()
    rate_decimal = float(rate_decimal)
    sip = float(sip)
    maturity_years = float(maturity_years)
    tenure_months = int(maturity_years * 12)
    rate_pct = round(rate_decimal * 100, 4) if rate_decimal < 1 else round(rate_decimal, 4)
    rate_for_calc = rate_decimal if rate_decimal < 1 else rate_decimal / 100

    end_dt = start_dt + relativedelta(months=tenure_months)

    # Determine type and payout period
    is_mis = "MIS" in name.upper()
    fd_type = "MIS" if is_mis else "FD"
    period_months = _payout_to_period(interest_payout)
    if is_mis:
        period_months = 1  # MIS always pays monthly
    periods_per_year = _payout_periods_per_year(period_months)

    # ── Generate installments ─────────────────────────────
    today = date.today()
    installments = []
    total_interest_earned = 0.0
    total_interest_projected = 0.0

    for m in range(1, tenure_months + 1):
        inst_date = start_dt + relativedelta(months=m - 1)
        is_past = inst_date <= today

        # Investment: lump-sum in month 1 only
        invested = sip if m == 1 else 0.0

        # Interest: paid every period_months, starting from that period
        # e.g. Monthly → month 2,3,4,...; Quarterly → month 3,6,9,...
        interest = 0.0
        if m > 1 and m % period_months == 0:
            interest = round(sip * rate_for_calc / periods_per_year, 2)

        if is_past:
            total_interest_earned += interest
        else:
            total_interest_projected += interest

        installments.append({
            "month": m,
            "date": inst_date.strftime("%Y-%m-%d"),
            "amount_invested": round(invested, 2),
            "interest_earned": round(interest, 2) if is_past else 0.0,
            "interest_projected": round(interest, 2) if not is_past else 0.0,
            "is_past": is_past,
        })

    # ── Totals ────────────────────────────────────────────
    total_invested = sip  # lump-sum
    total_interest = total_interest_earned + total_interest_projected
    maturity_amount = round(total_invested + total_interest, 2)

    status = "Matured" if end_dt <= today else "Active"
    days_to_maturity = max(0, (end_dt - today).days)

    installments_paid = sum(
        1 for i in installments
        if i["is_past"] and (i["amount_invested"] > 0 or i["interest_earned"] > 0)
    )

    return {
        "id": _gen_fd_id(name),
        "name": name,
        "bank": bank,
        "type": fd_type,
        "principal": round(total_invested, 2),
        "interest_rate": rate_pct,
        "interest_payout": interest_payout,
        "sip_amount": round(sip, 2),
        "tenure_months": tenure_months,
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "maturity_date": end_dt.strftime("%Y-%m-%d"),
        "maturity_amount": maturity_amount,
        "interest_earned": round(total_interest_earned, 2),
        "interest_projected": round(total_interest_projected, 2),
        "total_invested": round(total_invested, 2),
        "status": status,
        "days_to_maturity": days_to_maturity,
        "source": "xlsx",
        "remarks": "",
        "tds": 0,
        "installments": installments,
        "installments_paid": installments_paid,
        "installments_total": len(installments),
    }


def _parse_all_xlsx(xlsx_dir: Path = None) -> list:
    """Parse all xlsx files from dumps/FD/ directory."""
    xlsx_dir = xlsx_dir or FD_XLSX_DIR
    results = []
    if not xlsx_dir.exists():
        return results

    for f in sorted(xlsx_dir.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        try:
            parsed = _parse_fd_xlsx(f)
            results.append(parsed)
        except Exception as e:
            logger.error(f"[FD] Error parsing {f.name}: {e}")
    return results


# ═══════════════════════════════════════════════════════════
#  XLSX CREATION (for new entries from UI)
# ═══════════════════════════════════════════════════════════

def _create_fd_xlsx(name: str, bank: str, principal: float, rate_pct: float,
                    tenure_months: int, start_date: str, interest_payout: str,
                    fd_type: str = "FD", xlsx_dir: Path = None):
    """Create an xlsx file following the FD template structure."""
    xlsx_dir = xlsx_dir or FD_XLSX_DIR
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    filepath = xlsx_dir / f"{name}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Index"

    rate_dec = rate_pct / 100
    maturity_years = tenure_months / 12
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")

    # Row 1: Start Date, Latest Amount, Maturity Years, Bank
    ws.cell(1, 2, start_dt)
    ws.cell(1, 5, principal)       # E1 = total invested
    ws.cell(1, 8, maturity_years)  # H1
    ws.cell(1, 11, bank)           # K1

    # Row 2: End Date, Interest Earned, Payout Type
    end_dt = start_dt + relativedelta(months=tenure_months)
    ws.cell(2, 2, end_dt)
    ws.cell(2, 8, interest_payout)  # H2

    # Row 3: Rate (decimal), Maturity Amount, SIP
    ws.cell(3, 2, rate_dec)   # B3
    ws.cell(3, 8, principal)  # H3 = principal/SIP

    # Row 5: Headers
    headers = ['S.No', '', '#', 'Date', 'Amount Invested', 'Interest Earned', 'Interest to be Earned']
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)

    # Data rows: compute values using payout frequency
    period_months = _payout_to_period(interest_payout)
    periods_per_year = _payout_periods_per_year(period_months)
    for m in range(1, tenure_months + 1):
        row = m + 5
        inst_date = start_dt + relativedelta(months=m - 1)
        ws.cell(row, 1, m)
        ws.cell(row, 3, m)
        ws.cell(row, 4, inst_date)
        ws.cell(row, 5, principal if m == 1 else 0)

        interest = 0.0
        if m > 1 and m % period_months == 0:
            interest = round(principal * rate_dec / periods_per_year, 2)
        ws.cell(row, 6, interest)
        ws.cell(row, 7, interest)

    # Compute and set E2 (total interest) and E3 (maturity amount)
    num_periods = tenure_months // period_months
    total_interest = round(principal * rate_dec / periods_per_year * num_periods, 2)

    ws.cell(2, 5, total_interest)  # E2
    ws.cell(3, 5, round(principal + total_interest, 2))  # E3

    wb.save(str(filepath))
    wb.close()
    _sync_to_drive(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════
#  JSON LOAD / SAVE (manual entries — fallback)
# ═══════════════════════════════════════════════════════════

def _load_json(json_file: Path = None) -> list:
    json_file = json_file or FD_JSON_FILE
    if not json_file.exists():
        return []
    try:
        with open(json_file, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, Exception):
        return []


def _save_json(data: list, json_file: Path = None, dumps_dir: Path = None):
    dumps_dir = dumps_dir or DUMPS_DIR
    json_file = json_file or FD_JSON_FILE
    dumps_dir.mkdir(parents=True, exist_ok=True)
    with open(json_file, "w") as f:
        json.dump(data, f, indent=2)
    _sync_to_drive(json_file)


# ═══════════════════════════════════════════════════════════
#  CALCULATIONS (for manual/JSON entries)
# ═══════════════════════════════════════════════════════════

def _calc_maturity(principal: float, rate: float, tenure_months: int,
                   interest_payout: str = "Quarterly") -> dict:
    """Calculate total interest and maturity amount using payout frequency."""
    rate_dec = rate / 100
    period_months = _payout_to_period(interest_payout)
    periods_per_year = _payout_periods_per_year(period_months)
    num_periods = tenure_months // period_months
    interest_per_period = principal * rate_dec / periods_per_year
    total_interest = round(interest_per_period * num_periods, 2)
    return {
        "maturity_amount": round(principal + total_interest, 2),
        "interest_earned": total_interest,
    }


def _calc_maturity_date(start_date: str, tenure_months: int) -> str:
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        maturity = start + relativedelta(months=tenure_months)
        return maturity.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return ""


def _generate_installments(principal: float, rate_pct: float, tenure_months: int,
                           start_date: str, interest_payout: str = "Quarterly") -> list:
    """Generate installment schedule for a manual entry."""
    today = date.today()
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    rate_dec = rate_pct / 100
    period_months = _payout_to_period(interest_payout)
    periods_per_year = _payout_periods_per_year(period_months)
    installments = []

    for m in range(1, tenure_months + 1):
        inst_date = start_dt + relativedelta(months=m - 1)
        is_past = inst_date <= today
        invested = principal if m == 1 else 0.0

        interest = 0.0
        if m > 1 and m % period_months == 0:
            interest = round(principal * rate_dec / periods_per_year, 2)

        installments.append({
            "month": m,
            "date": inst_date.strftime("%Y-%m-%d"),
            "amount_invested": round(invested, 2),
            "interest_earned": round(interest, 2) if is_past else 0.0,
            "interest_projected": round(interest, 2) if not is_past else 0.0,
            "is_past": is_past,
        })

    return installments


def _enrich_json_item(item: dict) -> dict:
    """Add computed fields to a JSON-based FD item."""
    today = date.today()
    fd_type = item.get("type", "FD")
    principal = item.get("principal", 0)
    rate = item.get("interest_rate", 0)
    tenure = item.get("tenure_months", 60)
    start_date = item.get("start_date", "")
    payout = item.get("interest_payout", "Monthly" if fd_type == "MIS" else "Quarterly")

    # Generate installments
    if start_date and principal > 0:
        installments = _generate_installments(principal, rate, tenure, start_date, payout)
    else:
        installments = []

    # Calculate totals
    calcs = _calc_maturity(principal, rate, tenure, payout)

    try:
        mat = datetime.strptime(item.get("maturity_date", ""), "%Y-%m-%d").date()
        days_to_maturity = max(0, (mat - today).days)
        status = "Matured" if mat <= today else "Active"
    except (ValueError, KeyError):
        days_to_maturity = 0
        status = item.get("status", "Active")

    item["source"] = "manual"
    item["type"] = fd_type
    item["name"] = item.get("name", f"{item.get('bank', 'FD')} {fd_type}")
    item["sip_amount"] = principal
    item["interest_payout"] = item.get("interest_payout", "Quarterly" if fd_type == "FD" else "Monthly")
    item["total_invested"] = principal
    item["maturity_amount"] = calcs["maturity_amount"]
    item["interest_earned"] = sum(i["interest_earned"] for i in installments)
    item["interest_projected"] = sum(i["interest_projected"] for i in installments)
    item["installments"] = installments
    item["installments_paid"] = sum(
        1 for i in installments
        if i["is_past"] and (i["amount_invested"] > 0 or i["interest_earned"] > 0)
    )
    item["installments_total"] = len(installments)
    item["days_to_maturity"] = days_to_maturity
    item["status"] = status
    return item


# ═══════════════════════════════════════════════════════════
#  PUBLIC API
# ═══════════════════════════════════════════════════════════

def get_all(base_dir=None) -> list:
    """Return all FDs: xlsx-parsed + JSON manual entries."""
    xlsx_dir = (Path(base_dir) / "FD") if base_dir else FD_XLSX_DIR
    json_file = (Path(base_dir) / "fixed_deposits.json") if base_dir else FD_JSON_FILE
    with _lock:
        xlsx_items = _parse_all_xlsx(xlsx_dir=xlsx_dir)
        json_items = _load_json(json_file=json_file)

    for item in json_items:
        _enrich_json_item(item)

    return xlsx_items + json_items


def get_dashboard(base_dir=None) -> dict:
    """Aggregate FD summary for dashboard."""
    items = get_all(base_dir=base_dir)
    active = [i for i in items if i.get("status") == "Active"]

    maturing_soon = sum(1 for i in active if 0 < i.get("days_to_maturity", 0) <= 90)

    return {
        "total_invested": round(sum(i.get("total_invested", i.get("principal", 0)) for i in active), 2),
        "total_maturity_value": round(sum(i.get("maturity_amount", 0) for i in active), 2),
        "total_interest": round(sum(i.get("interest_earned", 0) for i in active), 2),
        "total_interest_projected": round(sum(i.get("interest_projected", 0) for i in active), 2),
        "total_tds": round(sum(i.get("tds", 0) for i in active), 2),
        "active_count": len(active),
        "total_count": len(items),
        "maturing_soon": maturing_soon,
    }


def add(data: dict, base_dir=None) -> dict:
    """Add a new FD — creates xlsx file + JSON entry."""
    xlsx_dir = (Path(base_dir) / "FD") if base_dir else FD_XLSX_DIR
    fd_type = data.get("type", "FD")
    principal = data["principal"]
    rate = data["interest_rate"]
    tenure_months = data["tenure_months"]
    start_date = data["start_date"]
    bank = data["bank"]
    interest_payout = data.get("interest_payout", "Quarterly" if fd_type == "FD" else "Monthly")

    calcs = _calc_maturity(principal, rate, tenure_months, interest_payout)
    maturity_date = data.get("maturity_date") or _calc_maturity_date(start_date, tenure_months)
    name = data.get("name", f"{bank} {fd_type}")

    # Create xlsx file (this is the only storage — parser picks it up on next load)
    _create_fd_xlsx(
        name=name,
        bank=bank,
        principal=principal,
        rate_pct=rate,
        tenure_months=tenure_months,
        start_date=start_date,
        interest_payout=interest_payout,
        fd_type=fd_type,
        xlsx_dir=xlsx_dir,
    )

    return {
        "id": _gen_fd_id(name),
        "name": name,
        "bank": bank,
        "type": fd_type,
        "principal": principal,
        "interest_rate": rate,
        "interest_payout": interest_payout,
        "tenure_months": tenure_months,
        "start_date": start_date,
        "maturity_date": maturity_date,
        "maturity_amount": calcs["maturity_amount"],
        "interest_earned": calcs["interest_earned"],
        "tds": data.get("tds", 0),
        "status": data.get("status", "Active"),
        "remarks": data.get("remarks", ""),
    }


def update(fd_id: str, data: dict, base_dir=None) -> dict:
    """Update an existing FD — supports both manual (JSON) and xlsx-imported."""
    xlsx_dir = (Path(base_dir) / "FD") if base_dir else FD_XLSX_DIR
    json_file = (Path(base_dir) / "fixed_deposits.json") if base_dir else FD_JSON_FILE
    dumps_dir = Path(base_dir) if base_dir else DUMPS_DIR

    # Try xlsx first
    if xlsx_dir.exists():
        for f in xlsx_dir.glob("*.xlsx"):
            if f.name.startswith("~$"):
                continue
            if _gen_fd_id(f.stem) == fd_id:
                return _update_xlsx_fd(f, data, dumps_dir)

    # Fall back to JSON (manual)
    with _lock:
        items = _load_json(json_file=json_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == fd_id), None)
        if idx is None:
            raise ValueError(f"FD {fd_id} not found")

        item = items[idx]
        for key, val in data.items():
            if val is not None and key in item:
                item[key] = val

        payout = item.get("interest_payout", "Quarterly")
        calcs = _calc_maturity(item["principal"], item["interest_rate"], item["tenure_months"], payout)
        item["maturity_amount"] = calcs["maturity_amount"]
        item["interest_earned"] = calcs["interest_earned"]

        if "start_date" in data or "tenure_months" in data:
            if not data.get("maturity_date"):
                item["maturity_date"] = _calc_maturity_date(item["start_date"], item["tenure_months"])

        items[idx] = item
        _save_json(items, json_file=json_file, dumps_dir=dumps_dir)
        return item


def _update_xlsx_fd(filepath: Path, data: dict, dumps_dir: Path) -> dict:
    """Update an xlsx-imported FD by editing cells in the Index sheet.
    Supports: bank, principal, interest_rate, start_date, tenure_months, interest_payout."""
    import openpyxl as xl

    wb = xl.load_workbook(str(filepath))
    ws = wb["Index"]

    if "start_date" in data:
        try:
            dt = datetime.strptime(data["start_date"], "%Y-%m-%d")
            ws.cell(1, 2, value=dt)  # B1
        except ValueError:
            pass
    if "tenure_months" in data:
        years = float(data["tenure_months"]) / 12
        ws.cell(1, 8, value=years)  # H1
    if "bank" in data:
        ws.cell(1, 11, value=data["bank"])  # K1
    if "interest_payout" in data:
        ws.cell(2, 8, value=data["interest_payout"])  # H2
    if "interest_rate" in data:
        rate = float(data["interest_rate"])
        ws.cell(3, 2, value=rate / 100 if rate > 1 else rate)  # B3 as decimal
    if "principal" in data:
        ws.cell(3, 8, value=float(data["principal"]))  # H3
    if "maturity_date" in data:
        try:
            dt = datetime.strptime(data["maturity_date"], "%Y-%m-%d")
            ws.cell(2, 2, value=dt)  # B2 = end_date
        except ValueError:
            pass

    wb.save(str(filepath))
    wb.close()

    # Re-parse and return updated data
    return _parse_fd_xlsx(filepath)


def delete(fd_id: str, base_dir=None) -> dict:
    """Delete an FD — removes xlsx file and/or JSON entry."""
    xlsx_dir = (Path(base_dir) / "FD") if base_dir else FD_XLSX_DIR
    json_file = (Path(base_dir) / "fixed_deposits.json") if base_dir else FD_JSON_FILE
    dumps_dir = Path(base_dir) if base_dir else DUMPS_DIR
    # Try xlsx first
    if xlsx_dir.exists():
        for f in xlsx_dir.glob("*.xlsx"):
            if _gen_fd_id(f.stem) == fd_id:
                _delete_from_drive(f)
                f.unlink()
                return {"message": f"FD {fd_id} deleted (xlsx)", "item": {"id": fd_id, "name": f.stem}}

    # Try JSON
    with _lock:
        items = _load_json(json_file=json_file)
        idx = next((i for i, x in enumerate(items) if x["id"] == fd_id), None)
        if idx is None:
            raise ValueError(f"FD {fd_id} not found")
        removed = items.pop(idx)
        _save_json(items, json_file=json_file, dumps_dir=dumps_dir)
        return {"message": f"FD {fd_id} deleted", "item": removed}
