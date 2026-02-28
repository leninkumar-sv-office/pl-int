"""
NPS (National Pension System) database layer.

Parses NPS Transaction Statement PDFs, stores data in xlsx format
in DUMPS_DIR/NPS/.  Handles deduplication across overlapping PDF periods.

xlsx layout:
    Row 1: [_, pran(B), _, _, current_value(E), _, _, reg_date(H), _, _, subscriber_name(K)]
    Row 2: [_, tier(B), _, _, xirr(E), _, _, scheme_pref(H), _, _, fund_manager(K)]
    Row 3: [_, status(B), _, _, _(E), _, _, nominee(H), _, _, remarks(K)]
    Row 4: [_, _(B), _, _, _(E), _, _, scheme_splits_json(H), _, _, _(K)]
    Row 5: headers [S.No, Date, Scheme, Description, Amount, NAV, Units]
    Row 6+: transaction data rows sorted by (date, scheme)
"""

import json
import hashlib
import re
import threading
import uuid
from datetime import datetime, date
from pathlib import Path

import openpyxl

from app.config import DUMPS_DIR

# ═══════════════════════════════════════════════════════════
#  FILE PATHS
# ═══════════════════════════════════════════════════════════

NPS_DIR = DUMPS_DIR / "NPS"
NPS_JSON_FILE = DUMPS_DIR / "nps_accounts.json"  # legacy, for migration
PDF_IMPORT_DIR = Path("/Users/lenin/Downloads/NPS")

_lock = threading.Lock()
_imported = False  # track whether PDF import has run this session

# Scheme short codes
SCHEME_MAP = {"E": "Equity", "C": "Corporate Bonds", "G": "Government Securities"}


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _gen_id(name: str) -> str:
    return hashlib.md5(name.encode()).hexdigest()[:8]


def _parse_num(s: str) -> float:
    """Parse a number string, handling (parentheses) for negatives and commas."""
    s = s.strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace(",", "")
    try:
        val = float(s)
        return -val if neg else val
    except (ValueError, TypeError):
        return 0.0


def _parse_date(s: str) -> date | None:
    """Parse DD-Mon-YYYY or DD-Mon-YY date strings."""
    for fmt in ("%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _to_str(val, default=""):
    if val is None:
        return default
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


# ═══════════════════════════════════════════════════════════
#  PDF PARSING
# ═══════════════════════════════════════════════════════════

def _extract_pdf_text(pdf_path: str) -> str:
    """Extract full text from a PDF using pdfplumber."""
    import pdfplumber
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n\n".join(pages)


def _parse_subscriber_info(text: str) -> dict:
    """Extract subscriber metadata from PDF text."""
    info = {}

    m = re.search(r"PRAN\s+(\d{12})", text)
    if m:
        info["pran"] = m.group(1)

    m = re.search(r"Registration Date\s+(\d{2}-\w{3}-\d{2,4})", text)
    if m:
        d = _parse_date(m.group(1))
        if d:
            info["registration_date"] = d.strftime("%Y-%m-%d")

    m = re.search(r"Subscriber Name\s+(.+?)(?:\s+Tier)", text)
    if m:
        info["subscriber_name"] = m.group(1).strip()

    m = re.search(r"Tier I Status\s+(\w+)", text)
    if m:
        info["status"] = m.group(1)

    # Scheme preference
    m = re.search(r"Scheme Choice\s*-\s*(.+?)(?:\s*\()", text)
    if m:
        info["scheme_preference"] = m.group(1).strip()
    else:
        m = re.search(r"Scheme Choice\s*-\s*(.+?)$", text, re.MULTILINE)
        if m:
            info["scheme_preference"] = m.group(1).strip()

    # Fund manager (from scheme details)
    m = re.search(r"(SBI|LIC|UTI|HDFC|ICICI|Kotak|Aditya|Tata|Max|Axis)\s+(?:PENSION FUND|Pension Fund)", text, re.IGNORECASE)
    if m:
        fund_names = {
            "SBI": "SBI Pension Fund", "LIC": "LIC Pension Fund",
            "UTI": "UTI Retirement Solutions", "HDFC": "HDFC Pension Management",
            "ICICI": "ICICI Prudential Pension Fund", "Kotak": "Kotak Mahindra Pension Fund",
        }
        info["fund_manager"] = fund_names.get(m.group(1).upper(), m.group(0).strip())

    # Scheme splits
    splits = []
    for code in ("E", "C", "G"):
        pat = rf"SCHEME {code}\s*-\s*TIER I\s+([\d.]+)%"
        m2 = re.search(pat, text)
        if m2:
            splits.append({"scheme": code, "pct": float(m2.group(1))})
    if splits:
        info["scheme_splits"] = splits

    # Nominee
    m = re.search(r"Nominee Name/s\s+Percentage\s*\n(.+?)\s+(\d+%)", text)
    if m:
        info["nominee"] = f"{m.group(1).strip()} ({m.group(2)})"

    # Investment summary - holdings value
    m = re.search(r"₹\s*([\d,]+\.\d{2})\s+\d+\s+₹\s*([\d,]+\.\d{2})", text)
    if m:
        info["holdings_value"] = _parse_num(m.group(1))
        info["total_contribution"] = _parse_num(m.group(2))

    return info


def _parse_scheme_transactions(text: str) -> list:
    """Parse transaction lines from a single scheme section.

    Returns list of dicts: {date, description, amount, nav, units}
    """
    lines = text.strip().split("\n")
    transactions = []
    pending_prefix = []
    i = 0

    # Regex for date-prefixed lines
    date_re = re.compile(r"^(\d{2}-\w{3}-\d{4})\s+(.*)")
    # Regex for numbers with decimals (positive or parenthesized negative)
    num_re = re.compile(r"(\([\d,]+\.\d+\)|[\d,]+\.\d+)")

    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        dm = date_re.match(line)
        if dm:
            date_str = dm.group(1)
            rest = dm.group(2)

            # Collect suffix lines (non-date, non-prefix lines after this)
            suffix_parts = []
            j = i + 1
            while j < len(lines):
                nl = lines[j].strip()
                if not nl:
                    j += 1
                    continue
                # Stop at next date line
                if date_re.match(nl):
                    break
                # Stop at what looks like a new prefix (To/By at start)
                if re.match(r"^(To |By )", nl):
                    break
                # Stop at scheme header
                if "PENSION FUND SCHEME" in nl and "TIER" in nl and "Date" not in nl:
                    break
                # Stop at section markers
                if nl.startswith("Notes") or nl.startswith("View More"):
                    break
                suffix_parts.append(nl)
                j += 1

            # Build full description from prefix + rest + suffix
            full_text = " ".join(pending_prefix + [rest] + suffix_parts)
            pending_prefix = []

            # Parse the transaction
            d = _parse_date(date_str)
            if not d:
                i = j
                continue

            # Check for Opening/Closing balance
            if "Opening balance" in full_text or "Closing Balance" in full_text:
                nums = num_re.findall(full_text)
                units = _parse_num(nums[-1]) if nums else 0
                txn_type = "opening_balance" if "Opening" in full_text else "closing_balance"
                transactions.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "description": "Opening balance" if txn_type == "opening_balance" else "Closing Balance",
                    "amount": 0, "nav": 0, "units": units,
                    "type": txn_type,
                })
            else:
                # Extract last 3 numbers: amount, NAV, units
                nums = num_re.findall(full_text)
                if len(nums) >= 3:
                    amount = _parse_num(nums[-3])
                    nav = _parse_num(nums[-2])
                    units = _parse_num(nums[-1])

                    # Description is everything before the third-to-last number
                    # Find position of the third-to-last number in full_text
                    positions = list(num_re.finditer(full_text))
                    desc_end = positions[-3].start()
                    description = full_text[:desc_end].strip()

                    # Determine transaction type
                    if "Contribution" in description:
                        txn_type = "contribution"
                    elif "Billing" in description:
                        txn_type = "billing"
                    elif "persistency" in description.lower():
                        txn_type = "persistency_charge"
                    elif "Trail Commission" in description:
                        txn_type = "trail_commission"
                    elif "Switch out" in description:
                        txn_type = "switch_out"
                    elif "Switch In" in description or "Switch in" in description:
                        txn_type = "switch_in"
                    else:
                        txn_type = "other"

                    transactions.append({
                        "date": d.strftime("%Y-%m-%d"),
                        "description": description,
                        "amount": round(amount, 2),
                        "nav": round(nav, 4),
                        "units": round(units, 4),
                        "type": txn_type,
                    })

            i = j
        else:
            # Non-date line - buffer as potential prefix for next transaction
            pending_prefix.append(line)
            i += 1

    return transactions


def _parse_pdf(pdf_path: str) -> dict:
    """Parse a single NPS PDF and return structured data.

    Returns: {
        subscriber_info: {...},
        scheme_transactions: {"E": [...], "C": [...], "G": [...]},
        contributions: [{date, amount, remarks}, ...],
    }
    """
    text = _extract_pdf_text(pdf_path)

    # Extract subscriber info
    info = _parse_subscriber_info(text)

    # Split text into scheme sections
    scheme_txns = {}
    for code in ("E", "C", "G"):
        # Find section for this scheme
        pattern = rf"SBI PENSION FUND SCHEME {code}\s*-\s*TIER I\s*\n\s*Date\s+Description.*?\n(.*?)(?=SBI PENSION FUND SCHEME [ECG]\s*-\s*TIER I|Notes\n|View More|$)"
        m = re.search(pattern, text, re.DOTALL)
        if m:
            section_text = m.group(1)
            txns = _parse_scheme_transactions(section_text)
            scheme_txns[code] = txns

    # Extract contribution summary from "Contribution/Redemption Details" section
    contributions = []
    contrib_section = re.search(
        r"Contribution/Redemption Details.*?\n(.*?)(?=Transaction Details|$)",
        text, re.DOTALL
    )
    if contrib_section:
        contrib_text = contrib_section.group(1)
        # Fix wrapped dates: "17-Sep-\n<text>\n2019" → "17-Sep-2019 <text>"
        contrib_text = re.sub(
            r"(\d{2}-\w{3})-\s*\n(.*?)\n\s*(\d{4})\b",
            r"\1-\3 \2",
            contrib_text
        )

        for cm in re.finditer(
            r"(\d{2}-\w{3}-\d{4})\s+By\s+(Contribution|Voluntary Contributions)\s+.*?([\d,]+\.\d{2})\s+[\d,.]+\s+([\d,]+\.\d{2})",
            contrib_text
        ):
            d = _parse_date(cm.group(1))
            if d:
                contributions.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "amount": _parse_num(cm.group(4)),
                    "remarks": f"By {cm.group(2)}",
                })

    return {
        "subscriber_info": info,
        "scheme_transactions": scheme_txns,
        "contributions": contributions,
    }


def _merge_pdf_data(all_parsed: list) -> dict:
    """Merge data from multiple PDFs into a single account record.

    Deduplicates transactions by (date, scheme, type, amount).
    Uses latest PDF's subscriber info and holdings value.
    """
    # Find the latest PDF (highest holdings value = most recent period)
    sorted_parsed = sorted(all_parsed, key=lambda p: p.get("subscriber_info", {}).get("holdings_value", 0))
    latest_pdf = sorted_parsed[-1] if sorted_parsed else None

    # Use subscriber info from the latest PDF
    latest_info = {}
    for parsed in all_parsed:
        info = parsed.get("subscriber_info", {})
        if info:
            latest_info.update({k: v for k, v in info.items() if v})
    # Override with latest PDF values for fields that change over time
    if latest_pdf:
        li = latest_pdf.get("subscriber_info", {})
        for key in ("holdings_value", "total_contribution", "status"):
            if key in li:
                latest_info[key] = li[key]

    # Merge all scheme transactions with deduplication
    all_txns = {}  # key -> txn
    for parsed in all_parsed:
        for scheme_code, txns in parsed.get("scheme_transactions", {}).items():
            for txn in txns:
                # Skip opening/closing balance entries (not real transactions)
                if txn.get("type") in ("opening_balance", "closing_balance"):
                    continue
                # Dedup key: (date, scheme, amount, nav, units)
                key = (txn["date"], scheme_code, txn["amount"], txn["nav"], txn["units"])
                if key not in all_txns:
                    all_txns[key] = {**txn, "scheme": scheme_code}

    merged_txns = sorted(all_txns.values(), key=lambda t: (t["date"], t["scheme"]))

    # Merge contributions with deduplication
    # Use (date, amount, remarks) as key to handle same-day different-source contributions
    contrib_keys = set()
    merged_contribs = []
    for parsed in all_parsed:
        for c in parsed.get("contributions", []):
            key = (c["date"], c["amount"], c.get("remarks", ""))
            if key not in contrib_keys:
                contrib_keys.add(key)
                merged_contribs.append(c)
    merged_contribs.sort(key=lambda c: c["date"])

    # Get closing balances from the latest PDF (highest holdings)
    schemes_summary = []
    if latest_pdf:
        for code in ("E", "C", "G"):
            txns = latest_pdf.get("scheme_transactions", {}).get(code, [])
            closing = [t for t in txns if t.get("type") == "closing_balance"]
            if closing:
                units = closing[-1]["units"]
                # Get latest NAV from last real transaction before closing
                real_txns = [t for t in txns if t.get("type") not in ("opening_balance", "closing_balance") and t.get("nav", 0) > 0]
                nav = real_txns[-1]["nav"] if real_txns else 0
                schemes_summary.append({
                    "scheme": code,
                    "units": round(units, 4),
                    "nav": round(nav, 4),
                    "value": round(units * nav, 2),
                })

    current_value = latest_info.get("holdings_value", sum(s["value"] for s in schemes_summary))

    return {
        "info": latest_info,
        "transactions": merged_txns,
        "contributions": merged_contribs,
        "schemes_summary": schemes_summary,
        "current_value": round(current_value, 2),
    }


# ═══════════════════════════════════════════════════════════
#  XLSX READ / WRITE
# ═══════════════════════════════════════════════════════════

def _write_xlsx(filepath: Path, account: dict, transactions: list):
    """Write NPS account data to xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPS"

    # Row 1: pran, current_value, reg_date, subscriber_name
    ws.cell(1, 2, account.get("pran", ""))
    ws.cell(1, 5, account.get("current_value", 0))
    ws.cell(1, 8, account.get("start_date", ""))
    ws.cell(1, 11, account.get("account_name", ""))

    # Row 2: tier, xirr, scheme_pref, fund_manager
    ws.cell(2, 2, account.get("tier", "Tier I"))
    ws.cell(2, 5, account.get("xirr", ""))
    ws.cell(2, 8, account.get("scheme_preference", ""))
    ws.cell(2, 11, account.get("fund_manager", ""))

    # Row 3: status, _, nominee, remarks
    ws.cell(3, 2, account.get("status", "Active"))
    ws.cell(3, 8, account.get("nominee", ""))
    ws.cell(3, 11, account.get("remarks", ""))

    # Row 4: scheme_splits_json, contributions_json
    ws.cell(4, 8, json.dumps(account.get("scheme_splits", [])))
    ws.cell(4, 5, json.dumps(account.get("contributions", [])))
    ws.cell(4, 11, json.dumps(account.get("schemes_summary", [])))

    # Row 5: headers
    headers = ["S.No", "Date", "Scheme", "Description", "Amount", "NAV", "Units"]
    for ci, h in enumerate(headers, 1):
        ws.cell(5, ci, h)

    # Row 6+: transactions
    for ri, txn in enumerate(transactions, 6):
        ws.cell(ri, 1, ri - 5)  # S.No
        d = txn.get("date", "")
        if isinstance(d, str) and d:
            try:
                ws.cell(ri, 2, datetime.strptime(d, "%Y-%m-%d").date())
            except ValueError:
                ws.cell(ri, 2, d)
        else:
            ws.cell(ri, 2, d)
        ws.cell(ri, 3, txn.get("scheme", ""))
        ws.cell(ri, 4, txn.get("description", ""))
        ws.cell(ri, 5, txn.get("amount", 0))
        ws.cell(ri, 6, txn.get("nav", 0))
        ws.cell(ri, 7, txn.get("units", 0))

    # Format date column
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "DD-MMM-YYYY"

    # Format amount columns
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.0000"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)


def _read_xlsx(filepath: Path) -> dict:
    """Read NPS account data from xlsx."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    account = {
        "pran": _to_str(ws.cell(1, 2).value),
        "current_value": _to_float(ws.cell(1, 5).value),
        "start_date": _to_str(ws.cell(1, 8).value),
        "account_name": _to_str(ws.cell(1, 11).value) or "NPS Account",
        "tier": _to_str(ws.cell(2, 2).value) or "Tier I",
        "xirr": _to_str(ws.cell(2, 5).value),
        "scheme_preference": _to_str(ws.cell(2, 8).value),
        "fund_manager": _to_str(ws.cell(2, 11).value),
        "status": _to_str(ws.cell(3, 2).value) or "Active",
        "nominee": _to_str(ws.cell(3, 8).value),
        "remarks": _to_str(ws.cell(3, 11).value),
    }

    # Parse scheme splits from H4
    try:
        splits_raw = ws.cell(4, 8).value
        if splits_raw:
            account["scheme_splits"] = json.loads(splits_raw)
    except (json.JSONDecodeError, TypeError):
        pass

    # Parse contributions from E4
    try:
        contribs_raw = ws.cell(4, 5).value
        if contribs_raw:
            account["contributions"] = json.loads(contribs_raw)
    except (json.JSONDecodeError, TypeError):
        account["contributions"] = []

    # Parse schemes summary from K4
    try:
        schemes_raw = ws.cell(4, 11).value
        if schemes_raw and schemes_raw.startswith("["):
            account["schemes_summary"] = json.loads(schemes_raw)
    except (json.JSONDecodeError, TypeError):
        pass

    # Read transactions from row 6+
    transactions = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        if row[1].value is None:
            continue
        d = row[1].value
        if isinstance(d, (datetime, date)):
            d = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else d.strftime("%Y-%m-%d")
        else:
            d = str(d).strip()

        txn = {
            "date": d,
            "scheme": _to_str(row[2].value),
            "description": _to_str(row[3].value),
            "amount": _to_float(row[4].value),
            "nav": _to_float(row[5].value),
            "units": _to_float(row[6].value),
        }
        transactions.append(txn)

    account["_transactions"] = transactions
    account["id"] = _gen_id(account.get("pran") or filepath.stem)

    return account


# ═══════════════════════════════════════════════════════════
#  PDF IMPORT
# ═══════════════════════════════════════════════════════════

def _import_from_pdfs():
    """Scan PDF_IMPORT_DIR for NPS statement PDFs, parse, and create xlsx files."""
    global _imported
    if _imported:
        return
    _imported = True

    if not PDF_IMPORT_DIR.exists():
        return

    pdfs = sorted(PDF_IMPORT_DIR.glob("*.pdf"))
    if not pdfs:
        return

    # Check if we already have xlsx files
    NPS_DIR.mkdir(parents=True, exist_ok=True)
    existing = list(NPS_DIR.glob("*.xlsx"))
    if existing:
        return  # already imported

    print(f"[NPS] Importing {len(pdfs)} PDF statements from {PDF_IMPORT_DIR}...")

    # Parse all PDFs
    all_parsed = []
    for pdf in pdfs:
        try:
            parsed = _parse_pdf(str(pdf))
            all_parsed.append(parsed)
            contribs = len(parsed.get("contributions", []))
            txns = sum(len(v) for v in parsed.get("scheme_transactions", {}).values())
            print(f"  [NPS] Parsed {pdf.name}: {contribs} contributions, {txns} transactions")
        except Exception as e:
            print(f"  [NPS] ERROR parsing {pdf.name}: {e}")

    if not all_parsed:
        return

    # Merge all PDFs into one account
    merged = _merge_pdf_data(all_parsed)
    info = merged["info"]

    pran = info.get("pran", "unknown")
    account = {
        "pran": pran,
        "account_name": info.get("subscriber_name", "NPS Account"),
        "tier": "Tier I",
        "fund_manager": info.get("fund_manager", ""),
        "scheme_preference": info.get("scheme_preference", ""),
        "start_date": info.get("registration_date", ""),
        "current_value": merged["current_value"],
        "status": info.get("status", "Active"),
        "remarks": "",
        "nominee": info.get("nominee", ""),
        "contributions": merged["contributions"],
        "scheme_splits": info.get("scheme_splits", []),
        "schemes_summary": merged.get("schemes_summary", []),
    }

    xlsx_path = NPS_DIR / f"{pran}.xlsx"
    _write_xlsx(xlsx_path, account, merged["transactions"])

    total_contrib = sum(c["amount"] for c in merged["contributions"])
    print(f"  [NPS] Created {xlsx_path.name}: {len(merged['transactions'])} transactions, "
          f"{len(merged['contributions'])} contributions, "
          f"total contributed: ₹{total_contrib:,.0f}, "
          f"current value: ₹{merged['current_value']:,.0f}")


# ═══════════════════════════════════════════════════════════
#  ENRICH
# ═══════════════════════════════════════════════════════════

def _enrich(account: dict) -> dict:
    """Compute derived fields for an NPS account."""
    contributions = account.get("contributions", [])
    total_contributed = sum(c.get("amount", 0) for c in contributions)
    current_value = account.get("current_value", 0)
    gain = round(current_value - total_contributed, 2)
    gain_pct = round((gain / total_contributed) * 100, 2) if total_contributed > 0 else 0

    today = date.today()
    try:
        start = datetime.strptime(account.get("start_date", ""), "%Y-%m-%d").date()
        years_active = max(0, round((today - start).days / 365.25, 1))
    except (ValueError, TypeError):
        years_active = 0

    account["total_contributed"] = round(total_contributed, 2)
    account["gain"] = gain
    account["gain_pct"] = gain_pct
    account["years_active"] = years_active
    account["status"] = account.get("status", "Active")

    # Remove internal transaction list from API response
    account.pop("_transactions", None)

    return account


# ═══════════════════════════════════════════════════════════
#  LOAD / SAVE (xlsx-based)
# ═══════════════════════════════════════════════════════════

def _load_all_xlsx() -> list:
    """Load all NPS accounts from xlsx files in NPS_DIR."""
    NPS_DIR.mkdir(parents=True, exist_ok=True)
    items = []
    for xlsx_path in sorted(NPS_DIR.glob("*.xlsx")):
        try:
            account = _read_xlsx(xlsx_path)
            account["_xlsx_path"] = str(xlsx_path)
            items.append(account)
        except Exception as e:
            print(f"[NPS] Error reading {xlsx_path.name}: {e}")
    return items


def _save_account(account: dict, transactions: list | None = None):
    """Save an NPS account to its xlsx file."""
    pran = account.get("pran", "")
    xlsx_name = pran if pran else account.get("id", str(uuid.uuid4())[:8])
    xlsx_path = Path(account.get("_xlsx_path", str(NPS_DIR / f"{xlsx_name}.xlsx")))

    if transactions is None:
        # Read existing transactions
        if xlsx_path.exists():
            existing = _read_xlsx(xlsx_path)
            transactions = existing.get("_transactions", [])
        else:
            transactions = []

    _write_xlsx(xlsx_path, account, transactions)


# ═══════════════════════════════════════════════════════════
#  PUBLIC API
# ═══════════════════════════════════════════════════════════

def get_all() -> list:
    """Return all NPS accounts with computed fields."""
    with _lock:
        _import_from_pdfs()
        items = _load_all_xlsx()

    for item in items:
        _enrich(item)
    return items


def get_dashboard() -> dict:
    """Aggregate NPS summary for dashboard."""
    items = get_all()
    active = [i for i in items if i.get("status") == "Active"]

    total_contributed = round(sum(i.get("total_contributed", 0) for i in active), 2)
    current_value = round(sum(i.get("current_value", 0) for i in active), 2)
    total_gain = round(current_value - total_contributed, 2)
    gain_pct = round((total_gain / total_contributed) * 100, 2) if total_contributed > 0 else 0

    return {
        "total_contributed": total_contributed,
        "current_value": current_value,
        "total_gain": total_gain,
        "gain_pct": gain_pct,
        "active_count": len(active),
        "total_count": len(items),
    }


def add(data: dict) -> dict:
    """Add a new NPS account."""
    with _lock:
        account = {
            "id": str(uuid.uuid4())[:8],
            "account_name": data.get("account_name", "NPS Account"),
            "pran": data.get("pran", ""),
            "tier": data.get("tier", "Tier I"),
            "fund_manager": data.get("fund_manager", ""),
            "scheme_preference": data.get("scheme_preference", "Auto Choice"),
            "start_date": data["start_date"],
            "current_value": data.get("current_value", 0),
            "status": data.get("status", "Active"),
            "remarks": data.get("remarks", ""),
            "contributions": [],
        }
        _save_account(account, [])
        return account


def update(nps_id: str, data: dict) -> dict:
    """Update an existing NPS account."""
    with _lock:
        items = _load_all_xlsx()
        item = next((x for x in items if x["id"] == nps_id), None)
        if item is None:
            raise ValueError(f"NPS account {nps_id} not found")

        transactions = item.pop("_transactions", [])

        for key, val in data.items():
            if val is not None and key not in ("contributions", "_transactions", "_xlsx_path", "id"):
                item[key] = val

        _save_account(item, transactions)
        return item


def delete(nps_id: str) -> dict:
    """Delete an NPS account."""
    with _lock:
        items = _load_all_xlsx()
        item = next((x for x in items if x["id"] == nps_id), None)
        if item is None:
            raise ValueError(f"NPS account {nps_id} not found")

        xlsx_path = Path(item.get("_xlsx_path", ""))
        if xlsx_path.exists():
            xlsx_path.unlink()

        return {"message": f"NPS {nps_id} deleted", "item": item}


def add_contribution(nps_id: str, contribution: dict) -> dict:
    """Add a contribution to an NPS account."""
    with _lock:
        items = _load_all_xlsx()
        item = next((x for x in items if x["id"] == nps_id), None)
        if item is None:
            raise ValueError(f"NPS account {nps_id} not found")

        transactions = item.pop("_transactions", [])

        if "contributions" not in item:
            item["contributions"] = []

        amount = contribution.get("amount", 0)
        contrib_date = contribution.get("date", datetime.now().strftime("%Y-%m-%d"))

        try:
            datetime.strptime(contrib_date, "%Y-%m-%d")
        except ValueError:
            raise ValueError("Invalid date format")

        item["contributions"].append({
            "date": contrib_date,
            "amount": amount,
            "remarks": contribution.get("remarks", ""),
        })

        _save_account(item, transactions)
        return item
