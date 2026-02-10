"""
XLSX-per-stock database layer.

Each stock is stored as an individual .xlsx file under dumps/Stocks/.
Holdings and sold positions are derived on-the-fly via FIFO matching
of Buy/Sell rows in each file's "Trading History" sheet.
"""

import hashlib
import json
import os
import threading
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import Holding, SoldPosition, Transaction

# ═══════════════════════════════════════════════════════════
#  SYMBOL RESOLUTION  (dynamic from Zerodha + NSE, no hardcoding)
# ═══════════════════════════════════════════════════════════
#
# Symbols are resolved dynamically from Zerodha instruments CSV
# and NSE EQUITY_L.csv via the shared symbol_resolver module.
# No hardcoded symbol map — all lookups go through Zerodha/NSE.
#
from . import symbol_resolver as _sym_resolver

# Resolved symbol map: populated at runtime during _build_file_map().
# Maps company name (xlsx stem) → trading symbol.
SYMBOL_MAP: Dict[str, str] = {}

# Reverse map: symbol → company name (for display)
_REVERSE_MAP: Dict[str, str] = {}


# ═══════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════

def _gen_id(symbol: str, exchange: str, buy_date: str, buy_price: float, row_idx: int) -> str:
    """Deterministic holding ID from transaction data + row position."""
    key = f"{symbol}|{exchange}|{buy_date}|{buy_price:.4f}|{row_idx}"
    return hashlib.md5(key.encode()).hexdigest()[:8]


def _parse_date(val) -> Optional[str]:
    """Convert xlsx cell value to YYYY-MM-DD string (e.g. 2026-01-27)."""
    _FMT = "%Y-%m-%d"
    if isinstance(val, datetime):
        return val.strftime(_FMT)
    if isinstance(val, date):
        return val.strftime(_FMT)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%B-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime(_FMT)
            except ValueError:
                continue
    return None


def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0) -> int:
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════════════════════
#  FIFO MATCHING  (extracted from import_dump.py)
# ═══════════════════════════════════════════════════════════

def fifo_match(buys: list, sells: list):
    """
    FIFO-match sells against buys.
    Returns (remaining_lots, sold_positions).
    Each lot in remaining_lots has a "remaining" key.
    """
    buys_sorted = sorted(buys, key=lambda x: x["date"])
    sells_sorted = sorted(sells, key=lambda x: x["date"])

    buy_lots = [{**b, "remaining": b["quantity"]} for b in buys_sorted]
    sold_positions = []

    for sell in sells_sorted:
        sell_qty = sell["quantity"]
        for lot in buy_lots:
            if sell_qty <= 0:
                break
            if lot["remaining"] <= 0:
                continue
            matched = min(lot["remaining"], sell_qty)
            realized_pl = (sell["price"] - lot["price"]) * matched
            sold_positions.append({
                "buy_price": lot["price"],
                "buy_date": lot["date"],
                "buy_exchange": lot.get("exchange", "NSE"),
                "sell_price": sell["price"],
                "sell_date": sell["date"],
                "quantity": matched,
                "realized_pl": round(realized_pl, 2),
                "row_idx": lot.get("row_idx", 0),
            })
            lot["remaining"] -= matched
            sell_qty -= matched

    remaining = [l for l in buy_lots if l["remaining"] > 0]
    return remaining, sold_positions


# ═══════════════════════════════════════════════════════════
#  XLSX PARSING
# ═══════════════════════════════════════════════════════════

def _extract_index_data(wb) -> dict:
    """Read metadata from the Index sheet."""
    data = {"code": None, "exchange": "NSE", "symbol": None,
            "current_price": 0, "week_52_high": 0, "week_52_low": 0}
    if "Index" not in wb.sheetnames:
        return data
    ws = wb["Index"]
    max_row = ws.max_row or 0
    if max_row == 0:
        return data
    for row in ws.iter_rows(min_row=1, max_row=min(15, max_row), values_only=False):
        vals = [c.value for c in row]
        if len(vals) >= 3:
            label = vals[1]
            value = vals[2]
            if label == "Code" and value:
                code = str(value)
                data["code"] = code
                if ":" in code:
                    parts = code.split(":")
                    data["exchange"] = parts[0]
                    data["symbol"] = parts[1]
            elif label == "Current Price" and isinstance(value, (int, float)):
                data["current_price"] = float(value)
            elif label == "52 Week High" and isinstance(value, (int, float)):
                data["week_52_high"] = float(value)
            elif label == "52 Week Low" and isinstance(value, (int, float)):
                data["week_52_low"] = float(value)
    return data


def _parse_excel_serial_date(val) -> Optional[str]:
    """Convert Excel serial date number to YYYY-MM-DD string."""
    if isinstance(val, (int, float)) and val > 1000:
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(val))).strftime("%Y-%m-%d")
    return _parse_date(val)


def _find_realised_columns(ws, header_row: int) -> dict:
    """Dynamically find the Realised section column positions.

    The "Realised" marker in row 2 tells us where the sold section starts.
    Under that marker, row 4 has sub-headers: Price, Date, [Gain%], Gain, Gross, Units.
    Column positions vary between file formats (new vs archive).
    """
    cols = {"sell_price": None, "sell_date": None, "sell_gain": None, "sold_units": None}

    # Step 1: Find "Realised" (but NOT "UnRealised") in row 2
    realised_col = None
    for c in range(10, ws.max_column + 1):
        val = ws.cell(2, c).value
        if val:
            val_str = str(val).strip()
            if val_str == "Realised" or val_str == "Realized":
                realised_col = c
                break

    if realised_col is None:
        return cols

    # Step 2: Read row 4 (header row) from the Realised column onward
    # and map sub-headers to column positions
    for c in range(realised_col, min(realised_col + 10, ws.max_column + 1)):
        header = ws.cell(header_row, c).value
        if not header:
            continue
        header = str(header).strip()

        if header == "Price" and cols["sell_price"] is None:
            cols["sell_price"] = c
        elif header == "Date" and cols["sell_date"] is None:
            cols["sell_date"] = c
        elif header == "Gain" and cols["sell_gain"] is None:
            cols["sell_gain"] = c
        elif header == "Units" and cols["sold_units"] is None:
            cols["sold_units"] = c

    return cols


def _parse_trading_history(wb) -> Tuple[list, list, list]:
    """Parse Buy and Sell rows from Trading History sheet.

    ALL Buy rows go into held[]; ALL Sell rows go into sell_rows[].
    Held/sold determination is done purely via FIFO matching of Sell rows
    against Buy rows (in _parse_and_match_symbol).

    The Realised columns (W–AB) in dump files often have broken formula
    references (stale row numbers after row insertions) and cannot be
    trusted for held/sold assignment.  They are still written by the app's
    add_sell_transaction for display in Excel.

    Returns (held_lots, column_sold_lots, sell_rows, dividends):
      - held_lots: ALL Buy rows (non-DIV)
      - column_sold_lots: always empty (FIFO handles sold determination)
      - sell_rows: ALL Sell action rows
      - dividends: List of {date, amount, remarks} for dividend rows
    """
    held, sold, sell_rows, dividends = [], [], [], []
    if "Trading History" not in wb.sheetnames:
        return held, sold, sell_rows, dividends
    ws = wb["Trading History"]
    max_row = ws.max_row or 0
    if max_row < 5:
        return held, sold, sell_rows, dividends

    # Find header row
    header_row = None
    for r in range(1, min(11, max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, 5)]
        if "DATE" in vals and "ACTION" in vals:
            header_row = r
            break
    if header_row is None:
        return held, sold, sell_rows, dividends

    for row_idx in range(header_row + 1, max_row + 1):
        date_val = ws.cell(row_idx, 1).value        # A: DATE
        exch = ws.cell(row_idx, 2).value             # B: EXCH
        action = ws.cell(row_idx, 3).value           # C: ACTION
        qty = ws.cell(row_idx, 4).value              # D: QTY
        price = ws.cell(row_idx, 5).value            # E: PRICE

        if not action or not date_val:
            continue

        action = str(action).strip()
        exch = str(exch).strip() if exch else ""

        # Collect dividends
        if exch == "DIV":
            tx_date = _parse_date(date_val)
            cost_col_f = _safe_float(ws.cell(row_idx, 6).value)  # F = total amount (qty * price)
            per_share = _safe_float(price)                        # E = dividend per share
            div_qty = _safe_float(qty) or 0                       # D = number of units
            # Prefer column F (total amount); fall back to price or qty
            amount = cost_col_f or per_share or div_qty
            if amount and amount > 0:
                remarks_val = str(ws.cell(row_idx, 7).value or "").strip()
                dividends.append({
                    "date": tx_date or "",
                    "amount": amount,
                    "units": int(div_qty) if div_qty > 0 else 0,
                    "remarks": remarks_val if remarks_val != "~" else "",
                })
            continue

        # Collect ALL Sell rows for FIFO matching
        if action == "Sell":
            tx_date = _parse_date(date_val)
            if not tx_date:
                continue
            qty_int = _safe_int(qty)
            price_f = _safe_float(price)
            if qty_int > 0 and price_f > 0:
                exchange = exch if exch in ("NSE", "BSE") else "NSE"
                sell_rows.append({
                    "date": tx_date,
                    "quantity": qty_int,
                    "price": price_f,
                    "exchange": exchange,
                    "row_idx": row_idx,
                })
            continue

        # Only process Buy rows
        if action != "Buy":
            continue

        tx_date = _parse_date(date_val)
        if not tx_date:
            continue

        qty_int = _safe_int(qty)
        price_e = _safe_float(price)                         # E: transaction price per share
        cost_f = _safe_float(ws.cell(row_idx, 6).value)      # F: COST (value at cost incl. charges)
        if qty_int <= 0 or (price_e <= 0 and cost_f <= 0):
            continue

        # buy_price = per-unit cost from column F; fallback to column E
        if cost_f > 0 and qty_int > 0:
            buy_price = cost_f / qty_int
        else:
            buy_price = price_e
            cost_f = round(price_e * qty_int, 2)

        exchange = exch if exch in ("NSE", "BSE") else "NSE"

        # Every Buy row goes into held; FIFO matching (later) moves sold lots out
        held.append({
            "date": tx_date,
            "exchange": exchange,
            "quantity": qty_int,
            "price": buy_price,
            "raw_price": price_e,
            "cost": cost_f,
            "row_idx": row_idx,
        })

    return held, sold, sell_rows, dividends


# ═══════════════════════════════════════════════════════════
#  MAIN CLASS
# ═══════════════════════════════════════════════════════════

class XlsxPortfolio:
    """File-per-stock xlsx database with FIFO-derived holdings."""

    def __init__(self, stocks_dir: str | Path):
        self.stocks_dir = Path(stocks_dir)
        self.stocks_dir.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()  # Reentrant — nested calls OK

        # Caches keyed by symbol (combined data from all files)
        self._cache: Dict[str, Tuple[float, List[Holding], List[SoldPosition]]] = {}
        # symbol → primary filepath (for writes)
        self._file_map: Dict[str, Path] = {}
        # symbol → ALL file paths (primary + archives, for reads)
        self._all_files: Dict[str, List[Path]] = {}
        # symbol → company name
        self._name_map: Dict[str, str] = {}
        # holding_id → Holding  (rebuilt on every full scan)
        self._holding_index: Dict[str, Holding] = {}
        # holding_id → filepath  (to find file for sell ops)
        self._holding_file: Dict[str, Path] = {}

        # Manual prices
        self._manual_prices_file = Path(stocks_dir).parent.parent / "backend" / "data" / "manual_prices.json"
        self._ensure_manual_prices()

        # Build file map on init
        self._build_file_map()

    # ── Initialisation ────────────────────────────────────

    def _ensure_manual_prices(self):
        self._manual_prices_file.parent.mkdir(parents=True, exist_ok=True)
        if not self._manual_prices_file.exists():
            with open(self._manual_prices_file, "w") as f:
                json.dump({}, f)

    def _build_file_map(self):
        """Scan xlsx files and build symbol ↔ filepath map.

        Multiple files for the same symbol (main + archive + duplicates)
        are all stored so that transactions can be merged during parsing.
        The non-archive file is treated as the primary (for writes).

        Symbol resolution order (no hardcoded map):
          1. Zerodha/NSE name→symbol lookup (dynamic, always fresh)
          2. Index sheet in the xlsx file (has "Code" like "NSE:RELIANCE")
          3. Derive from filename as last resort
        """
        # Ensure Zerodha/NSE symbol data is loaded (cached to disk, fast)
        _sym_resolver.ensure_loaded()

        for fp in sorted(self.stocks_dir.glob("*.xlsx")):
            if fp.name.startswith("~") or fp.name.startswith("."):
                continue

            stem = fp.stem

            # Skip "(1)" files — they are near-duplicates with rounding diffs
            if "(1)" in stem:
                continue

            clean = stem.replace("Archive_", "").replace(" - Archive", "").strip()

            # 1. Dynamic lookup: Zerodha/NSE name → symbol
            symbol = _sym_resolver.resolve_by_name(clean)

            # 2. Fallback: read Index sheet from xlsx (has Code like "NSE:RELIANCE")
            if not symbol:
                try:
                    wb = openpyxl.load_workbook(fp, data_only=True)
                    idx = _extract_index_data(wb)
                    wb.close()
                    symbol = idx.get("symbol")
                except Exception:
                    pass

            # 3. Last resort: derive from filename heuristic
            if not symbol:
                symbol = _sym_resolver.derive_symbol(clean)
                print(f"[XlsxDB] WARNING: Could not resolve '{clean}', derived: {symbol}")

            # Populate the runtime SYMBOL_MAP so other modules can reference it
            if clean not in SYMBOL_MAP:
                SYMBOL_MAP[clean] = symbol
                if symbol not in _REVERSE_MAP:
                    _REVERSE_MAP[symbol] = clean

            # Accumulate all files for this symbol
            if symbol not in self._all_files:
                self._all_files[symbol] = []
            self._all_files[symbol].append(fp)

            # Primary file = non-archive (for writes)
            is_archive = "Archive" in stem
            existing = self._file_map.get(symbol)
            if existing is None or (not is_archive):
                self._file_map[symbol] = fp
                self._name_map[symbol] = clean

        print(f"[XlsxDB] Indexed {len(self._file_map)} stock files "
              f"({sum(len(v) for v in self._all_files.values())} total including archives)")

    def reindex(self):
        """Re-scan the dumps folder for new/removed/modified xlsx files.

        Call this periodically (e.g. every refresh cycle) so that newly
        dropped xlsx files appear without a backend restart.
        """
        with self._lock:
            old_symbols = set(self._file_map.keys())
            # Clear maps and rebuild
            self._file_map.clear()
            self._all_files.clear()
            self._name_map.clear()
            self._build_file_map()
            new_symbols = set(self._file_map.keys())

            added = new_symbols - old_symbols
            removed = old_symbols - new_symbols

            # Invalidate caches for changed/new/removed symbols
            # Also invalidate existing symbols whose files may have changed
            for sym in new_symbols | removed:
                self._invalidate_symbol(sym)

            if added or removed:
                parts = []
                if added:
                    parts.append(f"+{len(added)}")
                if removed:
                    parts.append(f"-{len(removed)}")
                print(f"[XlsxDB] Reindex: {len(new_symbols)} stocks ({', '.join(parts)} changed)")
            return {"total": len(new_symbols), "added": list(added), "removed": list(removed)}

    def _find_file_for_symbol(self, symbol: str) -> Optional[Path]:
        """Find xlsx file for a given stock symbol."""
        symbol = symbol.upper()
        if symbol in self._file_map:
            return self._file_map[symbol]
        # Try glob fallback
        for fp in self.stocks_dir.glob("*.xlsx"):
            if symbol.lower() in fp.stem.lower():
                return fp
        return None

    # ── Cache Layer ───────────────────────────────────────

    def _get_stock_data(self, symbol: str) -> Tuple[List[Holding], List[SoldPosition]]:
        """Get holdings/sold for a symbol, combining all files, with mtime caching."""
        files = self._all_files.get(symbol, [])
        if not files:
            return [], [], []

        # Compute combined mtime (max of all files)
        try:
            combined_mtime = max(fp.stat().st_mtime for fp in files)
        except OSError:
            return [], [], []

        if symbol in self._cache:
            cached_mtime, cached_h, cached_s, cached_d = self._cache[symbol]
            if cached_mtime == combined_mtime:
                return cached_h, cached_s, cached_d

        holdings, sold, dividends = self._parse_and_match_symbol(symbol, files)
        self._cache[symbol] = (combined_mtime, holdings, sold, dividends)
        return holdings, sold, dividends

    def _invalidate_symbol(self, symbol: str):
        """Remove a symbol from cache so next read re-parses."""
        self._cache.pop(symbol, None)

    def _invalidate_all(self):
        """Clear all caches."""
        self._cache.clear()
        self._holding_index.clear()
        self._holding_file.clear()

    # ── Parse + FIFO ──────────────────────────────────────

    def _parse_and_match_symbol(self, symbol: str, files: List[Path]):
        """Parse ALL xlsx files for a symbol with hybrid sold detection.

        Column-based: Buy rows with Realised section data → sold via columns.
        FIFO-based: Explicit Sell rows are FIFO-matched against held lots.
        This ensures sells done through the app (which add Sell rows) are
        properly reflected alongside original column-tracked sells.
        """
        all_held, all_sold_raw, all_sell_rows, all_dividends = [], [], [], []
        exchange = "NSE"
        name = symbol

        for filepath in files:
            stem = filepath.stem
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
            except Exception as e:
                print(f"[XlsxDB] Failed to open {filepath.name}: {e}")
                continue

            idx = _extract_index_data(wb)
            held, sold, sell_rows, dividends = _parse_trading_history(wb)
            wb.close()

            # Use index data from the primary (non-archive) file
            if idx.get("exchange"):
                exchange = idx["exchange"]
            clean_name = stem.replace("Archive_", "").replace(" - Archive", "").strip()
            if not name or name == symbol:
                name = clean_name

            all_held.extend(held)
            all_sold_raw.extend(sold)
            all_sell_rows.extend(sell_rows)
            all_dividends.extend(dividends)

        if not all_held and not all_sold_raw and not all_sell_rows:
            return [], [], all_dividends

        # FIFO-match explicit Sell rows against held lots.
        # Column-tracked sells (Buy rows with Realised data) are already in
        # all_sold_raw and their lots are NOT in all_held, so no double-counting.
        if all_sell_rows and all_held:
            buys_for_fifo = [{
                "date": h["date"],
                "quantity": h["quantity"],
                "price": h["price"],
                "raw_price": h.get("raw_price", h["price"]),
                "cost": h.get("cost", 0),
                "exchange": h.get("exchange", exchange),
                "row_idx": h.get("row_idx", 0),
            } for h in all_held]

            sells_for_fifo = [{
                "date": s["date"],
                "quantity": s["quantity"],
                "price": s["price"],
            } for s in all_sell_rows]

            remaining, fifo_sold = fifo_match(buys_for_fifo, sells_for_fifo)

            # Replace held lots with remaining after FIFO
            all_held = [{
                "date": r["date"],
                "exchange": r.get("exchange", exchange),
                "quantity": r["remaining"],
                "price": r["price"],
                "raw_price": r.get("raw_price", r["price"]),
                "cost": round(r["price"] * r["remaining"], 2),
                "row_idx": r.get("row_idx", 0),
            } for r in remaining if r["remaining"] > 0]

            # Add FIFO-derived sold positions
            for fs in fifo_sold:
                all_sold_raw.append({
                    "buy_date": fs["buy_date"],
                    "buy_price": fs["buy_price"],
                    "sell_date": fs["sell_date"],
                    "sell_price": fs["sell_price"],
                    "quantity": fs["quantity"],
                    "realized_pl": fs["realized_pl"],
                    "exchange": fs.get("buy_exchange", exchange),
                    "row_idx": fs.get("row_idx", 0),
                })

        # Build Holding objects directly
        holdings = []
        for lot in all_held:
            h_id = _gen_id(symbol, lot.get("exchange", exchange),
                           lot["date"], lot["price"], lot.get("row_idx", 0))
            lot_cost = lot.get("cost", 0)
            if lot_cost <= 0:
                lot_cost = round(lot["price"] * lot["quantity"], 2)
            h = Holding(
                id=h_id,
                symbol=symbol,
                exchange=lot.get("exchange", exchange),
                name=name,
                quantity=lot["quantity"],
                price=round(lot.get("raw_price", lot["price"]), 2),
                buy_price=round(lot["price"], 2),
                buy_cost=round(lot_cost, 2),
                buy_date=lot["date"],
                notes="",
            )
            holdings.append(h)

        # Build SoldPosition objects directly
        sold = []
        for s in all_sold_raw:
            s_id = _gen_id(symbol + "_S", s.get("exchange", exchange),
                           s["sell_date"], s["sell_price"], s.get("row_idx", 0))
            sold.append(SoldPosition(
                id=s_id,
                symbol=symbol,
                exchange=s.get("exchange", exchange),
                name=name,
                quantity=s["quantity"],
                buy_price=round(s["buy_price"], 2),
                buy_date=s["buy_date"],
                sell_price=round(s["sell_price"], 2),
                sell_date=s["sell_date"],
                realized_pl=s["realized_pl"],
            ))

        return holdings, sold, all_dividends

    # ── Public READ API ───────────────────────────────────

    def get_all_holdings(self) -> List[Holding]:
        """Get all current holdings across every stock file."""
        all_holdings: List[Holding] = []
        self._holding_index.clear()
        self._holding_file.clear()

        # Snapshot to avoid "dict changed size during iteration" from bg reindex
        file_snapshot = dict(self._file_map)
        for symbol, primary_fp in file_snapshot.items():
            try:
                holdings, _, _ = self._get_stock_data(symbol)
                for h in holdings:
                    self._holding_index[h.id] = h
                    self._holding_file[h.id] = primary_fp
                all_holdings.extend(holdings)
            except Exception as e:
                print(f"[XlsxDB] Error reading {symbol}: {e}")

        return all_holdings

    def get_holding_by_id(self, holding_id: str) -> Optional[Holding]:
        """Lookup a specific holding by its deterministic ID."""
        # Try index first (populated by get_all_holdings)
        if holding_id in self._holding_index:
            return self._holding_index[holding_id]
        # Fallback: full scan
        self.get_all_holdings()
        return self._holding_index.get(holding_id)

    def get_all_sold(self) -> List[SoldPosition]:
        """Get all sold positions (FIFO-derived) across every stock file."""
        all_sold: List[SoldPosition] = []
        # Snapshot to avoid "dict changed size during iteration" from bg reindex
        symbols = list(self._file_map.keys())
        for symbol in symbols:
            try:
                _, sold, _ = self._get_stock_data(symbol)
                all_sold.extend(sold)
            except Exception as e:
                print(f"[XlsxDB] Error reading sold for {symbol}: {e}")
        return all_sold

    def get_dividends_by_symbol(self) -> dict:
        """Get dividend totals grouped by symbol.

        Returns {symbol: {"amount": total, "count": n_entries, "units": total_units}}.
        """
        result = {}
        # Snapshot to avoid "dict changed size during iteration" from bg reindex
        symbols = list(self._file_map.keys())
        for symbol in symbols:
            try:
                _, _, dividends = self._get_stock_data(symbol)
                if dividends:
                    result[symbol] = {
                        "amount": sum(d["amount"] for d in dividends),
                        "count": len(dividends),
                        "units": sum(d.get("units", 0) for d in dividends),
                    }
            except Exception as e:
                print(f"[XlsxDB] Error reading dividends for {symbol}: {e}")
        return result

    # ── Public WRITE API ──────────────────────────────────

    def add_holding(self, holding: Holding) -> Holding:
        """Add a Buy transaction to the stock's xlsx file."""
        symbol = holding.symbol.upper()
        exchange = holding.exchange.upper()
        name = holding.name or symbol

        filepath = self._find_file_for_symbol(symbol)
        if filepath is None:
            filepath = self._create_stock_file(symbol, exchange, name)

        tx = Transaction(
            date=holding.buy_date,
            exchange=exchange,
            action="Buy",
            quantity=holding.quantity,
            price=holding.buy_price,
            remarks=holding.notes or "~",
        )
        self._insert_transaction(filepath, tx)
        self._invalidate_symbol(symbol)

        # Re-parse to get the proper deterministic ID
        holdings, _, _ = self._get_stock_data(symbol)
        # Find the one we just added (most recent buy matching date+price+qty)
        for h in holdings:
            if (h.buy_date == holding.buy_date and
                    abs(h.buy_price - holding.buy_price) < 0.01 and
                    h.quantity == holding.quantity):
                return h

        # Fallback
        holding.id = _gen_id(symbol, exchange, holding.buy_date, holding.buy_price, 5)
        return holding

    def add_sell_transaction(self, symbol: str, exchange: str,
                             quantity: int, price: float, sell_date: str):
        """Write Realised data (columns W–AB) onto BUY rows in FIFO order.

        For each BUY row being sold, writes computed values into:
          W: sell price    X: sell date
          Y: gain %        Z: gain (net of commission)
          AA: gross         AB: units sold

        Also inserts a Sell record row at the top for tracking.
        Values are computed to match the original dump formulas:
          Z  = (1 - commission*2) * ((sell_price * units) - (E * units))
          AA = (E * units) + Z
          Y  = Z / SUM(F:I)
          AB = units
        """
        symbol = symbol.upper()
        filepath = self._find_file_for_symbol(symbol)
        if filepath is None:
            raise FileNotFoundError(f"No xlsx file for symbol {symbol}")

        with self._lock:
            # Convert Realised formulas → values so insert_rows won't break them
            self._convert_realised_formulas(filepath)

            wb = openpyxl.load_workbook(filepath)
            ws = wb["Trading History"]
            header_row = self._find_header_row(ws)

            # ── Ensure Realised section headers exist ───────────
            self._ensure_realised_headers(ws, header_row)
            rcols = _find_realised_columns(ws, header_row)

            col_w = rcols["sell_price"]     # W: Price
            col_x = rcols["sell_date"]      # X: Date
            col_z = rcols["sell_gain"]      # Z: Gain
            col_ab = rcols["sold_units"]    # AB: Units

            # Derive Gain% (between Date and Gain) and Gross (between Gain and Units)
            col_y = None
            if col_x and col_z and col_z - col_x == 2:
                col_y = col_x + 1
            col_aa = None
            if col_z and col_ab and col_ab - col_z == 2:
                col_aa = col_z + 1

            # ── Read commission rate from Index sheet ───────────
            commission = 0.007  # default
            if "Index" in wb.sheetnames:
                idx_ws = wb["Index"]
                comm_val = idx_ws.cell(2, 6).value  # Index!$F$2
                if comm_val and isinstance(comm_val, (int, float)):
                    commission = float(comm_val)

            # ── Step 1: Insert Sell record row at top ───────────
            max_row_before = ws.max_row or header_row
            insert_at = header_row + 1
            ws.insert_rows(insert_at)

            try:
                sell_dt = datetime.strptime(sell_date, "%Y-%m-%d")
            except (ValueError, TypeError):
                sell_dt = datetime.now()

            ws.cell(insert_at, 1, value=sell_dt)                          # A: DATE
            ws.cell(insert_at, 2, value=exchange)                         # B: EXCH
            ws.cell(insert_at, 3, value="Sell")                           # C: ACTION
            ws.cell(insert_at, 4, value=quantity)                         # D: QTY
            ws.cell(insert_at, 5, value=price)                            # E: PRICE
            ws.cell(insert_at, 6, value=round(price * quantity, 2))       # F: COST
            ws.cell(insert_at, 7, value="~")                              # G: REMARKS

            # ── Step 2: Find unsold BUY rows (after the insert) ─
            unsold = []
            for row_idx in range(insert_at + 1, max_row_before + 2):
                action = ws.cell(row_idx, 3).value
                exch_val = ws.cell(row_idx, 2).value
                if not action or str(action).strip() != "Buy":
                    continue
                if exch_val and str(exch_val).strip() == "DIV":
                    continue
                # Skip rows that already have Realised data (W column filled)
                sp = ws.cell(row_idx, col_w).value
                if sp is not None and sp != "" and _safe_float(sp) > 0:
                    continue
                qty = _safe_int(ws.cell(row_idx, 4).value)
                if qty <= 0:
                    continue
                dt = _parse_date(ws.cell(row_idx, 1).value)
                unsold.append({"row": row_idx, "qty": qty, "date": dt})

            # ── Step 3: FIFO sort (oldest first) and allocate ───
            unsold.sort(key=lambda x: (x["date"] or "", x["row"]))

            remaining_to_sell = quantity
            to_fill = []
            for u in unsold:
                if remaining_to_sell <= 0:
                    break
                sell_qty = min(u["qty"], remaining_to_sell)
                to_fill.append({
                    "row": u["row"],
                    "total_qty": u["qty"],
                    "sell_qty": sell_qty,
                    "is_full": sell_qty == u["qty"],
                })
                remaining_to_sell -= sell_qty

            if remaining_to_sell > 0:
                wb.close()
                raise ValueError(
                    f"Cannot sell {quantity} shares of {symbol}: "
                    f"only {quantity - remaining_to_sell} unsold shares available"
                )

            # ── Step 4: Write Realised data to BUY rows ─────────
            for info in to_fill:
                r = info["row"]
                sell_qty = info["sell_qty"]
                price_e = _safe_float(ws.cell(r, 5).value)   # E: PRICE per share
                cost_f = _safe_float(ws.cell(r, 6).value)     # F: COST
                stt = _safe_float(ws.cell(r, 8).value)        # H: STT
                add_chrg = _safe_float(ws.cell(r, 9).value)   # I: ADD CHRG
                total_cost = cost_f + stt + add_chrg           # = SUM(F:I)

                # Z: Gain = (1 - commission*2) * ((sell_price * units) - (E * units))
                gain = (1 - commission * 2) * ((price * sell_qty) - (price_e * sell_qty))
                gain = round(gain, 2)

                # AA: Gross = (E * units) + gain
                gross = round((price_e * sell_qty) + gain, 2)

                # Y: Gain% = Z / SUM(F:I)
                gain_pct = round(gain / total_cost, 6) if total_cost > 0 else 0

                # Write cells
                ws.cell(r, col_w, value=round(price, 2))           # W: Sell price
                ws.cell(r, col_x, value=sell_dt)                   # X: Sell date
                if col_y:
                    ws.cell(r, col_y, value=gain_pct)              # Y: Gain %
                if col_z:
                    ws.cell(r, col_z, value=gain)                  # Z: Gain
                if col_aa:
                    ws.cell(r, col_aa, value=gross)                # AA: Gross
                ws.cell(r, col_ab, value=sell_qty)                 # AB: Units

            wb.save(filepath)
        self._invalidate_symbol(symbol)

    def add_dividend(self, symbol: str, exchange: str, amount: float,
                     dividend_date: str, remarks: str = ""):
        """Insert a dividend row into the stock's xlsx file."""
        symbol = symbol.upper()
        filepath = self._find_file_for_symbol(symbol)
        if filepath is None:
            raise FileNotFoundError(f"No xlsx file for symbol {symbol}")

        tx = Transaction(
            date=dividend_date,
            exchange="DIV",
            action="Buy",         # action column, doesn't matter for DIV
            quantity=1,            # placeholder
            price=amount,          # dividend amount stored in price column
            remarks=remarks or "DIVIDEND",
        )
        self._insert_transaction(filepath, tx)
        self._invalidate_symbol(symbol)

    def remove_holding(self, holding_id: str) -> bool:
        """
        Remove a holding by deleting its Buy row from the xlsx.
        Use sparingly — selling is the normal workflow.
        """
        holding = self.get_holding_by_id(holding_id)
        if not holding:
            return False

        filepath = self._holding_file.get(holding_id)
        if not filepath:
            return False

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb["Trading History"]
            # Find the row
            header_row = self._find_header_row(ws)
            for row_idx in range(header_row + 1, ws.max_row + 1):
                tx_date = _parse_date(ws.cell(row_idx, 1).value)
                action = ws.cell(row_idx, 3).value
                price = _safe_float(ws.cell(row_idx, 5).value)
                if (action == "Buy" and tx_date == holding.buy_date and
                        abs(price - holding.buy_price) < 0.01):
                    ws.delete_rows(row_idx)
                    wb.save(filepath)
                    # Find symbol for this file to invalidate cache
                    for sym, fp in self._file_map.items():
                        if fp == filepath:
                            self._invalidate_symbol(sym)
                            break
                    return True
            wb.close()
        except Exception as e:
            print(f"[XlsxDB] Failed to remove holding: {e}")
        return False

    # ── XLSX Write Helpers ────────────────────────────────

    def _convert_realised_formulas(self, filepath: Path):
        """Replace Realised section formulas with cached values.

        Row insertion shifts cells; openpyxl updates formula text but
        drops cached computed values. Converting formulas → values first
        ensures the data survives row insertions intact.
        """
        # Read cached values
        wb_data = openpyxl.load_workbook(filepath, data_only=True)
        ws_data = wb_data["Trading History"]
        hr = self._find_header_row(ws_data)
        rcols = _find_realised_columns(ws_data, hr)
        cols = [c for c in [
            rcols.get("sell_price"), rcols.get("sell_date"),
            rcols.get("sell_gain"), rcols.get("sold_units"),
        ] if c is not None]
        # Derived: Gain%, Gross
        if rcols.get("sell_date") and rcols.get("sell_gain"):
            if rcols["sell_gain"] - rcols["sell_date"] == 2:
                cols.append(rcols["sell_date"] + 1)
        if rcols.get("sell_gain") and rcols.get("sold_units"):
            if rcols["sold_units"] - rcols["sell_gain"] == 2:
                cols.append(rcols["sell_gain"] + 1)

        cached: Dict[Tuple[int, int], object] = {}
        for r in range(hr + 1, (ws_data.max_row or hr) + 1):
            for c in cols:
                val = ws_data.cell(r, c).value
                if val is not None:
                    cached[(r, c)] = val
        wb_data.close()

        if not cached:
            return

        # Replace formulas with values
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Trading History"]
        changed = False
        for (r, c), val in cached.items():
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and str(cell.value).startswith("="):
                cell.value = val
                changed = True
        if changed:
            wb.save(filepath)
        else:
            wb.close()

    def _ensure_realised_headers(self, ws, header_row: int):
        """Create Realised section headers (W–AB) if they don't already exist."""
        rcols = _find_realised_columns(ws, header_row)
        if rcols["sell_price"] is not None:
            return  # Already has Realised section

        # Standard positions: W=23, X=24, Y=25, Z=26, AA=27, AB=28
        # Row 2: "Realised" marker
        ws.cell(2, 23, value="Realised")

        # Row header_row: sub-headers
        headers = {23: "Price", 24: "Date", 25: "Gain %", 26: "Gain", 27: "Gross", 28: "Units"}
        hdr_font = Font(bold=True)
        hdr_fill = PatternFill("solid", fgColor="FF967BB6")
        for col, name in headers.items():
            cell = ws.cell(header_row, col, value=name)
            cell.font = hdr_font
            cell.fill = hdr_fill

    def _find_header_row(self, ws) -> int:
        """Find the header row in a Trading History sheet."""
        for r in range(1, 11):
            vals = [ws.cell(r, c).value for c in range(1, 5)]
            if "DATE" in vals and "ACTION" in vals:
                return r
        return 4  # default

    def _insert_transaction(self, filepath: Path, tx: Transaction):
        """Insert a transaction row at the top of Trading History."""
        with self._lock:
            # Preserve cached formula values before row insertion
            self._convert_realised_formulas(filepath)

            wb = openpyxl.load_workbook(filepath)
            ws = wb["Trading History"]
            header_row = self._find_header_row(ws)
            insert_at = header_row + 1  # row 5 by default

            ws.insert_rows(insert_at)

            # A: DATE
            try:
                dt = datetime.strptime(tx.date, "%Y-%m-%d")
            except ValueError:
                dt = datetime.now()
            ws.cell(insert_at, 1, value=dt)

            # B: EXCH
            ws.cell(insert_at, 2, value=tx.exchange)
            # C: ACTION
            ws.cell(insert_at, 3, value=tx.action)
            # D: QTY
            ws.cell(insert_at, 4, value=tx.quantity)
            # E: PRICE
            ws.cell(insert_at, 5, value=tx.price)
            # F: COST (use explicit cost if provided, else compute from price * qty)
            cost_val = tx.cost if tx.cost > 0 else round(tx.price * tx.quantity, 2)
            ws.cell(insert_at, 6, value=cost_val)
            # G: REMARKS
            ws.cell(insert_at, 7, value=tx.remarks or "~")
            # H: STT
            if tx.stt:
                ws.cell(insert_at, 8, value=tx.stt)
            # I: ADD CHRG
            if tx.add_chrg:
                ws.cell(insert_at, 9, value=tx.add_chrg)

            # J: Current Price formula (only for Buy rows)
            if tx.action == "Buy":
                ws.cell(insert_at, 10, value="=Index!$C$2")

            wb.save(filepath)

    def _create_stock_file(self, symbol: str, exchange: str, company_name: str) -> Path:
        """Create a new xlsx file with proper template structure."""
        filename = f"{company_name}.xlsx"
        filepath = self.stocks_dir / filename

        wb = openpyxl.Workbook()

        # ── Trading History sheet ─────────────────────────
        ws = wb.active
        ws.title = "Trading History"

        headers = [
            "DATE", "EXCH", "ACTION", "QTY", "PRICE", "COST",
            "REMARKS", "STT", "ADD CHRG", "Current Price",
            "Gain %", "Gain", "Gross", "Units",
        ]
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="FF967BB6")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(4, col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Sub-headers (rows 2-3) to match existing format
        ws.cell(2, 10, value="UnRealised")
        ws.cell(3, 10, value="Total")

        # Realised section headers (W=23 through AB=28)
        ws.cell(2, 23, value="Realised")
        realised_headers = {23: "Price", 24: "Date", 25: "Gain %", 26: "Gain", 27: "Gross", 28: "Units"}
        for col, h in realised_headers.items():
            cell = ws.cell(4, col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Column widths
        widths = {1: 12, 2: 6, 3: 6, 4: 6, 5: 10, 6: 12,
                  7: 10, 8: 8, 9: 8, 10: 12, 11: 10, 12: 12, 13: 12, 14: 8}
        for col, w in widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        # ── Index sheet ───────────────────────────────────
        ws_idx = wb.create_sheet("Index")
        ws_idx.cell(1, 2, value="Code")
        ws_idx.cell(1, 3, value=f"{exchange}:{symbol}")
        ws_idx.cell(2, 2, value="Current Price")
        ws_idx.cell(2, 3, value=0)
        ws_idx.cell(2, 5, value="Commission")
        ws_idx.cell(2, 6, value=0.007)
        ws_idx.cell(3, 2, value="52 Week High")
        ws_idx.cell(3, 3, value=0)
        ws_idx.cell(4, 2, value="52 Week Low")
        ws_idx.cell(4, 3, value=0)
        ws_idx.cell(6, 2, value="UnRealised")
        ws_idx.cell(8, 1, value="All")
        ws_idx.cell(8, 2, value="Units")
        ws_idx.cell(8, 3, value=0)
        ws_idx.cell(11, 2, value="Invested")
        ws_idx.cell(11, 3, value=0)

        wb.save(filepath)

        # Register in maps
        self._file_map[symbol] = filepath
        self._name_map[symbol] = company_name
        if symbol not in self._all_files:
            self._all_files[symbol] = []
        self._all_files[symbol].append(filepath)

        print(f"[XlsxDB] Created new stock file: {filename}")
        return filepath

    # ── Manual Prices ─────────────────────────────────────

    def get_manual_price(self, symbol: str, exchange: str) -> Optional[float]:
        try:
            with open(self._manual_prices_file) as f:
                prices = json.load(f)
            return prices.get(f"{symbol}.{exchange}")
        except (FileNotFoundError, json.JSONDecodeError):
            return None

    def set_manual_price(self, symbol: str, exchange: str, price: float):
        try:
            with open(self._manual_prices_file) as f:
                prices = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            prices = {}
        prices[f"{symbol}.{exchange}"] = price
        with open(self._manual_prices_file, "w") as f:
            json.dump(prices, f, indent=2)

    def get_all_manual_prices(self) -> dict:
        try:
            with open(self._manual_prices_file) as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return {}


# ═══════════════════════════════════════════════════════════
#  MODULE-LEVEL SINGLETON
# ═══════════════════════════════════════════════════════════

# Path: relative to backend/ directory → ../dumps/Stocks
_STOCKS_DIR = Path(__file__).parent.parent.parent / "dumps" / "Stocks"

xlsx_db = XlsxPortfolio(_STOCKS_DIR)
