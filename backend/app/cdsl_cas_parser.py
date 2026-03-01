"""
CDSL CAS (Consolidated Account Statement) PDF parser for Mutual Funds.

Extracts fund sections and transaction lines from CDSL CAS PDFs
covering all AMCs (ICICI, Nippon, SBI, Tata, etc.) using pdfplumber
table extraction as the primary method.

The PDF has bilingual Hindi/English text that garbles plain-text extraction,
but pdfplumber table extraction works perfectly. Each MF transaction table
has the structure:
  Row 0: ['ISIN : INFxxx UCC : ...', None, None, ...]  ← ISIN identifier
  Row 1: ['Date', 'Transaction Description', 'Amount', 'NAV', 'Price', 'Units', ...]  ← header
  Row 2: ['', 'Opening Balance', '', '', '', '360.314', ...]
  Row 3: ['12-01-2026', 'SIP Purchase - ...', '4999.75', '41.04', '41.04', '121.826', '.25', '0', '0']
  Row N: ['', 'Closing Balance', '', '', '', '482.14', ...]
"""

import io
import re
from datetime import datetime
from typing import Optional

import pdfplumber

from .mf_xlsx_database import mf_db, _safe_float


# ── Regex patterns ────────────────────────────────────────

_ISIN_RE = re.compile(r"ISIN\s*:\s*(INF\w+)", re.IGNORECASE)
_FOLIO_RE = re.compile(r"Folio\s*No\s*:\s*(\S+)", re.IGNORECASE)
_AMC_HEADER_RE = re.compile(r"AMC\s+Name\s*:\s*(.+)", re.IGNORECASE)
_SCHEME_HEADER_RE = re.compile(r"^(\w{3,6})\s*[-–]\s*(.+)")
_SCHEME_CODE_NUM_RE = re.compile(r"^(\d{3,6})\s*[-–]\s*(.+)")
_PERIOD_RE = re.compile(
    r"(?:STATEMENT.*?PERIOD.*?FROM\s+)(\d{2}-\d{2}-\d{4})\s+TO\s+(\d{2}-\d{2}-\d{4})",
    re.IGNORECASE,
)
_PERIOD_ALT_RE = re.compile(
    r"(\d{2}-\d{2}-\d{4})\s+(?:से|to)\s+(\d{2}-\d{2}-\d{4})",
    re.IGNORECASE,
)

# Descriptions to skip
_SKIP_DESCRIPTIONS = {
    "opening balance", "opening bal", "closing balance", "closing bal",
    "stt", "total tax", "net assets",
}


def _parse_number(s: str) -> float:
    """Parse a number string, removing commas and whitespace."""
    if not s:
        return 0.0
    cleaned = s.replace(",", "").replace("\n", "").strip()
    if not cleaned or cleaned == "--" or cleaned == "-":
        return 0.0
    return float(cleaned)


def _parse_date_ddmmyyyy(s: str) -> str:
    """Convert DD-MM-YYYY to YYYY-MM-DD."""
    try:
        return datetime.strptime(s.strip(), "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return s


def _determine_action(description: str) -> str:
    """Determine Buy/Sell from transaction description.

    Handles:
    - Redemption / Switch Out → Sell
    - Reversal (e.g. "SIP Purchase (Reversal)") → Sell (undoes a previous buy)
    - Insufficient Balance (failed SIP attempt recorded in CAS) → Sell (paired with
      a matching Buy entry; together they cancel out to net zero)
    """
    desc_upper = description.upper()
    if "REDEMPTION" in desc_upper or "SWITCH OUT" in desc_upper:
        return "Sell"
    if "REVERSAL" in desc_upper:
        return "Sell"
    if "INSUFFICIENT" in desc_upper:
        return "Sell"
    return "Buy"


def _should_skip_row(description: str) -> bool:
    """Check if a transaction row should be skipped."""
    desc_lower = description.strip().lower()
    for skip in _SKIP_DESCRIPTIONS:
        if desc_lower.startswith(skip):
            return True
    return False


def _clean_description(desc: str) -> str:
    """Clean multi-line description from table cell."""
    if not desc:
        return ""
    # Table cells may have newlines from multi-line content
    return " ".join(desc.replace("\n", " ").split()).strip()


def _match_fund_code(isin: str, fund_name: str) -> Optional[str]:
    """Try to match a parsed fund's ISIN against existing MF database entries.

    First checks if ISIN already exists as a fund_code in the database.
    Falls back to name matching only if ISIN not found (threshold 0.85).
    """
    # Primary: match by ISIN directly
    if isin in mf_db._file_map:
        return isin

    # Fallback: name matching with strict threshold
    parsed_upper = fund_name.upper().strip()
    parsed_words = set(parsed_upper.replace("-", " ").split())
    significant_parsed = {w for w in parsed_words if len(w) > 2}

    if not significant_parsed:
        return None

    best_code = None
    best_overlap = 0.0

    for code, name in mf_db._name_map.items():
        name_upper = name.upper().strip()
        existing_words = set(name_upper.replace("-", " ").split())
        significant_existing = {w for w in existing_words if len(w) > 2}
        if not significant_existing:
            continue
        significant_common = significant_parsed & significant_existing
        overlap = len(significant_common) / max(len(significant_parsed), len(significant_existing))
        if overlap > best_overlap:
            best_overlap = overlap
            best_code = code

    return best_code if best_overlap >= 0.85 else None


def _check_duplicate(fund_code: str, tx_date: str, units: float, nav: float) -> bool:
    """Check if a transaction already exists in the fund's xlsx."""
    filepath = mf_db._file_map.get(fund_code)
    if not filepath or not filepath.exists():
        return False

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Trading History"]
        header_row = mf_db._find_header_row(ws)

        try:
            tx_dt = datetime.strptime(tx_date, "%Y-%m-%d").date()
        except ValueError:
            wb.close()
            return False

        for row in range(header_row + 1, (ws.max_row or 0) + 1):
            row_date = ws.cell(row, 1).value
            row_units = ws.cell(row, 4).value
            row_nav = ws.cell(row, 5).value

            if row_date is None:
                continue

            if isinstance(row_date, datetime):
                row_date = row_date.date()
            elif hasattr(row_date, "date"):
                pass
            elif isinstance(row_date, str):
                try:
                    row_date = datetime.strptime(row_date.strip(), "%Y-%m-%d").date()
                except ValueError:
                    continue
            else:
                continue

            if (row_date == tx_dt
                    and abs(_safe_float(row_units) - units) < 1e-4
                    and abs(_safe_float(row_nav) - nav) < 1e-2):
                wb.close()
                return True

        wb.close()
    except Exception:
        pass

    return False


def _extract_metadata_from_text(all_text: str) -> dict:
    """Extract fund metadata from the PDF text (both Account Details and MF sections).

    Builds a mapping of ISIN → {amc, scheme_name, scheme_code, folio}.

    Handles two PDF formats:
    1. New (2025+): Account Details section with "AMC Name:", "Scheme Name:", "Scheme Code:"
    2. Old (2019-2024): MF section with direct AMC names ("HDFC Mutual Fund"),
       scheme codes ("MCOGT - HDFC Mid-Cap..."), Folio/ISIN lines.

    In both formats, folio appears BEFORE ISIN.
    """
    isin_map = {}  # ISIN → metadata dict
    current_amc = ""
    current_scheme_name = ""
    current_scheme_code = ""
    pending_folio = ""
    awaiting_continuation = False

    for line in all_text.split("\n"):
        line = line.strip()
        if not line:
            continue

        # AMC header: "AMC Name : ICICI Prudential Mutual Fund"
        m = _AMC_HEADER_RE.match(line)
        if m:
            current_amc = m.group(1).strip()
            current_scheme_name = ""
            current_scheme_code = ""
            pending_folio = ""
            awaiting_continuation = False
            continue

        # Direct AMC name (old format): "HDFC Mutual Fund", "Kotak Mutual Fund"
        if line.endswith("Mutual Fund") and len(line) < 60 and ":" not in line:
            current_amc = line
            current_scheme_name = ""
            current_scheme_code = ""
            pending_folio = ""
            awaiting_continuation = False
            continue

        # "Scheme Name : ..." — may also have "Scheme Code : xxx" on same line
        if line.startswith("Scheme Name"):
            rest = line.split(":", 1)[1].strip() if ":" in line else ""
            sc_match = re.search(r"Scheme\s+Code\s*:\s*(\S+)", rest)
            if sc_match:
                current_scheme_code = sc_match.group(1)
                current_scheme_name = rest[:sc_match.start()].strip()
                awaiting_continuation = False
            else:
                current_scheme_name = rest
                awaiting_continuation = False
            pending_folio = ""
            continue

        # "Scheme Code : 9453" on its own line
        if line.startswith("Scheme Code"):
            current_scheme_code = line.split(":", 1)[1].strip() if ":" in line else ""
            awaiting_continuation = True
            continue

        # Continuation line for scheme name (after Scheme Code line)
        # MUST come before scheme header check, because lines like
        # "Plan - Growth" match _SCHEME_HEADER_RE but are actually continuations
        if awaiting_continuation and current_scheme_name and not line.startswith("KYC"):
            current_scheme_name = current_scheme_name + " " + line
            awaiting_continuation = False
            continue

        # Scheme header (old format): "MCOGT - HDFC Mid-Cap..." or "9453 - ICICI Prudential..."
        m = _SCHEME_CODE_NUM_RE.match(line)
        if not m:
            m = _SCHEME_HEADER_RE.match(line)
        if m and current_amc:
            current_scheme_code = m.group(1).strip()
            current_scheme_name = m.group(2).strip()
            pending_folio = ""
            awaiting_continuation = False
            continue

        # Folio line (appears BEFORE ISIN)
        m = _FOLIO_RE.search(line)
        if m and current_amc:
            pending_folio = m.group(1).strip()
            awaiting_continuation = False
            continue

        # ISIN line — finalizes the current fund entry
        m = _ISIN_RE.search(line)
        if m and current_amc and current_scheme_name:
            isin = m.group(1).strip()
            if isin not in isin_map:
                isin_map[isin] = {
                    "amc": current_amc,
                    "scheme_name": current_scheme_name,
                    "scheme_code": current_scheme_code,
                    "folio": pending_folio,
                }
            current_scheme_name = ""
            current_scheme_code = ""
            pending_folio = ""
            awaiting_continuation = False
            continue

    return isin_map


def _extract_statement_info(all_text: str) -> tuple[str, str]:
    """Extract CAS ID and statement period from early pages."""
    cas_id = ""
    statement_period = ""

    # CAS ID from the text (e.g., in filename or text like "AA00604621")
    m = re.search(r"\b(AA\d{8,})\b", all_text[:5000])
    if m:
        cas_id = m.group(1)

    # Statement period
    m = _PERIOD_RE.search(all_text[:10000])
    if m:
        statement_period = f"{m.group(1)} to {m.group(2)}"
    else:
        m = _PERIOD_ALT_RE.search(all_text[:10000])
        if m:
            statement_period = f"{m.group(1)} to {m.group(2)}"

    return cas_id, statement_period


def parse_cdsl_cas(pdf_bytes: bytes) -> dict:
    """Parse a CDSL CAS PDF and return structured fund/transaction data.

    Uses a table-first approach:
    1. Extract fund metadata (AMC, scheme, ISIN, folio) from text
    2. Extract MF transaction tables from all pages
    3. Each table's first row has ISIN to link to fund metadata
    4. Parse transaction rows from tables
    """
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))

    # ── Phase 1: Extract all text ──
    all_text = ""
    for page in pdf.pages:
        text = page.extract_text() or ""
        all_text += text + "\n"

    # ── Phase 2: Extract metadata from Account Details section ──
    isin_map = _extract_metadata_from_text(all_text)
    cas_id, statement_period = _extract_statement_info(all_text)

    # ── Phase 3: Extract MF transaction tables from ALL pages ──
    # Each MF transaction table starts with an ISIN row
    funds = {}  # ISIN → fund dict

    # Track current AMC/scheme from text for the MF transaction pages.
    # On transaction pages, AMC appears directly as e.g. "ICICI Prudential Mutual Fund"
    # (without "AMC Name :" prefix), and scheme as "9453 - ICICI Prudential ...".
    current_amc = ""
    current_scheme_name = ""
    current_scheme_code = ""
    last_text_isin = ""  # Track last ISIN seen in page text (for cross-page tables)

    for page in pdf.pages:
        page_text = page.extract_text() or ""
        tables = page.extract_tables()

        # Save ISIN from previous page BEFORE this page's text scan overwrites it.
        # Continuation tables on this page belong to the fund from the previous page,
        # not from any new fund header found in this page's text.
        prev_page_last_isin = last_text_isin

        if not tables:
            # Still scan text for ISIN tracking even without tables
            for line in page_text.split("\n"):
                m = _ISIN_RE.search(line.strip())
                if m:
                    last_text_isin = m.group(1).strip()
            continue

        # Update AMC/scheme from page text (MF transaction section)
        for line in page_text.split("\n"):
            line = line.strip()
            if not line:
                continue
            # Direct AMC name line: short line ending with "Mutual Fund"
            if line.endswith("Mutual Fund") and len(line) < 60 and ":" not in line:
                current_amc = line
                continue
            m = _AMC_HEADER_RE.match(line)
            if m:
                current_amc = m.group(1).strip()
                continue
            # Scheme header: "9453 - ICICI Prudential ..." (numeric code)
            m = _SCHEME_CODE_NUM_RE.match(line)
            if m:
                current_scheme_code = m.group(1).strip()
                current_scheme_name = m.group(2).strip()
                continue
            # Scheme header: "MCOGT - HDFC Mid-Cap ..." (alpha code)
            m = _SCHEME_HEADER_RE.match(line)
            if m:
                current_scheme_code = m.group(1).strip()
                current_scheme_name = m.group(2).strip()
                continue
            # Track last ISIN seen in text
            m = _ISIN_RE.search(line)
            if m:
                last_text_isin = m.group(1).strip()

        for ti, table in enumerate(tables):
            if not table or len(table) < 2:
                continue

            # Detect table format: new (ISIN in row 0) or old (Folio in row 0, ISIN in row 1)
            first_row = table[0]
            if not first_row:
                continue

            first_cell = str(first_row[0] or "").strip()
            second_cell = _clean_description(str(first_row[1])) if len(first_row) > 1 and first_row[1] else ""
            isin = None
            folio = ""
            data_start_row = 2  # default: skip row 0 (ISIN/Folio) + row 1 (header)

            # Format 1 (new, 2025+): Row 0 has "ISIN : INFxxx"
            if first_cell:
                isin_match = _ISIN_RE.search(first_cell)
                if isin_match:
                    isin = isin_match.group(1).strip()

            # Format 2 (old, 2019-2024): Row 0 has "Folio No :", Row 1 has "ISIN :"
            if not isin and first_cell.startswith("Folio No"):
                folio_m = _FOLIO_RE.search(first_cell)
                if folio_m:
                    folio = folio_m.group(1).strip()
                # Check row 1 for ISIN
                if len(table) > 1 and table[1] and table[1][0]:
                    row1_cell = str(table[1][0]).strip()
                    isin_match = _ISIN_RE.search(row1_cell)
                    if isin_match:
                        isin = isin_match.group(1).strip()
                        data_start_row = 3  # skip Folio row, ISIN row, header row

            # Format 3a (cross-page split): Table starts with header row "Date"
            # Always use prev_page_last_isin because cross-page splits only
            # come from the fund on the previous page. The current page's text
            # scan may have already overwritten last_text_isin with a new fund.
            if not isin and first_cell.lower().startswith("date") and prev_page_last_isin:
                isin = prev_page_last_isin
                data_start_row = 1  # skip header row only

            # Format 3b (cross-page split): Table starts with "Opening Balance" row
            # e.g. Row 0: ['', 'Opening Balance', '', '', '', '6482.653', ...]
            # Continuation from previous page's fund (nav tables with "Summary"
            # header won't match this pattern)
            if not isin and not first_cell and second_cell.lower().startswith("opening bal") and prev_page_last_isin:
                isin = prev_page_last_isin
                data_start_row = 0  # all rows are data (opening bal will be skipped by _should_skip_row)

            # Format 3c (cross-page split): Table starts directly with date data
            # e.g. Row 0: ['06-05-2024', 'Purchase - Systematic-...', ...]
            # Continuation from previous page (date pattern won't match nav headers)
            if not isin and re.match(r"\d{2}-\d{2}-\d{4}", first_cell) and prev_page_last_isin:
                isin = prev_page_last_isin
                data_start_row = 0  # all rows are data

            # Format 3d (cross-page split): Table starts with empty date + STT/Closing Balance
            # e.g. Row 0: ['', 'STT', '.25', ...] or ['', 'Closing Balance', ...]
            # Tail-end rows of a fund that split across pages
            if not isin and not first_cell and second_cell and prev_page_last_isin:
                if _should_skip_row(second_cell) or second_cell.upper().startswith("TOTAL TAX"):
                    isin = prev_page_last_isin
                    data_start_row = 0

            if not isin:
                continue

            # Look up metadata for this ISIN
            meta = isin_map.get(isin, {})
            amc = meta.get("amc", current_amc)
            scheme_name = meta.get("scheme_name", current_scheme_name)
            scheme_code = meta.get("scheme_code", current_scheme_code)
            if not folio:
                folio = meta.get("folio", "")

            # Create or get fund entry
            if isin not in funds:
                matched_code = _match_fund_code(isin, scheme_name)
                funds[isin] = {
                    "fund_code": matched_code or isin,
                    "fund_name": scheme_name,
                    "isin": isin,
                    "amc": amc,
                    "scheme_code": scheme_code,
                    "folio": folio,
                    "is_new_fund": matched_code is None,
                    "transactions": [],
                }

            fund = funds[isin]
            closing_balance = 0.0
            opening_balance = 0.0

            # Parse transaction rows (skip metadata + header rows)
            for row in table[data_start_row:]:
                if not row or len(row) < 6:
                    continue

                cells = [str(c).strip() if c else "" for c in row]
                date_str = cells[0].strip()
                description = _clean_description(cells[1]) if len(cells) > 1 else ""

                # Check for opening/closing balance rows
                if _should_skip_row(description):
                    if description.lower().startswith("closing bal"):
                        try:
                            closing_balance = _parse_number(cells[5]) if len(cells) > 5 else 0.0
                        except (ValueError, IndexError):
                            pass
                    elif description.lower().startswith("opening bal"):
                        try:
                            opening_balance = _parse_number(cells[5]) if len(cells) > 5 else 0.0
                        except (ValueError, IndexError):
                            pass
                    continue

                # Must have a date: DD-MM-YYYY
                if not re.match(r"\d{2}-\d{2}-\d{4}", date_str):
                    continue

                # Parse numeric fields
                try:
                    amount = _parse_number(cells[2]) if len(cells) > 2 else 0.0
                    nav = _parse_number(cells[3]) if len(cells) > 3 else 0.0
                    price = _parse_number(cells[4]) if len(cells) > 4 else 0.0
                    units = _parse_number(cells[5]) if len(cells) > 5 else 0.0
                    stamp_duty = _parse_number(cells[6]) if len(cells) > 6 else 0.0
                except (ValueError, IndexError):
                    continue

                if units == 0.0 or nav == 0.0:
                    continue

                fund["transactions"].append({
                    "date": _parse_date_ddmmyyyy(date_str),
                    "description": description,
                    "action": _determine_action(description),
                    "amount": abs(amount),
                    "nav": nav,
                    "units": round(abs(units), 4),
                    "balance_units": 0.0,
                    "stamp_duty": abs(stamp_duty),
                })

            # Attach opening/closing balance to fund
            if opening_balance and "opening_balance" not in fund:
                fund["opening_balance"] = round(abs(opening_balance), 4)
            if closing_balance and fund["transactions"]:
                fund["transactions"][-1]["balance_units"] = round(abs(closing_balance), 4)

    pdf.close()

    # ── Phase 4: Dedup check ──
    funds_list = list(funds.values())
    for fund in funds_list:
        for tx in fund["transactions"]:
            tx["isDuplicate"] = _check_duplicate(
                fund["fund_code"], tx["date"], tx["units"], tx["nav"]
            )

    # Filter out funds with no transactions
    funds_list = [f for f in funds_list if len(f["transactions"]) > 0]

    # ── Build summary ──
    total_purchases = sum(
        1 for f in funds_list for t in f["transactions"] if t["action"] == "Buy"
    )
    total_redemptions = sum(
        1 for f in funds_list for t in f["transactions"] if t["action"] == "Sell"
    )

    return {
        "cas_id": cas_id,
        "statement_period": statement_period,
        "source": "CDSL",
        "funds": funds_list,
        "summary": {
            "total_purchases": total_purchases,
            "total_redemptions": total_redemptions,
            "funds_count": len(funds_list),
        },
    }
