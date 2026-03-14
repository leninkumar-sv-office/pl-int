"""
Import stock data from xlsx dump files into the portfolio database.

Parses each xlsx file's "Trading History" and "Index" sheets to extract:
  - Buy transactions → current holdings (after FIFO sell matching)
  - Sell transactions → sold positions with realized P&L
  - Dividend entries are recorded as notes but skipped for holdings

Usage:
    python import_dump.py [--dump-dir /path/to/dumps/Stocks]
"""

import openpyxl
import json
import os
import sys
import uuid
from datetime import datetime, date
from pathlib import Path

# ── Paths ─────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
DUMP_DIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/Users/lenin/Google Drive/My Drive/pl/dumps") / "Stocks"
DB_FILE = SCRIPT_DIR / "data" / "portfolio.json"

# ── Symbol resolution: dynamic from Zerodha + NSE (no hardcoded map) ──
# import_dump.py runs as a standalone script, so we add app/ to sys.path
sys.path.insert(0, str(SCRIPT_DIR))
from app.symbol_resolver import ensure_loaded as _ensure_symbols, resolve_by_name as _resolve_name, derive_symbol as _derive


def get_short_id():
    return str(uuid.uuid4())[:8]


def extract_index_data(wb):
    """Extract metadata from the Index sheet."""
    index_data = {"code": None, "exchange": "NSE", "symbol": None,
                  "current_price": 0, "week_52_high": 0, "week_52_low": 0,
                  "total_units": 0, "total_invested": 0}

    if "Index" not in wb.sheetnames:
        return index_data

    ws = wb["Index"]
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False):
        values = [cell.value for cell in row]
        # Look for Code
        if len(values) >= 3 and values[1] == "Code" and values[2]:
            code = str(values[2])
            index_data["code"] = code
            if ":" in code:
                parts = code.split(":")
                index_data["exchange"] = parts[0]
                index_data["symbol"] = parts[1]
        # Current Price
        if len(values) >= 3 and values[1] == "Current Price" and values[2]:
            try:
                index_data["current_price"] = float(values[2])
            except (ValueError, TypeError):
                pass
        # 52 Week High
        if len(values) >= 3 and values[1] == "52 Week High" and values[2]:
            try:
                index_data["week_52_high"] = float(values[2])
            except (ValueError, TypeError):
                pass
        # 52 Week Low
        if len(values) >= 3 and values[1] == "52 Week Low" and values[2]:
            try:
                index_data["week_52_low"] = float(values[2])
            except (ValueError, TypeError):
                pass
        # Units (total holdings)
        if len(values) >= 3 and values[1] == "Units":
            try:
                index_data["total_units"] = int(float(values[2] or 0))
            except (ValueError, TypeError):
                pass
        # Invested
        if len(values) >= 3 and values[1] == "Invested":
            try:
                v = values[2]
                if isinstance(v, (int, float)):
                    index_data["total_invested"] = float(v)
            except (ValueError, TypeError):
                pass

    return index_data


def parse_trading_history(wb):
    """Parse Buy/Sell transactions from Trading History sheet."""
    buys = []
    sells = []
    dividends = []

    if "Trading History" not in wb.sheetnames:
        return buys, sells, dividends

    ws = wb["Trading History"]

    # Find the header row (contains DATE, EXCH, ACTION, QTY, PRICE, COST)
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        values = [cell.value for cell in row]
        if "DATE" in values and "ACTION" in values:
            header_row = row_idx
            break

    if header_row is None:
        return buys, sells, dividends

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 6:
            continue

        date_val = row[0]
        exch = row[1]
        action = row[2]
        qty = row[3]
        price = row[4]
        cost = row[5]
        remarks = row[6] if len(row) > 6 else ""

        # Skip empty/invalid rows
        if not action or not date_val:
            continue

        action = str(action).strip()
        exch = str(exch).strip() if exch else ""

        # Parse date
        if isinstance(date_val, datetime):
            tx_date = date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, date):
            tx_date = date_val.strftime("%Y-%m-%d")
        else:
            continue  # Skip rows with non-date values

        # Skip dividend rows for holdings
        if exch == "DIV":
            if qty and isinstance(qty, (int, float)):
                dividends.append({
                    "date": tx_date,
                    "quantity": int(qty),
                    "amount_per_share": float(price) if isinstance(price, (int, float)) else 0,
                    "total": float(cost) if isinstance(cost, (int, float)) else 0,
                    "remarks": str(remarks or ""),
                })
            continue

        # Parse quantity and price
        try:
            qty = int(float(qty)) if qty and isinstance(qty, (int, float)) else 0
            price = float(price) if price and isinstance(price, (int, float)) else 0
            cost = float(cost) if cost and isinstance(cost, (int, float)) else 0
        except (ValueError, TypeError):
            continue

        if qty <= 0 or price <= 0:
            continue

        tx = {
            "date": tx_date,
            "exchange": exch if exch in ("NSE", "BSE") else "NSE",
            "quantity": qty,
            "price": price,
            "cost": cost,
            "remarks": str(remarks or ""),
        }

        if action == "Buy":
            buys.append(tx)
        elif action == "Sell":
            sells.append(tx)

    return buys, sells, dividends


def fifo_match(buys, sells):
    """
    Apply FIFO matching of sells to buys.
    Returns: (remaining_holdings, sold_positions)
    """
    # Sort buys by date (oldest first) for FIFO
    buys_sorted = sorted(buys, key=lambda x: x["date"])
    # Sort sells by date
    sells_sorted = sorted(sells, key=lambda x: x["date"])

    # Create mutable copies with remaining qty
    buy_lots = []
    for b in buys_sorted:
        buy_lots.append({**b, "remaining": b["quantity"]})

    sold_positions = []

    for sell in sells_sorted:
        sell_qty = sell["quantity"]
        sell_price = sell["price"]
        sell_date = sell["date"]

        for lot in buy_lots:
            if sell_qty <= 0:
                break
            if lot["remaining"] <= 0:
                continue

            matched_qty = min(lot["remaining"], sell_qty)
            realized_pl = (sell_price - lot["price"]) * matched_qty

            sold_positions.append({
                "buy_price": lot["price"],
                "buy_date": lot["date"],
                "buy_exchange": lot["exchange"],
                "sell_price": sell_price,
                "sell_date": sell_date,
                "quantity": matched_qty,
                "realized_pl": round(realized_pl, 2),
            })

            lot["remaining"] -= matched_qty
            sell_qty -= matched_qty

    # Remaining holdings
    remaining = [lot for lot in buy_lots if lot["remaining"] > 0]

    return remaining, sold_positions


def process_file(filepath):
    """Process a single xlsx file and return holdings and sold positions."""
    filename = Path(filepath).stem  # Without extension
    print(f"  Processing: {filename}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"    ⚠ Failed to open: {e}")
        return [], [], []

    # Extract metadata from Index
    index_data = extract_index_data(wb)

    # Determine symbol and exchange
    symbol = index_data.get("symbol")
    exchange = index_data.get("exchange", "NSE")

    if not symbol:
        # Dynamic lookup: Zerodha/NSE name → symbol
        clean_name = filename.replace("Archive_", "").strip()
        symbol = _resolve_name(clean_name)

    if not symbol:
        # Generate from filename (rough fallback)
        symbol = _derive(clean_name)
        print(f"    ⚠ No symbol found, using derived: {symbol}")

    # Parse transactions
    buys, sells, dividends = parse_trading_history(wb)
    wb.close()

    if not buys and not sells:
        print(f"    ⚠ No transactions found, skipping")
        return [], [], []

    print(f"    Found {len(buys)} buys, {len(sells)} sells, {len(dividends)} dividends")

    # FIFO match
    remaining_lots, sold_matches = fifo_match(buys, sells)

    # Build holdings
    holdings = []
    for lot in remaining_lots:
        holdings.append({
            "id": get_short_id(),
            "symbol": symbol,
            "exchange": lot.get("exchange", exchange),
            "name": filename.replace("Archive_", "").replace(" - Archive", "").strip(),
            "quantity": lot["remaining"],
            "buy_price": round(lot["price"], 2),
            "buy_date": lot["date"],
            "notes": "",
        })

    # Build sold positions
    sold = []
    for s in sold_matches:
        sold.append({
            "id": get_short_id(),
            "symbol": symbol,
            "exchange": s.get("buy_exchange", exchange),
            "name": filename.replace("Archive_", "").replace(" - Archive", "").strip(),
            "quantity": s["quantity"],
            "buy_price": round(s["buy_price"], 2),
            "buy_date": s["buy_date"],
            "sell_price": round(s["sell_price"], 2),
            "sell_date": s["sell_date"],
            "realized_pl": s["realized_pl"],
        })

    total_remaining = sum(h["quantity"] for h in holdings)
    expected = index_data.get("total_units", 0)
    check = "✓" if expected == 0 or total_remaining == expected else f"⚠ expected {expected}"
    print(f"    → {total_remaining} units held ({len(holdings)} lots), {len(sold)} sold positions {check}")

    return holdings, sold, dividends


def main():
    print("═" * 60)
    print("  Stock Dump Importer")
    print(f"  Source: {DUMP_DIR}")
    print(f"  Target: {DB_FILE}")
    print("═" * 60)

    if not DUMP_DIR.exists():
        print(f"❌ Dump directory not found: {DUMP_DIR}")
        sys.exit(1)

    # Load symbol data from Zerodha/NSE (cached to disk)
    _ensure_symbols()

    # Find all xlsx files (skip duplicates with "(1)")
    xlsx_files = sorted(DUMP_DIR.glob("*.xlsx"))
    # Filter out duplicates
    seen_bases = set()
    filtered_files = []
    for f in xlsx_files:
        base = f.stem.replace("(1)", "").strip()
        if base in seen_bases:
            print(f"  Skipping duplicate: {f.name}")
            continue
        seen_bases.add(base)
        filtered_files.append(f)

    print(f"\nFound {len(filtered_files)} stock files to process\n")

    all_holdings = []
    all_sold = []
    errors = []

    for filepath in filtered_files:
        try:
            holdings, sold, dividends = process_file(filepath)
            all_holdings.extend(holdings)
            all_sold.extend(sold)
        except Exception as e:
            errors.append((filepath.name, str(e)))
            print(f"    ❌ Error: {e}")

    # Write to database
    db_data = {
        "holdings": all_holdings,
        "sold": all_sold,
        "manual_prices": {},
    }

    os.makedirs(DB_FILE.parent, exist_ok=True)

    # Backup existing database
    if DB_FILE.exists():
        backup = DB_FILE.with_suffix(".json.bak")
        with open(DB_FILE, "r") as f:
            old = f.read()
        with open(backup, "w") as f:
            f.write(old)
        print(f"\n📦 Backed up existing database to {backup.name}")

    with open(DB_FILE, "w") as f:
        json.dump(db_data, f, indent=2, default=str)

    # Summary
    total_invested = sum(h["buy_price"] * h["quantity"] for h in all_holdings)
    total_realized = sum(s["realized_pl"] for s in all_sold)
    unique_stocks = len(set(h["symbol"] for h in all_holdings))

    print(f"\n{'═' * 60}")
    print(f"  IMPORT COMPLETE")
    print(f"{'═' * 60}")
    print(f"  Holdings:       {len(all_holdings)} lots across {unique_stocks} stocks")
    print(f"  Total Invested:  ₹{total_invested:,.2f}")
    print(f"  Sold Positions: {len(all_sold)}")
    print(f"  Realized P&L:   ₹{total_realized:,.2f}")
    if errors:
        print(f"\n  Errors ({len(errors)}):")
        for name, err in errors:
            print(f"    - {name}: {err}")
    print(f"\n  Database saved to: {DB_FILE}")
    print()


if __name__ == "__main__":
    main()
