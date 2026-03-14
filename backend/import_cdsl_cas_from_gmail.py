"""
One-time script: Process downloaded CDSL CAS PDFs, parse, and import MF transactions.

Approach:
1. Parse ALL CAS PDFs and collect all transactions per ISIN
2. De-duplicate at the transaction level (same date + units + NAV + action)
3. Seed opening balances from earliest CAS for pre-2019 funds
4. Import unique transactions chronologically
5. Verify against latest CAS closing balances + user's actual portfolio

Usage:
    cd backend && python import_cdsl_cas_from_gmail.py
"""

import os
import sys
import io
import glob
import re
from pathlib import Path
from collections import defaultdict, Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.mf_xlsx_database import mf_db

PAN = "AEPPL3176B"
PDF_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "cas_pdfs")


def decrypt_pdf(pdf_bytes, password):
    import pikepdf
    try:
        reader = pikepdf.open(io.BytesIO(pdf_bytes), password=password)
        out = io.BytesIO()
        reader.save(out)
        reader.close()
        return out.getvalue()
    except Exception:
        return None


def main():
    import pdfplumber
    from app.cdsl_cas_parser import parse_cdsl_cas, _ISIN_RE, _parse_number

    # Step 1: Delete existing MF xlsx files
    mf_dir = mf_db.mf_dir
    existing_files = list(Path(mf_dir).glob("*.xlsx"))
    if existing_files:
        print(f"Deleting {len(existing_files)} existing MF xlsx files...")
        for f in existing_files:
            f.unlink()
        mf_db._file_map.clear()
        mf_db._name_map.clear()
        print()

    # Step 2: Get all PDFs sorted chronologically
    pdf_files = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    print(f"Found {len(pdf_files)} PDF files\n")

    # Step 3: Parse ALL CAS PDFs and collect transactions per ISIN
    print("Phase 1: Parsing all CAS PDFs...")
    fund_meta = {}    # ISIN → {fund_name, fund_code, amc, scheme_code, folio, is_new_fund}
    fund_cas_txns = defaultdict(list)  # ISIN → [(cas_index, tx_dict), ...]
    fund_opening = {}  # ISIN → opening_balance (ONLY from first CAS appearance)
    fund_seen = set()  # ISINs we've seen in any CAS (tracks first appearance)
    fund_closing = {}  # ISIN → closing_balance (from LAST CAS where fund appears)
    fund_cas_history = defaultdict(list)  # ISIN → [(period, opening, closing, first_tx_date, first_tx_nav)]
    parsed_count = 0
    seen_periods = set()
    skipped_pdfs = []

    for pdf_path in pdf_files:
        fname = os.path.basename(pdf_path)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
        decrypted = decrypt_pdf(pdf_bytes, PAN)
        if not decrypted:
            continue
        try:
            result = parse_cdsl_cas(decrypted)
        except Exception:
            continue

        tx_count = sum(len(f.get("transactions", [])) for f in result.get("funds", []))
        if tx_count == 0:
            continue

        period = result.get("statement_period", "")
        parsed_count += 1

        for fund in result.get("funds", []):
            isin = fund["isin"]

            # Store metadata (first seen wins for name, update fund_code)
            if isin not in fund_meta:
                fund_meta[isin] = {
                    "fund_name": fund["fund_name"],
                    "fund_code": fund.get("fund_code", isin),
                    "amc": fund.get("amc", ""),
                    "scheme_code": fund.get("scheme_code", ""),
                    "folio": fund.get("folio", ""),
                    "is_new_fund": fund.get("is_new_fund", True),
                }

            # Track opening balance ONLY from the FIRST CAS where this fund appears.
            # If the first appearance has opening > 0, it means pre-statement history.
            # If first appearance has opening == 0, fund started within CAS coverage.
            if isin not in fund_seen:
                fund_seen.add(isin)
                opening = fund.get("opening_balance", 0.0)
                if opening > 0:
                    fund_opening[isin] = opening

            # Track closing balance (latest CAS wins, even if 0 after full redemption)
            last_tx = fund["transactions"][-1] if fund["transactions"] else None
            if last_tx and "balance_units" in last_tx:
                fund_closing[isin] = last_tx["balance_units"]

            # Track CAS history for gap detection
            txns = fund.get("transactions", [])
            first_tx = txns[0] if txns else None
            fund_cas_history[isin].append({
                "period": period,
                "opening": fund.get("opening_balance", 0.0),
                "closing": fund_closing.get(isin, 0.0),  # just updated above
                "first_tx_date": first_tx["date"] if first_tx else None,
                "first_tx_nav": first_tx["nav"] if first_tx else None,
            })

            # Collect all transactions tagged with CAS index
            for tx in txns:
                fund_cas_txns[isin].append((parsed_count, tx))

        print(f"  [{parsed_count}] {fname[:55]} | {period} | {tx_count} txns")

    print(f"\n  Total parsed: {parsed_count} PDFs, {len(fund_meta)} funds")
    total_raw = sum(len(txns) for txns in fund_cas_txns.values())
    print(f"  Total raw transactions: {total_raw}")

    # Step 3b: Detect gaps from missing CAS PDFs
    # We find the May 2019 CAS closing and July 2019 CAS opening for each fund.
    # If they differ, the June 2019 CAS PDF is missing and we infer the transaction.
    from datetime import datetime, timedelta

    print("\nPhase 1b: Detecting CAS gaps (missing June 2019 PDF)...")
    gap_count = 0

    for isin, history in fund_cas_history.items():
        may_closing = None
        jul_entry = None
        for entry in history:
            period = entry["period"]
            if period.startswith("01-05-2019"):
                may_closing = entry["closing"]
            elif period.startswith("01-07-2019") and jul_entry is None:
                jul_entry = entry

        if may_closing is None or jul_entry is None or may_closing <= 0:
            continue
        jul_opening = jul_entry["opening"]
        if jul_opening <= 0 or abs(jul_opening - may_closing) < 0.5:
            continue

        gap = jul_opening - may_closing
        action = "Buy" if gap > 0 else "Sell"
        gap_units = round(abs(gap), 4)
        gap_nav = jul_entry["first_tx_nav"] or 10.0
        gap_amount = round(gap_units * gap_nav, 2)

        synthetic_tx = {
            "date": "2019-06-15",
            "description": "Inferred from CAS gap (missing June 2019 PDF)",
            "action": action,
            "amount": gap_amount,
            "nav": gap_nav,
            "units": gap_units,
            "balance_units": 0.0,
            "stamp_duty": 0.0,
        }
        fund_cas_txns[isin].append((0, synthetic_tx))
        gap_count += 1
        name = fund_meta.get(isin, {}).get("fund_name", isin)[:45]
        print(f"  Gap: {name} | {action} {gap_units:.4f} units @ ~{gap_nav:.4f} on ~2019-06-15")

    if gap_count:
        print(f"  Total gaps filled: {gap_count}")
    else:
        print("  No gaps detected")

    # Step 4: De-duplicate transactions per ISIN using per-CAS multi-set approach
    # Problem: Two identical SIP transactions on the same date in the SAME CAS are real.
    # Old set-based approach would remove one. Fix: for each key, keep max(count_per_cas).
    print("\nPhase 2: De-duplicating transactions (per-CAS multi-set)...")
    fund_unique_txns = {}  # ISIN → [unique tx dicts, sorted by date]

    def tx_key(tx):
        return (tx["date"], tx["action"], round(tx["units"], 3), round(tx["nav"], 2))

    total_unique = 0
    total_dups = 0
    for isin, tagged_txns in fund_cas_txns.items():
        # Count occurrences of each key per CAS
        cas_key_counts = defaultdict(Counter)  # cas_index → Counter(key → count)
        for cas_idx, tx in tagged_txns:
            cas_key_counts[cas_idx][tx_key(tx)] += 1

        # For each key, keep max count seen in any single CAS
        max_counts = Counter()  # key → max count across all CAS documents
        for cas_idx, counter in cas_key_counts.items():
            for key, count in counter.items():
                if count > max_counts[key]:
                    max_counts[key] = count

        # Collect unique transactions: pick up to max_counts[key] instances of each key
        used_counts = Counter()  # key → how many we've kept so far
        unique = []
        # Process in CAS order (lowest cas_idx first = chronological)
        tagged_txns.sort(key=lambda x: (x[0], x[1]["date"]))
        for cas_idx, tx in tagged_txns:
            key = tx_key(tx)
            if used_counts[key] < max_counts[key]:
                unique.append(tx)
                used_counts[key] += 1
                total_unique += 1
            else:
                total_dups += 1

        # Sort by date
        unique.sort(key=lambda t: t["date"])
        fund_unique_txns[isin] = unique

    print(f"  Unique transactions: {total_unique}")
    print(f"  Duplicates removed:  {total_dups}")

    # Step 5: Verify formula for each fund before import
    # Expected: opening_balance + sum(buys) - sum(sells) ≈ closing_balance
    print("\nPhase 3: Pre-import verification (formula check)...")
    print(f"{'Fund':<50} {'Open':>8} {'+ Buy':>10} {'- Sell':>10} {'= Calc':>10} {'CAS Cls':>10} {'Diff':>8}")
    print("-" * 110)

    formula_ok = 0
    formula_issue = 0
    seeds_to_remove = set()
    for isin in sorted(fund_unique_txns.keys()):
        meta = fund_meta[isin]
        txns = fund_unique_txns[isin]
        opening = fund_opening.get(isin, 0.0)
        closing = fund_closing.get(isin, 0.0)
        buy_units = sum(t["units"] for t in txns if t["action"] == "Buy")
        sell_units = sum(t["units"] for t in txns if t["action"] == "Sell")
        calc = opening + buy_units - sell_units
        diff = calc - closing

        # Auto-correct: if diff ≈ opening seed, the seed duplicates transactions
        # (transactions already cover the opening balance)
        if opening > 0 and closing > 0 and abs(diff - opening) < 1.0:
            seeds_to_remove.add(isin)
            calc -= opening
            diff = calc - closing

        status = "OK" if abs(diff) < 1.0 else "!!"
        if abs(diff) < 1.0:
            formula_ok += 1
        else:
            formula_issue += 1
        flag = " (seed removed)" if isin in seeds_to_remove else ""
        print(f"{meta['fund_name'][:49]:<50} {opening:>8.1f} {buy_units:>+10.1f} {sell_units:>-10.1f} {calc:>10.1f} {closing:>10.1f} {diff:>+8.1f} {status}{flag}")

    # Remove incorrect seeds
    for isin in seeds_to_remove:
        del fund_opening[isin]
        print(f"  >> Removed seed for {fund_meta[isin]['fund_name'][:50]} (transactions already cover opening)")

    print(f"\n  Formula OK: {formula_ok}, Issues: {formula_issue}")

    # Step 6: Extract cost basis from "Mutual Fund Units Held" summary table (earliest CAS)
    print("\nPhase 4: Extracting cost basis from 'Mutual Fund Units Held' tables...")
    holdings_cost = {}  # ISIN → avg_cost_per_unit

    # Check earliest few CAS PDFs for summary tables
    for pdf_path in pdf_files:
        fname = os.path.basename(pdf_path)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
        decrypted = decrypt_pdf(pdf_bytes, PAN)
        if not decrypted:
            continue

        pdf = pdfplumber.open(io.BytesIO(decrypted))
        found_table = False
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                if not table or not table[0] or not table[0][0]:
                    continue
                header = str(table[0][0]).upper()
                if "MUTUAL FUND" not in header or "HELD" not in header:
                    continue
                found_table = True
                for row in table[3:]:
                    if not row or len(row) < 7:
                        continue
                    # Find ISIN in row
                    isin_val = None
                    for cell in row:
                        if cell:
                            m = re.search(r"INF\w{9,}", str(cell))
                            if m:
                                isin_val = m.group(0)
                                break
                    if not isin_val:
                        continue
                    try:
                        units = _parse_number(str(row[4] or ""))
                        invested = _parse_number(str(row[6] or ""))
                        if units > 0 and invested > 0:
                            avg_cost = round(invested / units, 4)
                            if isin_val not in holdings_cost:
                                holdings_cost[isin_val] = avg_cost
                    except (ValueError, IndexError):
                        pass
        pdf.close()
        if found_table:
            # Only need the earliest CAS with this table
            break

    print(f"  Found cost basis for {len(holdings_cost)} funds from summary table")
    for isin, cost in holdings_cost.items():
        name = fund_meta.get(isin, {}).get("fund_name", "?")[:40]
        units = fund_opening.get(isin, 0.0)
        print(f"    {name}: {units:.2f} units @ avg cost {cost:.4f}")

    # Step 7: Seed opening balances for funds with pre-period history
    print("\nPhase 5: Seeding opening balances...")
    seeded = 0
    for isin, opening_units in fund_opening.items():
        if opening_units < 0.001:
            continue
        meta = fund_meta[isin]
        fc = meta["fund_code"]
        fn = meta["fund_name"]

        # Get cost basis
        if isin in holdings_cost:
            nav = holdings_cost[isin]
            cost_source = "summary table"
        else:
            # Fallback: use first transaction NAV
            first_tx = fund_unique_txns.get(isin, [None])[0]
            nav = first_tx["nav"] if first_tx else 10.0
            cost_source = "first tx NAV"

        # Use a date just before the earliest transaction
        txns = fund_unique_txns.get(isin, [])
        if txns:
            # Set opening balance date to day before first transaction
            from datetime import datetime, timedelta
            first_date = datetime.strptime(txns[0]["date"], "%Y-%m-%d")
            seed_date = (first_date - timedelta(days=1)).strftime("%Y-%m-%d")
        else:
            seed_date = "2019-01-01"

        try:
            mf_db.add_mf_holding(
                fund_code=fc, fund_name=fn,
                units=opening_units, nav=nav,
                buy_date=seed_date,
                remarks=f"Opening balance (pre-statement history)",
            )
            print(f"  Seeded {fn[:45]}: {opening_units:.4f} units @ {nav:.4f} ({cost_source})")
            seeded += 1
        except Exception as e:
            print(f"  WARN: Failed to seed {fn[:40]}: {e}")

    print(f"  Total seeded: {seeded} funds")

    # Step 8: Import unique transactions chronologically
    print(f"\nPhase 6: Importing {total_unique} unique transactions...")
    total_buys, total_sells, total_skipped = 0, 0, 0
    total_errors = []

    # Flatten and sort ALL transactions across all funds by date
    all_txns = []
    for isin, txns in fund_unique_txns.items():
        meta = fund_meta[isin]
        for tx in txns:
            all_txns.append((isin, meta, tx))
    all_txns.sort(key=lambda x: x[2]["date"])

    for isin, meta, tx in all_txns:
        fc = meta["fund_code"]
        fn = meta["fund_name"]
        try:
            if tx["action"] == "Buy":
                mf_db.add_mf_holding(fund_code=fc, fund_name=fn,
                    units=float(tx["units"]), nav=float(tx["nav"]),
                    buy_date=tx["date"], remarks=tx.get("description", ""),
                    skip_dup_check=True)
                total_buys += 1
            elif tx["action"] == "Sell":
                mf_db.add_mf_sell_transaction(fund_code=fc,
                    units=float(tx["units"]), nav=float(tx["nav"]),
                    sell_date=tx["date"], remarks=tx.get("description", ""))
                total_sells += 1
        except ValueError as e:
            if "Duplicate" in str(e):
                total_skipped += 1
            else:
                total_errors.append(f"  {fn[:40]}: {tx.get('date','?')} - {e}")
        except Exception as e:
            total_errors.append(f"  {fn[:40]}: {tx.get('date','?')} - {e}")

    print(f"  Buys imported:  {total_buys}")
    print(f"  Sells imported: {total_sells}")
    print(f"  Dups skipped:   {total_skipped}")
    print(f"  Errors:         {len(total_errors)}")

    # Step 9: Verify against latest CAS closing balances
    import openpyxl

    print("\n" + "=" * 120)
    print("VERIFICATION: DB units vs Latest CAS closing balances")
    print("=" * 120)
    print(f"{'Fund':<50} {'CAS Close':>10} {'DB Units':>10} {'Diff':>10} {'Status':>8}")
    print("-" * 92)

    match_count = 0
    mismatch_count = 0
    for isin in sorted(fund_closing.keys()):
        meta = fund_meta.get(isin, {})
        fn = meta.get("fund_name", isin)
        cas_close = fund_closing[isin]

        # Calculate DB units
        fc = meta.get("fund_code", isin)
        filepath = mf_db._file_map.get(fc)
        db_units = 0.0
        if filepath and filepath.exists():
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb["Trading History"]
            hr = mf_db._find_header_row(ws)
            for row in range(hr + 1, (ws.max_row or 0) + 1):
                action = ws.cell(row, 3).value
                u = ws.cell(row, 4).value
                if u is None:
                    continue
                u = float(u)
                if action == "Buy":
                    db_units += u
                elif action == "Sell":
                    db_units -= u
            wb.close()

        diff = db_units - cas_close
        status = "OK" if abs(diff) < 0.5 else "DIFF"
        if abs(diff) < 0.5:
            match_count += 1
        else:
            mismatch_count += 1
        print(f"{fn[:49]:<50} {cas_close:>10.4f} {db_units:>10.4f} {diff:>+10.4f} {status:>8}")

    print("-" * 92)
    print(f"  OK:   {match_count} funds | DIFF: {mismatch_count} funds")

    # Step 10: Compare with user's actual portfolio (known values)
    user_portfolio = {
        "INF200K01RY0": ("SBI Equity Hybrid", 299.43),
        "INF200K01TS8": ("SBI Conservative Hybrid", 1129.09),
        "INF200K01T51": ("SBI Small Cap", 82.07),
        "INF109K01Y31": ("ICICI ELSS", 68.96),
        "INF109K1A252": ("ICICI Nifty200", 4485.13),
        "INF109KC1RH9": ("ICICI Opportunities", 602.15),
        "INF204KC1378": ("Nippon Silver ETF FOF", 431.387),
        "INF204KB19V4": ("Nippon Multi Asset", 636.477),
        "INF204K01XF9": ("Nippon Multi Cap", 44.540),
    }

    print("\n" + "=" * 120)
    print("USER PORTFOLIO COMPARISON (actual holdings as of Mar 2026)")
    print("Note: Latest CAS is Jan 2026, so ~1 month of SIPs may be missing")
    print("=" * 120)
    print(f"{'Fund':<30} {'Actual':>10} {'CAS Jan':>10} {'DB':>10} {'DB-CAS':>10} {'Actual-CAS':>10}")
    print("-" * 85)

    for isin, (name, actual) in user_portfolio.items():
        cas_close = fund_closing.get(isin, 0.0)
        fc = fund_meta.get(isin, {}).get("fund_code", isin)
        filepath = mf_db._file_map.get(fc)
        db_units = 0.0
        if filepath and filepath.exists():
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb["Trading History"]
            hr = mf_db._find_header_row(ws)
            for row in range(hr + 1, (ws.max_row or 0) + 1):
                action = ws.cell(row, 3).value
                u = ws.cell(row, 4).value
                if u is None:
                    continue
                u = float(u)
                if action == "Buy":
                    db_units += u
                elif action == "Sell":
                    db_units -= u
            wb.close()

        db_vs_cas = db_units - cas_close
        actual_vs_cas = actual - cas_close
        print(f"{name:<30} {actual:>10.3f} {cas_close:>10.3f} {db_units:>10.3f} {db_vs_cas:>+10.3f} {actual_vs_cas:>+10.3f}")

    # Summary
    print("\n" + "=" * 120)
    print("IMPORT SUMMARY")
    print("=" * 120)
    print(f"  PDFs parsed:       {parsed_count}")
    print(f"  Dups removed:      {total_dups} (transaction-level de-dup)")
    print(f"  Opening seeds:     {seeded}")
    print(f"  Buys imported:     {total_buys}")
    print(f"  Sells imported:    {total_sells}")
    print(f"  Xlsx dups:         {total_skipped}")
    print(f"  Errors:            {len(total_errors)}")
    if total_errors:
        for e in total_errors[:20]:
            print(f"    {e}")
        if len(total_errors) > 20:
            print(f"    ... and {len(total_errors) - 20} more")
    print("=" * 120)


if __name__ == "__main__":
    main()
