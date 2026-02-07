#!/usr/bin/env python3
"""
Stock Portfolio Duplicate Row Analysis Script

Analyzes xlsx files in dumps/Stocks/ to identify potential duplicate rows
created during the migration process from .xls load files.
"""

import os
import glob
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Column mapping for Trading History sheet
COLUMNS = {
    'A': 'DATE',
    'B': 'EXCH',
    'C': 'ACTION',
    'D': 'QTY',
    'E': 'PRICE',
    'F': 'COST',
    'G': 'REMARKS',
    'H': 'STT',
    'I': 'ADD CHRG'
}

REVERSE_COLUMNS = {v: k for k, v in COLUMNS.items()}


class TradeRecord:
    """Represents a single trade record"""
    def __init__(self, row_num: int, data: dict):
        self.row_num = row_num
        self.date = data.get('DATE')
        self.exch = data.get('EXCH')
        self.action = data.get('ACTION')
        self.qty = data.get('QTY')
        self.price = data.get('PRICE')
        self.cost = data.get('COST')
        self.remarks = data.get('REMARKS')
        self.stt = data.get('STT')
        self.add_chrg = data.get('ADD CHRG')

    def __repr__(self):
        return f"Row {self.row_num}: {self.date} {self.action} {self.qty} @ {self.price}"


def safe_value(cell_value):
    """Safely convert cell value, handling None and various types"""
    if cell_value is None:
        return None
    if isinstance(cell_value, str):
        val = cell_value.strip()
        return val if val else None
    if isinstance(cell_value, (int, float)):
        return cell_value
    if isinstance(cell_value, datetime):
        return cell_value
    return cell_value


def parse_numeric(value):
    """Parse numeric value, return None if not a number"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except (ValueError, AttributeError):
            return None
    return None


def parse_date(value):
    """Parse date value"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), '%d-%m-%Y').date()
        except ValueError:
            try:
                return datetime.strptime(value.strip(), '%d/%m/%Y').date()
            except ValueError:
                return None
    return None


def is_buy_or_sell(action):
    """Check if action is Buy or Sell"""
    if action is None:
        return False
    action_upper = str(action).strip().upper()
    return action_upper in ['BUY', 'SELL']


def prices_similar(price1, price2, tolerance_percent=5):
    """Check if two prices are similar within tolerance"""
    p1 = parse_numeric(price1)
    p2 = parse_numeric(price2)

    if p1 is None or p2 is None:
        return False

    if p1 == 0 or p2 == 0:
        return p1 == p2

    # Calculate percentage difference
    avg = (abs(p1) + abs(p2)) / 2
    diff_percent = (abs(p1 - p2) / avg) * 100

    return diff_percent <= tolerance_percent


def fuzzy_key(record: TradeRecord) -> Tuple:
    """
    Create a fuzzy key for duplicate detection
    Components: date, action, quantity
    Price similarity is checked separately
    """
    date_str = str(record.date) if record.date else 'NONE'
    action_upper = str(record.action).strip().upper() if record.action else 'NONE'
    qty = record.qty

    return (date_str, action_upper, qty)


def find_duplicates(records: List[TradeRecord]) -> List[Tuple[TradeRecord, TradeRecord]]:
    """
    Find potential duplicate pairs
    Returns list of (record1, record2) tuples
    """
    duplicates = []

    for i, rec1 in enumerate(records):
        for rec2 in records[i+1:]:
            # Check if fuzzy key matches (date + action + qty)
            if fuzzy_key(rec1) == fuzzy_key(rec2):
                # Check if prices are similar (within 5%)
                if prices_similar(rec1.price, rec2.price, tolerance_percent=5):
                    duplicates.append((rec1, rec2))

    return duplicates


def analyze_file(filepath: str) -> Dict:
    """
    Analyze a single xlsx file for duplicates
    """
    result = {
        'filepath': filepath,
        'filename': os.path.basename(filepath),
        'total_rows': 0,
        'buy_sell_rows': 0,
        'duplicates': [],
        'error': None
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        if 'Trading History' not in wb.sheetnames:
            result['error'] = "No 'Trading History' sheet found"
            return result

        ws = wb['Trading History']
        records = []

        # Skip header row (row 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_letter, col_name in COLUMNS.items():
                cell = ws[f'{col_letter}{row_idx}']
                row_data[col_name] = safe_value(cell.value)

            # Only consider Buy/Sell rows
            if is_buy_or_sell(row_data.get('ACTION')):
                result['total_rows'] += 1
                record = TradeRecord(row_idx, row_data)
                records.append(record)
                result['buy_sell_rows'] += 1

        # Find duplicates
        duplicates = find_duplicates(records)
        result['duplicates'] = duplicates

        wb.close()

    except Exception as e:
        result['error'] = str(e)

    return result


def format_value(val):
    """Format value for display"""
    if val is None:
        return "None"
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


def main():
    """Main execution"""
    # Find all xlsx files in dumps/Stocks
    stocks_dir = '/sessions/vibrant-nice-lamport/mnt/pl/dumps/Stocks'

    if not os.path.isdir(stocks_dir):
        print(f"Error: Directory not found: {stocks_dir}")
        return

    # Get all xlsx files, excluding those starting with ~, . or in _backup
    xlsx_files = []
    for filepath in glob.glob(os.path.join(stocks_dir, '*.xlsx')):
        filename = os.path.basename(filepath)
        if not filename.startswith('~') and not filename.startswith('.') and '_backup' not in filepath:
            xlsx_files.append(filepath)

    xlsx_files.sort()

    if not xlsx_files:
        print(f"No xlsx files found in {stocks_dir}")
        return

    print(f"\n{'='*100}")
    print(f"STOCK PORTFOLIO DUPLICATE ANALYSIS")
    print(f"Directory: {stocks_dir}")
    print(f"Total files to analyze: {len(xlsx_files)}")
    print(f"{'='*100}\n")

    total_files = len(xlsx_files)
    files_with_duplicates = 0
    total_duplicates = 0

    for idx, filepath in enumerate(xlsx_files, 1):
        filename = os.path.basename(filepath)

        print(f"\n[{idx}/{total_files}] Analyzing: {filename}")
        print("-" * 100)

        result = analyze_file(filepath)

        if result['error']:
            print(f"  ERROR: {result['error']}")
            continue

        print(f"  Total Buy/Sell rows: {result['buy_sell_rows']}")
        print(f"  Potential duplicates found: {len(result['duplicates'])}")

        if result['duplicates']:
            files_with_duplicates += 1
            total_duplicates += len(result['duplicates'])

            print(f"\n  DUPLICATE PAIRS:")
            for dup_idx, (rec1, rec2) in enumerate(result['duplicates'], 1):
                print(f"\n    Duplicate Pair {dup_idx}:")
                print(f"      Row {rec1.row_num} (older):  {rec1.date} | {rec1.action:4} | Qty: {format_value(rec1.qty):>8} | Price: {format_value(rec1.price):>10} | Cost: {format_value(rec1.cost):>10} | Exch: {rec1.exch}")
                print(f"      Row {rec2.row_num} (newer):  {rec2.date} | {rec2.action:4} | Qty: {format_value(rec2.qty):>8} | Price: {format_value(rec2.price):>10} | Cost: {format_value(rec2.cost):>10} | Exch: {rec2.exch}")

                # Highlight differences
                differences = []
                if rec1.exch != rec2.exch:
                    differences.append(f"Exchange: '{rec1.exch}' vs '{rec2.exch}'")
                if rec1.cost != rec2.cost:
                    differences.append(f"Cost: {format_value(rec1.cost)} vs {format_value(rec2.cost)}")

                if differences:
                    print(f"      Notable differences: {', '.join(differences)}")

    # Summary
    print(f"\n\n{'='*100}")
    print(f"SUMMARY")
    print(f"{'='*100}")
    print(f"Total files analyzed: {total_files}")
    print(f"Files with duplicates: {files_with_duplicates}")
    print(f"Total duplicate pairs found: {total_duplicates}")
    print(f"{'='*100}\n")


if __name__ == '__main__':
    main()
