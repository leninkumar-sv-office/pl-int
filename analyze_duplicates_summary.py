#!/usr/bin/env python3
"""
Stock Portfolio Duplicate Analysis - Summary Report
"""

import os
import glob
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
from datetime import datetime

COLUMNS = {
    'A': 'DATE',
    'B': 'EXCH',
    'C': 'ACTION',
    'D': 'QTY',
    'E': 'PRICE',
    'F': 'COST',
    'G': 'REMARKS',
    'H': 'STT',
    'I': 'ADD CHRG'
}


class TradeRecord:
    def __init__(self, row_num: int, data: dict):
        self.row_num = row_num
        self.date = data.get('DATE')
        self.exch = data.get('EXCH')
        self.action = data.get('ACTION')
        self.qty = data.get('QTY')
        self.price = data.get('PRICE')
        self.cost = data.get('COST')

    def __repr__(self):
        return f"Row {self.row_num}: {self.date} {self.action} {self.qty} @ {self.price}"


def safe_value(cell_value):
    if cell_value is None:
        return None
    if isinstance(cell_value, str):
        val = cell_value.strip()
        return val if val else None
    if isinstance(cell_value, (int, float)):
        return cell_value
    if isinstance(cell_value, datetime):
        return cell_value
    return cell_value


def parse_numeric(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip())
        except (ValueError, AttributeError):
            return None
    return None


def is_buy_or_sell(action):
    if action is None:
        return False
    action_upper = str(action).strip().upper()
    return action_upper in ['BUY', 'SELL']


def prices_similar(price1, price2, tolerance_percent=5):
    p1 = parse_numeric(price1)
    p2 = parse_numeric(price2)
    if p1 is None or p2 is None:
        return False
    if p1 == 0 or p2 == 0:
        return p1 == p2
    avg = (abs(p1) + abs(p2)) / 2
    diff_percent = (abs(p1 - p2) / avg) * 100
    return diff_percent <= tolerance_percent


def fuzzy_key(record: TradeRecord) -> Tuple:
    date_str = str(record.date) if record.date else 'NONE'
    action_upper = str(record.action).strip().upper() if record.action else 'NONE'
    qty = record.qty
    return (date_str, action_upper, qty)


def find_duplicates(records: List[TradeRecord]) -> List[Tuple[TradeRecord, TradeRecord]]:
    duplicates = []
    for i, rec1 in enumerate(records):
        for rec2 in records[i+1:]:
            if fuzzy_key(rec1) == fuzzy_key(rec2):
                if prices_similar(rec1.price, rec2.price, tolerance_percent=5):
                    duplicates.append((rec1, rec2))
    return duplicates


def analyze_file(filepath: str) -> Dict:
    result = {
        'filepath': filepath,
        'filename': os.path.basename(filepath),
        'total_rows': 0,
        'duplicates': [],
        'error': None
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        if 'Trading History' not in wb.sheetnames:
            result['error'] = "No 'Trading History' sheet found"
            return result

        ws = wb['Trading History']
        records = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_letter, col_name in COLUMNS.items():
                cell = ws[f'{col_letter}{row_idx}']
                row_data[col_name] = safe_value(cell.value)

            if is_buy_or_sell(row_data.get('ACTION')):
                record = TradeRecord(row_idx, row_data)
                records.append(record)
                result['total_rows'] += 1

        duplicates = find_duplicates(records)
        result['duplicates'] = duplicates
        wb.close()

    except Exception as e:
        result['error'] = str(e)

    return result


def main():
    stocks_dir = '/sessions/vibrant-nice-lamport/mnt/pl/dumps/Stocks'

    if not os.path.isdir(stocks_dir):
        print(f"Error: Directory not found: {stocks_dir}")
        return

    xlsx_files = []
    for filepath in glob.glob(os.path.join(stocks_dir, '*.xlsx')):
        filename = os.path.basename(filepath)
        if not filename.startswith('~') and not filename.startswith('.') and '_backup' not in filepath:
            xlsx_files.append(filepath)

    xlsx_files.sort()

    if not xlsx_files:
        print(f"No xlsx files found in {stocks_dir}")
        return

    print(f"\n{'='*120}")
    print(f"STOCK PORTFOLIO DUPLICATE ANALYSIS - SUMMARY REPORT")
    print(f"{'='*120}\n")

    results_by_file = {}
    total_files = len(xlsx_files)
    files_with_duplicates = 0
    total_duplicate_pairs = 0

    # Analyze all files
    for idx, filepath in enumerate(xlsx_files, 1):
        filename = os.path.basename(filepath)
        result = analyze_file(filepath)
        results_by_file[filename] = result

        if result['error']:
            continue

        if result['duplicates']:
            files_with_duplicates += 1
            total_duplicate_pairs += len(result['duplicates'])

    # Print summary table
    print(f"{'File':<50} {'Total Rows':<12} {'Duplicates':<12}")
    print("-" * 120)

    for filename in sorted(results_by_file.keys()):
        result = results_by_file[filename]
        if result['error']:
            print(f"{filename:<50} ERROR: {result['error']}")
        else:
            dup_count = len(result['duplicates'])
            if dup_count > 0:
                print(f"{filename:<50} {result['total_rows']:<12} {dup_count:<12} <- HAS DUPLICATES")
            else:
                print(f"{filename:<50} {result['total_rows']:<12} {dup_count:<12}")

    print("\n" + "=" * 120)
    print(f"STATISTICS:")
    print(f"  Total files analyzed: {total_files}")
    print(f"  Files with duplicates: {files_with_duplicates}")
    print(f"  Total duplicate pairs found: {total_duplicate_pairs}")
    print("=" * 120 + "\n")

    # Detailed duplicate report
    if files_with_duplicates > 0:
        print(f"\nDETAILED DUPLICATE REPORT (Top 10 files with most duplicates):\n")
        print("=" * 120)

        # Sort by number of duplicates
        sorted_results = sorted(
            [(fname, result) for fname, result in results_by_file.items() if len(result['duplicates']) > 0],
            key=lambda x: len(x[1]['duplicates']),
            reverse=True
        )

        for file_idx, (filename, result) in enumerate(sorted_results[:10], 1):
            print(f"\n[{file_idx}] {filename}")
            print("-" * 120)
            print(f"    Total rows: {result['total_rows']}, Duplicate pairs: {len(result['duplicates'])}\n")

            for dup_idx, (rec1, rec2) in enumerate(result['duplicates'][:5], 1):  # Show first 5 pairs
                print(f"    Pair {dup_idx}:")
                print(f"      Row {rec1.row_num}: {rec1.date} | {rec1.action} {rec1.qty} @ {rec1.price} | Cost: {rec1.cost} | Exch: {rec1.exch}")
                print(f"      Row {rec2.row_num}: {rec2.date} | {rec2.action} {rec2.qty} @ {rec2.price} | Cost: {rec2.cost} | Exch: {rec2.exch}")
                print()

            if len(result['duplicates']) > 5:
                print(f"    ... and {len(result['duplicates']) - 5} more duplicate pairs\n")

        print("=" * 120)


if __name__ == '__main__':
    main()
